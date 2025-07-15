from pathlib import Path
import pytest
import networkx as nx
import networkx.algorithms.isomorphism as iso
import pandas as pd
import numpy as np
import xgboost as xgb
import seaborn as sns
import sys, os
import shutil
import torch
import torch.nn as nn
import drawsvg as draw
import warnings
import importlib
import matplotlib
matplotlib.use('Agg')  # Set non-interactive backend before importing pyplot
import matplotlib.pyplot as plt
from matplotlib.collections import PathCollection, LineCollection
from matplotlib.patches import FancyArrowPatch
from unittest.mock import Mock, patch, MagicMock, mock_open
from torch_geometric.data import Data
from collections import Counter
from contextlib import contextmanager
from glycowork.glycan_data.data_entry import check_presence
from glycowork.motif.query import get_insight
from glycowork.motif.tokenization import (
    constrain_prot, prot_to_coded, string_to_labels, pad_sequence, mz_to_composition,
    get_core, get_modification, get_stem_lib, stemify_glycan, mask_rare_glycoletters,
    glycan_to_composition, calculate_adduct_mass, structure_to_basic, map_to_basic,
    compositions_to_structures, match_composition_relaxed, stemify_dataset, composition_to_mass,
    condense_composition_matching, get_unique_topologies, mz_to_structures, glycan_to_mass,
    get_random_glycan
)
from glycowork.motif.processing import (
    min_process_glycans, get_lib, expand_lib, get_possible_linkages,
    get_possible_monosaccharides, de_wildcard_glycoletter, canonicalize_iupac,
    glycoct_to_iupac, wurcs_to_iupac, oxford_to_iupac, glytoucan_to_glycan, canonicalize_composition, parse_glycoform,
    presence_to_matrix, process_for_glycoshift, linearcode_to_iupac, iupac_extended_to_condensed,
    in_lib, get_class, enforce_class, equal_repeats, get_matching_indices,
    bracket_removal, check_nomenclature, IUPAC_to_SMILES, get_mono, iupac_to_smiles
)
from glycowork.glycan_data.loader import (
    unwrap, find_nth, find_nth_reverse, remove_unmatched_brackets, lib, HashableDict, df_species,
    reindex, stringify_dict, replace_every_second, multireplace, count_nested_brackets,
    strip_suffixes, build_custom_df, DataFrameSerializer, Hex, linkages, glycan_binding, glycomics_data_loader
)
from glycowork.glycan_data.stats import (
    cohen_d, mahalanobis_distance, variance_stabilization, shannon_diversity_index,
    simpson_diversity_index, get_equivalence_test, sequence_richness,
    hotellings_t2, calculate_permanova_stat, omega_squared, anosim, alpha_biodiversity_stats, permanova_with_permutation,
    clr_transformation, alr_transformation, get_procrustes_scores,
    get_additive_logratio_transformation, get_BF, get_alphaN, mahalanobis_variance,
    pi0_tst, TST_grouped_benjamini_hochberg, compare_inter_vs_intra_group,
    correct_multiple_testing, partial_corr, estimate_technical_variance, MissForest, impute_and_normalize,
    variance_based_filtering, get_glycoform_diff, get_glm, process_glm_results, replace_outliers_with_IQR_bounds,
    replace_outliers_winsorization, perform_tests_monte_carlo
)
from glycowork.motif.graph import (
    glycan_to_graph, glycan_to_nxGraph, evaluate_adjacency,
    compare_glycans, subgraph_isomorphism, generate_graph_features,
    graph_to_string, largest_subgraph, get_possible_topologies,
    deduplicate_glycans, neighbor_is_branchpoint, try_string_conversion,
    subgraph_isomorphism_with_negation, categorical_node_match_wildcard,
    expand_termini_list, ensure_graph, possible_topology_check
)
from glycowork.motif.annotate import (
    annotate_glycan, annotate_dataset, get_molecular_properties,
    get_k_saccharides, get_terminal_structures, create_correlation_network,
    group_glycans_core, group_glycans_sia_fuc, group_glycans_N_glycan_type,
    Lectin, load_lectin_lib, create_lectin_and_motif_mappings,
    lectin_motif_scoring, deduplicate_motifs, quantify_motifs, get_size_branching_features,
    count_unique_subgraphs_of_size_k, annotate_glycan_topology_uncertainty
)
from glycowork.motif.regex import (preprocess_pattern, specify_linkages, process_occurrence,
                  convert_pattern_component, reformat_glycan_string,
                  motif_to_regex, get_match, process_question_mark, calculate_len_matches_comb,
                  process_main_branch, check_negative_look, get_match_batch,
                  filter_matches_by_location, parse_pattern, compile_pattern
)
from glycowork.motif.draw import (process_bonds, draw_hex, scale_in_range, process_per_residue, col_dict_base,
                 get_hit_atoms_and_bonds, add_colours_to_map, is_jupyter, process_repeat, draw_bracket,
                 display_svg_with_matplotlib, get_coordinates_and_labels, get_highlight_attribute, add_sugar, add_bond, draw_shape,
                 draw_chem2d, draw_chem3d, GlycoDraw, plot_glycans_excel, annotate_figure
)
from glycowork.motif.analysis import (preprocess_data, get_pvals_motifs, select_grouping, get_glycanova, get_differential_expression,
                     get_biodiversity, get_time_series, get_SparCC, get_roc, get_ma, get_volcano, get_meta_analysis,
                     get_representative_substructures, get_lectin_array, get_coverage, plot_embeddings, get_pval_distribution,
                     characterize_monosaccharide, get_heatmap, get_pca, get_jtk, multi_feature_scoring, get_glycoshift_per_site
)
from glycowork.network.biosynthesis import (safe_compare, safe_index, get_neighbors, create_neighbors,
                         find_diff, find_path, construct_network, prune_network, estimate_weights, network_alignment, export_network,
                         extend_glycans, highlight_network, infer_roots, deorphanize_nodes, get_edge_weight_by_abundance,
                         find_diamonds, trace_diamonds, get_maximum_flow, get_reaction_flow, process_ptm, get_differential_biosynthesis,
                         deorphanize_edge_labels, infer_virtual_nodes, retrieve_inferred_nodes, monolink_to_glycoenzyme, infer_network,
                         get_max_flow_path, edges_for_extension, choose_leaves_to_extend, evoprune_network, extend_network,
                         plot_network, add_high_man_removal, net_dic, find_shared_virtuals, create_adjacency_matrix, find_ptm
)
from glycowork.network.evolution import (calculate_distance_matrix, distance_from_embeddings,
                      jaccard, distance_from_metric, check_conservation, get_communities, dendrogram_from_distance
)
from glycowork.ml.train_test_split import (seed_wildcard_hierarchy, hierarchy_filter,
                            general_split, prepare_multilabel
)
from glycowork.ml.processing import (augment_glycan, AugmentedGlycanDataset,
                       dataset_to_graphs, dataset_to_dataloader, split_data_to_train
)
from glycowork.ml.model_training import (EarlyStopping, sigmoid, disable_running_stats,
                          enable_running_stats, train_model, SAM, Poly1CrossEntropyLoss, training_setup,
                          train_ml_model, analyze_ml_model, get_mismatch
)
from glycowork.ml.models import (SweetNet, NSequonPred, sigmoid_range, SigmoidRange, LectinOracle,
                          init_weights, prep_model, LectinOracle_flex
)
from glycowork.ml.inference import (SimpleDataset, sigmoid, glycans_to_emb, get_multi_pred, get_lectin_preds,
                          get_esmc_representations, get_Nsequon_preds
)
device = "cpu"
if torch.cuda.is_available():
    device = "cuda:0"

@contextmanager
def suppress_pydot_warnings():
    """Context manager to suppress specific pydot-related deprecation warnings"""
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="nx.nx_pydot.pydot_layout depends on the pydot package",
            category=DeprecationWarning
        )
        warnings.filterwarnings(
            "ignore",
            message="nx.nx_pydot.to_pydot depends on the pydot package",
            category=DeprecationWarning
        )
        yield


GLYCAN_TEST_CASES = [
    # linear glycans
    "Rha(a1-2)Glc(b1-2)GlcAOBut",
    "ManNAc1PP(a1-5)Ribf1N",
    "Glc3Me(b1-3)Xyl(b1-4)Qui",
    "D-Fuc2Me4NAc(a1-4)GlcA",
    "Man(b1-3)Man(b1-3)Ery-onic",
    "Man(b1-3)Glc(a1-4)Rha(a1-4)Man(b1-3)Glc(a1-4)D-Rha",
    "Rha(?1-?)Glc(?1-?)Glc(?1-?)Gal",
    "GlcA(b1-3)Gal(b1-3)GalNAc(b1-4)GlcNAc6P(b1-3)Man(b1-4)Glc",
    "Neu5Gc(a2-6)Glc(?1-?)Neu5Gc(a2-?)Glc",
    "Neu5Ac(b2-6)GalNAc",

    # single branching
    "Gal(a1-3)Gal(b1-4)GlcNAc(b1-3)[Gal(a1-3)Gal(b1-4)GlcNAc(b1-6)]Gal(b1-?)GlcNAc(b1-3)Gal(b1-4)Glc",
    "Glc(b1-2)[Glc(b1-3)]Gal",
    "Fuc(a1-3)[Fuc(b1-6)]GlcNAc1N",
    "Neu5Ac(a2-?)GalNAc(b1-4)GlcNAc(b1-2)Man(a1-?)[Man(a1-?)Man(a1-?)]Man(b1-4)GlcNAc(b1-4)GlcNAc",
    "GalNAc(a1-3)[Fuc(a1-2)]Gal(b1-4)Glc-ol",
    "QuiNAcOPGro(b1-3)[GlcNAcOAc(b1-2)]Gal(b1-3)GlcNAcOAc(b1-3)QuiNAcOPGro",
    "Neu5Ac(a2-?)GalNAc(b1-4)GlcNAc(b1-2)Man(a1-?)[Neu5Gc(a2-?)Gal(b1-4)GlcNAc(b1-2)Man(a1-?)]Man(b1-4)GlcNAc(b1-4)GlcNAc",
    "Glc(a1-2)[Glc(b1-6)]Glc(a1-3)Glc(b1-6)Glc",
    "Glc(a1-6)Glc(a1-6)[Glc(a1-2)]Glc",
    "Neu5Ac(a2-6)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)[Man(a1-3)]Man(b1-4)GlcNAc(b1-4)GlcNAc",

    # multiple branching
    "GalNAc(?1-?)[Neu5Ac(a2-?)]Gal(b1-?)GlcNAc(b1-?)Man(a1-?)[Neu5Ac(a2-?)Gal(b1-?)GlcNAc(b1-?)Man(a1-?)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-?)]GlcNAc",
    "Gal(?1-?)[Fuc(?1-?)]Gal(b1-?)GlcNAc(b1-?)Man(a1-?)[GlcNAc(b1-?)Man(a1-?)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-?)]GlcNAc",
    "Glc(b1-3)[Glc(b1-3)[Glc(b1-6)Glc(b1-6)]Glc(b1-6)]Glc(b1-6)Glc",
    "Neu5Gc(a2-?)Gal(b1-4)GlcNAc(b1-2)[Neu5Gc(a2-?)Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Neu5Ac(a2-?)Gal(b1-4)GlcNAc(b1-2)[Neu5Ac(a2-?)Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc",
    "Man(a1-3)[Gal3Me(b1-6)Gal3Me(b1-3)[Gal3Me(b1-6)]GlcNAc(b1-4)GlcNAc(b1-2)Man(a1-6)][Xyl(b1-2)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
    "Neu5Ac(a2-?)Hex(?1-?)HexNAc(?1-?)Hex(?1-?)[Neu5Ac(a2-?)Hex(?1-?)HexNAc(?1-?)Hex(?1-?)]Hex(?1-?)HexNAc(?1-?)[Fuc(a1-?)]HexNAc",
    "Neu5Ac(a2-?)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-3)Gal(b1-4)GlcNAc(b1-2)[Neu5Ac(a2-?)Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Man(a1-3)[Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc",
    "Fuc(a1-2)[GalNAc(a1-3)]Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-3)Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc",
    "Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)[Gal(b1-4)GlcNAc(b1-2)Man(a1-3)]Man",
    "GlcA(b1-2)Man(a1-4)[Araf(a1-3)]GlcA(b1-2)[Araf(a1-3)]Man(a1-4)GlcA",

    # double branching
    "Man(a1-3)[Man(a1-6)][Xylf(?1-?)]Man(b1-4)GlcNAc(b1-4)GlcNAc",
    "GlcNAc(b1-2)Man(a1-3)[GlcNAc(b1-4)][GlcNAc(b1-2)Man(a1-6)]Man",
    "HexOMe(?1-?)[HexOMe(?1-?)]HexOMe(?1-?)[HexOMe(?1-?)]GalNAc(b1-4)GlcNAc(b1-2)Man(a1-3)[Xyl(b1-2)][Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
    "Neu5Ac(a2-8)Neu5Ac(a2-?)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[GlcNAc(b1-4)][GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
    "Neu5Ac(a2-?)Gal(b1-?)GlcNAc(b1-2)[Neu5Ac(a2-?)Gal(b1-?)GlcNAc(b1-4)]Man(a1-3)[Neu5Ac(a2-?)Gal(b1-?)GlcNAc(b1-2)Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)GlcNAc",
    "Gal3Me(b1-6)Gal3Me(b1-3)[Gal3Me(b1-6)]GalNAc(b1-4)GlcNAc(b1-2)Man(a1-6)[Xyl(b1-2)][Man(a1-3)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
    "Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)[GlcNAc(b1-4)]Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-6)]Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
    "Neu5Gc(a2-?)Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[GlcNAc(b1-4)][Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
    "Neu5Ac(a2-?)Neu5Ac(a2-?)Gal(b1-?)GlcNAc(b1-?)Gal(b1-?)GlcNAc(b1-?)[Neu5Gc(a2-?)Gal(b1-?)[Fuc(a1-?)]GlcNAc(b1-?)Gal(b1-?)GlcNAc(b1-?)]Man(a1-?)[Man(a1-?)[Man(a1-?)]Man(a1-?)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)GlcNAc",
    "Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[GlcNAc(b1-4)][Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc",

    # nested branching
    "Glc(a1-3)Man(a1-2)Man(a1-2)Man(a1-3)[Man(a1-3)[Man(a1-2)Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc-ol",
    "Man(a1-?)Man(a1-3)[Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc-ol",
    "Gal(?1-?)Gal(b1-?)GlcNAc(?1-?)[Fuc(a1-?)[Gal(b1-?)]GlcNAc(b1-?)]Man(a1-?)[Fuc(a1-?)[Gal(b1-?)]GlcNAc(b1-?)[Fuc(a1-?)[Gal(b1-?)]GlcNAc(b1-?)]Man(a1-?)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-?)]GlcNAc",
    "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)[Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)]Man(b1-4)GlcNAc",
    "Man(a1-2)Man6P(a1-2)Man(a1-3)[Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc",
    "Neu5Ac(a2-6)Gal(b1-4)GlcNAc(b1-2)[Gal(b1-3)[Neu5Ac(a2-6)]GlcNAc(b1-4)]Man(a1-3)[Neu5Ac(a2-6)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)]Man",
    "Gal(b1-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Gal(b1-3)Gal(b1-4)GlcNAc(b1-2)[Gal(b1-3)Gal(b1-4)GlcNAc(b1-4)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
    "Neu5Ac(a2-?)Gal(b1-3)[Neu5Ac(a2-?)Gal(b1-?)[Fuc(a1-?)]GlcNAc(b1-6)]GalNAc",
    "Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Gal(b1-4)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
    "Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Gal(b1-4)GlcNAc(b1-3)[Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc",

    # floating elements
    "{Fuc(a1-?)}Neu5Ac(a2-?)Gal(b1-4)GlcNAc(b1-2)Man(a1-?)[Gal(b1-4)GlcNAc(b1-2)Man(a1-?)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
    "{Fuc(a1-?)}Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
    "{Gal(b1-4)GlcNAc(b1-?)}{Neu5Ac(a2-?)}{Neu5Ac(a2-?)}Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-?)]Man(a1-?)[Gal(b1-4)GlcNAc(b1-2)Man(a1-?)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
    "{Neu5Ac(a2-?)}Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
    "{Fuc(a1-2)}Gal(b1-4)[Neu5Ac(a2-6)]GlcNAc(b1-?)[Gal(b1-4)GlcNAc(b1-?)]Gal(b1-4)Glc-ol",
    "{Neu5Ac(a2-?)}{Neu5Ac(a2-?)}Gal(b1-?)GlcNAc(b1-2)[Gal(b1-?)GlcNAc(b1-4)]Man(a1-3)[Gal(b1-?)GlcNAc(b1-2)[Gal(b1-?)GlcNAc(b1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc",
    "{Fuc(a1-?)}Neu5Ac(a2-?)Gal(b1-?)GlcNAc(b1-3)Gal(b1-4)GlcNAc(b1-?)[Neu5Ac(a2-?)Gal(b1-4)GlcNAc(b1-?)]Man(a1-?)[Neu5Ac(a2-?)Gal(b1-4)GlcNAc(b1-2)[Neu5Ac(a2-?)Gal(b1-4)GlcNAc(b1-?)]Man(a1-?)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
    "{Fuc(a1-?)}{GlcNAc(b1-?)}{Hex(?1-?)}{Neu5Gc(a2-?)}GlcNAc(b1-2)Man(a1-3)[Man(a1-3)[Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
    "{Neu5Ac(a2-?)}Gal(b1-?)GlcNAc(b1-2)[Gal(b1-?)GlcNAc(b1-4)]Man(a1-3)[Gal(b1-?)GlcNAc(b1-2)Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)GlcNAc",
    "{Fuc(a1-?)}{Neu5Ac(a2-?)}Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc",
]


def test_torch_import_error():
  # Store original module states
  original_processing = sys.modules.get('glycowork.ml.processing')
  original_models = sys.modules.get('glycowork.ml.models')
  original_inference = sys.modules.get('glycowork.ml.inference')
  original_model_training = sys.modules.get('glycowork.ml.model_training')
  try:
    # Remove modules if they exist
    if 'glycowork.ml.processing' in sys.modules:
      del sys.modules['glycowork.ml.processing']
    if 'glycowork.ml.models' in sys.modules:
      del sys.modules['glycowork.ml.models']
    if 'glycowork.ml.inference' in sys.modules:
      del sys.modules['glycowork.ml.inference']
    if 'glycowork.ml.model_training' in sys.modules:
      del sys.modules['glycowork.ml.model_training']
    # Patch the imports to fail
    with patch.dict('sys.modules', {'torch': None, 'torch_geometric': None, 'torch_geometric.loader': None, 'torch_geometric.utils': None, 'torch_geometric.utils.convert': None}):
      with pytest.raises(ImportError, match="torch or torch_geometric missing"):
        importlib.import_module('glycowork.ml.processing')
      with pytest.raises(ImportError, match="torch or torch_geometric missing"):
        importlib.import_module('glycowork.ml.models')
      with pytest.raises(ImportError, match="torch missing;"):
        importlib.import_module('glycowork.ml.inference')
      with pytest.raises(ImportError, match="torch missing;"):
        importlib.import_module('glycowork.ml.model_training')
  finally:
    # Restore original states
    if original_processing is not None:
      sys.modules['glycowork.ml.processing'] = original_processing
    elif 'glycowork.ml.processing' in sys.modules:
      del sys.modules['glycowork.ml.processing']
    if original_models is not None:
      sys.modules['glycowork.ml.models'] = original_models
    elif 'glycowork.ml.models' in sys.modules:
      del sys.modules['glycowork.ml.models']
    if original_inference is not None:
      sys.modules['glycowork.ml.inference'] = original_inference
    elif 'glycowork.ml.inference' in sys.modules:
      del sys.modules['glycowork.ml.inference']
    if original_model_training is not None:
      sys.modules['glycowork.ml.model_training'] = original_model_training
    elif 'glycowork.ml.model_training' in sys.modules:
      del sys.modules['glycowork.ml.model_training']


@pytest.mark.parametrize("glycan", GLYCAN_TEST_CASES)
def test_graph_to_string_int(glycan: str):
    """This test assumes that the function gylcan_to_graph_int is correct."""
    label = glycan_to_nxGraph(glycan)
    target = glycan_to_nxGraph(graph_to_string(label))
    assert nx.is_isomorphic(label, target, node_match=iso.categorical_node_match("string_labels", ""))


def test_get_mono():
    try:
        get_mono("wrong_wrong")
        return False
    except:
        pass


def test_module_imports():
    from glycowork.motif import graph, annotate, draw, analysis


def test_graph_to_string():
    assert graph_to_string("Gal(b1-3)GlcNAc") == "Gal(b1-3)GlcNAc"
    # Complex cases
    assert graph_to_string(glycan_to_nxGraph("Gal(a1-3)[Fuc(a1-2)]Gal(b1-4)GlcNAc(b1-6)[Gal(a1-3)[Fuc(a1-2)]Gal(b1-4)GlcNAc(b1-3)]GalNAc")) == "Fuc(a1-2)[Gal(a1-3)]Gal(b1-4)GlcNAc(b1-3)[Fuc(a1-2)[Gal(a1-3)]Gal(b1-4)GlcNAc(b1-6)]GalNAc"
    result = graph_to_string(glycan_to_nxGraph("Gal(a1-3)[Fuc(a1-2)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-6)]Gal(b1-4)GlcNAc(b1-6)[Gal(a1-3)[Fuc(a1-2)]Gal(b1-4)GlcNAc(b1-3)]GalNAc"))
    assert result == "Fuc(a1-2)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-6)[Gal(a1-3)]Gal(b1-4)GlcNAc(b1-6)[Fuc(a1-2)[Gal(a1-3)]Gal(b1-4)GlcNAc(b1-3)]GalNAc"
    assert graph_to_string(glycan_to_nxGraph("Xyl(b1-2)[Man(a1-3)][Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)GlcNAc")) == "Xyl(b1-2)[Man(a1-3)][GlcNAc(b1-4)][Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert graph_to_string(glycan_to_nxGraph('Man(a1-3)[Xyl(b1-2)][Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)][Fuc(a1-6)]GlcNAc')) == "Xyl(b1-2)[Man(a1-3)][Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)][Fuc(a1-6)]GlcNAc"
    # Mode for GlycoDraw
    result = graph_to_string(glycan_to_nxGraph("Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"), order_by="linkage")
    assert result == "Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    result = graph_to_string(glycan_to_nxGraph("Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[Neu5Gc(a2-6)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"), order_by="linkage")
    assert result == "Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[Neu5Gc(a2-6)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    result = graph_to_string(glycan_to_nxGraph("Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)[GlcNAc(b1-4)]Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-6)]Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"), order_by="linkage")
    assert result == "Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)[GlcNAc(b1-4)]Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-6)]Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert graph_to_string(glycan_to_nxGraph("Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-6)[Gal(b1-3)]GalNAc"), order_by="linkage") == "Gal(b1-3)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-6)]GalNAc"
    assert graph_to_string(glycan_to_nxGraph("Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)[Man(a1-3)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"), order_by="linkage") == "Man(a1-3)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert graph_to_string(glycan_to_nxGraph("Neu5Ac(a2-?)Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Fuc(a1-2)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"), order_by="linkage") == "Neu5Ac(a2-?)Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Fuc(a1-2)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert graph_to_string(glycan_to_nxGraph("Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"),
                      order_by = "linkage") == "Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert graph_to_string(glycan_to_nxGraph("Fuc(a1-?)[Gal(b1-?)]GlcNAc(b1-2)Man(a1-6)[Man(a1-2)Man(a1-3)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"),
                      order_by = "linkage") == "Man(a1-2)Man(a1-3)[Fuc(a1-?)[Gal(b1-?)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert graph_to_string(glycan_to_nxGraph("Gal(b1-4)GlcNAc(b1-6)[GlcNAc(b1-3)]Gal(b1-4)Glc"), order_by="linkage") == "GlcNAc(b1-3)[Gal(b1-4)GlcNAc(b1-6)]Gal(b1-4)Glc"


def test_canonicalize_iupac():
    # Test basic cleanup
    assert canonicalize_iupac("Galb4GlcNAc") == "Gal(b1-4)GlcNAc"
    assert canonicalize_iupac("'Galb4GlcNAc'") == "Gal(b1-4)GlcNAc"
    assert canonicalize_iupac("{Gal(b1-4)}{Neu5Ac(a2-3)}Gal(b1-4)GlcNAc") == "{Neu5Ac(a2-3)}{Gal(b1-4)}Gal(b1-4)GlcNAc"
    assert canonicalize_iupac("{Gal(b1-4)}{Neu5Ac(a2-3)}Gal(b1-4)[Fuc(a1-3)]GlcNAc") == "{Neu5Ac(a2-3)}{Gal(b1-4)}Fuc(a1-3)[Gal(b1-4)]GlcNAc"
    assert canonicalize_iupac("Fucα2Galβ1-4GlcNAcβ1-3(NeuAcα2-3Galβ1-4GlcNAcβ1-6)Galβ1-4GlcNAcol") == "Fuc(a1-2)Gal(b1-4)GlcNAc(b1-3)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-6)]Gal(b1-4)GlcNAc"
    assert canonicalize_iupac("Fucα1-2Galβ1-4GlcNAcβ1-3(Fucα1-2Galβ1-4GlcNAcβ1-6)Galβ1-4GlcNAcβ1-3Galβ1-4Glcβ-") == "Fuc(a1-2)Gal(b1-4)GlcNAc(b1-3)[Fuc(a1-2)Gal(b1-4)GlcNAc(b1-6)]Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc"
    assert canonicalize_iupac("Neu5Aca2-3Galb1-3{Neu5Aca2-6}GalNAc") == "Neu5Ac(a2-3)Gal(b1-3)[Neu5Ac(a2-6)]GalNAc"
    assert canonicalize_iupac("Gal(a1-2)[Man(a1-3)D-Rha(a1-3)][Rha2Me3Me(a1-2)D-Ara(b1-3)Rha(b1-4)Xyl(b1-4)]Fuc(a1-3)[Xyl(b1-4)]Glc") == "Rha2Me3Me(a1-2)D-Ara(b1-3)Rha(b1-4)Xyl(b1-4)[Man(a1-3)D-Rha(a1-3)][Gal(a1-2)]Fuc(a1-3)[Xyl(b1-4)]Glc"
    assert canonicalize_iupac("bGal14GlcNAc") == "Gal(b1-4)GlcNAc"
    assert canonicalize_iupac("aMan13(aMan16)Man") == "Man(a1-3)[Man(a1-6)]Man"
    assert canonicalize_iupac("bGal13bGalNAc14(aNeuAc23)bGal14Glc") == "Gal(b1-3)GalNAc(b1-4)[Neu5Ac(a2-3)]Gal(b1-4)Glc"
    assert canonicalize_iupac("Mana12Man") == "Man(a1-2)Man"
    assert canonicalize_iupac("Rib5P-ol(5-4)Glc6PEtN(b1-3)Gal(b1-3)GalNAc(b1-4)Rib5P-ol") == "Rib5P-ol(5-4)Glc6PEtN(b1-3)Gal(b1-3)GalNAc(b1-4)Rib5P-ol"
    assert canonicalize_iupac("Neu5Ac-α-2,6-Gal-β-1,3-GlcNAc-β-Sp") == "Neu5Ac(a2-6)Gal(b1-3)GlcNAc"
    assert canonicalize_iupac("GlcNAc-α-1,3-(Glc-α-1,2-Glc-α-1,2)-Gal-α-1,3-Glc-α-Sp") == "Glc(a1-2)Glc(a1-2)[GlcNAc(a1-3)]Gal(a1-3)Glc"
    assert canonicalize_iupac('GlcNAcβ(1-2)Manα(1-3)[Neu5Acα(2-6)Galβ(1-4)GlcNAcβ(1-2)Manα(1-6)]Manβ(1-4)GlcNAcβ(1-4)[Fucα(1-6)]GlcNAc' ) ==  'Neu5Ac(a2-6)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)[GlcNAc(b1-2)Man(a1-3)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc'
    assert canonicalize_iupac("Sorf(a2-1b)[L-Glc(b1-6)]L-Tal") == "Sorf(a2-1)[L-Glc(b1-6)]L-Tal"
    assert canonicalize_iupac("Psif(a2-1b)Glc") == "Psif(a2-1)Glc"
    assert canonicalize_iupac("Glc(a1-4)2,3-Anhydro-Man(a1-4)Glc(a1-4)Glc") == "Glc(a1-4)2,3-Anhydro-Man(a1-4)Glc(a1-4)Glc"
    # Test linkage uncertainty
    assert canonicalize_iupac("Gal-GlcNAc") == "Gal(?1-?)GlcNAc"
    assert canonicalize_iupac("Gal(b1-3/4)Gal(b1-4)GlcNAc") == "Gal(b1-3/4)Gal(b1-4)GlcNAc"
    assert canonicalize_iupac("Gal+Gal(b1-4)GlcNAc") == "{Gal(?1-?)}Gal(b1-4)GlcNAc"
    assert canonicalize_iupac("NeuAcα2-6Galβ1-4(6S)GlcNAcβ1-2Manα1-3(Manα-Manα1-6)Manβ1-4GlcNAcβ1-4(Fucα1-6)GlcNAcol") == "Neu5Ac(a2-6)Gal(b1-4)GlcNAc6S(b1-2)Man(a1-3)[Man(a1-?)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("Neu5Ac{Galb1-3(Neu5Aca2-6)GalNAc") == "{Neu5Ac(a2-?)}Gal(b1-3)[Neu5Ac(a2-6)]GalNAc"
    # Test modification handling
    assert canonicalize_iupac("Neu5,9Ac2a2-6Galb1-4GlcNAcb-Sp8") == "Neu5Ac9Ac(a2-6)Gal(b1-4)GlcNAc"
    assert canonicalize_iupac("Neu4,5Ac2a2-6Galb1-4GlcNAcb-Sp8") == "Neu4Ac5Ac(a2-6)Gal(b1-4)GlcNAc"
    assert canonicalize_iupac("6SGal(b1-4)GlcNAc") == "Gal6S(b1-4)GlcNAc"
    assert canonicalize_iupac("(6S)Galb1-4GlcNAcb-Sp0") == "Gal6S(b1-4)GlcNAc"
    assert canonicalize_iupac("(6S)(4S)Galb1-4GlcNAcb-Sp0") == "Gal4S6S(b1-4)GlcNAc"
    assert canonicalize_iupac("S+NeuAcα2-6(Galβ1-3)GlcNAcβ1-3Galβ1-4Glcol") == "{OS}Gal(b1-3)[Neu5Ac(a2-6)]GlcNAc(b1-3)Gal(b1-4)Glc-ol"
    assert canonicalize_iupac("Gal(b1-4)Glc-olS") == "Gal(b1-4)GlcOS-ol"
    assert canonicalize_iupac("SGalNAc(b1-4)GlcNAc") == "GalNAcOS(b1-4)GlcNAc"
    assert canonicalize_iupac("S-Gal(b1-4)Glc-ol") == "GalOS(b1-4)Glc-ol"
    assert canonicalize_iupac("SGaNAcb1-4SGlcNac") == "GalNAcOS(b1-4)GlcNAcOS"
    assert canonicalize_iupac("Rha(a1-2)Ara4S") == "Rha(a1-2)Ara4S"
    assert canonicalize_iupac("GalNAc(a1-3)GalNAc(b1-3)[D-Fuc3NAc(a1-4)]Gal(a1-4)Glc") == "GalNAc(a1-3)GalNAc(b1-3)[D-Fuc3NAc(a1-4)]Gal(a1-4)Glc"
    # Test sanitization
    assert canonicalize_iupac("GlcNAc(b1-2)[GlcNAc(b1-2)]Man") == "GlcNAc(b1-?)[GlcNAc(b1-?)]Man"
    assert canonicalize_iupac("Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-2)]Man") == "Gal(b1-4)GlcNAc(b1-?)[Gal(b1-4)GlcNAc(b1-?)]Man"
    assert canonicalize_iupac("Gal(b1-2)GlcNAc") == "Gal(b1-?)GlcNAc"
    assert canonicalize_iupac("GlcNAc(b1-3)Gal3S") == "GlcNAc(b1-?)Gal3S"
    # Test lookup
    assert canonicalize_iupac("LacNAc") == "Gal(b1-4)GlcNAc"
    assert canonicalize_iupac("LEWISX") == "Fuc(a1-3)[Gal(b1-4)]GlcNAc"
    assert canonicalize_iupac("3'-FL") == "Fuc(a1-3)[Gal(b1-4)]Glc-ol"
    # Test branch ordering
    assert canonicalize_iupac("GalNAcβ1-4(NeuAcα2-3)GlcNAcβ1-3(NeuAcα2-3Galβ1-4GlcNAcβ1-6)Galβ1-4Glcol") == "Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-6)[Neu5Ac(a2-3)[GalNAc(b1-4)]GlcNAc(b1-3)]Gal(b1-4)Glc-ol"
    assert canonicalize_iupac("Fucα1-2Galβ1-4GlcNAcβ1-3[NeuAcα2-3Galβ1-4(Fucα1-3)GlcNAcβ1-6]Galβ1-4GlcNAcβ1-3(Fucα1-2Galβ1-4GlcNAcβ1-6)Galβ1-4Glcol") == "Fuc(a1-2)Gal(b1-4)GlcNAc(b1-3)[Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-6)]Gal(b1-4)GlcNAc(b1-3)[Fuc(a1-2)Gal(b1-4)GlcNAc(b1-6)]Gal(b1-4)Glc-ol"
    assert canonicalize_iupac("Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)Man(a1-6)][Xyl(b1-2)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)][Fuc(a1-3)]GlcNAc") == "Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)Man(a1-6)][Xyl(b1-2)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)][Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)[Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc") == "Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("Galβ1–4GlcNAcβ1–2(Fucα1–3(Galβ1–4)GlcNAcβ1–4)Manα1–3(Galβ1–4GlcNAcβ1–2(Galβ1–4GlcNAcβ1–6)Manα1–6)Manβ1–4GlcNAcβ1–4GlcNAc") == "Gal(b1-4)GlcNAc(b1-2)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-4)]Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)[GlcNAc(b1-4)]Man(a1-3)[Man(a1-3)[Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc") == "Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)[GlcNAc(b1-4)]Man(a1-3)[Man(a1-3)[Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Man(a1-3)[Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc") == "Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Man(a1-3)[Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("GalNAc(b1-4)[Neu5Ac(a2-3)]Gal(b1-4)GlcNAc(b1-6)[Fuc(a1-2)[GalNAc(a1-3)]Gal(b1-3)]GalNAc") == "Neu5Ac(a2-3)[GalNAc(b1-4)]Gal(b1-4)GlcNAc(b1-6)[Fuc(a1-2)[GalNAc(a1-3)]Gal(b1-3)]GalNAc"
    assert canonicalize_iupac("Gal(b1-3)[Fuc(a1-4)]GlcNAc(b1-2)Man(a1-3)[Xyl(b1-2)][Gal(b1-3)[Fuc(a1-4)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)]GlcNAc") == "Gal(b1-3)[Fuc(a1-4)]GlcNAc(b1-2)Man(a1-3)[Gal(b1-3)[Fuc(a1-4)]GlcNAc(b1-2)Man(a1-6)][Xyl(b1-2)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)]GlcNAc"
    assert canonicalize_iupac("Glc(?1-?)[Gal(?1-?)]2,5-Anhydro-Tal") == "Gal(?1-?)[Glc(?1-?)]2,5-Anhydro-Tal"
    # Test other nomenclatures
    assert canonicalize_iupac("aDMan(1-2)bDGlcp(1-1)Me") == "Man(a1-2)Glc1Me"
    assert canonicalize_iupac("aDGlcp(1-2)bDFruf") == "Glc(a1-2)Fruf"
    assert canonicalize_iupac("-1)bDFruf(2-3)bDFruf(2-") == "Fruf(b2-3)Fruf(b2-1)Fruf"
    assert canonicalize_iupac("-1)[Ac(1-3)]bDFruf(2-3)bDFruf(2-") == "Fruf3Ac(b2-3)Fruf(b2-1)Fruf3Ac"
    assert canonicalize_iupac("-2)[aDGalp(1-3)]aDRhap(1-3)bDRhap(1-4)bDGlcp(1-") == "[Gal(a1-3)]D-Rha(a1-3)D-Rha(b1-4)Glc(b1-2)D-Rha"
    assert canonicalize_iupac("-2)aLRhap(1-3)[bDGlcp(1-2)]aLRhap(1-") == "Rha(a1-3)[Glc(b1-2)]Rha(a1-2)Rha"
    assert canonicalize_iupac("-3)[65%Ac(1-2)]bDRibf(1-2)bDRibf(1-6)bXKdof(2-") == "Ribf2Ac(b1-2)Ribf(b1-6)Kdof(b2-3)Ribf2Ac"
    assert canonicalize_iupac("-2)aLRhap(1-P-4)[Ac(1-2)]bDManpN(1-4)aDGlcp(1-") == "Rha1P(a1-4)ManNAc(b1-4)Glc(a1-2)Rha1P"
    assert canonicalize_iupac("-2)bDGlcpA(1-3)[%Ac(1-2)aL6dTalpN(1-4)bDGlcp(1-4),%Ac(1-2)]aLFucpN(1-") == "GlcA(b1-3)[L-6dTalNAc(a1-4)Glc(b1-4)]FucNAc(a1-2)GlcA"
    assert canonicalize_iupac("-3)aLRhap(1-2)aLRhap(1-5)[<<Ac(1-8)|Ac(1-7)>>]bXKdo(2-") == "Rha(a1-2)Rha(a1-5)KdoOAc(b2-3)Rha"
    assert canonicalize_iupac("-P-2)bDRibf(1-2)xDRib-ol(5-") == "Ribf(b1-2)Rib5P-ol(?5-2)Ribf"
    assert canonicalize_iupac("-8)[%Ac(1-7),Ac(1-5)]aXNeup(2-") == "Neu5Ac7Ac(a2-8)Neu5Ac7Ac"
    assert canonicalize_iupac("-6)[x?Rib-ol(1-P-4),80%Ac(1-3),Ac(1-2)]aDGlcpN(1-4)[40%Ac(1-2)]aDGalp(1-3)bDGalp(1-4)bDGlcp(1-") == "[Rib1P-ol(?1-4)]GlcNAc3Ac(a1-4)Gal2Ac(a1-3)Gal(b1-4)Glc(b1-6)GlcNAc3Ac"
    assert canonicalize_iupac("-6)bDGalf(1-1)xDMan-ol(6-P-3)[aDGlcp(1-2)]bDGalp(1-3)bDGalf(1-3)bDGlcp(1-") == "Galf(b1-1)Man6P-ol(?6-3)[Glc(a1-2)]Gal(b1-3)Galf(b1-3)Glc(b1-6)Galf"
    assert canonicalize_iupac("-6)aDGlcp(1-6)aDGlcp(1-6)[aDGlcp(1-6)/aDGlcp(1-4)/n=?/aDGlcp(1-4)]aDGlcp(1-6)aDGlcp(1-") == "Glc(a1-6)Glc(a1-6)[Glc(a1-6)Glc(a1-4)Glc(a1-4)]Glc(a1-6)Glc(a1-6)Glc"
    assert canonicalize_iupac("bDGlcp(1-3)/bDGlcpA(1-4)bDGlcp(1-3)/n=1-2/bDGlcpA(1-4)bDGlcp") == "Glc(b1-3)GlcA(b1-4)Glc(b1-3)GlcA(b1-4)Glc"
    assert canonicalize_iupac("DManpa1-6DManpb1-4DGlcpNAcb1-4[LFucpa1-6]DGlcpNAcb1-OH") == "Man(a1-6)Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("Neup5Aca2-3DGalpb1-4DGlcpNAcb1-3DGalpb1-3DGalpb1-4DGlcpb1-OH") == "Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-3)Gal(b1-3)Gal(b1-4)Glc"
    assert canonicalize_iupac("DGalp[6S]b1-3DGalpNAca1-OH") == "Gal6S(b1-3)GalNAc"
    assert canonicalize_iupac("DGalpNAcb1-4[LFucpa1-3]DGlcpNAc[6PC]b1-2DManpa1-3[DManpa1-6]DManpb1-4DGlcpNAcb1-4[LFucpa1-6]DGlcpNAc") == "Fuc(a1-3)[GalNAc(b1-4)]GlcNAc6PCho(b1-2)Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("DKDNpa2-3DGalpb1-3DGlcpNAc") == "Kdn(a2-3)Gal(b1-3)GlcNAc"
    assert canonicalize_iupac("DGalp[3S,6S]b1-4DGlcpNAca1-OH") == "Gal3S6S(b1-4)GlcNAc"
    assert canonicalize_iupac("DGalpa1-2[DManpa1-3DRhapa1-3][LRhap[2Me,3Me]a1-2[DArapb1-3]LRhapb1-4DXylpb1-4]LFucpa1-3[DXylpb1-4]DGlcpa1-OH") == "Rha2Me3Me(a1-2)[D-Ara(b1-3)]Rha(b1-4)Xyl(b1-4)[Man(a1-3)D-Rha(a1-3)][Gal(a1-2)]Fuc(a1-3)[Xyl(b1-4)]Glc"
    assert canonicalize_iupac("DGlcpAb1-4DGlcpNAca1-4DGlcpA[2S]b1-4DGlcpNAc") == "GlcA(b1-4)GlcNAc(a1-4)GlcA2S(b1-4)GlcNAc"
    assert canonicalize_iupac("DManpa1-2DGlcpA[4Me]b1-4DGalpAa1-4DGlcpAb1-4DGlcp") == "Man(a1-2)GlcA4Me(b1-4)GalA(a1-4)GlcA(b1-4)Glc"
    assert canonicalize_iupac("DNeup5Ac[9A]a2-3DGalpb1-4[LFucpa1-3]DGlcpNAc") == "Neu5Ac9Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc"
    assert canonicalize_iupac("DGlcpNAcb1-2[DGlcpa1-3]LRhapa1-2LRhapa1-3LRhap[2A]a1-OH") == "GlcNAc(b1-2)[Glc(a1-3)]Rha(a1-2)Rha(a1-3)Rha2Ac"
    assert canonicalize_iupac("LDmanpHepa1-OME") == "LDManHep1Me"
    assert canonicalize_iupac("Ma3(Ma6)Mb4GNb4GN;") == "Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("GNb2Ma3(Ab4GNb2Ma6)Mb4GNb4(Fa6)GNb;") == "Gal(b1-4)GlcNAc(b1-2)Man(a1-6)[GlcNAc(b1-2)Man(a1-3)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("01Y41Y41M(31M21M21M)61M(31M21M)61M21M") == "Man(a1-2)Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("01Y41Y41M(31M21M21M31G)61M(31M21M)61M") == "Glc(a1-3)Man(a1-2)Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-3)[Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("β-D-Galp-(1→4)-β-D-GlcpNAc-(1→") == "Gal(b1-4)GlcNAc"
    assert canonicalize_iupac("β-L-Galp-(1→4)-β-D-GlcpNAc-(1→") == "L-Gal(b1-4)GlcNAc"
    assert canonicalize_iupac("α-D-Neup5Ac-(2→3)-β-D-Galp-(1→4)-β-D-GlcpNAc-(1→") == "Neu5Ac(a2-3)Gal(b1-4)GlcNAc"
    assert canonicalize_iupac("α-D-Manp-(1→3)[α-D-Manp-(1→6)]-β-D-Manp-(1→4)-β-D-GlcpNAc-(1→4)-β-D-GlcpNAc-(1→") == "Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("M3") == "Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("FA4G3F2") == "Fuc(a1-3/4)[Gal(b1-3/4)]GlcNAc(b1-?)[Fuc(a1-3/4)[Gal(b1-3/4)]GlcNAc(b1-?)]Man(a1-3/6)[Gal(b1-3/4)GlcNAc(b1-?)[GlcNAc(b1-?)]Man(a1-3/6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("A4G4") == "Gal(b1-3/4)GlcNAc(b1-?)[Gal(b1-3/4)GlcNAc(b1-?)]Man(a1-3)[Gal(b1-3/4)GlcNAc(b1-?)[Gal(b1-3/4)GlcNAc(b1-?)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("FA4G3") == "Gal(b1-3/4)GlcNAc(b1-?)[GlcNAc(b1-?)]Man(a1-3/6)[Gal(b1-3/4)GlcNAc(b1-?)[Gal(b1-3/4)GlcNAc(b1-?)]Man(a1-3/6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("FA2BiG2") == "Gal(b1-3/4)GlcNAc(b1-2)Man(a1-3)[Gal(b1-3/4)GlcNAc(b1-2)Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("F(3)XA2") == "GlcNAc(b1-2)Man(a1-3)[GlcNAc(b1-2)Man(a1-6)][Xyl(b1-2)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)]GlcNAc"
    assert canonicalize_iupac("F(6)A2G(4)1Sg(6)1") == "Neu5Gc(a2-6)Gal(b1-4)GlcNAc(b1-2)Man(a1-3/6)[GlcNAc(b1-2)Man(a1-3/6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("FA3F1G3S[3,3]2") == "Neu5Ac(a2-3)Gal(b1-3/4)[Fuc(a1-3/4)]GlcNAc(b1-?)[Gal(b1-3/4)GlcNAc(b1-?)]Man(a1-3/6)[Neu5Ac(a2-3)Gal(b1-3/4)GlcNAc(b1-2)Man(a1-3/6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("A2G2S2(2,3)") == "Neu5Ac(a2-3/6)Gal(b1-3/4)GlcNAc(b1-2)Man(a1-3/6)[Neu5Ac(a2-3)Gal(b1-3/4)GlcNAc(b1-2)Man(a1-3/6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("FA4G3S1B") == "Neu5Ac(a2-3/6)Gal(b1-3/4)GlcNAc(b1-?)[Gal(b1-3/4)GlcNAc(b1-?)]Man(a1-3/6)[Gal(b1-3/4)GlcNAc(b1-?)[GlcNAc(b1-?)]Man(a1-3/6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("A2G1G[SO4-2]1S1") == "Neu5Ac(a2-3/6)GalOS(b1-3/4)GlcNAc(b1-2)Man(a1-3/6)[GlcNAc(b1-2)Man(a1-3/6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac('A2[6]G1') == 'Gal(b1-3/4)GlcNAc(b1-2)Man(a1-6)[GlcNAc(b1-2)Man(a1-3)]Man(b1-4)GlcNAc(b1-4)GlcNAc'
    assert canonicalize_iupac('A2B[3]G1') == 'Gal(b1-3/4)GlcNAc(b1-2)Man(a1-3)[GlcNAc(b1-2)Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)GlcNAc'
    assert canonicalize_iupac('F(6)A2[6]G(4)1Sg(6)1') == 'Neu5Gc(a2-6)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)[GlcNAc(b1-2)Man(a1-3)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc'
    assert canonicalize_iupac('D0H0') == '4uHexA(?1-?)GlcN'
    assert canonicalize_iupac('D2S9') == '4uHexA2S(?1-?)GlcNS3S6S'
    assert canonicalize_iupac('D2a4') == '4uHexA2S(?1-?)GalNAc4S'
    assert canonicalize_iupac("M4") == "Man(a1-2/3/6)Man(a1-3/6)[Man(a1-3/6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("man 4") == "Man(a1-2/3/6)Man(a1-3/6)[Man(a1-3/6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac('m7') == '{Man(a1-?)}{Man(a1-?)}{Man(a1-?)}{Man(a1-?)}Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc'
    assert canonicalize_iupac('Man-7') == '{Man(a1-?)}{Man(a1-?)}{Man(a1-?)}{Man(a1-?)}Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc'
    assert canonicalize_iupac('Man7') == '{Man(a1-?)}{Man(a1-?)}{Man(a1-?)}{Man(a1-?)}Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc'
    assert canonicalize_iupac("(Hex)3 (HexNAc)1 (NeuAc)1 + (Man)3(GlcNAc)2") == "{HexNAc(?1-?)}{Neu5Ac(a2-?)}{Hex(?1-?)}{Hex(?1-?)}{Hex(?1-?)}Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("(Hex)2 (HexNAc)3 (Deoxyhexose)1 (NeuAc)2 + (Man)3(GlcNAc)2") == "{HexNAc(?1-?)}{HexNAc(?1-?)}{HexNAc(?1-?)}{Neu5Ac(a2-?)}{Neu5Ac(a2-?)}{Hex(?1-?)}{Hex(?1-?)}{Fuc(a1-?)}Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("G2S2") == "Neu5Ac(a2-3/6)Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Neu5Ac(a2-3/6)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("G1FN") == "Gal(b1-3/4)GlcNAc(b1-2)Man(a1-3/6)[GlcNAc(b1-2)Man(a1-3/6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("redEnd--??1D-GalNAc,p--??2D-KDN,p$MONO,Und,-H,0,redEnd") == "Kdn(a2-?)GalNAc"
    assert canonicalize_iupac("redEnd--??1D-Glc,p--4b1D-Gal,p(--3a2D-NeuGc,p@270)--4b1D-GalNAc,p$MONO,Und,-H,0,redEnd") == "Neu5Gc(a2-3)[GalNAc(b1-4)]Gal(b1-4)Glc"
    assert canonicalize_iupac("redEnd--??1D-GalNAc,p(--3b1D-Gal,p(--3a2D-NeuAc,p)--6?1S)--6b1D-GlcNAc,p--4b1D-Gal,p$MONO,Und,-2H,0,redEnd") == "Neu5Ac(a2-3)Gal6S(b1-3)[Gal(b1-4)GlcNAc(b1-6)]GalNAc"
    assert canonicalize_iupac("redEnd--?b1D-GlcNAc,p--4b1D-GlcNAc,p--4b1D-Man,p((--3a1D-Man,p--2b1D-GlcNAc,p)--4b1D-GlcNAc,p)--6a1D-Man,p--2b1D-GlcNAc,p$MONO,Und,-H,0,redEnd") == "GlcNAc(b1-2)Man(a1-3)[GlcNAc(b1-2)Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("redEnd--?b1D-GlcNAc,p--4b1D-GlcNAc,p--4b1D-Man,p((--3a1D-Man,p--??1D-GlcNAc,p)--4b1D-GlcNAc,p)--6a1D-Man,p--??1D-GlcNAc,p}--??1D-GlcNAc,p$MONO,Und,-H,0,redEnd") == "{GlcNAc(?1-?)}GlcNAc(?1-?)Man(a1-3)[GlcNAc(?1-?)Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("redEnd--?a1D-GalNAc,p(--6b1D-GlcNAc,p)--3b1D-Gal,p--??1D-GlcNAc,p(--??1L-Fuc,p)--??1S$MONO,Und,-H,0,redEnd") == "Fuc(a1-?)GlcNAcOS(?1-?)Gal(b1-3)[GlcNAc(b1-6)]GalNAc"
    assert canonicalize_iupac("freeEnd--?b1D-GlcNAc,p(--6a1L-Fuc,p)--4b1D-GlcNAc,p--4b1D-Man,p(--3a1D-Man,p(--??1D-GlcNAc,p--??1D-Gal,p--??2D-NeuAc,p)--??1D-GlcNAc,p--??1D-Gal,p--??2D-NeuAc,p--??2D-NeuAc,p)--6a1D-Man,p(--??1D-Man,p)--??1D-Man,p$MONO,Und,-2H,0,freeEnd") == "Neu5Ac(a2-?)Neu5Ac(a2-?)Gal(?1-?)GlcNAc(?1-?)[Neu5Ac(a2-?)Gal(?1-?)GlcNAc(?1-?)]Man(a1-3)[Man(?1-?)[Man(?1-?)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc-ol"
    assert canonicalize_iupac('freeEnd--?b1D-GlcNAc,p--4b1D-GlcNAc,p--4b1D-Man,p((--3a1D-Man,p--2b1D-GlcNAc,p@270--4b1D-Gal,p--3a2D-NeuAc,p@315)--4b1D-GlcNAc,p)--6a1D-Man,p--2b1D-GlcNAc,p@270--4b1D-Gal,p--3a2D-NeuAc,p@315$MONO,Und,-2H,0,freeEnd')  == 'Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)GlcNAc-ol'
    assert canonicalize_iupac("freeEnd--?D-GalNAc--3b1D-GlcNAc,p(--3a1D-Gal,p)--4a1D-Fuc,p$MONO,perMe,Na,0,freeEnd") == "Gal(a1-3)[D-Fuc(a1-4)]GlcNAc(b1-3)GalNAc-ol"
    assert canonicalize_iupac("freeEnd--??1D-Qui(--2b1D-Glc,p)--4b1D-Qui,p") == "Glc(b1-2)[Qui(b1-4)]Qui-ol"
    assert canonicalize_iupac("freeEnd--?[--4b1L-Rha,p--3a1D-Glc,p((--3a1L-Rha,p)--4b1D-Glc,p--?])--6b1D-Glc,p") == "Rha(a1-3)[Glc(b1-4)][Glc(b1-6)]Glc(a1-3)Rha-ol"
    assert canonicalize_iupac("freeEnd--?b1D-GlcNAc,p(--4b1D-GlcNAc,p--4b1D-Man,p(--3a1D-Man,p--2b1D-GlcNAc,p--4b1D-GalNAc,p--??1S)--6a1D-Man,p(--3a1D-Man,p)--6a1D-Man,p)--6a1L-Fuc,p$MONO,perMe,Na,0,freeEnd") == "GalNAcOS(b1-4)GlcNAc(b1-2)Man(a1-3)[Man(a1-3)[Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc-ol"
    assert canonicalize_iupac("freeEnd--??1L-Ara--5[--5a1L-Ara,f--?]--5a1L-Ara,f$MONO,perMe,Na,0,freeEnd") == "Araf(a1-5)Araf(a1-5)Araf(a1-5)Araf(a1-5)Araf(a1-5)Araf(a1-5)Ara-ol"
    assert canonicalize_iupac("WURCS=2.0/5,7,6/[u2122h_2*NCC/3=O][a2122h-1b_1-5_2*NCC/3=O][a1122h-1b_1-5][a1122h-1a_1-5][a2112h-1b_1-5_2*NCC/3=O_4*OSO/3=O/3=O]/1-2-3-4-2-5-4/a4-b1_b4-c1_c3-d1_c6-g1_d2-e1_e4-f1") == "GalNAc4S(b1-4)GlcNAc(b1-2)Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("WURCS=2.0/8,15,14/[u2122h_2*NCC/3=O][a2122h-1b_1-5_2*NCC/3=O][a1122h-1b_1-5][a1122h-1a_1-5][a1221m-1a_1-5][a2112h-1b_1-5][Aad21122h-2a_2-6_5*NCCO/3=O][Aad21122h-2a_2-6_5*NCC/3=O]/1-2-3-4-2-5-6-7-8-4-2-5-6-8-5/a4-b1_a6-o1_b4-c1_c3-d1_c6-j1_d2-e1_e3-f1_e4-g1_h8-i2_j2-k1_k3-l1_k4-m1_h2-g3|g6_n2-m3|m6 ") == "Neu5Ac(a2-8)Neu5Gc(a2-3/6)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[Neu5Ac(a2-3/6)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("WURCS=2.0/3,5,4/[a2122h-1b_1-5_2*NCC/3=O][a1122h-1b_1-5][a1122h-1a_1-5]/1-1-2-3-3/a4-b1_b4-c1_c3-d1_c6-e1") == "Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("WURCS=2.0/3,5,4/[a2122h-1b_1-5_2*NCC/3=O][a1221m-1b_1-5][a1122h-1a_1-5]/1-1-2-3-3/a4-b1_b4-c1_c3-d1_c6-e1") == "Man(a1-3)[Man(a1-6)]Fuc(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("WURCS=2.0/5,19,18/[a2122h-1x_1-5_2*NCC/3=O][a1122h-1x_1-5][a2112h-1x_1-5][Aad21122h-2x_2-6_5*NCC/3=O][a1221m-1x_1-5]/1-1-2-2-1-3-4-1-3-4-2-1-3-4-1-3-4-5-4/a?-b1_a?-r1_b?-c1_c?-d1_c?-k1_d?-e1_d?-h1_e?-f1_f?-g2_h?-i1_i?-j2_k?-l1_k?-o1_l?-m1_m?-n2_o?-p1_p?-q2_s2-a?|b?|c?|d?|e?|f?|g?|h?|i?|j?|k?|l?|m?|n?|o?|p?|q?|r?}") == "{Neu5Ac(a2-?)}Neu5Ac(a2-?)Gal(?1-?)GlcNAc(?1-?)[Neu5Ac(a2-?)Gal(?1-?)GlcNAc(?1-?)]Man(?1-?)[Neu5Ac(a2-?)Gal(?1-?)GlcNAc(?1-?)[Neu5Ac(a2-?)Gal(?1-?)GlcNAc(?1-?)]Man(?1-?)]Man(?1-?)GlcNAc(?1-?)[Fuc(a1-?)]GlcNAc"
    assert canonicalize_iupac("WURCS=2.0/4,8,7/[u2122h_2*NCC/3=O][a2121A-1a_1-5][a2122h-1a_1-5_2*NCC/3=O][a2122A-1b_1-5]/1-2-3-4-3-2-3-4/a4-b1_b4-c1_c4-d1_d4-e1_e4-f1_f4-g1_g4-h1") == "GlcA(b1-4)GlcNAc(a1-4)IdoA(a1-4)GlcNAc(a1-4)GlcA(b1-4)GlcNAc(a1-4)IdoA(a1-4)GlcNAc"
    assert canonicalize_iupac("WURCS=2.0/6,9,8/[a2122h-1a_1-5_2*NCC/3=O][a1221m-1a_1-5][a2122h-1b_1-5_2*NCC/3=O][a1122h-1b_1-5][a212h-1b_1-5][a1122h-1a_1-5]/1-2-3-4-5-6-3-6-3/a3-b1_a4-c1_c4-d1_d2-e1_d3-f1_d6-h1_f2-g1_h2-i1") == "GlcNAc(b1-2)Man(a1-3)[GlcNAc(b1-2)Man(a1-6)][Xyl(b1-2)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)]GlcNAc"
    assert canonicalize_iupac("WURCS=2.0/7,19,18/[u2122h_2*NCC/3=O][a2122h-1b_1-5_2*NCC/3=O][a1122h-1b_1-5][a1122h-1a_1-5][a1221m-1a_1-5][a2112h-1b_1-5][Aad21122h-2a_2-6_5*NCCO/3=O]/1-2-3-4-2-5-6-7-2-6-2-4-2-5-6-2-5-6-7/a4-b1_b4-c1_c3-d1_c4-k1_c6-l1_d2-e1_d4-i1_e3-f1_e4-g1_i4-j1_l2-m1_l6-p1_m3-n1_m4-o1_p3-q1_p4-r1_h2-g3|g6_s2-r3|r6 ") == "Neu5Gc(a2-3/6)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Neu5Gc(a2-3/6)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-6)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)]Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("WURCS=2.0/2,2,1/[a1122h-1x_1-5][a2122h-1a_1-5_2*NCC/3=O]/1-2/a6-b1*OPO*/3O/3=O") == "GlcNAc1P(a1-6)Man"
    assert canonicalize_iupac("WURCS=2.0/1,2,1/[a2122h-1b_1-5]/1-1/a4-b1*OSO*/3=O/3=O") == "Glc1S(b1-4)Glc"
    assert canonicalize_iupac("WURCS=2.0/4,8,7/[a1211h-1b_1-5_2*NCC/3=O][a1122h-1b_1-5][a2211h-1a_1-5][a1122h-1a_1-5]/1-2-3-3-3-3-4-4/a4-b1_b3-c1_b6-e1_c2-d1_e3-f1_e6-g1_g2-h1") == "Man(a1-2)Man(a1-6)[L-Man(a1-3)]L-Man(a1-6)[L-Man(a1-2)L-Man(a1-3)]Man(b1-4)L-GlcNAc"
    assert canonicalize_iupac("G07426YY") == "Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("G96417BZ") == "Man(a1-2)Man(a1-3)[Man(a1-3)[Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("G26039ES") == "Gal(b1-4)Glc-ol"
    assert canonicalize_iupac("G02923VP") == "GlcNAc(b1-2)Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac(782) == "Gal(?1-?)[GlcNAcOS(?1-?)]GalNAc"
    assert canonicalize_iupac(72) == "Fuc(a1-2)Gal(b1-3/4)GlcNAc(b1-?)[Fuc(a1-2)Gal(b1-3/4)GlcNAc(b1-?)]Gal(b1-3)[GlcNAc(b1-6)]GalNAc"
    assert canonicalize_iupac("RES 1b:a-lgal-HEX-1:5|6:d") == "Fuc"
    assert canonicalize_iupac("""RES
1b:b-dglc-HEX-1:5
2s:n-acetyl
3b:b-dglc-HEX-1:5
4s:n-acetyl
5b:b-dman-HEX-1:5
6b:a-dman-HEX-1:5
7b:a-dman-HEX-1:5
LIN
1:1d(2+1)2n
2:1o(4+1)3d
3:3d(2+1)4n
4:3o(4+1)5d
5:5o(3+1)6d
6:5o(6+1)7d""") == "Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert canonicalize_iupac("""RES
1b:b-dglc-HEX-1:5
2s:n-acetyl
3b:b-dglc-HEX-1:5
4s:n-acetyl
5b:b-dman-HEX-1:5
6b:a-dman-HEX-1:5
7b:b-dglc-HEX-1:5
8s:n-acetyl
9b:b-dgal-HEX-1:5
10b:b-dglc-HEX-1:5
11s:n-acetyl
12b:b-dgal-HEX-1:5
13b:a-dman-HEX-1:5
14b:b-dglc-HEX-1:5
15s:n-acetyl
16b:b-dgal-HEX-1:5
17b:b-dglc-HEX-1:5
18s:n-acetyl
19b:b-dgal-HEX-1:5
20b:b-dglc-HEX-1:5
21s:n-acetyl
22b:b-dgal-HEX-1:5
23b:b-dglc-HEX-1:5
24s:n-acetyl
25b:b-dgal-HEX-1:5
26b:a-lgal-HEX-1:5|6:d
LIN
1:1d(2+1)2n
2:1o(4+1)3d
3:3d(2+1)4n
4:3o(4+1)5d
5:5o(3+1)6d
6:6o(2+1)7d
7:7d(2+1)8n
8:7o(4+1)9d
9:6o(4+1)10d
10:10d(2+1)11n
11:10o(4+1)12d
12:5o(6+1)13d
13:13o(2+1)14d
14:14d(2+1)15n
15:14o(4+1)16d
16:16o(3+1)17d
17:17d(2+1)18n
18:17o(4+1)19d
19:13o(6+1)20d
20:20d(2+1)21n
21:20o(4+1)22d
22:22o(3+1)23d
23:23d(2+1)24n
24:23o(4+1)25d
25:1o(6+1)26d""") == "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)[Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("""RES
1b:b-dglc-HEX-1:5
2s:n-acetyl
3b:b-dglc-HEX-1:5
4s:n-acetyl
5b:b-dman-HEX-1:5
6b:a-dman-HEX-1:5
7b:b-dglc-HEX-1:5
8s:n-acetyl
9b:b-dgal-HEX-1:5
10b:a-dman-HEX-1:5
11b:b-dglc-HEX-1:5
12s:n-acetyl
13b:b-dgal-HEX-1:5
14b:b-dglc-HEX-1:5
15s:n-acetyl
16b:b-dgal-HEX-1:5
17b:a-lgal-HEX-1:5|6:d
LIN
1:1d(2+1)2n
2:1o(4+1)3d
3:3d(2+1)4n
4:3o(4+1)5d
5:5o(-1+1)6d
6:6o(2+1)7d
7:7d(2+1)8n
8:7o(4+1)9d
9:5o(-1+1)10d
10:10o(-1+1)11d
11:11d(2+1)12n
12:11o(4+1)13d
13:10o(2+1)14d
14:14d(2+1)15n
15:14o(4+1)16d
16:1o(6+1)17d
UND
UND1:100.0:100.0
ParentIDs:1|3|5|6|7|9|10|11|13|14|16|17
SubtreeLinkageID1:o(3+2)d
RES
18b:a-dgro-dgal-NON-2:6|1:a|2:keto|3:d
19s:n-acetyl
LIN
17:18d(5+1)19n
UND2:100.0:100.0
ParentIDs:1|3|5|6|7|9|10|11|13|14|16|17
SubtreeLinkageID1:o(3+2)d
RES
20b:a-dgro-dgal-NON-2:6|1:a|2:keto|3:d
21s:n-acetyl
LIN
18:20d(5+1)21n
UND3:100.0:100.0
ParentIDs:1|3|5|6|7|9|10|11|13|14|16|17
SubtreeLinkageID1:o(6+2)d
RES
22b:a-dgro-dgal-NON-2:6|1:a|2:keto|3:d
23s:n-acetyl
LIN
19:22d(5+1)23n""") == "{Neu5Ac(a2-?)}{Neu5Ac(a2-?)}{Neu5Ac(a2-?)}Gal(b1-4)GlcNAc(b1-2)Man(a1-?)[Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-?)]Man(a1-?)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    assert canonicalize_iupac("""RES
1b:a-dgal-HEX-1:5
2s:n-acetyl
3b:b-dgal-HEX-1:5
4b:b-dglc-HEX-1:5
5s:n-acetyl
6s:sulfate
LIN
1:1d(2+1)2n
2:1o(3+1)3d
3:1o(6+1)4d
4:4d(2+1)5n
5:4o(6+1)6n""") == "Gal(b1-3)[GlcNAc6S(b1-6)]GalNAc"
    assert canonicalize_iupac("""RES
1b:b-dglc-HEX-1:5
2s:n-acetyl
3b:b-dgal-HEX-1:5
4b:a-lxyl-HEX-1:5|3:d|6:d
5b:a-lxyl-HEX-1:5|3:d|6:d
LIN
1:1d(2+1)2n
2:1o(3+1)3d
3:3o(2+1)4d
4:1o(4+1)5d""") == "Col(a1-2)Gal(b1-3)[Col(a1-4)]GlcNAc"
    assert glycoct_to_iupac("""RES
1b:a-dman-OCT-2:6|1:a|2:keto|3:d
2b:a-dman-OCT-2:6|1:a|2:keto|3:d
3b:a-dman-OCT-2:6|1:a|2:keto|3:d
LIN
1:1o(8+2)2d
2:2o(8+2)3d""") == "Kdo(a2-8)Kdo(a2-8)Kdo"
    # Test raised errors
    with pytest.raises(ValueError, match="Mismatching brackets in formatted glycan string"):
        canonicalize_iupac("Fuc(a1-3)[Gal(b1-4)Glc-ol")
    with pytest.raises(ValueError, match="Seems like you're using SMILES. We currently can only convert IUPAC-->SMILES; not the other way around."):
        canonicalize_iupac("O[C@@H]1[C@H](O)[C@@H](O)[C@H](O)[C@@]([H])(CO)O1")


def test_constrain_prot():
    # Test replacing non-library characters with 'z'
    proteins = ['ABC', 'XYZ', 'A1C']
    chars = {'A': 1, 'B': 2, 'C': 3}
    result = constrain_prot(proteins, chars)
    assert result == ['ABC', 'zzz', 'AzC']
    result = constrain_prot(proteins)
    assert result == ['ABC', 'XYZ', 'AzC']


def test_prot_to_coded():
    proteins = ['ABC', 'DE']
    chars = {'A': 0, 'B': 1, 'C': 2, 'D': 3, 'E': 4, 'z': 5}
    pad_len = 5
    result = prot_to_coded(proteins, chars, pad_len)
    # ABC should be [0,1,2,5,5] where len(chars)-1 is padding
    # DE should be [3,4,5,5,5]
    assert result[0] == [0, 1, 2, 5, 5]
    assert result[1] == [3, 4, 5, 5, 5]
    result = prot_to_coded(proteins)
    assert len(result[0]) == 1000


def test_string_to_labels():
    chars = {'A': 1, 'B': 2, 'C': 3}
    assert string_to_labels('ABC', chars) == [1, 2, 3]
    assert string_to_labels('BAC', chars) == [2, 1, 3]
    assert string_to_labels('', chars) == []
    assert len(string_to_labels('ABC')) == 3


def test_pad_sequence():
    seq = [1, 2, 3]
    max_length = 5
    pad_label = 0
    result = pad_sequence(seq, max_length, pad_label)
    assert result == [1, 2, 3, 0, 0]
    result = pad_sequence(seq, max_length)
    # Test when sequence is longer than max_length
    seq = [1, 2, 3, 4, 5, 6]
    max_length = 4
    result = pad_sequence(seq, max_length, pad_label)
    assert result == [1, 2, 3, 4, 5, 6]  # Should not truncate


def test_get_core():
    assert get_core('Neu5Ac') == 'Neu5Ac'
    assert get_core('GlcNAc') == 'GlcNAc'
    assert get_core('Gal6S') == 'Gal'
    assert get_core('Man3S') == 'Man'
    assert get_core('6dAlt') == '6dAlt'


def test_get_modification():
    assert get_modification('Neu5Ac9Ac') == '9Ac'
    assert get_modification('GlcNAc') == ''
    assert get_modification('Gal6S') == '6S'
    assert get_modification('Man3S') == '3S'
    assert get_modification('Gal3S6S') == '3S6S'


def test_get_stem_lib():
    lib = {'Neu5Ac': 1, 'Gal6S': 2, 'GlcNAc': 3, 'Neu5Gc8S': 4}
    result = get_stem_lib(lib)
    assert result == {'Neu5Ac': 'Neu5Ac', 'Gal6S': 'Gal', 'GlcNAc': 'GlcNAc', 'Neu5Gc8S': 'Neu5Gc'}


def test_stemify_glycan():
    # Test basic glycan
    result = stemify_glycan("Neu5Ac9Ac(a2-3)Gal(b1-4)GlcNAc")
    assert result == "Neu5Ac(a2-3)Gal(b1-4)GlcNAc"
    result = stemify_glycan("Neu5Ac9Ac")
    assert result == "Neu5Ac"
    # Test glycan with multiple modifications
    result = stemify_glycan("Neu5Ac9Ac(a2-3)Gal6S(b1-4)GlcNAc")
    assert result == "Neu5Ac(a2-3)Gal(b1-4)GlcNAc"
    assert stemify_glycan("Neu5Ac9Ac(a2-3)Gal6S(b1-4)GlcNAc6S") == "Neu5Ac(a2-3)Gal(b1-4)GlcNAc"


def test_glycan_to_composition():
    # Test basic glycan
    result = glycan_to_composition("Neu5Ac(a2-3)Gal(b1-4)GlcNAc")
    assert result == {'Neu5Ac': 1, 'Hex': 1, 'HexNAc': 1}
    #result = glycan_to_composition('Neu5Ac(a2-3/6)Gal(b1-3)GalNAc')
    # Test glycan with sulfation
    result = glycan_to_composition("Neu5Ac(a2-3)Gal6S(b1-4)GlcNAc")
    assert result == {'Neu5Ac': 1, 'Hex': 1, 'HexNAc': 1, 'S': 1}
    # Test lactonization
    result = glycan_to_composition("Neu1,7lactone5,9Ac2(a2-3)Gal(b1-4)Glc")
    assert result == {'Neu5Ac': 1, 'Hex': 2, '-H2O': 1, 'Ac': 1}


def test_glycan_to_mass():
    assert abs(glycan_to_mass("Neu1,7lactone5,9Ac2(a2-3)Gal(b1-4)Glc") - 674.2149056) < 0.1
    assert abs(glycan_to_mass("Neu5Az(a2-3)Gal(b1-4)Glc") - 658.2181545999999) < 0.1
    assert abs(glycan_to_mass("Neu5Ac(a2-3)Gal(b1-4)Glc") - 633.2115546) < 0.1
    assert abs(glycan_to_mass("Neu5Az9Ac(a2-3)Gal(b1-4)Glc") - 717.2320056) < 0.1
    assert abs(glycan_to_mass("Neu5Ac(a2-3)Gal(b1-4)Glc", adduct = "C2H4O2") - 693.2325546) < 0.1
    assert abs(glycan_to_mass("Neu5Ac(a2-3)Gal(b1-4)Glc", adduct = "-C2H4O2") - 573.1905546) < 0.1
    assert abs(glycan_to_mass("GalOS(b1-3)GalNAc4/6S") - 543.0563546) < 0.1


def test_calculate_adduct_mass():
    # Test simple molecules
    assert abs(calculate_adduct_mass('H2O') - 18.0106) < 0.001
    assert abs(calculate_adduct_mass('+H2O') - 18.0106) < 0.001
    assert abs(calculate_adduct_mass('CH3COOH') - 60.0211) < 0.001
    assert abs(calculate_adduct_mass('-H2O') + 18.0106) < 0.001
    # Test with different mass types
    h2o_mono = calculate_adduct_mass('H2O', mass_value='monoisotopic')
    h2o_avg = calculate_adduct_mass('H2O', mass_value='average')
    assert h2o_mono != h2o_avg
    # Test enforce_sign
    assert calculate_adduct_mass('CH3COOH', enforce_sign=True) < 0.001


def test_structure_to_basic():
    # Basic glycan
    result = structure_to_basic("Neu5Ac(a2-3)Gal(b1-4)GlcNAc")
    assert result == "Neu5Ac(?1-?)Hex(?1-?)HexNAc"
    result = structure_to_basic("GlcNAc")
    assert result == "HexNAc"
    # Glycan with sulfation
    result = structure_to_basic("Gal6S(b1-4)GlcNAc")
    assert result == "HexOS(?1-?)HexNAc"
    # Glycan with branch
    result = structure_to_basic("Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc")
    assert result == "Neu5Ac(?1-?)Hex(?1-?)[dHex(?1-?)]HexNAc"


def test_map_to_basic():
    assert map_to_basic("Gal") == "Hex"
    assert map_to_basic("Man") == "Hex"
    assert map_to_basic("Glc") == "Hex"
    assert map_to_basic("GlcNAc") == "HexNAc"
    assert map_to_basic("Fuc") == "dHex"
    assert map_to_basic("GlcA") == "HexA"
    assert map_to_basic("GlcA3S") == "HexAOS"
    assert map_to_basic("Gal6S") == "HexOS"
    assert map_to_basic("GlcNAc6S") == "HexNAcOS"
    assert map_to_basic("GlcNS") == "HexNS"
    assert map_to_basic("Man6P") == "HexOP"
    assert map_to_basic("GlcNAc6P") == "HexNAcOP"
    assert map_to_basic("a2-3/6/8/9") == "?1-?"


def test_mask_rare_glycoletters():
    glycans = [
        "Neu5Ac(a2-3)Gal(b1-4)GlcNAc",
        "Neu5Ac(a2-3)Gal(b1-4)GlcNAc",
        "Neu5Ac(a2-6)Gal(b1-4)GlcNAc",
        "Gal(b1-4)GlcNAc",
        "Man6P(a1-2)Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc",
        "RareSugar(a1-2)GlcNAc",
        "GlcNAc(a1-2)WeirdSugar"
    ]

    # Test with default thresholds
    result = mask_rare_glycoletters(glycans)
    assert "RareSugar" not in result[5]
    assert "Monosaccharide" in result[5]
    # Test with custom thresholds
    result = mask_rare_glycoletters(glycans, thresh_monosaccharides=2)
    assert "Neu5Ac" in result[0]  # Occurs twice, should remain
    assert "Man6P" not in result[4]  # Occurs once, should be masked


def test_mz_to_composition():
    # Test O-glycan mass
    result = mz_to_composition(
        675,
        mode='negative',
        mass_value='monoisotopic',
        glycan_class='O',
        mass_tolerance=0.5,
        reduced=True,
        filter_out = {'Kdn'}
    )
    expected = [{'Neu5Ac': 1, 'Hex': 1, 'HexNAc': 1}]
    assert result == expected
    result = mz_to_composition(
        675,
        mode='negative',
        mass_value='monoisotopic',
        glycan_class='all',
        mass_tolerance=0.5,
        reduced=True,
        adduct="H2O",
        extras=["doubly_charged", "adduct"]
    )
    result = mz_to_composition(
        698,
        mode='positive',
        mass_value='monoisotopic',
        glycan_class='all',
        mass_tolerance=0.5,
        reduced=True,
        filter_out = {'Kdn'},
        extras=["doubly_charged", "adduct"]
    )


def test_compositions_to_structures():
    composition = {'Hex': 3, 'HexNAc': 2}
    result = compositions_to_structures(
        [composition],
        glycan_class='N',
        kingdom='Animalia'
    )
    # Should return DataFrame with N-glycan structures matching composition
    assert isinstance(result, pd.DataFrame)
    assert 'glycan' in result.columns
    # Verify returned glycans have correct composition
    for glycan in result['glycan']:
        comp = glycan_to_composition(glycan)
        assert comp['Hex'] == 3
        assert comp['HexNAc'] == 2
    composition = {'Hex': 3, 'HexNAc': 2, 'dHex': 10}
    result = compositions_to_structures(
        [composition],
        glycan_class='N',
        kingdom='Animalia',
        verbose=True
    )


def test_match_composition_relaxed():
    composition = {'Hex': 3, 'HexNAc': 2}
    result = match_composition_relaxed(
        composition,
        glycan_class='N',
        kingdom='Animalia'
    )
    # Should return list of glycan strings
    assert isinstance(result, list)
    # Each glycan should match the composition
    for glycan in result:
        comp = glycan_to_composition(glycan)
        assert comp['Hex'] == 3
        assert comp['HexNAc'] == 2


def test_stemify_dataset():
    df = pd.DataFrame({
        'glycan': [
            "Neu5Ac(a2-3)Gal(b1-4)GlcNAc",
            "Neu5Ac9Ac(a2-3)Gal6S(b1-4)GlcNAc",
            "Man(a1-3)[Man(a1-6)]Man"
        ]
    })

    result = stemify_dataset(df, rarity_filter=1)

    # Check modifications are removed
    assert "Neu5Ac(a2-3)Gal(b1-4)GlcNAc" in result['glycan'].values
    assert "6S" not in result['glycan'].str.cat()
    assert "5Ac9Ac" not in result['glycan'].str.cat()


def test_composition_to_mass():
    # Test basic composition
    comp = {'Hex': 1, 'HexNAc': 1}
    mass = composition_to_mass(comp)
    expected_mass = 383.1322  # Theoretical mass for Hex-HexNAc
    assert abs(mass - expected_mass) < 0.1
    mass = composition_to_mass(comp, adduct="H2O")
    assert abs(mass - expected_mass - 18.01) < 0.1
    # Test with modifications
    comp = {'Hex': 1, 'HexNAc': 1, 'S': 1}
    mass = composition_to_mass(comp)
    assert mass > composition_to_mass({'Hex': 1, 'HexNAc': 1})


def test_condense_composition_matching():
    glycans = [
        "Gal(b1-4)GlcNAc",
        "Gal(b1-4)GlcNAc",  # Duplicate
        "Gal(?1-4)GlcNAc",  # Similar but with wildcard
        "Man(a1-3)GlcNAc"   # Different structure
    ]
    result = condense_composition_matching(glycans)
    # Should remove duplicates and similar structures with wildcards
    assert len(result) < len(glycans)
    assert "Gal(b1-4)GlcNAc" in result
    assert "Man(a1-3)GlcNAc" in result


def test_get_unique_topologies():
    composition = {'Hex': 3, 'HexNAc': 2}
    result = get_unique_topologies(
        composition,
        glycan_type='N',
        taxonomy_value='Animalia'
    )
    # Should return list of unique topology strings
    assert isinstance(result, list)
    # Each topology should match the original composition
    for topology in result:
        comp = glycan_to_composition(topology)
        assert comp['Hex'] == 3
        assert comp['HexNAc'] == 2
    # with universal replacer
    composition = {'Hex': 3, 'HexNAc': 2, 'dHex': 1}
    result = get_unique_topologies(
        composition,
        glycan_type='N',
        taxonomy_value='Animalia',
        universal_replacers={'dHex': 'Fuc'}
    )
    assert isinstance(result, list)
    for topology in result:
        comp = glycan_to_composition(topology)
        assert comp['Hex'] == 3
        assert comp['HexNAc'] == 2
        assert 'Fuc' in topology


def test_mz_to_structures():
    # Test single mass
    mz_values = [530]
    result = mz_to_structures(
        mz_values,
        glycan_class='O',
        mode='negative',
        mass_value='monoisotopic',
        mass_tolerance=0.5,
        reduced=True
    )
    assert isinstance(result, pd.DataFrame)
    # Verify returned structures match the mass
    for glycan in result['glycan']:
        mass = glycan_to_mass(glycan) + 1.0078
        assert abs(mass - mz_values[0]) < 0.5
    # Test multiple masses
    mz_values = [530, 675]
    abundances = pd.DataFrame({
        'mz': mz_values,
        'intensity': [100, 50]
    })
    result = mz_to_structures(
        mz_values,
        glycan_class='O',
        mode='negative',
        reduced=True,
        filter_out={'Kdn'},
        abundances=abundances
    )
    assert len(result) > 0
    assert 'abundance' in result.columns
    # Test filtering
    result = mz_to_structures(
        mz_values,
        glycan_class='O',
        reduced=True,
        filter_out={'Kdn', 'Neu5Ac'}
    )
    # Verify no structures contain Neu5Ac
    for glycan in result['glycan']:
        assert 'Neu5Ac' not in glycan


# Test the rescue_glycans decorator
def test_rescue_glycans():
    result = glycan_to_mass("Fuc(a1-2)Gal(b1-3)GalNAc") + 1.0078
    assert abs(result - 530) < 0.5
    result = glycan_to_mass("Fuc(a1-2Gal(b1-3)GalNAc") + 1.0078
    assert abs(result - 530) < 0.5
    result = glycan_to_mass("Fuca2Galb3GalNAc") + 1.0078
    assert abs(result - 530) < 0.5
    result = glycan_to_mass("Fuca2Galb3GalNac") + 1.0078
    assert abs(result - 530) < 0.5


# Test the rescue_compositions decorator
def test_rescue_compositions():
    result = composition_to_mass({'HexNAc': 1, 'Hex': 1, 'dHex': 1}) + 1.0078
    assert abs(result - 530) < 0.5
    result = composition_to_mass('Hex1HexNAc1Fuc1') + 1.0078
    assert abs(result - 530) < 0.5
    result = composition_to_mass('H1N1F1') + 1.0078
    assert abs(result - 530) < 0.5


# Test for get_random_glycan
@pytest.mark.parametrize("n,glycan_class,kingdom,expected_type", [
    (1, "all", "Animalia", str),
    (3, "N", "Animalia", list),
    (2, "O", "Bacteria", list),
    (1, "lipid", "Fungi", str),
])
def test_get_random_glycan(n, glycan_class, kingdom, expected_type):
    # Mock the DataFrame and sample function
    mock_df = Mock()
    mock_df.glycan.values.tolist.return_value = ["Glycan1", "Glycan2", "Glycan3", "Glycan4"]
    with patch("glycowork.glycan_data.loader.df_glycan", mock_df):
        with patch("random.sample", lambda x, n: x[:n]):
            result = get_random_glycan(n, glycan_class, kingdom)
            assert isinstance(result, expected_type)
            if n > 1:
                assert len(result) == n


def test_min_process_glycans():
    glycans = [
        "Neu5Ac(a2-3)Gal(b1-4)GlcNAc",
        "Gal(b1-4)[Fuc(a1-3)]GlcNAc"
    ]
    result = min_process_glycans(glycans)
    assert result[0] == ['Neu5Ac', 'a2-3', 'Gal', 'b1-4', 'GlcNAc']
    assert result[1] == ['Gal', 'b1-4', 'Fuc', 'a1-3', 'GlcNAc']


def test_get_lib():
    glycans = ["Neu5Ac(a2-3)Gal(b1-4)GlcNAc"]
    lib = get_lib(glycans)
    assert 'Neu5Ac' in lib
    assert 'Gal' in lib
    assert 'GlcNAc' in lib
    assert 'a2-3' in lib
    assert 'b1-4' in lib


def test_expand_lib():
    existing_lib = {'Gal': 0, 'GlcNAc': 1, 'b1-4': 2}
    new_glycans = ["Neu5Ac(a2-3)Gal(b1-4)GlcNAc"]
    expanded = expand_lib(existing_lib, new_glycans)
    assert 'Neu5Ac' in expanded
    assert 'a2-3' in expanded
    assert expanded['Gal'] == 0  # Should keep original indices
    assert expanded['GlcNAc'] == 1


def test_get_possible_linkages():
    # Test exact match
    assert 'a1-3' in get_possible_linkages('a1-3')
    # Test wildcard
    result = get_possible_linkages('a?-3')
    assert 'a1-3' in result
    assert 'a2-3' in result
    # Test multiple wildcards
    result = get_possible_linkages('?1-?')
    assert 'a1-3' in result
    assert 'b1-4' in result
    # Test narrow linkage ambiguities
    result = get_possible_linkages("b1-3/4")
    assert 'b1-3' in result
    assert 'b1-4' in result
    assert 'b1-?' in result
    result = get_possible_linkages("?1-3/4")
    assert 'b1-3' in result
    result = get_possible_linkages("?1-2/4/6")
    assert '?1-2/6' in result
    result = get_possible_linkages("a1-2/4/6")
    assert 'a1-4/6' in result
    assert 'a1-4' in result
    assert 'a1-?' in result


def test_get_possible_monosaccharides():
    # Test Hex category
    hex_result = get_possible_monosaccharides('Hex')
    assert 'Glc' in hex_result
    assert 'Gal' in hex_result
    assert 'Man' in hex_result
    # Test HexNAc category
    hexnac_result = get_possible_monosaccharides('HexNAc')
    assert 'GlcNAc' in hexnac_result
    assert 'GalNAc' in hexnac_result


def test_de_wildcard_glycoletter():
    # Test linkage wildcard
    result = de_wildcard_glycoletter('?1-?')
    assert result in linkages
    result = de_wildcard_glycoletter('a1-3/4')
    assert result in linkages
    # Test monosaccharide wildcard
    result = de_wildcard_glycoletter('Hex')
    assert result in Hex
    # Test quick return
    result = de_wildcard_glycoletter('Man')
    assert result == "Man"


def test_glycoct_to_iupac():
    glycoct = """RES
1b:x-dglc-HEX-1:5
2s:n-acetyl
3b:b-dglc-HEX-1:5
4s:n-acetyl
5b:b-dman-HEX-1:5
6b:a-dman-HEX-1:5
7b:x-dglc-HEX-1:5
8s:n-acetyl
9b:a-dman-HEX-1:5
10b:a-lgal-HEX-1:5|6:d
LIN
1:1d(2+1)2n
2:1o(4+1)3d
3:3d(2+1)4n
4:3o(4+1)5d
5:5o(3+1)6d
6:6o(-1+1)7d
7:7d(2+1)8n
8:5o(6+1)9d
9:1o(6+1)10d"""
    result = glycoct_to_iupac(glycoct)
    assert "Man" in result
    assert "GlcNAc" in result


def test_wurcs_to_iupac():
    wurcs = "WURCS=2.0/6,10,9/[a2122h-1x_1-5_2*NCC/3=O][a2122h-1b_1-5_2*NCC/3=O][a1122h-1b_1-5][a1122h-1a_1-5][a2112h-1b_1-5][Aad21122h-2a_2-6_5*NCC/3=O]/1-2-3-4-2-5-6-4-2-5/a4-b1_b4-c1_c3-d1_c6-h1_d2-e1_e4-f1_f6-g2_h2-i1_i4-j1"
    result = wurcs_to_iupac(wurcs)
    assert "GlcNAc" in result
    assert "Man" in result
    assert "Gal" in result
    assert "Neu5Ac" in result


def test_oxford_to_iupac():
    # Test basic N-glycan
    assert "GlcNAc" in oxford_to_iupac("M3")
    # Test complex glycan
    complex_result = oxford_to_iupac("FA2G2")
    assert "Gal" in complex_result
    assert "GlcNAc" in complex_result
    assert "Fuc" in complex_result


def test_canonicalize_composition():
    # Test H5N4F1A2 format
    result = canonicalize_composition("H5N4F1A2")
    assert result["Hex"] == 5
    assert result["HexNAc"] == 4
    assert result["dHex"] == 1
    assert result["Neu5Ac"] == 2
    # Test Hex5HexNAc4Fuc1Neu5Ac2 format
    result = canonicalize_composition("Hex5HexNAc4Fuc1Neu5Ac2")
    assert result["Hex"] == 5
    assert result["HexNAc"] == 4
    assert result["dHex"] == 1
    assert result["Neu5Ac"] == 2
    result = canonicalize_composition("H8N2H1")
    assert result["Hex"] == 9
    # Test 9_2_0_0 format
    result = canonicalize_composition("9_2_0_0")
    assert result["Hex"] == 9
    assert result["HexNAc"] == 2
    result = canonicalize_composition("9200")
    assert result["Hex"] == 9
    assert result["HexNAc"] == 2
    result = canonicalize_composition("9 2 0 0")
    assert result["Hex"] == 9
    assert result["HexNAc"] == 2
    result = canonicalize_composition("9 2 0 1 0")
    assert result["Hex"] == 9
    assert result["HexNAc"] == 2
    assert result["Neu5Gc"] == 1


def test_parse_glycoform():
    # Test string input
    result = parse_glycoform("H5N4F1A2")
    assert result["H"] == 5
    assert result["N"] == 4
    assert result["F"] == 1
    assert result["A"] == 2
    result = parse_glycoform("H9N2")
    assert result["high_Man"] == 1
    result = parse_glycoform("H5N4F2A2")
    assert result["antennary_Fuc"] == 1
    # Test dict input
    result = parse_glycoform({"Hex": 5, "HexNAc": 4, "dHex": 1, "Neu5Ac": 2})
    print(result)
    assert result["complex"] == 1
    assert result["high_Man"] == 0


def test_presence_to_matrix():
    df = pd.DataFrame({
        'glycan': ['Gal(b1-4)GlcNAc', 'Man(a1-3)Man', 'Gal(b1-4)GlcNAc'],
        'Species': ['Human', 'Mouse', 'Human']
    })
    result = presence_to_matrix(df)
    assert result.loc['Human', 'Gal(b1-4)GlcNAc'] == 2
    assert result.loc['Mouse', 'Man(a1-3)Man'] == 1


def test_process_for_glycoshift():
    df = pd.DataFrame(
        index=['PROT1_123_H5N4F1A2', 'PROT1_124_H3N3F1A1'],
        data={'abundance': [1.0, 2.0]}
    )
    result, features = process_for_glycoshift(df)
    assert 'Glycosite' in result.columns
    assert 'Glycoform' in result.columns
    assert 'HexNAc' in features
    assert 'Hex' in features
    df = pd.DataFrame(
        index=['PROT1_123_[H5N4A2F1]', 'PROT1_124_[H3N3A1F1]'],
        data={'abundance': [1.0, 2.0]}
    )
    result, features = process_for_glycoshift(df)


def test_linearcode_to_iupac():
    assert linearcode_to_iupac("Ma3(Ma6)Mb4GNb4GN;") == 'Mana3(Mana6)Manb4GlcNAcb4GlcNAc'


def test_iupac_extended_to_condensed():
    assert iupac_extended_to_condensed("β-D-Galp-(1→4)-β-D-GlcpNAc-(1→") == 'D-Galp(β1→4)D-GlcpNAc'
    assert iupac_extended_to_condensed("α-D-Neup5Ac-(2→3)-β-D-Galp-(1→4)-β-D-GlcpNAc-(1→") == 'D-Neup5Ac(α2→3)D-Galp(β1→4)D-GlcpNAc'


def test_get_class():
    assert get_class("GalNAc(b1-3)GalNAc") == "O"
    assert get_class("GlcNAc(b1-4)GlcNAc") == "N"
    assert get_class("Glc-ol") == "free"
    assert get_class("Glc1Cer") == "lipid"
    assert get_class("Gal(b1-4)Glc") == "lipid/free"
    assert get_class("UnknownGlycan") == ""
    assert get_class("[Glc(b1-3)]Glc(b1-6)Glc") == "repeat"


def test_enforce_class():
    # Test N-glycan
    assert enforce_class("GlcNAc(b1-4)GlcNAc", "N") == True
    assert enforce_class("GalNAc(b1-3)GalNAc", "N") == False
    # Test O-glycan
    assert enforce_class("GalNAc(b1-3)GalNAc", "O") == True
    # Test with confidence override
    assert enforce_class("GalNAc(b1-3)GalNAc", "N", conf=0.9, extra_thresh=0.3) == True
    # Test with mismatch
    assert enforce_class("GalNAc(b1-3)GalNAc", "N") == False
    assert enforce_class("GalNAc(b1-3)GalNAc", "glycoRNA") == False


def test_in_lib():
    library = {'Gal': 0, 'GlcNAc': 1, 'b1-4': 2, 'a2-3': 3, 'Neu5Ac': 4}
    assert in_lib("Gal(b1-4)GlcNAc", library) == True
    assert in_lib("Gal(b1-4)Unknown", library) == False
    assert in_lib("Neu5Ac(a2-3)Gal(b1-4)GlcNAc", library) == True


def test_equal_repeats():
    # Test identical repeats
    assert equal_repeats("Gal(b1-4)GlcNAc(b1-3)", "Gal(b1-4)GlcNAc(b1-3)") == True
    # Test shifted repeats
    assert equal_repeats("GlcNAc(b1-3)Gal(b1-3)", "Gal(b1-3)GlcNAc(b1-3)") == True
    # Test different repeats
    assert equal_repeats("Gal(b1-4)GlcNAc(b1-3)", "Man(a1-3)Man(a1-2)") == False


def test_get_matching_indices2():
    # Test simple brackets
    assert list(get_matching_indices("[abc]", '[', ']')) == [(1, 4, 0)]
    # Test nested brackets
    assert list(get_matching_indices("a[b[c]d]e", '[', ']')) == [(4, 5, 1), (2, 7, 0)]
    # Test multiple brackets
    result = list(get_matching_indices("[a][b][c]", '[', ']'))
    assert len(result) == 3
    assert result[0] == (1, 2, 0)
    # Test unmatched brackets
    assert len(list(get_matching_indices("[a[b", '[', ']'))) == 0
    # Test escaped delimiters
    assert len(list(get_matching_indices(r"[a\[b]", '[', ']'))) == 1
    assert list(get_matching_indices(r"[a\[b]", '[', ']'))[0] == (1, 5, 0)
    # Test extraneous closing delimiter
    assert len(list(get_matching_indices("]abc[", '[', ']'))) == 0
    # Test escaped closing delimiter
    assert len(list(get_matching_indices(r"[a\]b]", '[', ']'))) == 1
    assert list(get_matching_indices(r"[a\]b]", '[', ']'))[0] == (1, 5, 0)


def test_bracket_removal():
    # Test simple branch removal
    assert bracket_removal("Gal(b1-4)[Fuc(a1-3)]GlcNAc") == "Gal(b1-4)GlcNAc"
    # Test nested branch removal
    assert bracket_removal("Neu5Ac(a2-3)[Neu5Ac(a2-6)[Gal(b1-3)]Gal(b1-4)]GlcNAc") == "Neu5Ac(a2-3)GlcNAc"
    # Test multiple branches
    assert bracket_removal("Gal(b1-4)[Fuc(a1-3)][Man(a1-6)]GlcNAc") == "Gal(b1-4)GlcNAc"


def test_check_nomenclature():
    # Test SMILES format (should print warning)
    with pytest.raises(ValueError) as exc_info:
        check_nomenclature("C@C(=O)NC1C(O)OC(CO)C(O)C1O")
    assert "can only convert IUPAC-->SMILES" in str(exc_info.value)
    # Test non-string input
    with pytest.raises(TypeError) as exc_info:
        check_nomenclature(123)
    assert "must be formatted as strings" in str(exc_info.value)
    # Test valid IUPAC format
    assert check_nomenclature("Gal(b1-4)GlcNAc") is None


def test_IUPAC_to_SMILES():
    try:
        # Test basic conversion
        glycans = ["Gal(b1-4)GlcNAc"]
        smiles = IUPAC_to_SMILES(glycans)
        assert isinstance(smiles, list)
        assert len(smiles) == 1
        assert all('@' in s for s in smiles)  # SMILES should contain stereochemistry
        # Test multiple glycans
        glycans = ["Gal(b1-4)GlcNAc", "Man(a1-3)Man"]
        smiles = IUPAC_to_SMILES(glycans)
        assert len(smiles) == 2
        smiles = iupac_to_smiles(glycans)
        # Test string input type
        smiles = IUPAC_to_SMILES("Gal(b1-4)GlcNAc")
        smiles = iupac_to_smiles("F(3)XA2")
        assert '@' in smiles[0]
    except ImportError:
        pytest.skip("glyles package not installed")
    # Mock ImportError for chem dependencies
    import builtins
    original_import = builtins.__import__

    def mock_import(name, *args, **kwargs):
        if name == 'glyles':
            raise ImportError()
        return original_import(name, *args, **kwargs)

    try:
        builtins.__import__ = mock_import
        with pytest.raises(ImportError, match=r"You must install the 'chem' dependencies to use this feature\. Try 'pip install glycowork\[chem\]'\."):
            IUPAC_to_SMILES(["Gal(b1-4)GlcNAc"])
    finally:
        builtins.__import__ = original_import


def test_unwrap():
    # Test nested list flattening
    nested = [[1, 2], [3, 4], [5, 6]]
    assert unwrap(nested) == [1, 2, 3, 4, 5, 6]
    # Test with strings
    nested_str = [['a', 'b'], ['c', 'd']]
    assert unwrap(nested_str) == ['a', 'b', 'c', 'd']
    # Test empty lists
    assert unwrap([[], []]) == []
    # Test single-level list
    assert unwrap([[1, 2, 3]]) == [1, 2, 3]


def test_find_nth():
    # Test basic string search
    assert find_nth("abcabc", "abc", 1) == 0
    assert find_nth("abcabc", "abc", 2) == 3
    # Test when pattern doesn't exist
    assert find_nth("abcabc", "xyz", 1) == -1
    # Test when n is too large
    assert find_nth("abcabc", "abc", 3) == -1
    # Test with glycan-like string
    glycan = "Gal(b1-4)Gal(b1-4)GlcNAc"
    assert find_nth(glycan, "Gal", 1) == 0
    assert find_nth(glycan, "Gal", 2) == 9


def test_find_nth_reverse():
    # Test basic reverse search
    assert find_nth_reverse("abcabc", "abc", 1) == 3
    assert find_nth_reverse("abcabc", "abc", 2) == 0
    # Test with ignore_branches=True
    glycan = "Gal(b1-4)[Fuc(a1-3)]GlcNAc"
    assert find_nth_reverse(glycan, "Gal", 1, ignore_branches=True) == 0
    # Test with nested branches
    glycan = "Neu5Ac(a2-3)[Neu5Ac(a2-6)[Gal(b1-3)]Gal(b1-4)]GlcNAc"
    assert find_nth_reverse(glycan, "Gal", 2, ignore_branches=True) == 26
    # Test missing
    assert find_nth_reverse("hello", "j", 1) == -1


def test_remove_unmatched_brackets():
    # Test balanced brackets
    assert remove_unmatched_brackets("[abc]") == "[abc]"
    # Test unmatched opening bracket
    assert remove_unmatched_brackets("[abc") == "abc"
    # Test unmatched closing bracket
    assert remove_unmatched_brackets("abc]") == "abc"
    # Test nested brackets
    assert remove_unmatched_brackets("[a[bc]]") == "[a[bc]]"
    # Test glycan-like string
    glycan = "Gal(b1-4)[Fuc(a1-3]GlcNAc"  # Missing closing bracket
    result = remove_unmatched_brackets(glycan)
    assert result.count('[') == result.count(']')


def test_hashable_dict():
    # Test __hash__
    d1 = HashableDict({'a': 1, 'b': 2})
    d2 = HashableDict({'b': 2, 'a': 1})
    assert hash(d1) == hash(d2)  # Tests hash() and sorted items
    # Test __eq__
    assert d1 == d2  # Same content, different order
    d3 = {'a': 1, 'b': 2}
    assert not isinstance(d3, HashableDict)
    assert d1.__eq__(d3) is False


def test_reindex():
    # Create test dataframes
    df_old = pd.DataFrame({
        'id': ['a', 'b', 'c'],
        'value': [1, 2, 3]
    })
    df_new = pd.DataFrame({
        'id': ['c', 'a', 'b']
    })
    # Test reindexing
    result = reindex(df_new, df_old, 'value', 'id', 'id')
    assert result == [3, 1, 2]
    df_old = pd.DataFrame({
        'id2': ['a', 'b', 'c'],
        'value': [1, 2, 3]
    })
    result = reindex(df_new, df_old, 'value', 'id2', 'id')


def test_stringify_dict():
    # Test basic dictionary
    d = {'a': 1, 'b': 2, 'c': 3}
    assert stringify_dict(d) == 'a1b2c3'
    # Test empty dictionary
    assert stringify_dict({}) == ''
    # Test dictionary with various types
    d = {'Hex': 5, 'HexNAc': 4, 'Neu5Ac': 2}
    result = stringify_dict(d)
    assert isinstance(result, str)
    assert 'Hex5' in result


def test_replace_every_second():
    # Test basic replacement
    assert replace_every_second("aaa", "a", "b") == "aba"
    # Test with mixed characters
    assert replace_every_second("ababa", "a", "c") == "abcba"
    # Test empty string
    assert replace_every_second("", "a", "b") == ""
    # Test no matches
    assert replace_every_second("xyz", "a", "b") == "xyz"


def test_multireplace():
    # Test basic replacements
    replacements = {'a': 'x', 'b': 'y'}
    assert multireplace("abc", replacements) == "xyc"
    # Test glycan-specific replacements
    replacements = {'Nac': 'NAc', 'AC': 'Ac'}
    assert multireplace("GlcNac", replacements) == "GlcNAc"
    # Test empty string
    assert multireplace("", {'a': 'b'}) == ""
    # Test no matches
    assert multireplace("xyz", {'a': 'b'}) == "xyz"


def test_strip_suffixes():
    # Test numerical suffixes
    cols = ['name.1', 'value.2', 'data.10']
    assert strip_suffixes(cols) == ['name', 'value', 'data']
    # Test mixed columns
    cols = ['name.1', 'value', 'data.10']
    assert strip_suffixes(cols) == ['name', 'value', 'data']
    # Test no suffixes
    cols = ['name', 'value', 'data']
    assert strip_suffixes(cols) == ['name', 'value', 'data']


def test_build_custom_df():
    # Create test DataFrame
    df = pd.DataFrame({
        'glycan': ['Gal', 'GlcNAc'],
        'Species': [['Human', 'Mouse'], ['Mouse']],
        'Genus': [['Homo', 'Mus'], ['Mus']],
        'Family': [['Hominidae', 'Muridae'], ['Muridae']],
        'Order': [['Primates', 'Rodentia'], ['Rodentia']],
        'Class': [['Mammalia', 'Mammalia'], ['Mammalia']],
        'Phylum': [['Chordata', 'Chordata'], ['Chordata']],
        'Kingdom': [['Animalia', 'Animalia'], ['Animalia']],
        'Domain': [['Eukaryota', 'Eukaryota'], ['Eukaryota']],
        'ref': [['ref1', 'ref2'], ['ref3']]
    })
    # Test df_species creation
    result = build_custom_df(df, 'df_species')
    assert len(result) > len(df)  # Should expand due to explode
    assert 'glycan' in result.columns
    assert 'Species' in result.columns
    assert len(df_species[df_species.Genus=='Diceros'].glyco_filter("Neu5Ac(a2-3)[GalNAc(b1-4)]Gal", min_count=2)) > 0
    try:
        result = build_custom_df(df, "wrong")
        return False
    except:
        pass


def test_dataframe_serializer():
    # Create test DataFrame with various types
    df = pd.DataFrame({
        'strings': ['a', 'b', 'c'],
        'lists': [[1, 2], [3, 4], [5, 6]],
        'dicts': [{'x': 1}, {'y': 2}, {'z': 3}],
        'nulls': [None, pd.NA, None],
        'stringified_lists': ["['d', 'e', 'f']", "['d2', 'e2', 'f2']", "['d3', 'e3', 'f3']"],
        'stringified_dicts': ["{'Hex': 1, 'HexNAc': 1}", "{'Hex: 2'}", "{'Neu5Ac': 1}"]
    })
    # Test serialization and deserialization
    serializer = DataFrameSerializer()
    # Test full cycle
    with pytest.raises(Exception):  # Should fail without proper file path
        serializer.serialize(df, "")
    # Test cell serialization
    cell = serializer._serialize_cell([1, 2, 3])
    assert cell['type'] == 'list'
    assert cell['value'] == ['1', '2', '3']
    # Test cell deserialization
    cell_data = {'type': 'list', 'value': ['1', '2', '3']}
    result = serializer._deserialize_cell(cell_data)
    assert result == ['1', '2', '3']
    serializer.serialize(df, "test.json")
    df2 = serializer.deserialize("test.json")
    assert isinstance(df2["stringified_lists"].tolist()[0], list)
    assert isinstance(df2["stringified_dicts"].tolist()[0], dict)


def test_glycan_binding():
    assert len(glycan_binding) > 0


def test_glycomics_data_loader():
    dir(glycomics_data_loader)
    assert len(glycomics_data_loader.human_brain_GSL_PMID38343116) > 0


def test_count_nested_brackets():
    # Basic cases
    assert count_nested_brackets('') == 0
    assert count_nested_brackets('[]') == 0
    assert count_nested_brackets('[a]') == 0
    assert count_nested_brackets('[[]]') == 1
    assert count_nested_brackets('[[][]]') == 2
    assert count_nested_brackets('[a][b]') == 0  # Parallel brackets don't count
    # Complex nested cases
    assert count_nested_brackets('[a[b][c]]') == 2
    assert count_nested_brackets('[a[b[c]]]') == 2
    assert count_nested_brackets('[a][b[c][d]]') == 2
    # Real glycan examples
    glycan1 = '[Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc]'
    assert count_nested_brackets(glycan1) == 1
    glycan2 = '[a[b[c]][d[e]]]'
    assert count_nested_brackets(glycan2) == 4
    # Edge cases
    assert count_nested_brackets('[][][]') == 0
    assert count_nested_brackets('[[[][]]]') == 3


def test_real_glycan_structures():
    s1 = 'Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-4)[Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)]Man(a1-3)[Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc'
    s2 = "Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    # These should have different nested bracket counts
    assert count_nested_brackets(s1) != count_nested_brackets(s2)


def test_cohen_d():
    # Test with clearly different groups
    group1 = [1, 2, 3, 4, 5]
    group2 = [6, 7, 8, 9, 10]
    d, _ = cohen_d(group1, group2)
    assert d < 0  # Effect size should be negative (group1 < group2)
    # Test with paired samples
    d, _ = cohen_d(group1, group2, paired=True)
    assert d < 0  # Should still be negative
    # Test with identical groups
    d, _ = cohen_d(group1, group1)
    assert abs(d) < 1e-10  # Effect size should be approximately 0


def test_mahalanobis_distance():
    # Test with 2D data
    x = np.array([[1, 2], [3, 4], [5, 6]])
    y = np.array([[7, 8], [9, 10], [11, 12]])
    dist = mahalanobis_distance(x, y)
    assert dist > 0  # Distance should be positive
    dist = mahalanobis_distance(pd.DataFrame(x), pd.DataFrame(y))
    # Test with paired data
    dist_paired = mahalanobis_distance(x, y, paired=True)
    assert dist_paired >= 0  # Should still be positive


def test_mahalanobis_variance():
    # Create simple test data
    x = np.array([[1, 2, 3], [4, 5, 6]]).T  # 3 samples, 2 features
    y = np.array([[2, 3, 4], [5, 6, 7]]).T  # 3 samples, 2 features
    # Call the function
    variance = mahalanobis_variance(x, y, paired=False)
    # Check that the output is a float
    assert isinstance(variance, float)
    # Check that the variance is non-negative
    assert variance >= 0
    # Check that the function works with paired=True as well
    variance_paired = mahalanobis_variance(x, y, paired=True)
    assert isinstance(variance_paired, float)
    assert variance_paired >= 0


def test_variance_stabilization():
    # Create test data
    data = pd.DataFrame({
        'sample1': [1, 2, 3],
        'sample2': [4, 5, 6],
        'sample3': [7, 8, 9]
    })
    # Test global normalization
    result = variance_stabilization(data)
    assert (result.std() - 1).abs().mean() < 0.1  # Should have ~unit variance
    # Test group-specific normalization
    groups = [['sample1', 'sample2'], ['sample3']]
    result = variance_stabilization(data, groups)
    assert result.shape == data.shape


def test_biodiversity_metrics():
    # Test data
    counts = np.array([10, 20, 30, 40])
    # Test sequence richness
    assert sequence_richness(counts) == 4
    assert sequence_richness(np.array([1, 0, 3, 0])) == 2
    # Test Shannon diversity
    shannon = shannon_diversity_index(counts)
    assert 0 <= shannon <= np.log(len(counts))
    # Test Simpson diversity
    simpson = simpson_diversity_index(counts)
    assert 0 <= simpson <= 1


def test_get_equivalence_test():
    # Test data
    group1 = np.array([1, 2, 3, 4, 5])
    group2 = np.array([1.1, 2.1, 3.1, 4.1, 5.1])
    # Test unpaired
    p_val = get_equivalence_test(group1, group2)
    assert 0 <= p_val <= 1
    # Test paired
    p_val_paired = get_equivalence_test(group1, group2, paired=True)
    assert 0 <= p_val_paired <= 1


def test_hotellings_t2():
    # Create multivariate test data
    group1 = np.array([[1, 2], [3, 4], [5, 6]])
    group2 = np.array([[7, 8], [9, 10], [11, 12]])
    # Test unpaired
    F_stat, p_val = hotellings_t2(group1, group2)
    assert F_stat >= 0
    assert 0 <= p_val <= 1
    # Test paired
    F_stat_paired, p_val_paired = hotellings_t2(group1, group2, paired=True)
    assert F_stat_paired >= 0
    assert 0 <= p_val_paired <= 1


def test_calculate_permanova_stat():
    # Create distance matrix
    dist_matrix = pd.DataFrame([
        [0, 1, 2, 3],
        [1, 0, 2, 3],
        [2, 2, 0, 1],
        [3, 3, 1, 0]
    ])
    group_labels = ['A', 'A', 'B', 'B']
    # Test calculation
    F_stat = calculate_permanova_stat(dist_matrix, group_labels)
    assert F_stat >= 0


def test_omega_squared():
    # Create test data
    data = np.array([1, 2, 3, 4, 5, 6, 7, 8])
    groups = ['A', 'A', 'A', 'A', 'B', 'B', 'B', 'B']
    # Calculate effect size
    omega_sq = omega_squared(data, groups)
    assert -1 <= omega_sq <= 1  # Omega squared should be between -1 and 1


def test_anosim():
    # Create test distance matrix
    dist_matrix = pd.DataFrame([
        [0, 1, 2, 3],
        [1, 0, 2, 3],
        [2, 2, 0, 1],
        [3, 3, 1, 0]
    ])
    group_labels = ['A', 'A', 'B', 'B']
    # Test ANOSIM
    R, p_val = anosim(dist_matrix, group_labels)
    assert -1 <= R <= 1  # R should be between -1 and 1
    assert 0 <= p_val <= 1


def test_alpha_biodiversity_stats():
    # Create test data
    distances = pd.DataFrame([1.0, 2.0, 1.5, 2.5])
    group_labels = ['A', 'A', 'B', 'B']
    # Test statistics
    result = alpha_biodiversity_stats(distances, group_labels)
    if result:  # Will be None if groups have only 1 sample
        F_stat, p_val = result
        assert F_stat >= 0
        assert 0 <= p_val <= 1


def test_permanova_with_permutation():
    # Create test distance matrix
    dist_matrix = pd.DataFrame([
        [0, 1, 2, 3],
        [1, 0, 2, 3],
        [2, 2, 0, 1],
        [3, 3, 1, 0]
    ])
    group_labels = ['A', 'A', 'B', 'B']
    # Test PERMANOVA
    F_stat, p_val = permanova_with_permutation(dist_matrix, group_labels, permutations=99)
    assert F_stat >= 0
    assert 0 <= p_val <= 1


def test_clr_transformation():
    # Create test data
    data = pd.DataFrame({
        'sample1': [1, 2, 3],
        'sample2': [4, 5, 6],
        'sample3': [7, 8, 9]
    })
    group1 = ['sample1', 'sample2']
    group2 = ['sample3']
    # Test CLR transformation
    result = clr_transformation(data, group1, group2)
    assert result.shape == data.shape
    # Test with scale model
    result_scaled = clr_transformation(data, group1, group2, custom_scale=2.0)
    assert result_scaled.shape == data.shape


def test_alr_transformation():
    # Create test data
    data = pd.DataFrame({
        'sample1': [1, 2, 3],
        'sample2': [4, 5, 6],
        'sample3': [7, 8, 9]
    })
    group1 = ['sample1', 'sample2']
    group2 = ['sample3']
    # Test ALR transformation
    result = alr_transformation(data, 0, group1, group2)
    assert result.shape[0] == data.shape[0] - 1  # One fewer row due to reference
    # Test with scale model
    result_scaled = alr_transformation(data, 0, group1, group2, custom_scale=2.0)
    assert result_scaled.shape[0] == data.shape[0] - 1
    data.columns = ['1_control', '2_control', '3_disease']
    result_scaled = alr_transformation(data, 0, ['1_control', '2_control', '3_disease'], [], custom_scale={'control':2.0, 'disease':1.0})


def test_get_procrustes_scores():
    # Create test data
    data = pd.DataFrame({
        'glycan': ['Gal', 'GlcNAc', 'Man'],
        'sample1': [1, 2, 3],
        'sample2': [4, 5, 6],
        'sample3': [7, 8, 9],
        'sample4': [10, 11, 12]
    })
    group1 = ['sample1', 'sample2']
    group2 = ['sample3', 'sample4']
    # Test Procrustes analysis
    scores, corr, var = get_procrustes_scores(data, group1, group2)
    assert len(scores) == len(corr) == len(var) == data.shape[0]
    scores, corr, var = get_procrustes_scores(data, [1, 2], [3, 4])
    scores, corr, var = get_procrustes_scores(data, group1, group2, paired=True)


def test_get_additive_logratio_transformation():
    # Create test data
    data = pd.DataFrame({
        'glycan': ['Gal', 'GlcNAc', 'Man'],
        'sample1': [1, 2, 3],
        'sample2': [4, 5, 6],
        'sample3': [7, 8, 9],
        'sample4': [10, 11, 12]
    })
    group1 = ['sample1', 'sample2']
    group2 = ['sample3', 'sample4']
    # Test ALR transformation with reference selection
    result = get_additive_logratio_transformation(data, group1, group2)
    assert 'glycan' in result.columns


def test_bayes_factor_functions():
    # Test get_BF
    bf = get_BF(n=20, p=0.05)
    assert bf > 0
    bf = get_BF(n=20, p=0.05, method="balanced")
    # Test get_alphaN
    alpha = get_alphaN(n=20)
    assert 0 < alpha < 1
    alpha = get_alphaN(n=20, method="balanced")


def test_pi0_tst():
    # Create test p-values
    p_values = np.array([0.01, 0.02, 0.03, 0.5, 0.6, 0.7, 0.8, 0.9])
    # Test pi0 estimation
    pi0 = pi0_tst(p_values)
    assert 0 <= pi0 <= 1


def test_tst_grouped_benjamini_hochberg():
    # Create test data
    identifiers = {
        'group1': ['g1', 'g2'],
        'group2': ['g3', 'g4']
    }
    p_values = {
        'group1': [0.01, 0.02],
        'group2': [0.03, 0.04]
    }
    # Test correction
    adj_pvals, sig_dict = TST_grouped_benjamini_hochberg(identifiers, p_values, 0.05)
    assert len(adj_pvals) == len(sig_dict) == 4
    assert all(0 <= p <= 1 for p in adj_pvals.values())
    assert all(isinstance(s, bool) for s in sig_dict.values())


def test_compare_inter_vs_intra_group():
    # Create test data
    cohort_a = pd.DataFrame({
        'sample1': [1, 2, 3],
        'sample2': [4, 5, 6]
    })
    cohort_b = pd.DataFrame({
        'sample3': [7, 8, 9],
        'sample4': [10, 11, 12]
    })
    glycans = ['Gal', 'GlcNAc', 'Man']
    grouped_glycans = {'group1': ['Gal'], 'group2': ['GlcNAc', 'Man']}
    # Test correlation analysis
    intra, inter = compare_inter_vs_intra_group(cohort_b, cohort_a, glycans, grouped_glycans)
    assert -1 <= intra <= 1
    assert -1 <= inter <= 1
    intra, inter = compare_inter_vs_intra_group(cohort_b, cohort_a, glycans, grouped_glycans, paired=True)


def test_correct_multiple_testing():
    # Create test p-values
    p_values = [0.01, 0.02, 0.03, 0.04, 0.05]
    # Test correction
    corrected_pvals, significance = correct_multiple_testing(p_values, 0.05)
    assert len(corrected_pvals) == len(significance) == len(p_values)
    assert all(0 <= p <= 1 for p in corrected_pvals)
    assert all(isinstance(s, bool) for s in significance)
    corrected_pvals, significance = correct_multiple_testing([], 0.05)
    assert len(corrected_pvals) == 0


def test_partial_corr():
    # Create test data
    x = np.array([1, 2, 3, 4, 5])
    y = np.array([2, 4, 6, 8, 10])
    controls = np.array([[1, 2, 3, 4, 5], [2, 3, 4, 5, 6]]).T
    # Test partial correlation
    corr, p_val = partial_corr(x, y, controls)
    assert -1 <= corr <= 1
    assert 0 <= p_val <= 1
    corr, p_val = partial_corr(x, y, np.empty(shape=(0,0)))


def test_estimate_technical_variance():
    # Create test data
    data = pd.DataFrame({
        'sample1': [1, 2, 3],
        'sample2': [4, 5, 6],
        'sample3': [7, 8, 9]
    })
    group1 = ['sample1', 'sample2']
    group2 = ['sample3']
    # Test variance estimation
    result = estimate_technical_variance(data, group1, group2, num_instances=10)
    assert result.shape[1] == (len(group1) + len(group2)) * 10


def test_missforest():
    # Test MissForest imputation
    data = pd.DataFrame({
        'A': [1, np.nan, 3, 4],
        'B': [np.nan, 2, 3, 4],
        'C': [1, 2, np.nan, 4]
    })
    mf = MissForest()
    imputed = mf.fit_transform(data)
    assert not imputed.isnull().any().any()
    assert imputed.shape == data.shape


def test_impute_and_normalize():
    # Test imputation and normalization
    data = pd.DataFrame({
        'glycan': ['Gal', 'GlcNAc', 'Man'],
        'sample1': [1, 0, 3],
        'sample2': [0, 2, 0],
        'sample3': [3, 0, 3]
    })
    groups = [['sample1', 'sample2'], ['sample3']]
    result = impute_and_normalize(data, groups)
    assert result.shape == data.shape
    assert not (result.iloc[:, 1:] == 0).any().any()
    assert abs(result.iloc[:, 1:].sum().sum() - 300) < 1e-5  # Check normalization to 100%
    data.columns = [0] + data.columns.tolist()[1:]
    result = impute_and_normalize(data, groups)


def test_variance_based_filtering():
    # Test variance-based filtering
    data = pd.DataFrame({
        'A': [1, 2, 3],
        'B': [1.1, 1.9, 3.1],
        'C': [2, 2, 2]  # Low variance
    }).T
    filtered, discarded = variance_based_filtering(data, min_feature_variance=0.1)
    assert len(filtered) + len(discarded) == len(data)
    assert len(discarded) == 1  # The constant feature should be discarded


def test_get_glycoform_diff():
    # Test glycoform differential expression
    result_df = pd.DataFrame({
        'Glycosite': ['Prot1_123_H5N4', 'Prot1_123_H6N5', 'Prot2_456_H4N3'],
        'corr p-val': [0.01, 0.02, 0.03],
        'Effect size': [0.5, 0.6, 0.7]
    })
    result = get_glycoform_diff(result_df, alpha=0.05, level='peptide')
    assert 'Glycosite' in result.columns
    assert 'corr p-val' in result.columns
    assert 'significant' in result.columns
    assert 'Effect size' in result.columns
    result = get_glycoform_diff(result_df, alpha=0.05, level='protein')


def test_get_glm():
    # Test GLM fitting
    data = pd.DataFrame({
        'Glycosite': ['Site1'] * 8 + ['Site2'] * 8,
        'Abundance': [1.0, 1.2, 0.8, 1.1, 1.3, 0.9, 1.4, 0.7,  # Site1 abundances
                     2.0, 2.2, 1.8, 2.1, 2.3, 1.9, 2.4, 1.7],  # Site2 abundances
        'Condition': [0, 0, 0, 0, 1, 1, 1, 1] * 2,
        'H': [1, 2, 1, 2, 1, 2, 1, 2] * 2,
        'N': [2, 3, 2, 3, 2, 3, 2, 3] * 2
    })
    model, variables = get_glm(data)
    if not isinstance(model, str):  # If model fitting succeeded
        assert hasattr(model, 'params')
        assert hasattr(model, 'pvalues')
        assert len(variables) > 0


def test_process_glm_results():
    # Test GLM results processing
    data = pd.DataFrame({
        'Glycosite': ['Site1'] * 8 + ['Site2'] * 8,
        'Abundance': [1.0, 1.2, 0.8, 1.1, 1.3, 0.9, 1.4, 0.7,  # Site1 abundances
                     2.0, 2.2, 1.8, 2.1, 2.3, 1.9, 2.4, 1.7],  # Site2 abundances
        'Condition': [0, 0, 0, 0, 1, 1, 1, 1] * 2,
        'H': [1, 2, 1, 2, 1, 2, 1, 2] * 2,
        'N': [2, 3, 2, 3, 2, 3, 2, 3] * 2
    })
    result = process_glm_results(data, 0.05, ['H', 'N'])
    assert 'Condition_coefficient' in result.columns
    assert 'Condition_corr_pval' in result.columns
    assert 'Condition_significant' in result.columns
    assert result.shape[0] == len(['Site1', 'Site2'])


def test_replace_outliers_with_iqr_bounds():
    # Test IQR-based outlier replacement
    data = pd.Series([1, 2, 3, 10, 2, 3, 1])  # 10 is an outlier
    result = replace_outliers_with_IQR_bounds(data)
    assert max(result) < 10  # Outlier should be replaced
    assert len(result) == len(data)
    data = pd.Series(['Neu5Ac', 1, 2, 3, 10, 2, 3, 1])  # 10 is an outlier
    result = replace_outliers_with_IQR_bounds(data)
    data = pd.Series(['Neu5Ac', 10, 20, 30, 1, 20, 30, 10])  # 1 is an outlier
    result = replace_outliers_with_IQR_bounds(data)
    data = pd.Series(['Neu5Ac', 10, 20, 30, -100, 20, 30, 10])  # -100 is an outlier
    result = replace_outliers_with_IQR_bounds(data)


def test_replace_outliers_winsorization():
    # Test Winsorization-based outlier replacement
    data = pd.Series([1, 2, 3, 10, 2, 3, 1, 2, 3, 1])  # 10 is an outlier
    result = replace_outliers_winsorization(data)
    assert max(result) < 10  # Outlier should be replaced
    assert len(result) == len(data)
    result = replace_outliers_winsorization(data, cap_side='lower')
    result = replace_outliers_winsorization(data, cap_side='upper')
    try:
        result = replace_outliers_winsorization(data, cap_side='wrong')
        return False
    except:
        pass


def test_perform_tests_monte_carlo():
    # Test Monte Carlo testing
    group_a = pd.DataFrame(np.random.normal(0, 1, (5, 100)))
    group_b = pd.DataFrame(np.random.normal(0.5, 1, (5, 100)))
    raw_p, adj_p, effect = perform_tests_monte_carlo(
        group_a, group_b, num_instances=10
    )
    assert len(raw_p) == len(adj_p) == len(effect) == 5
    assert all(0 <= p <= 1 for p in raw_p)
    assert all(0 <= p <= 1 for p in adj_p)


def test_evaluate_adjacency():
    assert evaluate_adjacency("(", 0)
    assert evaluate_adjacency(")", 0)
    assert not evaluate_adjacency("(1)[", 0)
    assert not evaluate_adjacency(")[", 0)


def test_glycan_to_graph():
    glycan = "Gal(b1-4)GlcNAc"
    node_dict, adj_matrix = glycan_to_graph(glycan)
    assert len(node_dict) == 3 # Gal, b1-4, GlcNAc
    assert adj_matrix.shape == (3, 3)
    assert adj_matrix[0, 1] == 1  # Gal connected to linkage
    assert adj_matrix[1, 2] == 1  # Linkage connected to GlcNAc


def test_glycan_to_nxGraph():
    glycan = "Gal(b1-4)GlcNAc"
    graph = glycan_to_nxGraph(glycan)
    assert len(graph.nodes) == 3
    assert len(graph.edges) == 2
    assert all('string_labels' in data for _, data in graph.nodes(data=True))


def test_compare_glycans():
    # Test identical glycans
    assert compare_glycans("Gal(b1-4)GlcNAc", "Gal(b1-4)GlcNAc")
    # Test order-specificity of comparison
    assert not compare_glycans("Gal(b1-4)GlcNAc", "GlcNAc(b1-4)Gal")
    # Test non-equivalent glycans
    assert not compare_glycans("Gal(b1-4)GlcNAc", "Man(a1-3)GlcNAc")
    # Test with PTM wildcards
    assert compare_glycans("Gal6S(b1-4)GlcNAc", "GalOS(b1-4)GlcNAc")
    assert compare_glycans("Gal6S(b1-3)GalNAc4S", "GalOS(b1-3)GalNAc4/6S")
    res, mappy = compare_glycans('Fuc(a1-2)Gal(b1-4)GlcNAc6S(b1-6)[Neu5Ac(a2-3)Gal(b1-3)]GalNAc', graph_to_string(glycan_to_nxGraph('Fuc(a1-2)Gal(b1-4)GlcNAc6S(b1-6)[Neu5Ac(a2-3)Gal(b1-3)]GalNAc'), order_by='linkage'), return_matches=True)


def test_subgraph_isomorphism():
    # Test motif presence
    assert subgraph_isomorphism("Gal(b1-4)GlcNAc", "GlcNAc")
    # Test motif count
    assert subgraph_isomorphism("Gal(b1-4)GlcNAc(b1-4)GlcNAc", "GlcNAc", count=True) == 2
    # Test with negation
    assert subgraph_isomorphism("Gal(b1-4)GlcNAc", "!Man(a1-3)GlcNAc")
    # Test with termini constraints
    assert subgraph_isomorphism("Gal(b1-4)GlcNAc", "GlcNAc", termini_list=['terminal'])
    assert subgraph_isomorphism(glycan_to_nxGraph("Gal(b1-4)GlcNAc"), glycan_to_nxGraph("GlcNAc"), termini_list=['terminal'])
    # Test with narrow linkage ambiguity
    assert subgraph_isomorphism("Gal(b1-4)GlcNAc(b1-6)[Gal(b1-3)]GalNAc", "Gal(b1-3/4)GlcNAc") == True
    assert subgraph_isomorphism("Gal(a1-4)GlcNAc(b1-6)[Gal(b1-3)]GalNAc", "Gal(b1-3/4)GlcNAc") == False
    assert subgraph_isomorphism("Gal(b1-?)GlcNAc(b1-6)[Gal(b1-3)]GalNAc", "Gal(b1-3/4)GlcNAc") == True


def test_generate_graph_features():
    glycan = "Gal(b1-4)GlcNAc"
    features = generate_graph_features(glycan)
    assert isinstance(features, pd.DataFrame)
    assert len(features) == 1
    assert 'diameter' in features.columns
    assert 'branching' in features.columns
    assert 'avgDeg' in features.columns
    features = generate_graph_features("GalNAc", glycan_to_nxGraph("GalNAc"))


def test_largest_subgraph():
    glycan1 = "Gal(b1-4)GlcNAc(b1-4)GlcNAc"
    glycan2 = "Man(a1-3)GlcNAc(b1-4)GlcNAc"
    result = largest_subgraph(glycan1, glycan2)
    assert "GlcNAc(b1-4)GlcNAc" in result


def test_get_possible_topologies():
    glycan = "{Neu5Ac(a2-?)}Gal(b1-3)GalNAc"
    topologies = get_possible_topologies(glycan)
    topologies = get_possible_topologies(glycan, return_graphs=True)
    assert len(topologies) > 0
    assert all(isinstance(g, nx.Graph) for g in topologies)
    assert get_possible_topologies("{6S}Neu5Ac(a2-3)Gal(b1-3)[Neu5Ac(a2-6)]GalNAc")[0] == "Neu5Ac(a2-3)Gal6S(b1-3)[Neu5Ac(a2-6)]GalNAc"
    assert get_possible_topologies("{OS}Gal(b1-3)GalNAc") == ['GalOS(b1-3)GalNAc', 'Gal(b1-3)GalNAcOS']
    with pytest.raises(ValueError, match="This glycan already has a defined topology; please don't use this function."):
        get_possible_topologies("Gal(b1-4)GlcNAc")


def test_deduplicate_glycans():
    glycans = [
        "Gal(b1-4)[Fuc(a1-3)]GlcNAc",
        "Fuc(a1-3)[Gal(b1-4)]GlcNAc",
        "Man(a1-3)GlcNAc"
    ]
    result = deduplicate_glycans(glycans)
    assert len(result) == 2
    assert deduplicate_glycans(["Man(a1-3)GlcNAc"]) == ["Man(a1-3)GlcNAc"]


def test_neighbor_is_branchpoint():
    # Create test graph
    G = nx.Graph()
    G.add_nodes_from(range(6))
    G.add_edges_from([(0,1), (1,2), (2,3), (2,4), (2,5)])
    assert neighbor_is_branchpoint(G, 1)  # Node 1 connected to branch point 2
    assert not neighbor_is_branchpoint(G, 3)  # Node 3 is leaf


def test_try_string_conversion():
    assert try_string_conversion(glycan_to_nxGraph("Neu5Ac(a2-3)Gal(b1-3)[Neu5Ac(a2-6)]GalNAc")) == "Neu5Ac(a2-3)Gal(b1-3)[Neu5Ac(a2-6)]GalNAc"


def test_subgraph_isomorphism_with_negation():
    glycan = "Gal(b1-4)GlcNAc"
    motif = "!Man(a1-3)GlcNAc"
    # Should match since glycan doesn't contain Man(a1-3)GlcNAc
    assert subgraph_isomorphism_with_negation(glycan, motif)
    # Test with counting
    result, matches = subgraph_isomorphism_with_negation(glycan, motif, count=True, return_matches=True)
    assert isinstance(result, int)
    assert isinstance(matches, list)
    assert subgraph_isomorphism_with_negation("Neu5Ac(a2-3)Gal(b1-3)[Fuc(a1-4)]GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc", '!Neu5Ac(a2-?)Gal(?1-?)GlcNAc', return_matches=True) == (1, [[12, 13, 14]])
    assert subgraph_isomorphism_with_negation("Neu5Ac(a2-3)Gal(b1-3)[Fuc(a1-4)]GlcNAc(b1-2)Man(a1-3)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc", 'Neu5Ac(a2-?)!Gal(b1-4)GlcNAc', return_matches=True) == (1, [[0, 1, 2, 3, 6]])


def test_categorical_node_match_wildcard():
    narrow_wildcard_list = {"Hex": ["Gal", "Glc", "Man"]}
    matcher = categorical_node_match_wildcard(
        'string_labels', 'unknown', narrow_wildcard_list, 'termini', 'flexible'
    )
    # Test wildcard matching
    assert matcher(
        {'string_labels': 'Hex', 'termini': 'terminal'},
        {'string_labels': 'Gal', 'termini': 'terminal'}
    )
    # Test exact matching
    assert matcher(
        {'string_labels': 'GlcNAc', 'termini': 'terminal'},
        {'string_labels': 'GlcNAc', 'termini': 'terminal'}
    )
    # Test flexible position
    assert matcher(
        {'string_labels': 'GlcNAc', 'termini': 'flexible'},
        {'string_labels': 'GlcNAc', 'termini': 'terminal'}
    )


def test_expand_termini_list():
    motif = "Gal(b1-4)GlcNAc"
    termini_list = ['t', 'i']  # terminal, internal
    result = expand_termini_list(motif, termini_list)
    assert len(result) == 3  # Two monosaccharides + one linkage
    assert 'terminal' in result
    assert 'internal' in result
    assert 'flexible' in result


def test_ensure_graph():
    # Test with string input
    glycan = "Gal(b1-4)GlcNAc"
    result = ensure_graph(glycan)
    assert isinstance(result, nx.Graph)
    # Test with graph input
    G = nx.Graph()
    G.add_node(0, string_labels="Gal")
    result = ensure_graph(G)
    assert result is G


def test_possible_topology_check():
    float_glycan = "{Neu5Ac(a2-?)}Gal(b1-4)GlcNAc"
    glycans = [
        "Neu5Ac(a2-3)Gal(b1-4)GlcNAc",
        "Neu5Ac(a2-6)Gal(b1-4)GlcNAc",
        "Gal(b1-4)GlcNAc"
    ]
    matches = possible_topology_check(float_glycan, glycans)
    assert len(matches) == 2  # Should match first two glycans
    assert "Gal(b1-4)GlcNAc" not in matches


def test_annotate_glycan():
    glycan = "Gal(b1-4)GlcNAc"
    result = annotate_glycan(glycan)
    assert isinstance(result, pd.DataFrame)
    assert len(result) == 1
    assert result.index[0] == glycan


@pytest.fixture
def test_glycans():
    return ['Gal(a1-4)Gal(b1-?)GlcNAc(b1-2)Man(a1-6)[Gal(b1-?)GlcNAc(b1-2)Man(a1-3)]Man(b1-4)GlcNAc(b1-4)GlcNAc', 'Fuc(a1-2)Gal(b1-4)GlcNAc(b1-3)[Fuc(a1-2)]Gal(b1-3)GlcNAc(b1-6)[Fuc(a1-2)Gal(b1-3)]GalNAc',
               'Gal(b1-4)GlcNAc(b1-2)Man(a1-?)[GlcNAc(b1-2)Man(a1-?)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc', 'Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-3)[Gal(b1-3)GlcNAc(b1-6)]Gal(b1-4)Glc-ol',
               'Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)[Neu5Ac(a2-6)Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc',
               'Glc(a1-3)Man(a1-2)[GlcNAc(b1-6)]Man6P(a1-2)Man(a1-3)[Man(a1-2)Man(a1-3)[Man(a1-2)Man6P(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc',
               'Neu5Ac(a2-3)[GalNAc(b1-4)]Gal(b1-4)GlcNAc(b1-6)[Gal(b1-3)]GalNAc', 'Neu5Ac(a2-?)GalNAc(b1-4)GlcNAc(b1-2)Man(a1-3)[Man(a1-?)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc',
               'Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)[GlcNAc(b1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc', 'Fuc2S4S(a1-3)Fuc2S4S(a1-3)Fuc2S4S',
               'Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)GlcNAc(b1-3)Gal(b1-3)[Fuc(a1-2)Gal(b1-4)GlcNAc(b1-6)]GalNAc', 'Gal(b1-3)GalN2Suc-ol'
               'Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)GlcNAc(b1-4)[Gal(b1-4)GlcNAc(b1-2)]Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc',
               'Gal(a1-3)Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Gal(a1-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc',
               'Gal(b1-?)GlcNAc(b1-3)Gal(b1-?)[Fuc(a1-?)]GlcNAc6S(b1-3)[GlcNAc(b1-6)]GalNAc', 'Neu5Gc(a2-?)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)[GlcNAc(?1-?)GlcNAc(b1-2)Man(a1-3)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc',
               'Neu5Ac(a2-?)GalNAc(b1-4)GlcNAc(b1-2)Man(a1-?)[Man(a1-?)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc', 'ManOP(a1-2)Man(a1-3)[Man(a1-3)[Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc',
               'Gal(b1-4)GlcNAc(b1-6)[Gal3S(b1-3)]GalNAc', 'Neu5Ac(a2-3)Gal(b1-2)GlcNAc(b1-4)Man(a1-?)[Gal(b1-4)GlcNAc(b1-2)Man(a1-?)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc']


def test_size_branching_features(test_glycans):
    result = get_size_branching_features(test_glycans, n_bins=3)
    assert result.shape[1] == 6
    result = get_size_branching_features(test_glycans, n_bins=0)
    assert result.shape[1] == 2


def test_annotate_dataset(test_glycans):
    few_glycans = ["Gal(b1-4)GlcNAc", "Man(a1-3)GlcNAc"]
    # Test known motifs
    result = annotate_dataset(few_glycans, feature_set=['known', 'chemical'])
    assert isinstance(result, pd.DataFrame)
    assert len(result) == 2
    result = annotate_dataset(few_glycans+["{Fuc(a1-?)}Gal(b1-4)GlcNAc"], feature_set=['known'])
    assert isinstance(result, pd.DataFrame)
    assert len(result) == 3
    # Test graph features
    result = annotate_dataset(few_glycans, feature_set=['graph'])
    assert 'diameter' in result.columns
    assert 'branching' in result.columns
    result = annotate_dataset(test_glycans, feature_set=['size_branch'])
    assert result.shape[1] == 6
    result = annotate_dataset(test_glycans, feature_set=['custom', 'terminal2', 'terminal3'], custom_motifs=["Gal(b1-4)GlcNAc"])
    few_glycans = ["LacNAc", "Ma3(Ma6)Mb4GNb4GN;"]
    result = annotate_dataset(few_glycans, feature_set=['exhaustive', 'wrong'])


def test_get_molecular_properties():
    try:
        glycans = ["Gal(b1-4)GlcNAc", "Neu4Ac5Ac7Ac9Ac(a2-6)Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc"]
        result = get_molecular_properties(glycans, verbose=True, placeholder=True)
        assert isinstance(result, pd.DataFrame)
        assert 'molecular_weight' in result.columns
        assert 'xlogp' in result.columns
    except ImportError:
        pytest.skip("Skipping test due to missing dependencies")


def test_get_k_saccharides():
    glycans = ["Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"]
    # Test basic functionality
    result = get_k_saccharides(glycans, size=2)
    assert isinstance(result, pd.DataFrame)
    # Test with up_to=True
    result = get_k_saccharides(glycans, size=2, up_to=True)
    assert isinstance(result, pd.DataFrame)
    assert len(result.columns) > 0
        # Simple glycan
    result = get_k_saccharides(["Gal(b1-4)Gal(b1-4)GlcNAc"], just_motifs=True)[0]
    assert "Gal(b1-4)GlcNAc" in result
    # Branched glycan
    result = get_k_saccharides(["Gal(b1-4)[Fuc(a1-3)]GlcNAc"], just_motifs=True)[0]
    assert "Gal(b1-4)GlcNAc" in result
    assert "Fuc(a1-3)GlcNAc" in result
    result = get_k_saccharides(["Fuc(a1-2)Gal(b1-3)[Fuc(a1-2)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-6)]GalNAc"], just_motifs=True)[0]
    assert "Fuc(a1-3)GlcNAc" in result
    assert "Gal(b1-4)Fuc" not in result
    assert "Gal(b1-3)GalNAc" in result
    assert "Gal(b1-3)Fuc" not in result
    result = get_k_saccharides(["Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"], just_motifs=True)[0]
    assert "Fuc(a1-3)GlcNAc" in result
    assert "Man(a1-3)Man" in result
    assert "Man(a1-3)GlcNAc" not in result
    assert "Gal(b1-4)Fuc" not in result
    assert "GlcNAc(b1-2)Neu5Ac" not in result
    assert not any('[' in k for k in result)
    result = get_k_saccharides(["GlcNAc(b1-2)[GlcNAc(b1-4)][GlcNAc(b1-6)]GalNAc"], just_motifs=True)[0]
    assert "GlcNAc(b1-2)GalNAc" in result
    assert "GlcNAc(b1-4)GalNAc" in result
    assert "GlcNAc(b1-6)GalNAc" in result
    assert "GlcNAc(b1-2)GlcNAc" not in result
    assert get_k_saccharides(['Gal(b1-3)GalNAc'], size=3, just_motifs=True) == []
    few_glycans = ["LacNAc", "Ma3(Ma6)Mb4GNb4GN;"]
    result = get_k_saccharides(few_glycans, just_motifs=True)
    try:
        get_k_saccharides(pd.DataFrame(), just_motifs=True)
        return False
    except TypeError:
        pass


def test_get_terminal_structures():
    glycan = "Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"
    # Test size=1
    result = get_terminal_structures(glycan, size=1)
    assert isinstance(result, list)
    assert "Man(b1-4)" in result[0]
    # Test size=2
    result = get_terminal_structures(glycan, size=2)
    assert isinstance(result, list)
    assert "Man(b1-4)GlcNAc(b1-4)" in result
    try:
        result = get_terminal_structures(glycan, size=3)
        return False
    except ValueError:
        pass


def test_create_correlation_network():
    data = pd.DataFrame({
        'glycan1': [1, 2, 3],
        'glycan2': [1, 2, 3],
        'glycan3': [-1, -2, -3]
    })
    clusters = create_correlation_network(data, 0.9)
    assert isinstance(clusters, list)
    assert isinstance(clusters[0], set)
    assert len(clusters) > 0


def test_group_glycans_core():
    glycans = ["GlcNAc(b1-6)GalNAc", "Gal(b1-3)GalNAc"]
    p_values = [0.01, 0.02]
    glycan_groups, _ = group_glycans_core(glycans, p_values)
    assert "core2" in glycan_groups
    assert "core1" in glycan_groups
    assert len(glycan_groups["core2"]) == 1
    assert len(glycan_groups["core1"]) == 1


def test_group_glycans_sia_fuc():
    glycans = ["Neu5Ac(a2-3)Gal", "Fuc(a1-3)GlcNAc", "Gal(b1-4)GlcNAc", "Fuc(a1-2)[Neu5Ac(a2-3)]Gal"]
    p_values = [0.01, 0.02, 0.03, 0.01]
    glycan_groups, _ = group_glycans_sia_fuc(glycans, p_values)
    assert "Sia" in glycan_groups
    assert "Fuc" in glycan_groups
    assert "rest" in glycan_groups


def test_group_glycans_N_glycan_type():
    glycans = [
        "Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc",  # complex
        "Man(a1-2)Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc",  # high mannose
        "Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Man(a1-3)[Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc",  # hybrid
        "Man(b1-4)GlcNAc(b1-4)GlcNAc"  # other
    ]
    p_values = [0.01, 0.02, 0.03, 0.01]
    glycan_groups, _ = group_glycans_N_glycan_type(glycans, p_values)
    assert "complex" in glycan_groups
    assert "high_man" in glycan_groups
    assert "rest" in glycan_groups


def test_Lectin():
    lectin = Lectin(
        abbr=["WGA"],
        name=["Wheat Germ Agglutinin"],
        specificity={
            "primary": {"GlcNAc": []},
            "secondary": {},
            "negative": {}
        }
    )

    # Test basic attributes
    assert lectin.abbr == ["WGA"]
    assert lectin.name == ["Wheat Germ Agglutinin"]
    # Test binding check
    result = lectin.check_binding("GlcNAc(b1-4)GlcNAc")
    assert result in [0, 1, 2]
    result = lectin.get_all_binding_motifs()
    lectin.show_info()


def test_load_lectin_lib():
    lectin_lib = load_lectin_lib()
    assert isinstance(lectin_lib, dict)
    assert all(isinstance(v, Lectin) for v in lectin_lib.values())


def test_create_lectin_and_motif_mappings():
    lectin_lib = load_lectin_lib()
    lectin_list = ["WGA", "ConA", "wrong"]
    lectin_mapping, motif_mapping = create_lectin_and_motif_mappings(
        lectin_list, lectin_lib
    )
    assert isinstance(lectin_mapping, dict)
    assert isinstance(motif_mapping, dict)

def test_lectin_motif_scoring():
    lectin_lib = load_lectin_lib()
    lectin_mapping = {"WGA": 0, "ConA": 1}
    motif_mapping = {"GlcNAc": {"WGA": 0, "ConA": 1}}
    lectin_scores = {"WGA": 1.0, "ConA": -0.5}
    idf = {"WGA": 1.0, "ConA": 1.0}
    result = lectin_motif_scoring(
        lectin_mapping, motif_mapping, lectin_scores,
        lectin_lib, idf
    )
    assert isinstance(result, pd.DataFrame)
    assert "motif" in result.columns
    assert "score" in result.columns


def test_deduplicate_motifs():
    data = pd.DataFrame({
        'sample1': [1, 1],
        'sample2': [2, 2]
    }, index=['motif1', 'motif2'])
    result = deduplicate_motifs(data)
    assert isinstance(result, pd.DataFrame)
    assert len(result) <= len(data)


def test_quantify_motifs():
    df = pd.DataFrame({
        'sample1': [1, 2],
        'sample2': [2, 3]
    })
    glycans = ["Gal(b1-4)GlcNAc", "Man(a1-3)GlcNAc"]
    result = quantify_motifs(
        df, glycans, feature_set=['exhaustive'],
        remove_redundant=True
    )
    assert isinstance(result, pd.DataFrame)
    assert len(result.columns) > 0


def test_count_unique_subgraphs_of_size_k():
    glycan = "Man(a1-2)Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    graph = glycan_to_nxGraph(glycan)
    # Test size 2
    result = count_unique_subgraphs_of_size_k(graph, size=2)
    assert isinstance(result, dict)
    assert len(result) > 0
    # Test terminal only
    result = count_unique_subgraphs_of_size_k(graph, size=2, terminal=True)
    assert isinstance(result, dict)


def test_annotate_glycan_topology_uncertainty():
    glycan = "{Neu5Ac(a2-?)}Fuc(a1-3)[Gal(b1-4)]GlcNAc"
    result = annotate_glycan_topology_uncertainty(glycan)
    assert isinstance(result, pd.DataFrame)
    assert len(result) == 1
    assert result.index[0] == glycan


def test_preprocess_pattern():
    # Test basic pattern splitting
    assert preprocess_pattern("Hex-HexNAc") == ["Hex-HexNAc"]
    # Test pattern with quantifiers
    assert preprocess_pattern("Hex-[HexNAc|Fuc]{1,2}-Gal") == ["Hex", "[HexNAc|Fuc]{1,2}", "Gal"]
    # Test wildcard replacement
    assert preprocess_pattern("Hex-.-Gal") == ['Hex-Monosaccharide-Gal']
    # Test empty pattern
    assert preprocess_pattern("") == []


def test_specify_linkages():
    # Test basic linkage conversion
    assert specify_linkages("Gala6(?1-?)Gal") == "Gal(a1-6)Gal"
    assert specify_linkages("Manb4(?1-?)GlcNAc") == "Man(b1-4)GlcNAc"
    # Test sialic acid special cases
    assert specify_linkages("Neu5Ac(a1-?)Gal") == "Neu5Ac(a2-?)Gal"
    assert specify_linkages("Sia(a1-?)Gal") == "Sia(a2-?)Gal"
    # Test no modification needed
    assert specify_linkages("Gal(b1-4)GlcNAc") == "Gal(b1-4)GlcNAc"
    assert specify_linkages("Man(a1-6)Man") == "Man(a1-6)Man"


def test_process_occurrence():
    # Test single number
    assert process_occurrence("2") == [2, 2]
    # Test range
    assert process_occurrence("1,3") == [1, 3]
    # Test open-ended range
    assert process_occurrence(",3") == [0, 3]
    assert process_occurrence("2,") == [2, 5]


def test_convert_pattern_component():
    # Test simple component
    result = convert_pattern_component("Hex")
    assert isinstance(result, str)
    # Test component with alternatives
    result = convert_pattern_component("[Hex|Fuc]{1}")
    assert isinstance(result, dict)
    assert all(isinstance(k, str) for k in result.keys())
    assert all(isinstance(v, list) for v in result.values())
    # Test component with quantifier
    result = convert_pattern_component("[Hex|Fuc]{1,2}")
    assert isinstance(result, dict)
    assert list(result.values())[0] == [1, 2]
    result = convert_pattern_component("([Hex|Fuc])*")
    assert isinstance(result, dict)
    assert list(result.values())[0] == [0, 1, 2, 3, 4, 5, 6, 7]
    result = convert_pattern_component("([Hex|Fuc])?")
    assert isinstance(result, dict)
    assert list(result.values())[0] == [0, 1]
    result = convert_pattern_component("([Hex|Fuc])+")
    assert isinstance(result, dict)
    assert list(result.values())[0] == [1, 2, 3, 4, 5, 6, 7]
    result = convert_pattern_component("[Hex|Fuc]?{1}")
    assert isinstance(result, dict)
    assert convert_pattern_component(".-.") == ".(?1-?)."


def test_reformat_glycan_string():
    # Test basic reformatting
    assert reformat_glycan_string("Gal(b1-4)GlcNAc") == "Galb4-GlcNAc"
    # Test with branch
    input_glycan = "Gal(b1-4)GlcNAc(b1-2)[Fuc(a1-6)]Man"
    result = reformat_glycan_string(input_glycan)
    assert "([" in result and "]){1}" in result
    # Test with uncertain linkage
    assert reformat_glycan_string("Gal(b1-?)GlcNAc") == "Galb?-GlcNAc"


def test_get_match():
    # Test exact match
    pattern = "Hex-HexNAc"
    glycan = "Gal(b1-4)GlcNAc"
    assert get_match(pattern, glycan, return_matches=False) is True
    # Test no match
    pattern = "Hex-Hex-Hex"
    glycan = "Gal(b1-4)GlcNAc"
    assert get_match(pattern, glycan, return_matches=False) is False
    # Test no match with positions
    pattern = "'^HexNAc-HexNAc"
    glycan = "Gal(b1-4)GlcNAc(b1-4)GlcNAc"
    assert get_match(pattern, glycan, return_matches=False) is False
    # Test pattern with quantifier
    pattern = "Hex-[HexNAc]{1,2}"
    glycan = "Gal(b1-4)GlcNAc(b1-4)GlcNAc"
    assert get_match(pattern, glycan, return_matches=False) is True
    # Test graph as input
    pattern = "Hex-[HexNAc]{1,2}"
    glycan = "Gal(b1-4)GlcNAc(b1-4)GlcNAc"
    assert get_match(pattern, glycan_to_nxGraph(glycan), return_matches=False) is True
    # Test other nomenclature as input
    pattern = "Hex-[HexNAc]{1,2}"
    assert get_match(pattern, "β-D-Galp-(1→4)-β-D-GlcpNAc-(1→", return_matches=False) is True
    # Complex cases
    pattern = "r[Sia]{,1}-Monosaccharide-([dHex]){,1}-Monosaccharide(?=-Mana6-Monosaccharide)"
    result = get_match(pattern, "GalNAc(b1-4)GlcNAc(b1-2)Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc")
    assert result == ['Gal(b1-4)GlcNAc']
    result = get_match(pattern, "GalNAc(b1-4)GlcNAc(b1-2)Man(a1-3)[GalNAc(b1-4)GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc")
    assert result == ['GalNAc(b1-4)GlcNAc']
    result = get_match(pattern, "GalNAc(b1-4)GlcNAc(b1-2)Man(a1-3)[Neu5Ac(a2-6)GalNAc(b1-4)GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc")
    assert result == ['Neu5Ac(a2-6)GalNAc(b1-4)GlcNAc']
    result = get_match(pattern, "GalNAc(b1-4)GlcNAc(b1-2)Man(a1-3)[Neu5Gc(a2-6)GalNAc(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc")
    assert result == ['Neu5Gc(a2-6)GalNAc(b1-4)[Fuc(a1-3)]GlcNAc']
    pattern = "r[Sia]{,1}-[.]{,1}-([dHex]){,1}-.b3(?=-GalNAc)"
    assert get_match(pattern, "Gal(b1-4)GlcNAc(b1-6)[Gal(b1-3)]GalNAc") == ['Gal']
    assert get_match(pattern, "Neu5Gc(a2-3)Gal(b1-3)[Gal(b1-4)GlcNAc(b1-6)]GalNAc") == ["Neu5Gc(a2-3)Gal"]
    assert get_match("Hex-HexNAc-([Hex|Fuc])?-HexNAc", "GalNAc(b1-4)GlcNAc(b1-2)Man(a1-3)[Neu5Gc(a2-6)GalNAc(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc") == ["Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"]
    assert get_match("Hex-HexNAc-([Hex|Fuc])*-HexNAc", "GalNAc(b1-4)GlcNAc(b1-2)Man(a1-3)[Neu5Gc(a2-6)GalNAc(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc") == ["Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc"]
    assert get_match(".-.-([Hex|Fuc])+-.", "GalNAc(b1-4)GlcNAc(b1-2)Man(a1-3)[Neu5Gc(a2-6)GalNAc(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc") == ['Neu5Gc(a2-6)GalNAc(b1-4)[Fuc(a1-3)]GlcNAc', 'Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc']
    assert get_match("Fuc-Galb3/4-([Hex|Fuc])*-HexNAc", "Fuc(a1-2)Gal(b1-?)[Fuc(a1-?)]GlcNAc(b1-6)[Gal(b1-3)]GalNAc") == ["Fuc(a1-2)Gal(b1-?)[Fuc(a1-?)]GlcNAc"]
    assert get_match("Fuc-([^Gal])+-GlcNAc", "Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc") == ['Fuc(a1-3)[Gal(b1-4)]GlcNAc']
    assert get_match(".-HexNAc$", "Neu5Ac(a2-3)Gal(b1-3)[Fuc(a1-4)]GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc") == ['GlcNAc(b1-4)GlcNAc']
    assert get_match("!Neu5Ac-Gal-GlcNAc", "Neu5Ac(a2-3)Gal(b1-3)[Fuc(a1-4)]GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc") == ["Gal(b1-4)GlcNAc"]
    assert get_match("Hex-HexNAc-([Hex|Fuc]){1,2}-HexNAc(?=-HexNAc)", "Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc(b1-4)GlcNAc") == ['Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc']
    assert get_match("Hex-HexNAc-([Hex|Fuc]){1,2}-HexNAc(?=-HexNAc)", "Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)][Fuc(a1-6)]GlcNAc(b1-4)GlcNAc") == ['Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)][Fuc(a1-6)]GlcNAc']
    assert get_match("Hex-HexNAc-([Hex|Fuc]){1,2}-HexNAc(?!-HexNAc)", "Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)][Fuc(a1-6)]GlcNAc") == ['Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)][Fuc(a1-6)]GlcNAc']
    assert get_match("(?<=Xyl-)Hex-HexNAc-([Hex|Fuc]){1,2}-HexNAc", "Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)][Xyl(b1-2)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)]GlcNAc") == ['Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)]GlcNAc']
    assert get_match("(?<!Xyl-)Hex-HexNAc-([Hex|Fuc]){1,2}-HexNAc", "Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)]GlcNAc") == ['Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)]GlcNAc']
    # Cases where no match is present
    assert get_match("Fuc-([^Gal])+-GlcNAc", "Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Fuc(a1-2)Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc") == []
    assert get_match("Man-HexNAc", "Gal(b1-4)GlcNAc(b1-6)[Gal(b1-3)]GalNAc") == []
    assert get_match("Hex-HexNAc-([Hex|Fuc]){1,2}-HexNAc(?=-HexNAc)", "Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc") == []
    assert get_match("Hex-HexNAc-([Hex|Fuc]){1,2}-HexNAc(?!-HexNAc)", "Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)][Fuc(a1-6)]GlcNAc(b1-4)GlcNAc") == []
    assert get_match("(?<=Xyl-)Hex-HexNAc-([Hex|Fuc]){1,2}-HexNAc", "Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)]GlcNAc") == []
    assert get_match("(?<!Xyl-)Hex-HexNAc-([Hex|Fuc]){1,2}-HexNAc", "Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)][Xyl(b1-2)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)]GlcNAc") == []


def test_motif_to_regex():
    # Test simple motif conversion
    motif = "Gal(b1-4)GlcNAc"
    pattern = motif_to_regex(motif)
    assert isinstance(pattern, str)
    assert get_match(pattern, motif, return_matches=False) is True
    # Test motif with branch
    motif = "Gal(b1-4)GlcNAc(b1-2)[Fuc(a1-6)]Man"
    pattern = motif_to_regex(motif)
    assert isinstance(pattern, str)
    assert get_match(pattern, motif, return_matches=False) is True


def test_compile_pattern():
    assert compile_pattern("r[Sia]{,1}-[.]{,1}-([dHex]){,1}-.b3(?=-GalNAc)") == ['[Sia]{,1}', '[Monosaccharide]{,1}', '([dHex]){,1}', 'Monosaccharideb3(?=-GalNAc)']


def test_get_match_batch():
    glycan_list = ["Gal(b1-4)GlcNAc(b1-6)[Gal(b1-3)]GalNAc", "Gal(b1-4)GlcNAc"]
    pattern = "r[Sia]{,1}-[.]{,1}-([dHex]){,1}-.b3(?=-GalNAc)"
    assert len(get_match_batch(pattern, glycan_list)) == 2


def test_process_question_mark():
    # Test basic lookahead
    result, pattern = process_question_mark("(?=Hex)", "Hex")
    assert result == [1]
    assert pattern == ["Hex"]
    # Test lookbehind
    result, pattern = process_question_mark("(?<=Hex)", "Hex")
    assert result == [1]
    assert pattern == ["Hex"]
    # Test simple question mark
    result, pattern = process_question_mark("Hex?", "Hex")
    assert result == [0, 1]
    assert pattern == ["Hex"]


def test_calculate_len_matches_comb():
    # Test single list of matches
    len_matches = [[2, 3]]
    result = calculate_len_matches_comb(len_matches)
    assert 0 in result
    assert 2 in result
    assert 3 in result
    # Test multiple lists
    len_matches = [[2], [3]]
    result = calculate_len_matches_comb(len_matches)
    assert 2 in result
    assert 3 in result
    assert 5 in result  # Combined length


def test_process_main_branch():
    glycan = "Gal(b1-4)GlcNAc(b1-2)[GlcNAc(b1-4)]Man"
    glycan_parts = min_process_glycans(["Gal(b1-4)GlcNAc(b1-2)Man"])[0]
    result = process_main_branch(glycan, glycan_parts)
    assert len(result) == len(glycan_parts)
    assert all(isinstance(x, str) for x in result)


def test_check_negative_look():
    # Test negative lookahead
    matches = [[0, 2, 4]]
    pattern = "(?!Hex)"
    glycan = "Gal(b1-4)GlcNAc(b1-2)Man"
    result = check_negative_look(matches, pattern, glycan)
    assert isinstance(result, list)
    # Test negative lookbehind
    pattern = "(?<!Hex)"
    result = check_negative_look(matches, pattern, glycan)
    assert isinstance(result, list)


def test_filter_matches_by_location():
    # Create a simple test graph
    ggraph = glycan_to_nxGraph("Fuc(a1-2)Gal(b1-4)Glc")
    # Test start location
    matches = [[0, 1], [1, 2]]
    result = filter_matches_by_location(matches, ggraph, "start")
    assert len(result) == 1
    assert result[0] == [0, 1]
    # Test end location
    result = filter_matches_by_location(matches, ggraph, "end")
    assert len(result) == 0  # None of the matches end at the last node
    # Test internal location
    matches = [[1, 2]]
    result = filter_matches_by_location(matches, ggraph, "internal")
    assert len(result) == 1
    assert result[0] == [1, 2]
    result = filter_matches_by_location(matches, ggraph, "nonsense")
    assert len(result) == 1


def test_parse_pattern():
    # Test exact occurrence
    pattern = "Hex{2}"
    min_occur, max_occur = parse_pattern(pattern)
    assert min_occur == 2
    assert max_occur == 2
    # Test range occurrence
    pattern = "Hex{1,3}"
    min_occur, max_occur = parse_pattern(pattern)
    assert min_occur == 1
    assert max_occur == 3
    # Test wildcard
    pattern = "Hex*"
    min_occur, max_occur = parse_pattern(pattern)
    assert min_occur == 0
    assert max_occur == 8
    # Test plus
    pattern = "Hex+"
    min_occur, max_occur = parse_pattern(pattern)
    assert min_occur == 1
    assert max_occur == 8
    # Test question mark
    pattern = "Hex?"
    min_occur, max_occur = parse_pattern(pattern)
    assert min_occur == 0
    assert max_occur == 1


def test_get_matching_indices():
    # Test nested parentheses matching
    result = list(get_matching_indices("(a(b)c)", "(", ")"))
    assert len(result) == 2
    assert result[0][2] == 1  # Nested level
    assert result[1][2] == 0  # Outer level
    # Test square brackets
    result = list(get_matching_indices("[a[b]c]", "[", "]"))
    assert len(result) == 2
    assert result[0][2] == 1  # Nested level
    assert result[1][2] == 0  # Outer level
    # Test empty string
    result = list(get_matching_indices("", "(", ")"))
    assert len(result) == 0


def test_process_bonds_1():
    # Test basic linkages
    assert process_bonds(["a1-2"]) == ["α 2"]
    assert process_bonds(["b1-4"]) == ["β 4"]
    # Test uncertain linkages
    assert process_bonds(["?1-?"]) == ["?"]
    assert process_bonds(["b1-2"]) == ["β 2"]
    assert process_bonds(["a1-?"]) == ["α"]
    assert process_bonds(["x1-?"]) == ["-"]
    assert process_bonds(["1-3"]) == ["1 - 3"]
    assert process_bonds(["x1-3"]) == ["-"]
    # Test multiple bonds
    result = process_bonds([["a1-2", "b1-4"]])
    assert result == [["α 2", "β 4"]]


def test_draw_chem2d():
    missing_deps = []
    try:
        from rdkit.Chem import MolFromSmiles
        from rdkit.Chem.Draw import PrepareMolForDrawing
        from rdkit.Chem.Draw.rdMolDraw2D import MolDraw2DSVG
    except ImportError:
        missing_deps.append("RDKit")
    if missing_deps:
        pytest.skip(f"Required dependencies not installed: {', '.join(missing_deps)}")
    # Test basic 2D drawing
    result = draw_chem2d("GlcNAc(b1-4)GlcA", ["GlcNAc"])
    plt.close('all')
    # Test with filepath
    draw_chem2d("GlcNAc(b1-4)GlcA", ["GlcNAc"], filepath="test.svg")
    assert Path("test.svg").exists()
    draw_chem2d("GlcNAc(b1-4)GlcA", ["GlcNAc"], filepath="test.pdf")
    assert Path("test.pdf").exists()
    # Test with unsupported glycan
    with patch('glycowork.motif.processing.IUPAC_to_SMILES', side_effect=Exception), pytest.raises(Exception):
        draw_chem2d("InvalidGlycan", ["GlcNAc"])


def test_draw_chem3d():
    missing_deps = []
    try:
        from rdkit.Chem import MolFromSmiles, AddHs, RemoveHs, MolToPDBFile
        from rdkit.Chem.AllChem import EmbedMolecule, MMFFOptimizeMolecule
    except ImportError:
        missing_deps.append("RDKit")
    if missing_deps:
        pytest.skip(f"Required dependencies not installed: {', '.join(missing_deps)}")
    # Test basic 3D drawing
    draw_chem3d("GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc", ["GlcNAc"])
    plt.close('all')
    # Test with filepath
    draw_chem3d("GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc", ["GlcNAc"], filepath="test.pdb")
    # Test with non-PDB filepath
    with patch('builtins.print') as mock_print:
        draw_chem3d("GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc", ["GlcNAc"], filepath="test.svg")


def test_glycodraw():
    # Test basic drawing
    result = GlycoDraw("GlcNAc(b1-4)GlcA", suppress=True)
    assert result is not None
    result = GlycoDraw("{Gal(b1-4)}GlcNAc(b1-4)GlcA", suppress=True)
    assert result is not None
    result = GlycoDraw("Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)Man(a1-6)][Xyl(b1-2)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)]GlcNAc", suppress=True)
    assert result is not None
    result = GlycoDraw("Terminal_Fuc(a1-2)Gal", suppress=True)
    assert result is not None
    result = GlycoDraw("Neu5Ac(a2-?)Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man(a1-6)][Xyl(b1-2)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc", suppress=True)
    assert result is not None
    result = GlycoDraw("Neu5Ac(a2-3)[GalNAc(b1-4)]Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-4)]Man(a1-3)[Fuc(a1-2)Gal(b1-4)[Fuc(a1-3)]GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-6)]Man(a1-6)][Xyl(b1-2)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-3)][Fuc(a1-6)]GlcNAc", suppress=True)
    assert result is not None
    result = GlycoDraw("Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-?)]Man(a1-?)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-?)]Man(a1-?)][GlcNAc(b1-4)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc", suppress=True)
    assert result is not None
    result = GlycoDraw("{Neu5Ac(a2-3)}{Neu5Ac(a2-6)}Gal", suppress=True)
    assert result is not None
    # Test vertical orientation
    result = GlycoDraw("GlcNAc(b1-4)GlcA", vertical=True, suppress=True)
    assert result is not None
    # Test compact mode
    result = GlycoDraw("GlcNAc(b1-4)GlcA", compact=True, suppress=True)
    assert result is not None
    # Test with highlighting
    result = GlycoDraw("GlcNAc(b1-4)GlcA", highlight_motif="GlcNAc", suppress=True)
    assert result is not None
    result = GlycoDraw("GlcNAc(b1-4)GlcA", per_residue=[0.8, 0.2], suppress=True)
    assert result is not None
    result = GlycoDraw("Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-6)[Neu5Ac(a2-3)Gal(b1-3)]GalNAc", highlight_motif="r[Sia]{,1}-[.]{,1}-([dHex]){,1}-.b3(?=-GalNAc)", suppress=True)
    assert result is not None
    result = GlycoDraw("GlcNAc(b1-2)Man(a1-3)[GlcNAc(b1-2)Man(a1-6)][Xyl(b1-2)]Man(b1-4)GlcNAc(b1-4)GlcNAc", highlight_motif="Xyl(b1-2)Man", reverse_highlight=True, suppress=True)
    assert result is not None
    # Test with repeat unit
    result = GlycoDraw("GlcNAc(b1-4)GlcA(b1-3)", repeat=True, suppress=True)
    assert result is not None
    result = GlycoDraw("GlcNAc(b1-4)GlcA(b1-3)", repeat=3, suppress=True)
    assert result is not None
    result = GlycoDraw("GlcNAc(b1-4)GlcA(b1-3)", repeat='2-3', suppress=True)
    assert result is not None
    result = GlycoDraw("Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)GlcNAc", repeat=True, repeat_range=[1,2], suppress=True)
    assert result is not None
    # Test chemical drawing modes
    result = GlycoDraw("Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-6)[Neu5Ac(a2-3)Gal(b1-3)]GalNAc", draw_method="chem2d", suppress=True)
    # Test name recognition
    result = GlycoDraw("Terminal_LewisX", suppress=True)
    assert result is not None
    result = GlycoDraw("DManpa1-3[DManpa1-6][DXylpb1-2]DManpb1-4DGlcpNAcb1-4[LFucpa1-3]DGlcpNAca1-OH", suppress=True)
    assert result is not None
    # Test file saving
    GlycoDraw("GlcNAc(b1-4)GlcA", filepath="test.svg")
    assert Path("test.svg").exists()
    GlycoDraw("GlcNAc(b1-4)GlcA", filepath="test.pdf")
    assert Path("test.pdf").exists()
    GlycoDraw("GlcNAc(b1-4)GlcA", filepath="test.png")
    assert Path("test.png").exists()
    # Test invalid glycan
    with pytest.raises(Exception):
        GlycoDraw("InvalidGlycan")


@pytest.mark.parametrize("glycan", GLYCAN_TEST_CASES)
def test_glycodraw_complex(glycan):
    result = GlycoDraw(glycan, suppress=True)
    assert result is not None


def test_plot_glycans_excel(tmp_path):
    # Create test directory
    test_dir = tmp_path / "test_folder"
    test_dir.mkdir()
    # Create test DataFrame
    df = pd.DataFrame({
        'Glycans': ['GlcNAc(b1-4)GlcA', 'Man(a1-3)Man'],
        'Values': [1, 2]
    })
    # Make sure PIL.Image are available
    from PIL import Image
    # Run the actual function to plot glycans and save to Excel
    plot_glycans_excel(df, str(test_dir))
    # Assert that the output file exists
    output_file = test_dir / "output.xlsx"
    assert output_file.exists(), f"Expected output file {output_file} to be created."
    # Load the output Excel file and check if image was inserted
    from openpyxl import load_workbook
    # Load workbook
    wb = load_workbook(output_file)
    sheet = wb.active
    # Check if the image is inserted by checking for a non-empty cell
    # This assumes the image is inserted in the third column and second row
    cell = sheet.cell(row=2, column=3)
    assert cell.value is None, "Expected an image to be inserted into the Excel sheet."
    # Test with different scaling
    plot_glycans_excel(df, str(test_dir), scaling_factor=0.5)
    # Test with compact mode
    plot_glycans_excel(df, str(test_dir), compact=True)
    # Test with CSV input
    csv_path = test_dir / "test.csv"
    df.to_csv(csv_path, index = False)
    plot_glycans_excel(str(csv_path), str(test_dir))
    # Test with Excel input
    excel_path = test_dir / "test.xlsx"
    df.to_excel(excel_path, index = False)
    plot_glycans_excel(str(excel_path), str(test_dir))
    # Test with different glycan column
    df2 = pd.DataFrame({
            'Other': ['GlcNAc(b1-4)GlcA', 'Man(a1-3)Man'],
            'Values': [1, 2]
    })
    plot_glycans_excel(df2, str(test_dir), glycan_col_num=0)
    # Clean up the test folder after the test
    shutil.rmtree(test_dir)


@pytest.fixture
def mock_svg_file():
    return """<?xml version="1.0" encoding="utf-8" standalone="no"?>
<svg xmlns:xlink="http://www.w3.org/1999/xlink" width="500" height="500" xmlns="http://www.w3.org/2000/svg" version="1.1">
 <!-- nontransfected_1 -->
 <g transform="translate(158.979875 703.017187) scale(0.1 -0.1)">
  <path d="M 0 0 L 10 10" />
 </g>
 <!-- nontransfected_2 -->
 <g transform="translate(343.687875 703.017187) scale(0.1 -0.1)">
  <path d="M 5 5 L 15 15" />
 </g>
 <!-- Neu5Ac(a2-3)Gal(b1-3)GalNAc -->
 <g transform="translate(486.77 252.656302) scale(0.1 -0.1)">
  <path d="M 5 5 L 15 15" />
 </g>
 <!-- Gal(b1-3)[Neu5Ac(a2-6)]GalNAc -->
 <g transform="translate(486.77 428.480969) scale(0.1 -0.1)">
  <path d="M 5 5 L 15 15" />
 </g>
 <!-- Gal(b1-4)GlcNAc(b1-6)[Gal(b1-3)]GalNAc -->
 <g transform="translate(486.77 604.305635) scale(0.1 -0.1)">
  <path d="M 5 5 L 15 15" />
 </g>
 <!-- Samples -->
 <g transform="translate(32.023875 150.666938) scale(0.1 -0.1)">
  <path d="M 0 0 L 10 10" />
 </g>
 <!-- Glycans -->
 <g transform="translate(130.877438 92.943625) rotate(-90) scale(0.1 -0.1)">
  <path d="M 0 0 L 10 10" />
 </g>
 <!-- glycan -->
 <g transform="translate(702.904375 441.186437) rotate(-90) scale(0.1 -0.1)">
  <path d="M 0 0 L 10 10" />
 </g>
</svg>"""


def test_annotate_figure(mock_svg_file):
   # Test basic annotation
    result = annotate_figure(mock_svg_file)
    assert isinstance(result, str)
    assert "svg" in result
    # Test with glycan size
    result = annotate_figure(mock_svg_file, glycan_size="small")
    assert isinstance(result, str)
    # Test with compact mode
    result = annotate_figure(mock_svg_file, compact=True)
    assert isinstance(result, str)
    # Test with differential expression results
    de_results = pd.DataFrame({
        'Glycan': ['Neu5Ac(a2-3)Gal(b1-3)GalNAc', 'Gal(b1-3)[Neu5Ac(a2-6)]GalNAc',
                   'Gal(b1-4)GlcNAc(b1-6)[Gal(b1-3)]GalNAc'],
        'Log2FC': [2.0, 1.7, 1.4],
        'corr p-val': [0.01, 0.03, 0.001]
    })
    result = annotate_figure(mock_svg_file, scale_by_DE_res=de_results)
    assert isinstance(result, str)
    annotate_figure(mock_svg_file, filepath="test.pdf")
    assert Path("test.pdf").exists()
    annotate_figure(mock_svg_file, filepath="test.svg")
    assert Path("test.svg").exists()
    annotate_figure(mock_svg_file, filepath="test.png")
    assert Path("test.png").exists()


def test_scale_in_range():
    # Test basic scaling
    result = scale_in_range([1, 2, 3], 0, 1)
    assert result[0] == 0  # Min value
    assert result[-1] == 1  # Max value
    assert 0 < result[1] < 1  # Middle value
    # Test with negative numbers
    result = scale_in_range([-2, 0, 2], 0, 10)
    assert result[0] == 0
    assert result[1] == 5
    assert result[2] == 10
    # Test single value
    result = scale_in_range([5, 5, 5], 0, 1)
    assert all(x == 0 for x in result)  # All values map to min when input range is 0


def test_process_per_residue():
    # Test linear glycan
    values = [1, 2, 3]
    main, side, branched = process_per_residue("Gal(b1-4)GlcNAc(b1-2)Man", values, "Gal(b1-4)GlcNAc(b1-2)Man")
    assert len(main) == 3
    assert len(side) == 0
    assert len(branched) == 0
    # Test branched glycan
    values = [1, 2, 3, 4]
    main, side, branched = process_per_residue("Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man", values, "Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-2)Man")
    assert len(main) > 0
    assert len(side) > 0
    assert isinstance(side[0], list)
    # Test branch-branched glycan
    values = [1, 2, 3, 4, 5, 6]
    main, side, branched = process_per_residue("Gal(b1-4)GlcNAc(b1-2)[Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-4)]Man", values, "Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-4)[Gal(b1-4)GlcNAc(b1-2)]Man")


try:
    from rdkit import Chem
    RDKIT_AVAILABLE = True
except ImportError:
    RDKIT_AVAILABLE = False


@pytest.mark.skipif(not RDKIT_AVAILABLE, reason="RDKit not installed")
def test_get_hit_atoms_and_bonds():
    # Create a simple molecule
    mol = Chem.MolFromSmiles("CC(=O)O")
    atoms, bonds = get_hit_atoms_and_bonds(mol, "CC(=O)O")
    assert isinstance(atoms, list)
    assert isinstance(bonds, list)


@pytest.mark.skipif(not RDKIT_AVAILABLE, reason="RDKit not installed")
def test_add_colours_to_map():
    # Test color assignment
    cols = {}
    els = [1, 2, 3]
    add_colours_to_map(els, cols, 0, alpha=True)
    assert len(cols) == 3
    assert all(isinstance(v, list) for v in cols.values())


def get_clean_drawing():
    return draw.Drawing(200, 200)


def test_draw_hex():
    # Test basic hexagon drawing
    mock_drawing = get_clean_drawing()
    draw_hex(1.0, 1.0, 50, col_dict_base, mock_drawing)
    # Verify drawing has 1 element (the filled hexagon)
    assert len(mock_drawing.all_elements()) == 1
    # Test with custom color
    mock_drawing = get_clean_drawing()
    draw_hex(1.0, 1.0, 50, col_dict_base, mock_drawing, color='snfg_red')
    assert len(mock_drawing.all_elements()) == 1


def test_is_jupyter():
  # Test basic functionality
  result = is_jupyter()
  assert isinstance(result, bool)


def test_process_repeat():
    # Test basic repeat unit processing
    input_str = "GlcNAc(b1-4)GlcA(b1-3)"
    expected = "blank(?1-3)GlcNAc(b1-4)GlcA(b1-?)"
    assert process_repeat(input_str) == expected
    # Test with different linkage
    input_str = "GlcNAc(b1-4)Glc(a1-6)"
    expected = "blank(?1-6)GlcNAc(b1-4)Glc(a1-?)"
    assert process_repeat(input_str) == expected


def test_draw_bracket():
    # Test basic bracket drawing
    mock_drawing = get_clean_drawing()
    draw_bracket(1.0, [0.0, 2.0], mock_drawing, 'right', 50, 'show')
    # Verify drawing has expected elements (vertical line + 2 horizontal lines)
    assert len(mock_drawing.all_elements()) == 1
    # Test left-facing bracket
    mock_drawing = get_clean_drawing()
    draw_bracket(1.0, [0.0, 2.0], mock_drawing, 'left', 50, 'show')
    assert len(mock_drawing.all_elements()) == 1
    # Test with rotation
    mock_drawing = get_clean_drawing()
    draw_bracket(1.0, [0.0, 2.0], mock_drawing, 'right', 50, 'show', 45)
    assert len(mock_drawing.all_elements()) == 1


def test_display_svg_with_matplotlib():
    # Create simple SVG
    test_svg = """<?xml version="1.0" encoding="UTF-8"?>
<svg xmlns="http://www.w3.org/2000/svg" xmlns:xlink="http://www.w3.org/1999/xlink"
     width="200" height="110" viewBox="-150.0 -55.0 200 110">
<defs>
<path d="M-100,0 L0,0" stroke-width="4.0" stroke="#000000" id="d0" />
<path d="M-50,25.0 L50,25.0" stroke-width="0" id="d1" />
<path d="M-150,25.0 L-50,25.0" stroke-width="0" id="d2" />
</defs>
<g transform="rotate(0 -50.0 0.0)">
<use xlink:href="#d0" />
<text font-size="20.0" text-anchor="middle" fill="#000000" valign="middle"><textPath xlink:href="#d0" startOffset="50%">
<tspan dy="-0.5em">β 3</tspan>
</textPath></text>
<rect x="-25.0" y="-25.0" width="50" height="50" fill="#FCC326" stroke-width="2.0" stroke="#000000" />
<use xlink:href="#d1" />
<text font-size="17.5" fill="#000000" text-anchor="middle"><textPath xlink:href="#d1" startOffset="50%"></textPath></text>
<circle cx="-100" cy="0" r="25.0" fill="#FCC326" stroke-width="2.0" stroke="#000000" />
<use xlink:href="#d2" />
<text font-size="17.5" fill="#000000" text-anchor="middle"><textPath xlink:href="#d2" startOffset="50%"></textPath></text>
</g>
</svg>"""
    # Test with mocked matplotlib
    class MockPlt:
        @staticmethod
        def imread(*args, **kwargs):
            return np.zeros((100, 100))
        @staticmethod
        def imshow(*args, **kwargs):
            pass
        @staticmethod
        def axis(*args, **kwargs):
            pass
        @staticmethod
        def show():
            pass
    import matplotlib.pyplot as plt
    old_plt = plt
    plt = MockPlt()
    display_svg_with_matplotlib(test_svg)
    plt = old_plt


def test_draw_shape_hex():
    # Test basic hexose drawing
    mock_drawing = get_clean_drawing()
    draw_shape('Hex', 'snfg_white', 1.0, 1.0, col_dict_base, mock_drawing, dim=50)
    elements = mock_drawing.all_elements()
    # Should have: circle + text path + text for modification
    assert len(elements) == 3
    # Test with scalar
    mock_drawing = get_clean_drawing()
    draw_shape('Hex', 'snfg_white', 1.0, 1.0, col_dict_base, mock_drawing, dim=50, scalar=10)
    elements = mock_drawing.all_elements()
    assert len(elements) == 4
    # Test with modification
    mock_drawing = get_clean_drawing()
    draw_shape('Hex', 'snfg_white', 1.0, 1.0, col_dict_base, mock_drawing,
               modification='2S', dim=50)
    assert len(mock_drawing.all_elements()) == 3
    # Test with furanose
    mock_drawing = get_clean_drawing()
    draw_shape('Hex', 'snfg_white', 1.0, 1.0, col_dict_base, mock_drawing,
               furanose=True, dim=50)
    # Should have additional text elements for furanose indicator
    assert len(mock_drawing.all_elements()) == 5


def test_draw_shape_hexnac():
    # Test HexNAc drawing (square shape)
    mock_drawing = get_clean_drawing()
    draw_shape('HexNAc', 'snfg_yellow', 1.0, 1.0, col_dict_base, mock_drawing, dim=50)
    elements = mock_drawing.all_elements()
    # Should have: rectangle + text path + text for modification
    assert len(elements) == 3
    # Test furanose
    mock_drawing = get_clean_drawing()
    draw_shape('HexNAc', 'snfg_yellow', 1.0, 1.0, col_dict_base, mock_drawing, dim=50, furanose=True)
    elements = mock_drawing.all_elements()
    assert len(elements) == 5
    # Test with configuration
    mock_drawing = get_clean_drawing()
    draw_shape('HexNAc', 'snfg_yellow', 1.0, 1.0, col_dict_base, mock_drawing,
               conf='L-', dim=50)
    assert len(mock_drawing.all_elements()) == 5


def test_draw_shape_dHex():
    # Test dHex drawing (triangle)
    mock_drawing = get_clean_drawing()
    draw_shape('dHex', 'snfg_red', 1.0, 1.0, col_dict_base, mock_drawing, dim=50)
    elements = mock_drawing.all_elements()
    # Should have: triangle + text path + text for modification
    assert len(elements) == 3
    # Test furanose
    mock_drawing = get_clean_drawing()
    draw_shape('dHex', 'snfg_red', 1.0, 1.0, col_dict_base, mock_drawing, dim=50, furanose=True)
    elements = mock_drawing.all_elements()
    assert len(elements) == 5


def test_draw_shape_Pen():
    # Test Pentose drawing (star)
    mock_drawing = get_clean_drawing()
    draw_shape('Pen', 'snfg_red', 1.0, 1.0, col_dict_base, mock_drawing, dim=50)
    elements = mock_drawing.all_elements()
    # Should have: star + text path + text for modification
    assert len(elements) == 3
    # Test furanose
    mock_drawing = get_clean_drawing()
    draw_shape('Pen', 'snfg_red', 1.0, 1.0, col_dict_base, mock_drawing, dim=50, furanose=True)
    elements = mock_drawing.all_elements()
    assert len(elements) == 5


# Test cases structured as: (shape_name, color, expected_elements, description)
TEST_CASES = [
    ('HexN', 'snfg_green', 5, 'crossed square'),
    ('HexA_2', 'snfg_yellow', 6, 'divided diamond'),
    ('dHexNAc', 'snfg_yellow', 7, 'divided triangle'),
    ('ddHex', 'snfg_yellow', 3, 'flat rectangle'),
    ('dNon', 'snfg_red', 3, 'diamond'),
    ('ddNon', 'snfg_red', 3, 'flat diamond'),
    ('Unknown', 'snfg_red', 3, 'flat hexagon'),
    ('Assigned', 'snfg_red', 3, 'pentagon'),
    ('red_end', 'snfg_red', 2, 'reducing end'),
    ('free', 'snfg_red', 1, 'free end'),
    # Fragment drawings
    ('04X', 'snfg_red', 5, 'fragment'),
    ('15A', 'snfg_red', 5, 'fragment'),
    ('02A', 'snfg_red', 5, 'fragment'),
    ('13X', 'snfg_red', 5, 'fragment'),
    ('24X', 'snfg_red', 5, 'fragment'),
    ('35X', 'snfg_red', 5, 'fragment'),
    ('04A', 'snfg_red', 5, 'fragment'),
    ('15X', 'snfg_red', 5, 'fragment'),
    ('02X', 'snfg_red', 5, 'fragment'),
    ('13A', 'snfg_red', 5, 'fragment'),
    ('24A', 'snfg_red', 5, 'fragment'),
    ('35A', 'snfg_red', 5, 'fragment'),
    ('25A', 'snfg_red', 4, 'fragment'),
    ('03A', 'snfg_red', 4, 'fragment'),
    ('14X', 'snfg_red', 4, 'fragment'),
    ('25X', 'snfg_red', 4, 'fragment'),
    ('03X', 'snfg_red', 4, 'fragment'),
    ('14A', 'snfg_red', 4, 'fragment'),
    ('Z', 'snfg_red', 1, 'fragment'),
    ('Y', 'snfg_red', 1, 'fragment'),
    ('B', 'snfg_red', 1, 'fragment'),
    ('C', 'snfg_red', 2, 'fragment'),
]


@pytest.mark.parametrize("shape, color, expected_elements, description", TEST_CASES)
def test_draw_shape_simple(shape, color, expected_elements, description):
    mock_drawing = get_clean_drawing()
    draw_shape(shape, color, 1.0, 1.0, col_dict_base, mock_drawing, dim=50)
    elements = mock_drawing.all_elements()
    assert len(elements) == expected_elements, f"Shape {shape} ({description}) should have exactly {expected_elements} elements"


def test_add_sugar():
    # Test basic sugar addition
    mock_drawing = get_clean_drawing()
    add_sugar('Glc', mock_drawing, 1.0, 1.0, dim=50)
    assert len(mock_drawing.all_elements()) > 0
    # Test with modification
    mock_drawing = get_clean_drawing()
    add_sugar('GlcNAc', mock_drawing, 1.0, 1.0, modification='3S', dim=50)
    assert len(mock_drawing.all_elements()) > 0
    # Test with highlight
    mock_drawing = get_clean_drawing()
    add_sugar('Man', mock_drawing, 1.0, 1.0, highlight='show', dim=50)
    assert len(mock_drawing.all_elements()) > 0
    # Test unknown sugar
    mock_drawing = get_clean_drawing()
    add_sugar('UnknownSugar', mock_drawing, 1.0, 1.0, dim=50)
    # Should draw X shape for unknown sugars
    assert len(mock_drawing.all_elements()) == 1


def test_get_highlight_attribute():
    # Create test glycan graph
    G = glycan_to_nxGraph("Gal(b1-4)GlcNAc(b1-2)Man")
    # Test with no motif
    result = get_highlight_attribute(G, '')
    attrs = nx.get_node_attributes(result, 'highlight_labels')
    assert all(v == 'show' for v in attrs.values())
    # Test with specific motif
    result = get_highlight_attribute(G, 'GlcNAc(b1-2)Man')
    attrs = nx.get_node_attributes(result, 'highlight_labels')
    assert attrs[0] == 'hide'  # Gal should not be highlighted
    assert attrs[2] == 'show'  # GlcNAc should be highlighted
    assert attrs[4] == 'show'  # Man should be highlighted
    # Test with multiple matches
    G = glycan_to_nxGraph("Gal(b1-4)GlcNAc(b1-2)[Gal(b1-4)GlcNAc(b1-6)]Man")
    result = get_highlight_attribute(G, 'GlcNAc(b1-?)Man')
    attrs = nx.get_node_attributes(result, 'highlight_labels')
    assert attrs[0] == 'hide'  # Gal should not be highlighted
    assert attrs[2] == 'show'  # 1st GlcNAc should be highlighted
    assert attrs[6] == 'show'  # 2nd GlcNAc should be highlighted
    assert attrs[8] == 'show'  # Man should be highlighted


def test_get_coordinates_and_labels():
    # Test basic linear glycan
    glycan = "GlcNAc(b1-4)GlcA"
    result = get_coordinates_and_labels(glycan, None)
    assert len(result) == 4  # Should return main chain and branch data
    # Check main chain data structure
    main_data = result[0]
    assert len(main_data) == 8  # Should have 8 components
    assert main_data[0] == ['GlcA', 'GlcNAc']  # Sugars
    assert len(main_data[1]) == 2  # x positions
    assert len(main_data[2]) == 2  # y positions
    # Test branched glycan
    glycan = "GlcNAc(b1-4)[Fuc(a1-3)]GlcA"
    result = get_coordinates_and_labels(glycan, None)
    assert len(result[1][0]) > 0  # Should have branch data
    # Test with highlight motif
    result = get_coordinates_and_labels(glycan, "Fuc(a1-3)GlcA")
    assert 'show' in result[0][-2]  # Check highlight attributes


def test_add_bond():
    # Test basic bond
    mock_drawing = get_clean_drawing()
    add_bond(0, 1, 0, 0, mock_drawing, dim=50)
    assert len(mock_drawing.all_elements()) == 1  # Line
    # Test with label
    mock_drawing = get_clean_drawing()
    add_bond(0, 1, 0, 0, mock_drawing, label='β1-4', dim=50)
    assert len(mock_drawing.all_elements()) == 2  # Line + text path
    # Test with highlight
    mock_drawing = get_clean_drawing()
    add_bond(0, 1, 0, 0, mock_drawing, highlight='hide', dim=50)
    assert len(mock_drawing.all_elements()) == 1
    # Test compact mode
    mock_drawing = get_clean_drawing()
    add_bond(0, 1, 0, 0, mock_drawing, compact=True, dim=50)
    assert len(mock_drawing.all_elements()) == 1


def test_process_bonds_2():
    # Test alpha linkages
    assert process_bonds(['a1']) == ['α 1']
    # Test beta linkages
    assert process_bonds(['b1']) == ['β 1']
    # Test unknown linkages
    assert process_bonds(['?1']) == [' 1']
    # Test multiple linkages
    assert process_bonds(['a1', 'b2']) == ['α 1', 'β 2']
    # Test nested linkages
    assert process_bonds([['a1', 'b2'], ['a3']]) == [['α 1', 'β 2'], ['α 3']]


def test_preprocess_data():
    # Create test dataframe
    df = pd.DataFrame({
        'glycan': ['Gal(b1-4)GlcNAc', 'Man(a1-6)Man'],
        'sample1': [10, 20],
        'sample2': [15, 25],
        'sample3': [12, 22],
        'sample4': [18, 28]
    })
    group1 = ['sample1', 'sample2']
    group2 = ['sample3', 'sample4']
    # Test basic preprocessing
    df_trans, df_org, g1, g2 = preprocess_data(df, group1, group2, impute=False)
    assert isinstance(df_trans, pd.DataFrame)
    assert isinstance(df_org, pd.DataFrame)
    assert g1 == group1
    assert g2 == group2
    # Test with motifs
    df_trans, df_org, g1, g2 = preprocess_data(df, group1, group2, motifs=True)
    assert 'Terminal_LacNAc_type2' in df_trans.index
    assert 'Man' in df_trans.index
    try:
        df_trans, df_org, g1, g2 = preprocess_data(df, group1, group2, transform="wrong")
        return False
    except ValueError:
        pass


def test_get_pvals_motifs():
    # Create test data with multiple glycans and varied z-scores
    glycans = [
        'Gal(b1-4)GlcNAc',
        'Man(a1-6)Man',
        'Neu5Ac(a2-6)Gal(b1-4)GlcNAc',
        'Fuc(a1-3)GlcNAc',
        'GalNAc(b1-4)GlcNAc',
        'Man(a1-3)[Man(a1-6)]Man',
        'Gal(b1-4)GlcNAc(b1-2)Man',
        'Fuc(a1-2)Gal(b1-4)GlcNAc'
    ]
    # Create z-scores with clear separation for testing
    np.random.seed(42)  # For reproducibility
    n_samples = len(glycans) * 4  # Total number of rows needed
    n_high = n_samples // 2
    n_low = n_samples - n_high
    high_scores = np.random.normal(2.5, 0.5, n_high)  # Above threshold
    low_scores = np.random.normal(-2.5, 0.5, n_low)   # Below threshold
    # Create multiple samples per glycan
    df = pd.DataFrame({
        'glycan': glycans * 4,  # Replicate each glycan 4 times
        'target': np.concatenate([high_scores, low_scores])  # Mix high and low z-scores
    })
    # Test basic functionality with default parameters
    results = get_pvals_motifs(df, zscores=True)
    # Check results structure
    assert isinstance(results, pd.DataFrame)
    assert 'motif' in results.columns
    assert 'pval' in results.columns
    assert 'corr_pval' in results.columns
    assert 'effect_size' in results.columns
    # Should find some significant motifs
    assert len(results) > 0
    assert any(results['corr_pval'] < 0.05)
    # Effect sizes should be non-zero
    assert not all(results['effect_size'] == 0)
    # Test with different feature sets
    results_known = get_pvals_motifs(df, feature_set=['known'])
    assert len(results_known) > 0
    results_terminal = get_pvals_motifs(df, feature_set=['terminal1'])
    assert len(results_terminal) > 0
    # Test with different threshold
    results_thresh = get_pvals_motifs(df, thresh=1.0)
    assert isinstance(results_thresh, pd.DataFrame)
    assert len(results_thresh) > 0
    high_scores = np.random.normal(100, 10, n_high)  # Above threshold
    low_scores = np.random.normal(10, 2, n_low)   # Below threshold
    df = pd.DataFrame({
        'glycan': glycans * 4,  # Replicate each glycan 4 times
        'target': np.concatenate([high_scores, low_scores])  # Mix high and low z-scores
    })
    results = get_pvals_motifs(df, zscores=False, thresh=0.5)
    df = pd.DataFrame({
        'glycan': glycans * 4,  # Replicate each glycan 4 times
        'target1': np.concatenate([high_scores, low_scores]),  # Mix high and low z-scores
        'target2': np.concatenate([low_scores, high_scores]),  # Mix high and low z-scores
        'target3': np.concatenate([low_scores, high_scores])  # Mix high and low z-scores
    })
    results = get_pvals_motifs(df, zscores=False, thresh=0.5, multiple_samples=True)


def sample_comp_glycomics_data():
    """
    Creates a sample glycomics dataset for testing purposes.
    Contains 5 glycans across 6 pairs of control/tumor samples.
    """
    data = {
        'glycan': [
            'Gal(b1-3)GalNAc',
            'Neu5Ac(a2-3)Gal(b1-3)GalNAc',
            'Fuc(a1-2)Gal(b1-3)GalNAc',
            'Neu5Ac(a2-6)GalNAc',
            'Fuc(a1-2)Gal'
        ]
    }
    # Set random seed for reproducibility
    np.random.seed(42)
    # Generate control and tumor columns for 6 samples
    for i in range(1, 7):
        # Control samples: relatively stable values with some variation
        data[f'control_{i}'] = [
            np.random.normal(1.5, 0.2),  # Stable baseline
            np.random.normal(20.0, 2.0),  # High abundance
            np.random.normal(15.0, 1.5),  # Medium-high abundance
            np.random.normal(8.0, 1.0),   # Medium abundance
            np.random.normal(2.0, 0.3)    # Low abundance
        ]
        # Tumor samples: showing some consistent changes from control
        data[f'tumor_{i}'] = [
            np.random.normal(0.8, 0.2),   # Typically decreased
            np.random.normal(30.0, 3.0),  # Typically increased
            np.random.normal(7.0, 1.0),   # Typically decreased
            np.random.normal(12.0, 1.5),  # Typically increased
            np.random.normal(1.5, 0.3)    # Slightly decreased
        ]
    return pd.DataFrame(data)


def sample_grouped_BH_data():
    """
    Creates a sample glycomics dataset specifically for testing grouped BH correction.
    Contains glycans with clear Sia/Fuc groupings and enough samples per group.
    """
    data = {
        'glycan': [
            # Sia only glycans
            'Neu5Ac(a2-3)Gal(b1-4)GlcNAc',
            'Neu5Ac(a2-6)Gal(b1-4)GlcNAc',
            'Neu5Ac(a2-3)Gal(b1-3)GalNAc',
            # Fuc only glycans
            'Fuc(a1-2)Gal(b1-4)GlcNAc',
            'Fuc(a1-3)Gal(b1-4)GlcNAc',
            'Fuc(a1-2)Gal(b1-3)GalNAc',
            # Both Sia and Fuc
            'Neu5Ac(a2-3)Gal(b1-4)[Fuc(a1-3)]GlcNAc',
            'Neu5Ac(a2-6)Gal(b1-4)[Fuc(a1-3)]GlcNAc',
            'Fuc(a1-2)Gal(b1-3)[Neu5Ac(a2-6)]GalNAc',
            # Neither (rest)
            'Gal(b1-4)GlcNAc',
            'Gal(b1-3)GalNAc',
            'GlcNAc(b1-3)Gal'
        ]
    }
    # Generate control and tumor columns for 6 samples
    np.random.seed(42)  # For reproducibility
    for i in range(1, 7):
        # Control samples
        data[f'control_{i}'] = [
            # Sia group - generally high abundance
            np.random.normal(20.0, 2.0),
            np.random.normal(18.0, 2.0),
            np.random.normal(22.0, 2.0),
            # Fuc group - medium abundance
            np.random.normal(10.0, 1.0),
            np.random.normal(12.0, 1.0),
            np.random.normal(11.0, 1.0),
            # SiaFuc group - very high abundance
            np.random.normal(30.0, 3.0),
            np.random.normal(32.0, 3.0),
            np.random.normal(28.0, 3.0),
            # Rest group - low abundance
            np.random.normal(2.0, 0.3),
            np.random.normal(1.8, 0.3),
            np.random.normal(2.2, 0.3)
        ]
        # Tumor samples - showing consistent changes within groups
        data[f'tumor_{i}'] = [
            # Sia group - increased in tumor
            np.random.normal(30.0, 3.0),
            np.random.normal(28.0, 3.0),
            np.random.normal(32.0, 3.0),
            # Fuc group - decreased in tumor
            np.random.normal(5.0, 1.0),
            np.random.normal(6.0, 1.0),
            np.random.normal(5.5, 1.0),
            # SiaFuc group - slight increase
            np.random.normal(35.0, 3.0),
            np.random.normal(37.0, 3.0),
            np.random.normal(33.0, 3.0),
            # Rest group - minimal change
            np.random.normal(2.1, 0.3),
            np.random.normal(1.9, 0.3),
            np.random.normal(2.3, 0.3)
        ]
    return pd.DataFrame(data)


def sample_comp_glycomics_data_corr():
    np.random.seed(42)
    data={'glycan':['Neu5Ac(a2-8)Neu5Ac','GlcNAc(b1-4)Man','Gal(b1-4)Neu5Ac',
                    'Man(a1-3)Man','Fuc(a1-3)GlcNAc']}
    base_trends={
        'Neu5Ac(a2-8)Neu5Ac':(1.5,0.8),      # Correlates inversely with Gal(b1-3)GalNAc
        'GlcNAc(b1-4)Man':(18.0,28.0),       # Correlates positively with Neu5Ac(a2-3)Gal(b1-3)GalNAc
        'Gal(b1-4)Neu5Ac':(16.0,6.0),        # Correlates positively with Fuc(a1-2)Gal(b1-3)GalNAc
        'Man(a1-3)Man':(7.0,13.0),           # Correlates inversely with Neu5Ac(a2-6)GalNAc
        'Fuc(a1-3)GlcNAc':(2.5,1.2)          # Correlates positively with Fuc(a1-2)Gal
    }
    for i in range(1,7):
        control_vals=[]
        tumor_vals=[]
        for glycan in data['glycan']:
            control_base,tumor_base=base_trends[glycan]
            # Add coordinated noise to maintain correlations
            noise_factor=np.random.normal(1,0.1)
            if glycan in ['Neu5Ac(a2-8)Neu5Ac','Man(a1-3)Man']:
                # Inverse correlations
                control_vals.append(control_base*noise_factor+np.random.normal(0,control_base*0.05))
                tumor_vals.append(tumor_base/noise_factor+np.random.normal(0,tumor_base*0.05))
            else:
                # Positive correlations
                control_vals.append(control_base*noise_factor+np.random.normal(0,control_base*0.05))
                tumor_vals.append(tumor_base*noise_factor+np.random.normal(0,tumor_base*0.05))
        data[f'control_{i}']=control_vals
        data[f'tumor_{i}']=tumor_vals
    return pd.DataFrame(data)


def test_get_differential_expression():
    df = sample_comp_glycomics_data()
    df_grouped = sample_grouped_BH_data()
    group1 = [k for k in df.columns if 'control' in k]
    group2 = [k for k in df.columns if 'tumor' in k]
    # Test basic functionality
    results = get_differential_expression(df, group1, group2, impute=False)
    assert isinstance(results, pd.DataFrame)
    assert 'Glycan' in results.columns
    assert 'p-val' in results.columns
    assert 'Effect size' in results.columns
    assert 'corr p-val' in results.columns
    assert 'significant' in results.columns
    group1_int = [df.columns.tolist().index(k) for k in group1]
    group2_int = [df.columns.tolist().index(k) for k in group2]
    results = get_differential_expression(df, group1_int, group2_int, impute=False)
    results = get_differential_expression(df, group1, group2, impute=False, transform="ALR")
    # At least some results should be significant
    assert any(results['significant'])
    # Effect sizes should be reasonable
    assert any(abs(x) > 0. for x in results['Effect size'])
    assert any(results['significant'])  # Should find some significant results
    # Test with motifs
    results_motifs = get_differential_expression(df, group1, group2, motifs=True, impute=False)
    assert isinstance(results_motifs, pd.DataFrame)
    assert len(results_motifs) > 0
    results = get_differential_expression(df, group1, group2, motifs=False, impute=False,
                                                 monte_carlo=True)
    results = get_differential_expression(df, group1, group2, motifs=False, impute=False,
                                                 sets=True)
    results = get_differential_expression(df_grouped, [k for k in df_grouped.columns if 'control' in k],
                                          [k for k in df_grouped.columns if 'tumor' in k], motifs=False, impute=False,
                                                 grouped_BH=True)
    results = get_differential_expression(df, group1, group2, motifs=False, impute=False,
                                                 effect_size_variance=True)
    # Test with paired samples
    results_paired = get_differential_expression(df, group1, group2,
                                              paired=True, impute=False)
    assert isinstance(results_paired, pd.DataFrame)


def test_get_glycanova():
    # Create test data with THREE groups
    df = pd.DataFrame({
        'glycan': ['Gal(b1-4)GlcNAc', 'Man(a1-6)Man', 'Fuc(a1-2)Gal', 'Neu5Ac(a2-6)Gal'],
        'sample1': [100, 200, 150, 300],
        'sample2': [150, 250, 120, 280],  # Group 1
        'sample3': [120, 220, 140, 290],
        'sample4': [180, 280, 200, 350],
        'sample5': [200, 300, 180, 330],  # Group 2
        'sample6': [190, 290, 190, 340],
        'sample7': [220, 310, 210, 360],
        'sample8': [240, 330, 230, 380],  # Group 3
        'sample9': [230, 320, 220, 370]
    })
    # Three groups with three samples each
    groups = [1, 1, 1, 2, 2, 2, 3, 3, 3]
    # Test basic functionality
    results, posthoc = get_glycanova(df, groups, impute=False)
    # Check results structure
    assert isinstance(results, pd.DataFrame)
    assert 'Glycan' in results.columns
    assert 'F statistic' in results.columns
    assert 'p-val' in results.columns
    assert 'corr p-val' in results.columns
    assert 'significant' in results.columns
    assert 'Effect size' in results.columns
    # Check posthoc results for all pairwise comparisons
    assert isinstance(posthoc, dict)
    if len(posthoc) > 0:  # If any significant differences found
        for _, result in posthoc.items():
            assert isinstance(result, pd.DataFrame)
            # Should have pairwise comparisons between groups
            assert len(result) == 3  # Number of pairwise comparisons for 3 groups = 3
    # Test with motifs enabled
    results_motifs, _ = get_glycanova(df, groups, motifs=True, impute=False)
    assert isinstance(results_motifs, pd.DataFrame)
    assert len(results_motifs) > 0  # Should find some motifs
    # Test without posthoc
    results_no_posthoc, posthoc_empty = get_glycanova(df, groups, posthoc=False, impute=False)
    assert isinstance(results_no_posthoc, pd.DataFrame)
    assert posthoc_empty == {}  # Should be empty dict when posthoc=False
    clr_dict = {1: 2.0, 2: 1.0, 3: 1.5}
    results, posthoc = get_glycanova(df, groups, impute=False, custom_scale=clr_dict)


def test_select_grouping():
    # Create test data
    cohort_a = pd.DataFrame({
        'sample1': [10, 20],
        'sample2': [15, 25]
    })
    cohort_b = pd.DataFrame({
        'sample3': [12, 22],
        'sample4': [18, 28]
    })
    glycans = ['Gal(b1-4)GlcNAc', 'Man(a1-6)Man']
    p_values = [0.04, 0.02]
    # Test without grouped BH
    glycan_groups, pval_groups = select_grouping(cohort_b, cohort_a, glycans, p_values, grouped_BH=False)
    assert len(glycan_groups) == 1
    assert "group1" in glycan_groups
    assert len(glycan_groups["group1"]) == len(glycans)
    # Test with grouped BH
    glycan_groups, pval_groups = select_grouping(cohort_b, cohort_a, glycans, p_values, grouped_BH=True)
    assert isinstance(glycan_groups, dict)
    assert isinstance(pval_groups, dict)


def test_get_biodiversity():
    # Create more realistic glycan compositions based on biological patterns
    np.random.seed(42)  # For reproducibility
    # Helper function to generate realistic glycan proportions
    def generate_group_data(base_proportions, n_samples=3, noise_scale=0.15):
        samples = []
        for _ in range(n_samples):
            # Add multiplicative noise to preserve positivity
            noise = np.random.normal(1, noise_scale, len(base_proportions))
            noisy = base_proportions * noise
            # Normalize to sum to 100
            normalized = 100 * noisy / noisy.sum()
            samples.append(normalized)
        return np.array(samples)
    # Define realistic base proportions for two distinct glycan profiles
    # Based on patterns seen in real data where some glycans are dominant
    group1_base = np.array([35, 25, 15, 10, 8, 7])  # More even distribution
    group2_base = np.array([45, 30, 10, 8, 4, 3])   # More dominated by top glycans
    # Generate sample data
    group1_data = generate_group_data(group1_base, n_samples=3)
    group2_data = generate_group_data(group2_base, n_samples=3)
    # Create DataFrame with realistic glycan names
    df = pd.DataFrame({
        'glycan': [
            'Neu5Ac(a2-3)Gal(b1-3)GalNAc',
            'Neu5Ac(a2-3)Gal(b1-3)[Neu5Ac(a2-6)]GalNAc',
            'Gal(b1-4)GlcNAc(b1-6)[Gal(b1-3)]GalNAc',
            'Neu5Ac(a2-3)Gal(b1-3)[Gal(b1-4)GlcNAc(b1-6)]GalNAc',
            'Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-6)[Gal(b1-3)]GalNAc',
            'Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-6)[Neu5Ac(a2-3)Gal(b1-3)]GalNAc'
        ]
    })
    # Add sample columns
    for i in range(3):
        df[f'sample1_{i+1}'] = group1_data[i]
        df[f'sample2_{i+1}'] = group2_data[i]
    # Define groups for testing
    group1 = [f'sample1_{i+1}' for i in range(3)]
    group2 = [f'sample2_{i+1}' for i in range(3)]
    # Run biodiversity analysis
    results = get_biodiversity(df, group1, group2, metrics=['alpha', 'beta'])
    # Basic assertions
    assert isinstance(results, pd.DataFrame), "Results should be a DataFrame"
    assert 'Metric' in results.columns, "Results should have a Metric column"
    assert 'p-val' in results.columns, "Results should have a p-val column"
    # Additional assertions to verify realistic results
    assert len(results) >= 2, "Should have at least alpha and beta diversity results"
    assert all(0 <= p <= 1 for p in results['p-val']), "p-values should be between 0 and 1"
    # Optional: Check if the differences are detectable
    # The groups are designed to be different, so p-values should be < 0.05
    assert any(p < 0.05 for p in results['p-val']), "Should detect differences between groups"
    results = get_biodiversity(df, group1, group2, metrics=['alpha', 'beta'], paired=True)


@pytest.fixture
def sample_time_series_data():
    np.random.seed(42)
    glycans=['Gal(b1-4)GlcNAc','Man(a1-6)Man','Neu5Ac(a2-6)Gal','Fuc(a1-2)Gal',
             'GlcNAc(b1-4)GlcNAc','Neu5Ac(a2-3)Gal','Gal(b1-3)GalNAc']
    data={'glycan':glycans}
    timepoints=[0,4,8,12,24]
    replicates=3
    for t in timepoints:
        for r in range(1,replicates+1):
            col_name=f'T1_h{t}_r{r}'
            base_values=[]
            for i,glycan in enumerate(glycans):
                if i%3==0:
                    base=100+(t*10)
                elif i%3==1:
                    base=200-(t*5)
                else:
                    base=150+(t*np.sin(t/8))
                value=base+np.random.normal(0,base*0.05)
                base_values.append(max(0,value))
            data[col_name]=base_values
    return pd.DataFrame(data)


def test_get_time_series(sample_time_series_data):
    df = sample_time_series_data
    results = get_time_series(df, impute=False)
    assert isinstance(results, pd.DataFrame)
    assert 'Glycan' in results.columns
    assert 'p-val' in results.columns
    results = get_time_series(df, impute=False, degree=2)
    df2 = df.set_index('glycan').T
    results = get_time_series(df2, impute=False)
    results = get_time_series(df, impute=False, transform="ALR")
    results = get_time_series(df, impute=False, transform="Nothing")
    results = get_time_series(df, impute=False, motifs=True)
    try:
        results = get_time_series(df, impute=False, transform="wrong")
        return False
    except:
        pass


def test_get_SparCC():
    df1 = sample_comp_glycomics_data()
    df2 = sample_comp_glycomics_data_corr()

    for corr, pvals in [
        get_SparCC(df1, df2),
        get_SparCC(df1, df2, transform="ALR"),
        get_SparCC(df1, df2, transform="Nothing"),
        get_SparCC(df1, df2, motifs=True),
        get_SparCC(df1, df2, partial_correlations=True),
    ]:
        assert isinstance(corr, pd.DataFrame)
        assert isinstance(pvals, pd.DataFrame)


def test_get_SparCC_ec_csv():
    df1 = sample_comp_glycomics_data()
    df2 = sample_comp_glycomics_data_corr()

    tmp_dir = Path("test_tmp")
    tmp_dir.mkdir(exist_ok=True)
    df1.to_csv(tmp_dir / "df1.csv", index=False)
    df2.to_csv(tmp_dir / "df2.csv", index=False)

    corr, pvals = get_SparCC(str(tmp_dir / "df1.csv"), str(tmp_dir / "df2.csv"))
    assert isinstance(corr, pd.DataFrame)
    assert isinstance(pvals, pd.DataFrame)
    shutil.rmtree(tmp_dir, ignore_errors=True)


def test_get_SparCC_ec_xlsx():
    df1 = sample_comp_glycomics_data()
    df2 = sample_comp_glycomics_data_corr()

    tmp_dir = Path("test_tmp")
    tmp_dir.mkdir(exist_ok=True)
    df1.to_excel(tmp_dir / "df1.xlsx", index=False)
    df2.to_excel(tmp_dir / "df2.xlsx", index=False)

    corr, pvals = get_SparCC(str(tmp_dir / "df1.xlsx"), str(tmp_dir / "df2.xlsx"))
    assert isinstance(corr, pd.DataFrame)
    assert isinstance(pvals, pd.DataFrame)
    shutil.rmtree(tmp_dir, ignore_errors=True)


def test_get_SparCC_ec_col_mismatch():
    df1 = sample_comp_glycomics_data()
    df2 = sample_comp_glycomics_data_corr()
    df2 = df2[df2.columns[::-1]]
    corr, pvals = get_SparCC(df1, df2)
    assert isinstance(corr, pd.DataFrame)
    assert isinstance(pvals, pd.DataFrame)


def test_get_SparCC_ec_inval_transform():
    df1 = sample_comp_glycomics_data()
    df2 = sample_comp_glycomics_data_corr()
    df1.drop(columns=df1.columns[1], inplace=True)
    with pytest.raises(ValueError):
        get_SparCC(df1, df2, transform="Invalid")


def test_get_roc():
    df = pd.DataFrame({
        'glycan': ['Gal(b1-4)GlcNAc', 'Man(a1-6)Man', 'Neu5Ac(a2-6)Gal', 'Fuc(a1-3)GlcNAc'],
        'sample1': [100, 200, 150, 300],
        'sample2': [110, 220, 140, 280],
        'sample3': [120, 210, 160, 290],
        'sample4': [500, 600, 450, 700],
        'sample5': [480, 580, 470, 680],
        'sample6': [520, 620, 430, 720]
    })
    group1 = ['sample1', 'sample2', 'sample3']
    group2 = ['sample4', 'sample5', 'sample6']
    results = get_roc(df, group1, group2)
    assert isinstance(results, list)
    assert all(isinstance(x, tuple) for x in results)
    df = sample_comp_glycomics_data()
    group1 = [k for k in df.columns if 'control' in k]
    group2 = [k for k in df.columns if 'tumor' in k]
    results = get_roc(df, group1, group2)
    results = get_roc(df, group1, group2, multi_score=True)
    df['sample7'] = [1000, 200, 400, 300, 100]
    df['sample8'] = [1100, 250, 410, 380, 150]
    df['sample9'] = [1050, 180, 380, 290, 120]
    groups = [1, 2]*6 + [3]*3
    results = get_roc(df, groups, [])


def test_get_representative_substructures():
    # Create test data with both monosaccharides and disaccharides
    enrichment_df = pd.DataFrame({
        'motif': [
            'Gal',  # Monosaccharides
            'GlcNAc',
            'Neu5Ac',
            'Gal(b1-3)GalNAc',  # Disaccharides
            'Gal(b1-4)GlcNAc',
            'Neu5Ac(a2-6)Gal'
        ],
        'pval': [0.01, 0.015, 0.02, 0.03, 0.04, 0.045],
        'corr_pval': [0.02, 0.025, 0.03, 0.04, 0.045, 0.049],
        'effect_size': [2.0, 1.8, 1.5, 1.3, 1.2, 1.1]
    })
    results = get_representative_substructures(enrichment_df)
    assert isinstance(results, list)
    assert all(isinstance(x, str) for x in results)
    assert len(results) <= 10  # Function should return up to 10 structures


def test_get_lectin_array():
    df = pd.DataFrame({
        'sample_id': ['sample1', 'sample2', 'sample3', 'sample4', 'sample5', 'sample6'],
        'ConA': [100, 110, 120, 500, 480, 520],
        'WGA': [200, 220, 210, 600, 580, 620],
        'SNA': [150, 140, 160, 450, 470, 430]
    })
    group1 = ['sample1', 'sample2', 'sample3']
    group2 = ['sample4', 'sample5', 'sample6']
    results = get_lectin_array(df, group1, group2)
    assert isinstance(results, pd.DataFrame)
    assert 'score' in results.columns
    results = get_lectin_array(df, [1,2,3], [4,5,6])


@pytest.mark.parametrize("extension", ["csv", "xlsx"])
def test_get_lectin_array_ec_file_extension(extension):
    df = pd.DataFrame({
        'sample_id': ['sample1', 'sample2', 'sample3', 'sample4', 'sample5', 'sample6'],
        'ConA': [100, 110, 120, 500, 480, 520],
        'WGA': [200, 220, 210, 600, 580, 620],
        'SNA': [150, 140, 160, 450, 470, 430]
    })
    group1 = ['sample1', 'sample2', 'sample3']
    group2 = ['sample4', 'sample5', 'sample6']

    tmp_dir = Path("test_tmp")
    tmp_dir.mkdir(exist_ok=True)
    filepath = tmp_dir / f"df.{extension}"
    if extension == "csv":
        df.to_csv(filepath, index=False)
    else:
        df.to_excel(filepath, index=False)
    results = get_lectin_array(str(filepath), group1, group2)
    assert isinstance(results, pd.DataFrame)
    assert 'score' in results.columns
    shutil.rmtree(tmp_dir, ignore_errors=True)


def test_get_lectin_array_ec_duplicate():
    df = pd.DataFrame({
        'sample_id': ['sample1', 'sample2', 'sample3', 'sample4', 'sample5', 'sample6', 'sample6'],
        'ConA': [100, 110, 120, 500, 480, 520, 520],
        'WGA': [200, 220, 210, 600, 580, 620, 620],
        'SNA': [150, 140, 160, 450, 470, 430, 430]
    })
    df = pd.concat([df, pd.DataFrame({
        'sample_id': ['sample1', 'sample2', 'sample3', 'sample4', 'sample5', 'sample6', 'sample6'],
        'WGA': [170, 200, 240, 700, 480, 630, 580],
    })], axis=1)
    group1 = ['sample1', 'sample2', 'sample3']
    group2 = ['sample4', 'sample5', 'sample6']
    with pytest.raises(ValueError):
        get_lectin_array(df, group1, group2)


def test_get_lectin_array_ec_no_group2():
    df = pd.DataFrame({
        'sample_id': ['sample1', 'sample2', 'sample3', 'sample4', 'sample5', 'sample6'],
        'ConA': [100, 110, 120, 500, 480, 520],
        'WGA': [200, 220, 210, 600, 580, 620],
        'SNA': [150, 140, 160, 450, 470, 430]
    })
    group1 = ['treatment', 'treatment', 'treatment', 'control', 'control', 'control']
    results = get_lectin_array(df, group1, None)
    assert isinstance(results, pd.DataFrame)
    assert 'score' in results.columns


@pytest.fixture
def simple_glycans():
    return [
        "Gal(b1-4)Glc-ol",
        "Gal(b1-4)GlcNAc-ol",
        "Neu5Ac(a2-3)[Neu5Ac(a2-6)]Gal(b1-4)Glc-ol",
        "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol"
    ]


# Mock plt.show() to avoid displaying plots during tests
@pytest.fixture(autouse=True)
def mock_show():
    with patch('matplotlib.pyplot.show'):
        yield


# Sample data fixtures
@pytest.fixture
def sample_df():
    data = {
        'glycan': ['Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc',
                   'Man(a1-2)Man(a1-3)[Man(a1-6)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc',
                   'Man(a1-2)Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc'],
        'sample1': [10.0, 20.0, 30.0],
        'sample2': [15.0, 25.0, 35.0],
        'sample3': [12.0, 22.0, 32.0]
    }
    return pd.DataFrame(data)


@pytest.fixture
def sample_abundance_df():
    data = {
        'sample1': [0.2, 0.3, 0.5],
        'sample2': [0.25, 0.35, 0.4],
        'sample3': [0.3, 0.3, 0.4]
    }
    index = ['Man3GlcNAc2', 'Man5GlcNAc2', 'Man9GlcNAc2']
    return pd.DataFrame(data, index=index)


@pytest.fixture
def sample_diff_expr_results():
    data = {
        'Glycan': ['Man3GlcNAc2', 'Man5GlcNAc2', 'Man9GlcNAc2'],
        'Mean abundance': [0.3, 0.35, 0.4],
        'Log2FC': [1.5, -0.5, 0.8],
        'Effect size': [0.6, -0.3, 0.4],
        'p-val': [0.01, 0.04, 0.002],
        'corr p-val': [0.03, 0.06, 0.006]
    }
    return pd.DataFrame(data)


def test_get_coverage_basic(sample_df):
    """Test basic functionality of get_coverage"""
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        get_coverage(sample_df)
        mock_savefig.assert_not_called()


def test_get_coverage_with_filepath(sample_df):
    """Test get_coverage with filepath saving"""
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        get_coverage(sample_df, filepath='test.png')
        mock_savefig.assert_called_once()


def test_get_ma_basic(sample_diff_expr_results):
    """Test basic functionality of get_ma"""
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        get_ma(sample_diff_expr_results)
        mock_savefig.assert_not_called()


def test_get_ma_with_filepath(sample_diff_expr_results):
    """Test get_ma with filepath saving"""
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        get_ma(sample_diff_expr_results, filepath='test.png')
        mock_savefig.assert_called_once()


def test_get_ma_with_custom_thresholds(sample_diff_expr_results):
    """Test get_ma with custom thresholds"""
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        get_ma(sample_diff_expr_results, log2fc_thresh=2, sig_thresh=0.01)
        mock_savefig.assert_not_called()


def test_get_volcano_basic(sample_diff_expr_results):
    """Test basic functionality of get_volcano"""
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        get_volcano(sample_diff_expr_results, n=6)
        mock_savefig.assert_not_called()


def test_get_volcano_with_filepath(sample_diff_expr_results):
    """Test get_volcano with filepath saving"""
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        get_volcano(sample_diff_expr_results, filepath='test.png')
        mock_savefig.assert_called_once()


def test_get_volcano_with_custom_thresholds(sample_diff_expr_results):
    """Test get_volcano with custom thresholds"""
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        get_volcano(sample_diff_expr_results, y_thresh=0.01, x_thresh=1.0)
        mock_savefig.assert_not_called()


def test_get_volcano_with_effect_size(sample_diff_expr_results):
    """Test get_volcano using effect size instead of Log2FC"""
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        get_volcano(sample_diff_expr_results, x_metric='Effect size')
        mock_savefig.assert_not_called()


def test_get_meta_analysis_fixed():
    """Test meta-analysis with fixed effects model"""
    effect_sizes = [0.5, 0.3, 0.7]
    variances = [0.1, 0.15, 0.08]
    combined_effect, p_value = get_meta_analysis(effect_sizes, variances, model='fixed')
    assert isinstance(combined_effect, float)
    assert isinstance(p_value, float)
    assert 0 <= p_value <= 1


def test_get_meta_analysis_random():
    """Test meta-analysis with random effects model"""
    effect_sizes = [0.5, 0.3, 0.7]
    variances = [0.1, 0.15, 0.08]
    combined_effect, p_value = get_meta_analysis(effect_sizes, variances, model='random')
    assert isinstance(combined_effect, float)
    assert isinstance(p_value, float)
    assert 0 <= p_value <= 1


def test_get_meta_analysis_with_study_names():
    """Test meta-analysis with study names"""
    effect_sizes = [0.5, 0.3, 0.7]
    variances = [0.1, 0.15, 0.08]
    study_names = ['Study1', 'Study2', 'Study3']
    with patch('matplotlib.pyplot.subplots') as mock_subplots:
        mock_fig = MagicMock()
        mock_ax = MagicMock()
        mock_subplots.return_value = (mock_fig, mock_ax)
        with patch('matplotlib.pyplot.savefig') as mock_savefig:
            get_meta_analysis(effect_sizes, variances, study_names=study_names, filepath='test.png')
            mock_savefig.assert_called_once()


def test_get_meta_analysis_invalid_model():
    """Test meta-analysis with invalid model specification"""
    effect_sizes = [0.5, 0.3, 0.7]
    variances = [0.1, 0.15, 0.08]
    with pytest.raises(ValueError):
        get_meta_analysis(effect_sizes, variances, model='invalid')


def test_get_meta_analysis_mismatched_lengths():
    """Test meta-analysis with mismatched effect sizes and variances"""
    effect_sizes = [0.5, 0.3, 0.7]
    variances = [0.1, 0.15]
    with pytest.raises(ValueError):
        get_meta_analysis(effect_sizes, variances)


def test_plot_embeddings_basic():
    """Test basic functionality of plot_embeddings"""
    glycans = ['Man3GlcNAc2', 'Man5GlcNAc2', 'Man9GlcNAc2']
    emb = pd.DataFrame(np.random.rand(3, 10))
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        plot_embeddings(glycans, emb)
        mock_savefig.assert_not_called()


def test_plot_embeddings_with_labels():
    """Test plot_embeddings with group labels"""
    glycans = ['Man3GlcNAc2', 'Man5GlcNAc2', 'Man9GlcNAc2']
    emb = pd.DataFrame(np.random.rand(3, 10))
    labels = ['A', 'B', 'A']
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        plot_embeddings(glycans, emb, label_list=labels)
        mock_savefig.assert_not_called()


def test_plot_embeddings_with_shape_feature():
    """Test plot_embeddings with shape feature"""
    glycans = ['Man3GlcNAc2', 'Man5GlcNAc2', 'Man9GlcNAc2']
    emb = pd.DataFrame(np.random.rand(3, 10))
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        plot_embeddings(glycans, emb, shape_feature='Man')
        mock_savefig.assert_not_called()


def test_plot_embeddings_with_filepath(tmp_path):
    """Test plot_embeddings with file saving"""
    glycans = ['Man3GlcNAc2', 'Man5GlcNAc2', 'Man9GlcNAc2']
    emb = pd.DataFrame(np.random.rand(3, 10))
    filepath = tmp_path / "test.png"
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        plot_embeddings(glycans, emb, filepath=str(filepath))
        mock_savefig.assert_called_once()


def test_get_pval_distribution(tmp_path):
    """Test plot_embeddings with file saving"""
    df = pd.DataFrame(np.random.rand(3, 1), columns=['p-val'])
    filepath = tmp_path / "test.png"
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        get_pval_distribution(df, filepath=str(filepath))
        mock_savefig.assert_called_once()


@pytest.fixture
def sample_time_series_df():
    data = {
        'glycan': ['Man3GlcNAc2', 'Man5GlcNAc2', 'Man9GlcNAc2'],
        'T1_h0_r1': [10, 20, 30],
        'T1_h4_r1': [15, 25, 35],
        'T1_h8_r1': [12, 22, 32],
        'T1_h0_r2': [11, 21, 31],
        'T1_h4_r2': [14, 24, 34],
        'T1_h8_r2': [13, 23, 33]
    }
    return pd.DataFrame(data)


@pytest.fixture
def sample_glycoshift_df():
    data = {
        'protein_site_composition': ['ProtA_123_Man3GlcNAc2', 'ProtA_123_Man5GlcNAc2', 'ProtB_456_Man3GlcNAc2'],
        'sample1': [10, 20, 30],
        'sample2': [15, 25, 35],
        'sample3': [12, 22, 32],
        'sample4': [11, 21, 31]
    }
    return pd.DataFrame(data)


def test_characterize_monosaccharide_basic():
    """Test basic functionality of characterize_monosaccharide"""
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        characterize_monosaccharide('Man', rank="Class", focus="Actinopterygii")
        mock_savefig.assert_not_called()


def test_characterize_monosaccharide_with_custom_df(sample_df):
    """Test characterize_monosaccharide with custom DataFrame"""
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        characterize_monosaccharide('Man', df=sample_df, thresh=1)
        mock_savefig.assert_not_called()


def test_characterize_monosaccharide_with_bond_mode():
    """Test characterize_monosaccharide in bond mode"""
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        characterize_monosaccharide('a1-3', mode='bond', rank="Class", focus="Actinopterygii")
        mock_savefig.assert_not_called()


def test_characterize_monosaccharide_with_modifications():
    """Test characterize_monosaccharide with modifications enabled"""
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        characterize_monosaccharide('Man', modifications=True, rank="Class", focus="Actinopterygii")
        mock_savefig.assert_not_called()


def test_get_heatmap_basic(sample_df):
    """Test basic functionality of get_heatmap"""
    result = get_heatmap(sample_df)
    # Add tests for return_plot=True
    result_with_plot = get_heatmap(sample_df, return_plot=True)
    assert result_with_plot is not None
    # Add test for filepath saving
    result_with_save = get_heatmap(sample_df, filepath="test.png")
    presence_df = pd.DataFrame({
        'glycan': [
            'Glc(a1-2)Gal',
            'Man(a1-3)Man',
            'Fuc(a1-2)Gal',
            'Xyl(b1-2)Man'
        ],
        'Fabaceae': [1, 2, 1, 0],
        'Fagaceae': [0, 1, 0, 0],
        'Polygalaceae': [0, 0, 1, 0],
        'Quillajaceae': [0, 0, 0, 1]
    }).set_index('glycan')
    result_presence = get_heatmap(presence_df, datatype='presence')
    result_presence = get_heatmap(presence_df, datatype='presence', motifs=True, feature_set=['exhaustive'])
    try:
        result = get_heatmap(sample_df, motifs=True, feature_set=['custom'])
        return False
    except ValueError:
        pass
    plt.close('all')


def test_get_heatmap_with_motifs(sample_df):
    """Test get_heatmap with motif analysis"""
    result = get_heatmap(sample_df, motifs=True)
    plt.close('all')


def test_get_heatmap_with_transform(sample_df):
    """Test get_heatmap with data transformation"""
    result = get_heatmap(sample_df, transform='CLR')
    result_alr = get_heatmap(sample_df, transform='ALR')
    plt.close('all')


def test_get_heatmap_with_custom_feature_set(sample_df):
    """Test get_heatmap with custom feature set"""
    result = get_heatmap(sample_df, motifs=True, feature_set=['known'], show_all=True)
    plt.close('all')


def test_get_pca_basic(sample_df):
    """Test basic functionality of get_pca"""
    groups = [1, 1, 1]
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        get_pca(sample_df, groups, transform="CLR")
        get_pca(sample_df, groups, transform="ALR")
        mock_savefig.assert_not_called()


def test_get_pca_with_metadata(sample_df):
    """Test get_pca with metadata DataFrame"""
    metadata = pd.DataFrame({'id': ['sample1', 'sample2', 'sample3'],
                           'group': ['A', 'A', 'B']})
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        get_pca(sample_df, metadata)
        mock_savefig.assert_not_called()


def test_get_pca_with_motifs(sample_df):
    """Test get_pca with motif analysis"""
    groups = [1, 1, 1]
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        get_pca(sample_df, groups, motifs=True)
        mock_savefig.assert_not_called()


def test_get_pca_with_custom_components(sample_df):
    """Test get_pca with custom principal components"""
    groups = [1, 1, 1]
    with patch('matplotlib.pyplot.savefig') as mock_savefig:
        get_pca(sample_df, groups, pc_x=2, pc_y=3)
        mock_savefig.assert_not_called()


@pytest.fixture
def sample_jtk_df():
    """
    Creates a test dataset for JTK analysis with:
    - 24h coverage with 4h intervals (0,4,8,12,16,20)
    - 2 replicates per timepoint
    - 4 glycans with different rhythmic patterns:
      - Man3GlcNAc2: Strong 24h rhythm
      - Man5GlcNAc2: Weak 12h rhythm
      - Man9GlcNAc2: Extremely weak 12h rhythm
      - Biantennary: No rhythm (control)
    """
    # Generate timepoints every 4 hours for 24 hours
    timepoints = [0, 3, 6, 9, 12, 15, 18, 21]
    columns = ['glycan'] + [f'T_h{t}_r{r}' for t in timepoints for r in [1, 2, 3]]
    # Create rhythmic patterns
    # Man3GlcNAc2: 24h cycle (peak at 12h)
    h24_pattern = [10, 15, 25, 27, 33, 25, 15, 10]  # Base values for 24h rhythm
    # Man5GlcNAc2: 12h cycle (peaks at 0h and 12h)
    h12_pattern = [25, 15, 17, 20, 25, 18, 13, 19]  # Base values for 12h rhythm
    # Man9GlcNAc2: Weak rhythm (tiny 12h amplitude)
    weak_rhythm = [20, 21, 19, 21, 22, 20, 21, 20]    # Stable values with small variation
    # Biantennary: No rhythm (random variation around mean)
    no_rhythm = [20, 20, 20, 20, 20, 20, 20, 20]    # Stable values with small variation
    # Add some noise to replicates (±10% variation)
    data = {
        'glycan': ['Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc',
                   'Man(a1-2)Man(a1-3)[Man(a1-6)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc',
                   'Man(a1-2)Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-3)[Man(a1-2)Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc',
                   'Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc']
    }
    # Add two replicates for each timepoint with small random variation
    np.random.seed(42)  # For reproducibility
    for t_idx, t in enumerate(timepoints):
        for r in [1, 2, 3]:
            col = f'T_h{t}_r{r}'
            base_values = [h24_pattern[t_idx], h12_pattern[t_idx], weak_rhythm[t_idx], no_rhythm[t_idx]]
            # Add 5% random noise
            noise = np.random.normal(0, 0.05, 4) * np.array(base_values)
            data[col] = np.array(base_values) + noise
    return pd.DataFrame(data)


def test_get_jtk_basic(sample_jtk_df):
    """Test basic functionality of get_jtk"""
    periods = [12, 24]
    result = get_jtk(sample_jtk_df, timepoints=8, interval=3, periods=periods)
    assert isinstance(result, pd.DataFrame)
    assert 'Adjusted_P_value' in result.columns
    assert 'Period_Length' in result.columns
    result = get_jtk(sample_jtk_df, timepoints=8, interval=3, periods=periods, transform="ALR")
    result = get_jtk(sample_jtk_df, timepoints=8, interval=3, periods=periods, transform="Nothing")
    assert sum(result.significant) == 2
    try:
        result = get_jtk(sample_jtk_df, timepoints=8, interval=3, periods=periods, transform="wrong")
        return False
    except:
        pass


def test_get_jtk_with_motifs(sample_jtk_df):
    """Test get_jtk with motif analysis"""
    periods = [12, 24]
    result = get_jtk(sample_jtk_df, timepoints=8, interval=3, periods=periods, motifs=True)
    assert isinstance(result, pd.DataFrame)
    assert 'Adjusted_P_value' in result.columns


def test_get_jtk_with_transform(sample_jtk_df):
    """Test get_jtk with data transformation"""
    periods = [12, 24]
    result = get_jtk(sample_jtk_df, timepoints=8, interval=3, periods=periods, transform='CLR')
    assert isinstance(result, pd.DataFrame)


def test_multi_feature_scoring_basic(sample_df):
    """Test basic functionality of multi_feature_scoring"""
    np.random.seed(42)
    n_features = 5
    n_samples = 10
    data = np.random.rand(n_features, n_samples)
    # Make the first feature strongly predictive
    data[0, :5] = 0.1  # Clear signal for group 1
    data[0, 5:] = 0.9  # Clear signal for group 2
    # Make the second feature moderately predictive
    data[1, :5] = 0.3
    data[1, 5:] = 0.7
    df_transformed = pd.DataFrame(data)
    group1 = [0, 1, 2, 3, 4]
    group2 = [5, 6, 7, 8, 9]
    with patch('matplotlib.pyplot.subplots') as mock_subplots:
        mock_fig = MagicMock()
        mock_ax = MagicMock()
        mock_subplots.return_value = (mock_fig, mock_ax)
        model, roc_auc = multi_feature_scoring(df_transformed, group1, group2)
        assert hasattr(model, 'predict')
        assert 0 <= roc_auc <= 1
        assert roc_auc > 0.5


def test_multi_feature_scoring_with_filepath(sample_df):
    """Test multi_feature_scoring with file saving"""
    np.random.seed(42)
    n_features = 5
    n_samples = 10
    data = np.random.rand(n_features, n_samples)
    # Make the first feature strongly predictive
    data[0, :5] = 0.1
    data[0, 5:] = 0.9
    df_transformed = pd.DataFrame(data)
    group1 = [0, 1, 2, 3, 4]
    group2 = [5, 6, 7, 8, 9]
    with patch('matplotlib.pyplot.subplots') as mock_subplots:
        mock_fig = MagicMock()
        mock_ax = MagicMock()
        mock_subplots.return_value = (mock_fig, mock_ax)
        with patch('matplotlib.pyplot.savefig') as mock_savefig:
            multi_feature_scoring(df_transformed, group1, group2, filepath='test.png')
            mock_savefig.assert_called_once()


def test_multi_feature_scoring_imbalanced_groups(sample_df):
    """Test multi_feature_scoring with imbalanced groups"""
    np.random.seed(42)
    n_features = 5
    n_samples = 10
    data = np.random.rand(n_features, n_samples)
    # Make the first feature strongly predictive
    data[0, :4] = 0.1  # Group 1 (4 samples)
    data[0, 4:] = 0.9  # Group 2 (6 samples)
    df_transformed = pd.DataFrame(data)
    group1 = [0, 1, 2, 3]
    group2 = [4, 5, 6, 7, 8, 9]
    with patch('matplotlib.pyplot.figure'):
        model, roc_auc = multi_feature_scoring(df_transformed, group1, group2)
    assert hasattr(model, 'predict')
    assert 0 <= roc_auc <= 1
    assert roc_auc > 0.5


def test_multi_feature_scoring_no_group2(sample_df):
    """Test basic functionality of multi_feature_scoring"""
    np.random.seed(42)
    n_features = 5
    n_samples = 10
    data = np.random.rand(n_features, n_samples)
    # Make the first feature strongly predictive
    data[0, :5] = 0.1  # Clear signal for group 1
    data[0, 5:] = 0.9  # Clear signal for group 2
    # Make the second feature moderately predictive
    data[1, :5] = 0.3
    data[1, 5:] = 0.7
    df_transformed = pd.DataFrame(data)
    group1 = [0, 0, 0, 0, 0, 1, 1, 1, 1, 1]
    with patch('matplotlib.pyplot.subplots') as mock_subplots:
        mock_fig = MagicMock()
        mock_ax = MagicMock()
        mock_subplots.return_value = (mock_fig, mock_ax)
        model, roc_auc = multi_feature_scoring(df_transformed, group1, None)
        assert hasattr(model, 'predict')
        assert 0 <= roc_auc <= 1
        assert roc_auc > 0.5


def test_get_glycoshift_per_site_basic(sample_glycoshift_df):
    """Test basic functionality of get_glycoshift_per_site"""
    group1 = ['sample1', 'sample2']
    group2 = ['sample3', 'sample4']
    result = get_glycoshift_per_site(sample_glycoshift_df, group1, group2)
    assert isinstance(result, pd.DataFrame)
    assert 'Condition_corr_pval' in result.columns


def test_get_glycoshift_per_site_with_custom_params(sample_glycoshift_df):
    """Test get_glycoshift_per_site with custom parameters"""
    group1 = ['sample1', 'sample2']
    group2 = ['sample3', 'sample4']
    result = get_glycoshift_per_site(sample_glycoshift_df, group1, group2,
                                    min_samples=0.3, gamma=0.2)
    assert isinstance(result, pd.DataFrame)


def test_get_glycoshift_per_site_paired(sample_glycoshift_df):
    """Test get_glycoshift_per_site with paired samples"""
    group1 = ['sample1', 'sample2']
    group2 = ['sample3', 'sample4']
    result = get_glycoshift_per_site(sample_glycoshift_df, group1, group2,
                                    paired=True)
    assert isinstance(result, pd.DataFrame)


def test_get_glycoshift_per_site_no_imputation(sample_glycoshift_df):
    """Test get_glycoshift_per_site without imputation"""
    group1 = ['sample1', 'sample2']
    group2 = ['sample3', 'sample4']
    result = get_glycoshift_per_site(sample_glycoshift_df, group1, group2,
                                    impute=False)
    assert isinstance(result, pd.DataFrame)


@pytest.fixture
def simple_network():
    # Create a simple directed network
    G = nx.DiGraph()
    G.add_edge("Gal(b1-4)Glc-ol", "GlcNAc(b1-3)Gal(b1-4)Glc-ol", diffs="GlcNAc(b1-3)")
    G.add_edge("GlcNAc(b1-3)Gal(b1-4)GlcNAc-ol", "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol", diffs="Gal(b1-4)")
    nx.set_node_attributes(G, {
        "Gal(b1-4)Glc-ol": {"virtual": 0},
        "Gal(b1-4)GlcNAc-ol": {"virtual": 0},
        "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol": {"virtual": 0}
    })
    return G


def test_safe_compare():
    g1 = glycan_to_nxGraph("Gal(b1-4)Glc-ol")
    g2 = glycan_to_nxGraph("Gal(b1-4)Glc-ol")
    g3 = glycan_to_nxGraph("Gal(b1-4)GlcNAc-ol")
    assert safe_compare(g1, g2) == True
    assert safe_compare(g1, g3) == False
    assert safe_compare(g1, None) == False


def test_safe_index():
    glycan = "Gal(b1-4)Glc-ol"
    graph_dic = {}
    result = safe_index(glycan, graph_dic)
    assert isinstance(result, nx.Graph)
    assert glycan in graph_dic
    assert graph_dic[glycan] == result


def test_get_neighbors(simple_glycans):
    ggraph = glycan_to_nxGraph("GlcNAc(b1-3)Gal(b1-4)Glc-ol")
    graphs = [glycan_to_nxGraph(g) for g in simple_glycans]
    neighbors, indices = get_neighbors(ggraph, simple_glycans, graphs)
    assert len(neighbors) > 0
    assert all(isinstance(n, str) for n in neighbors)
    assert all(isinstance(idx, list) for idx in indices)
    neighbors, indices = get_neighbors(glycan_to_nxGraph("Glc-ol"), simple_glycans, graphs)
    assert neighbors == []


def test_create_neighbors():
    glycan = "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol"
    ggraph = glycan_to_nxGraph(glycan)
    neighbors = create_neighbors(ggraph)
    assert len(neighbors) > 0
    assert all(isinstance(n, nx.Graph) for n in neighbors)
    assert len(neighbors[0].nodes()) < len(ggraph.nodes())
    neighbors = create_neighbors(glycan_to_nxGraph("Glc-ol"))
    assert neighbors == []


def test_find_diff():
    graph_dic = {}
    glycan_a = "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol"
    glycan_b = "Gal(b1-4)Glc-ol"
    diff = find_diff(glycan_a, glycan_b, graph_dic)
    assert isinstance(diff, str)
    assert "GlcNAc" in diff
    assert find_diff(glycan_a, glycan_a, graph_dic) == ""
    assert find_diff(glycan_a, "Gal(b1-3)[GlcNAc(b1-6)]GalNAc", graph_dic) == "disregard"
    a = "{Fuc(a1-2/3/6)}Gal(b1-4)GlcNAc(b1-2)Man(a1-3/6)[Gal(b1-4)GlcNAc(b1-2/4/6)[Gal(b1-4)GlcNAc(b1-2/4/6)]Man(a1-3/6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    b = "{Fuc(a1-2/3/6)}{Neu5Ac(a2-3/6/8)}Gal(b1-4)GlcNAc(b1-2)Man(a1-3/6)[GlcNAc(b1-2/4/6)[GlcNAc(b1-2/4/6)]Man(a1-3/6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"
    assert find_diff(a, b, {}) == "disregard"


def test_find_path():
    graph_dic = {}
    glycan_a = "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol"
    glycan_b = "Gal(b1-4)Glc-ol"
    edges, labels = find_path(glycan_a, glycan_b, graph_dic)
    assert isinstance(edges, list)
    assert isinstance(labels, dict)
    assert len(edges) > 0
    assert all(isinstance(e, tuple) for e in edges)
    glycan_a = "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol"
    edges, labels = find_path(glycan_a, glycan_b, graph_dic)
    assert edges == []


def test_construct_network(simple_glycans):
    network = construct_network(simple_glycans)
    assert isinstance(network, nx.DiGraph)
    assert len(network.nodes()) >= len(simple_glycans)
    assert len(network.edges()) > 0
    assert all("diffs" in data for _, _, data in network.edges(data=True))
    network = construct_network(simple_glycans, edge_type="enzyme")
    network = construct_network(simple_glycans, edge_type="monosaccharide")


def test_prune_network(simple_glycans):
    net = construct_network(simple_glycans)
    nx.set_node_attributes(net, {
        "Gal(b1-4)Glc-ol": {"abundance": 0.5},
        "Neu5Ac(a2-6)Gal(b1-4)Glc-ol": {"abundance": 0.0},
        "Neu5Ac(a2-3)Gal(b1-4)Glc-ol": {"abundance": 0.5},
        "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol": {"abundance": 0.8}
    })
    pruned = prune_network(net, threshold=0.1)
    assert isinstance(pruned, nx.DiGraph)
    assert len(pruned.nodes()) < len(net.nodes())
    assert "Neu5Ac(a2-6)Gal(b1-4)Glc-ol" not in pruned.nodes()


def test_extend_glycans():
    glycans = {"Gal(b1-4)Glc-ol"}
    reactions = {"Fuc(a1-3)", "GlcNAc(b1-3)"}
    new_glycans = extend_glycans(glycans, reactions)
    assert isinstance(new_glycans, set)
    assert len(new_glycans) > 0
    assert all(isinstance(g, str) for g in new_glycans)
    assert all(g.endswith('Glc-ol') for g in new_glycans)


def test_highlight_network_motif(simple_network):
    motif = "Gal(b1-4)Glc"
    highlighted = highlight_network(simple_network, highlight="motif", motif=motif)
    assert isinstance(highlighted, nx.DiGraph)
    assert all("origin" in data for _, data in highlighted.nodes(data=True))
    assert all(data["origin"] in ["limegreen", "darkviolet"]
              for _, data in highlighted.nodes(data=True))
    highlighted = highlight_network(simple_network, highlight="motif", motif="Gal(b1-4)")
    highlighted = highlight_network(simple_network, highlight="motif", motif="r.-.")
    highlighted = highlight_network(simple_network, highlight="species", species='Lama_pacos')
    try:
        highlighted = highlight_network(simple_network, highlight="wrong")
        return False
    except ValueError:
        pass
    try:
        highlighted = highlight_network(simple_network, highlight="motif")
        return False
    except ValueError:
        pass
    try:
        highlighted = highlight_network(simple_network, highlight="species")
        return False
    except ValueError:
        pass
    try:
        highlighted = highlight_network(simple_network, highlight="abundance")
        return False
    except ValueError:
        pass
    try:
        highlighted = highlight_network(simple_network, highlight="conservation")
        return False
    except ValueError:
        pass


def test_highlight_network_abundance(simple_network):
    abundance_df = pd.DataFrame({
        "glycan": list(simple_network.nodes()),
        "rel_intensity": [0.5] * len(simple_network)
    })
    highlighted = highlight_network(simple_network, highlight="abundance",
                                  abundance_df=abundance_df)
    assert isinstance(highlighted, nx.DiGraph)
    assert all("abundance" in data for _, data in highlighted.nodes(data=True))
    assert all(isinstance(data["abundance"], (int, float))
              for _, data in highlighted.nodes(data=True))


@pytest.fixture
def sample_network():
    G = nx.DiGraph()
    # Create a diamond-like structure
    G.add_edge("Gal(b1-4)Glc-ol", "Neu5Ac(a2-3)Gal(b1-4)Glc-ol", diffs="Neu5Ac(a2-3)")
    G.add_edge("Gal(b1-4)Glc-ol", "Neu5Ac(a2-6)Gal(b1-4)Glc-ol", diffs="Neu5Ac(a2-6)")
    G.add_edge("Neu5Ac(a2-3)Gal(b1-4)Glc-ol", "Neu5Ac(a2-3)[Neu5Ac(a2-6)]Gal(b1-4)Glc-ol", diffs="Neu5Ac(a2-6)")
    G.add_edge("Neu5Ac(a2-6)Gal(b1-4)Glc-ol", "Neu5Ac(a2-3)[Neu5Ac(a2-6)]Gal(b1-4)Glc-ol", diffs="Neu5Ac(a2-3)")
    G.add_edge("Gal(b1-4)Glc-ol", "Fuc(a1-2)Gal(b1-4)Glc-ol", diffs="Fuc(a1-2)")
    G.add_edge("Fuc(a1-2)Gal(b1-4)Glc-ol", "Fuc(a1-2)Gal(b1-4)[Fuc(a1-3)]Glc-ol", diffs="Fuc(a1-3)")
    # Set node attributes
    nx.set_node_attributes(G, {
        "Gal(b1-4)Glc-ol": {"virtual": 0, "abundance": 10.0},
        "Neu5Ac(a2-3)Gal(b1-4)Glc-ol": {"virtual": 0, "abundance": 5.0},
        "Neu5Ac(a2-6)Gal(b1-4)Glc-ol": {"virtual": 1, "abundance": 3.0},
        "Neu5Ac(a2-3)[Neu5Ac(a2-6)]Gal(b1-4)Glc-ol": {"virtual": 0, "abundance": 7.0},
        "Fuc(a1-2)Gal(b1-4)Glc-ol": {"virtual": 0, "abundance": 1.0},
        "Fuc(a1-2)Gal(b1-4)[Fuc(a1-3)]Glc-ol": {"virtual": 0, "abundance": 0.5}
    })
    return G


@pytest.fixture
def abundance_data():
    return pd.DataFrame({
        'glycan': ["Gal(b1-4)Glc-ol", "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol",
                  "Fuc(a1-2)Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol"],
        'sample1': [10.0, 5.0, 7.0],
        'sample2': [8.0, 6.0, 8.0],
        'sample3': [9.0, 4.0, 6.0],
        'sample4': [11.0, 7.0, 9.0]
    })


def test_infer_roots():
    # Test O-linked glycans
    o_glycans = {"GalNAc(b1-3)Gal", "Fuc(a1-2)GalNAc"}
    roots = infer_roots(frozenset(o_glycans))
    assert "GalNAc" in roots
    assert "Fuc" in roots
    # Test free glycans
    free_glycans = {"Gal(b1-4)Glc-ol", "Gal(b1-4)GlcNAc-ol"}
    roots = infer_roots(frozenset(free_glycans))
    assert "Gal(b1-4)Glc-ol" in roots
    assert "Gal(b1-4)GlcNAc-ol" in roots
    # Test N-linked glycans
    n_glycans = {"Neu5Ac(a2-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc"}
    roots = infer_roots(frozenset(n_glycans))
    assert "Man(b1-4)GlcNAc(b1-4)GlcNAc" in roots
    # Test glycolipids
    glycolipids = {"Neu5Ac(a2-3)Gal(b1-4)Glc1Cer", "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc1Cer"}
    roots = infer_roots(frozenset(glycolipids))
    assert "Glc1Cer" in roots
    # Test poorly formatted glycolipids
    glycolipids = {"Neu5Ac(a2-3)Gal(b1-4)Glc", "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc"}
    roots = infer_roots(frozenset(glycolipids))
    assert "Gal(b1-4)Glc" in roots
    # Test fallback case (unrecognized glycan class)
    unknown_glycans = {"XyzUnknownStructure", "SomeOtherUnknownFormat"}
    roots = infer_roots(frozenset(unknown_glycans))
    assert len(roots) == 0  # Returns empty frozenset


def test_deorphanize_nodes(sample_network):
    # Add an orphaned node
    sample_network.add_node("GlcNAc(b1-3)Gal(b1-4)Glc-ol", virtual=0)
    deorphanized = deorphanize_nodes(sample_network, {})
    assert isinstance(deorphanized, nx.DiGraph)
    assert len(deorphanized.edges()) > len(sample_network.edges())
    # Check if orphaned node is now connected
    assert deorphanized.degree("GlcNAc(b1-3)Gal(b1-4)Glc-ol") > 0
    deorphanized = deorphanize_nodes(sample_network, {}, permitted_roots = frozenset("GalNAc"))


def test_get_edge_weight_by_abundance(sample_network):
    weighted = get_edge_weight_by_abundance(sample_network)
    assert isinstance(weighted, nx.DiGraph)
    assert all("capacity" in data for _, _, data in weighted.edges(data=True))
    assert all(isinstance(data["capacity"], (int, float))
              for _, _, data in weighted.edges(data=True))


def test_find_diamonds(simple_glycans):
    net = construct_network(simple_glycans)
    diamonds = find_diamonds(net)
    assert isinstance(diamonds, list)
    assert len(diamonds) > 0
    assert all(isinstance(d, dict) for d in diamonds)
    assert all(len(d) >= 4 for d in diamonds)  # Diamond should have at least 4 nodes
    diamonds = find_diamonds(net_dic['Lama_pacos'], nb_intermediates=4)
    try:
        diamonds = find_diamonds(net, nb_intermediates=3)
        return False
    except ValueError:
        pass


def test_trace_diamonds(simple_glycans):
    net = construct_network(simple_glycans)
    species_list = ["species1"]
    network_dic = {"species1": net}
    results = trace_diamonds(net, species_list, network_dic)
    assert isinstance(results, pd.DataFrame)
    assert "probability" in results.columns
    assert len(results) > 0


def test_get_maximum_flow(sample_network):
    sample_network = estimate_weights(sample_network, root = "Gal(b1-4)Glc-ol", min_default = 0.1)
    flow_results = get_maximum_flow(sample_network,
                                  source="Gal(b1-4)Glc-ol")
    assert isinstance(flow_results, dict)
    assert all("flow_value" in data for data in flow_results.values())
    assert all("flow_dict" in data for data in flow_results.values())
    assert all(isinstance(data["flow_value"], (int, float))
              for data in flow_results.values())


def test_get_reaction_flow(sample_network):
    sample_network = estimate_weights(sample_network, root = "Gal(b1-4)Glc-ol", min_default = 0.1)
    flow_results = get_maximum_flow(sample_network,
                                  source="Gal(b1-4)Glc-ol")
    reaction_flows = get_reaction_flow(sample_network, flow_results, aggregate="sum")
    assert isinstance(reaction_flows, dict)
    assert len(reaction_flows) > 0
    assert all(isinstance(v, (int, float)) for v in reaction_flows.values())
    reaction_flows = get_reaction_flow(sample_network, flow_results, aggregate="mean")


def test_process_ptm():
    glycans = ["Gal(b1-4)Glc3S-ol", "Gal(b1-4)Glc-ol"]
    stem_lib = get_stem_lib(get_lib(glycans))
    edges, labels = process_ptm(glycans, {}, stem_lib)
    if edges:  # If PTMs were found
        assert len(edges) == len(labels)
        assert all(isinstance(e, tuple) for e in edges)
        assert all(isinstance(l, str) for l in labels)


def test_get_differential_biosynthesis(abundance_data):
    results = get_differential_biosynthesis(
        abundance_data,
        group1=['sample1', 'sample2'],
        group2=['sample3', 'sample4'],
        analysis="reaction"
    )
    assert isinstance(results, pd.DataFrame)
    assert "Log2FC" in results.columns
    assert "p-val" in results.columns
    assert "significant" in results.columns
    assert len(results) > 0
    results = get_differential_biosynthesis(
        abundance_data,
        group1=[1,2],
        group2=[3,4],
        analysis="flow",
        paired=True
    )
    try:
        results = get_differential_biosynthesis(
            abundance_data,
            group1=[1,2],
            group2=[3,4],
            analysis="wrong",
            paired=True
            )
        return False
    except ValueError:
        pass


def test_deorphanize_edge_labels(sample_network):
    # Remove some edge labels
    edge_list = list(sample_network.edges())[0:2]
    for edge in edge_list:
        del sample_network.edges[edge]["diffs"]
    labeled = deorphanize_edge_labels(sample_network, {})
    assert isinstance(labeled, nx.DiGraph)
    assert all("diffs" in data for _, _, data in labeled.edges(data=True))


def test_get_differential_biosynthesis_longitudinal(abundance_data):
    # Add ID column for longitudinal analysis
    abundance_data = abundance_data.copy().set_index('glycan').T
    abundance_data['ID'] = ['p1_t1_1', 'p1_t2_1', 'p2_t1_1', 'p2_t2_1']
    results = get_differential_biosynthesis(
        abundance_data,
        group1=['t1', 't2'],
        longitudinal=True,
        id_column='ID'
    )
    assert isinstance(results, pd.DataFrame)
    assert "F-statistic" in results.columns
    assert "Direction" in results.columns
    assert "Average Slope" in results.columns
    assert len(results) > 0


def test_process_ptm_with_multiple_ptms():
    glycans = ["Gal(b1-4)Glc3S6S-ol", "Gal(b1-4)Glc-ol", "Gal(b1-4)Glc6S-ol"]
    stem_lib = get_stem_lib(get_lib(glycans))
    edges, labels = process_ptm(glycans, {}, stem_lib)
    if edges:
        assert len(edges) > 1  # Should find multiple PTM edges
        assert all(isinstance(e, tuple) for e in edges)
        assert all(isinstance(l, str) for l in labels)


def test_get_maximum_flow_with_no_path(sample_network):
    # Add isolated node with no path from source
    sample_network.add_node("IsolatedGlycan", virtual=0, abundance=1.0)
    sample_network = estimate_weights(sample_network, root = "Gal(b1-4)Glc-ol", min_default = 0.1)
    flow_results = get_maximum_flow(sample_network,
                                  source="Gal(b1-4)Glc-ol",
                                  sinks=["IsolatedGlycan"])
    assert isinstance(flow_results, dict)
    assert len(flow_results) == 0  # Should return empty dict for unreachable sink


@pytest.fixture
def simple_networks():
    # Create two simple test networks
    network_a = nx.Graph()
    network_a.add_nodes_from(['A', 'B', 'C'], virtual=0)
    network_a.add_edges_from([('A', 'B'), ('B', 'C')])
    network_b = nx.Graph()
    network_b.add_nodes_from(['B', 'C', 'D'], virtual=0)
    network_b.add_edges_from([('B', 'C'), ('C', 'D')])
    return network_a, network_b


@pytest.fixture
def networks_with_virtuals():
    # Create networks with virtual nodes
    network_a = nx.Graph()
    network_a.add_nodes_from(['A', 'B', 'V1', 'V3'])
    network_a.add_edges_from([('A', 'V1'), ('V1', 'B'), ('V3', 'V1')])
    nx.set_node_attributes(network_a, {'A': 0, 'B': 0, 'V1': 1, 'V3': 0}, 'virtual')
    network_b = nx.Graph()
    network_b.add_nodes_from(['B', 'C', 'V1', 'V2', 'V3'])
    network_b.add_edges_from([('B', 'V1'), ('V1', 'C'), ('C', 'V2'), ('V3', 'V1')])
    nx.set_node_attributes(network_b, {'B': 0, 'C': 0, 'V1': 1, 'V2': 1, 'V3': 1}, 'virtual')
    return network_a, network_b


@pytest.fixture
def mock_df_enzyme():
    # Create mock enzyme mapping data
    data = {
        'monolink': ['Gal(b1-4)', 'GlcNAc(b1-3)', 'Fuc(a1-2)'],
        'glycoenzyme': ['B4GALT1', 'B3GNT2', 'FUT2'],
        'glycoclass': ['B4GALT', 'B3GNT', 'FUT']
    }
    return pd.DataFrame(data)


def test_network_alignment_basic(simple_networks):
    network_a, network_b = simple_networks
    combined = network_alignment(network_a, network_b)
    # Test node presence
    assert set(combined.nodes()) == {'A', 'B', 'C', 'D'}
    # Test edge presence
    assert ('A', 'B') in combined.edges()
    assert ('B', 'C') in combined.edges()
    assert ('C', 'D') in combined.edges()
    # Test node colors
    node_colors = nx.get_node_attributes(combined, 'origin')
    assert node_colors['A'] == 'cornflowerblue'  # only in network_a
    assert node_colors['B'] == 'saddlebrown'     # in both networks
    assert node_colors['D'] == 'darkorange'      # only in network_b


def test_network_alignment_empty():
    empty_network = nx.Graph()
    network = nx.Graph()
    network.add_nodes_from(['A', 'B'])
    network.add_edge('A', 'B')
    combined = network_alignment(empty_network, network)
    assert set(combined.nodes()) == {'A', 'B'}
    assert ('A', 'B') in combined.edges()


def test_infer_virtual_nodes_basic(networks_with_virtuals):
    network_a, network_b = networks_with_virtuals
    inferred_a, inferred_b = infer_virtual_nodes(network_a, network_b)
    # Test structure of returned tuples
    assert len(inferred_a) == 2  # (inferred nodes, supported nodes)
    assert len(inferred_b) == 2
    # Test virtual node inference
    assert 'V1' in inferred_a[1]  # V1 is shared virtual between networks
    assert 'V3' in inferred_b[0]  # V3 is virtual in network_b and observed in network_a


def test_retrieve_inferred_nodes():
    network = nx.Graph()
    network.add_nodes_from(['A', 'B', 'V1', 'V2'])
    nx.set_node_attributes(network, {'A': 0, 'B': 0, 'V1': 2, 'V2': 2}, 'virtual')
    # Test without species parameter
    inferred = retrieve_inferred_nodes(network)
    assert set(inferred) == {'V1', 'V2'}
    # Test with species parameter
    inferred_with_species = retrieve_inferred_nodes(network, species='test_species')
    assert isinstance(inferred_with_species, dict)
    assert 'test_species' in inferred_with_species
    assert set(inferred_with_species['test_species']) == {'V1', 'V2'}


def test_monolink_to_glycoenzyme(mock_df_enzyme):
    # Test condensed mode
    assert monolink_to_glycoenzyme('Gal(b1-4)', mock_df_enzyme, mode='condensed') == 'B4GALT'
    assert monolink_to_glycoenzyme('GlcNAc(b1-3)', mock_df_enzyme, mode='condensed') == 'B3GNT'
    # Test full mode
    assert monolink_to_glycoenzyme('Gal(b1-4)', mock_df_enzyme, mode='full') == 'B4GALT1'
    assert monolink_to_glycoenzyme('Fuc(a1-2)', mock_df_enzyme, mode='full') == 'FUT2'
    # Test unknown monolink
    assert monolink_to_glycoenzyme('Unknown(x1-1)', mock_df_enzyme) == 'Unknown(x1-1)'


def test_monolink_to_glycoenzyme_empty_df():
    empty_df = pd.DataFrame(columns=['monolink', 'glycoenzyme', 'glycoclass'])
    edge_label = 'Gal(b1-4)'
    assert monolink_to_glycoenzyme(edge_label, empty_df) == edge_label


@pytest.fixture
def flow_network():
    """Create a simple directed network with capacities for flow testing"""
    network = nx.DiGraph()
    # Add nodes
    nodes = ['A', 'B', 'C', 'D', 'E']
    network.add_nodes_from(nodes)
    # Add edges with capacities
    edges_with_capacity = [
        ('A', 'B', {'capacity': 5.0}),
        ('A', 'C', {'capacity': 3.0}),
        ('B', 'D', {'capacity': 4.0}),
        ('C', 'D', {'capacity': 2.0}),
        ('C', 'E', {'capacity': 2.0}),
        ('D', 'E', {'capacity': 3.0})
    ]
    network.add_edges_from(edges_with_capacity)
    # Add abundance attributes
    nx.set_node_attributes(network, {
        'A': 10.0, 'B': 5.0, 'C': 3.0, 'D': 4.0, 'E': 2.0
    }, 'abundance')
    return network


@pytest.fixture
def glycan_extension_data():
    """Create test data for glycan extension functions"""
    # Sample glycans and their graph representations
    test_glycans = {
        "Gal(b1-4)Glc": glycan_to_nxGraph("Gal(b1-4)Glc"),
        "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc": glycan_to_nxGraph("Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc"),
        "Fuc(a1-2)Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc": glycan_to_nxGraph("Fuc(a1-2)Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc")
    }
    leaf_glycans = {"Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc"}
    new_glycans = {"Fuc(a1-2)Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc"}
    return test_glycans, leaf_glycans, new_glycans


def test_get_max_flow_path_basic(flow_network):
    """Test basic path finding in a flow network"""
    flow_dict = {
        'A': {'B': 4.0, 'C': 2.0},
        'B': {'D': 4.0},
        'C': {'D': 1.0, 'E': 1.0},
        'D': {'E': 2.0}
    }
    path = get_max_flow_path(flow_network, flow_dict, sink='E', source='A')
    # Should choose path A->B->D->E due to higher capacities
    assert path == [('A', 'B'), ('B', 'D'), ('D', 'E')]


def test_get_max_flow_path_no_path(flow_network):
    """Test behavior when no path exists"""
    flow_dict = {
        'A': {'B': 0.0, 'C': 0.0},
        'B': {'D': 0.0},
        'C': {'D': 0.0, 'E': 0.0},
        'D': {'E': 0.0}
    }
    flow_network.remove_edge('B', 'D')
    with pytest.raises(ValueError, match="No path found"):
        get_max_flow_path(flow_network, flow_dict, sink='E', source='A')


def test_edges_for_extension(glycan_extension_data):
    """Test creation of edges between leaf and extended glycans"""
    graphs, leaf_glycans, new_glycans = glycan_extension_data
    edges, labels = edges_for_extension(leaf_glycans, new_glycans, graphs)
    # Should create edge from Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc to Fuc(a1-2)Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc
    assert len(edges) == 1
    assert edges[0][0] == "Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc"
    assert edges[0][1] == "Fuc(a1-2)Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc"
    # Check edge label (should be Fuc(a1-2))
    assert len(labels) == 1
    assert "Fuc(a1-2)" in labels[0]


def test_edges_for_extension_no_connections(glycan_extension_data):
    """Test behavior when no valid connections exist"""
    graphs, leaf_glycans, _ = glycan_extension_data
    # Try to connect to incompatible glycans
    incompatible_glycans = {"GalNAc(a1-3)GlcNAc"}
    edges, labels = edges_for_extension(leaf_glycans, incompatible_glycans, graphs)
    assert len(edges) == 0
    assert len(labels) == 0


def test_choose_leaves_to_extend_exact_match():
    """Test selection of leaves when exact composition match exists"""
    leaf_glycans = {
        "Gal(b1-4)Glc",
        "GlcNAc(b1-3)Gal(b1-4)Glc",
        "Fuc(a1-2)Gal(b1-4)Glc"
    }
    target_composition = {"Hex": 2, "HexNAc": 1}
    selected_leaves = choose_leaves_to_extend(leaf_glycans, target_composition)
    assert "GlcNAc(b1-3)Gal(b1-4)Glc" in selected_leaves
    assert len(selected_leaves) == 1


def test_choose_leaves_to_extend_closest_match():
    """Test selection of leaves closest to target composition"""
    leaf_glycans = {
        "Gal(b1-4)Glc",  # 2 Hex
        "GlcNAc(b1-3)Gal(b1-4)Glc"  # 2 Hex, 1 HexNAc
    }
    target_composition = {"Hex": 2, "HexNAc": 2}
    selected_leaves = choose_leaves_to_extend(leaf_glycans, target_composition)
    # Should select the glycan with HexNAc as it's closer to target
    assert "GlcNAc(b1-3)Gal(b1-4)Glc" in selected_leaves
    assert len(selected_leaves) == 1


def test_choose_leaves_to_extend_invalid_composition():
    """Test behavior with incompatible target composition"""
    leaf_glycans = {"Gal(b1-4)Glc"}
    target_composition = {"NeuAc": 1}  # Not present in any leaf
    selected_leaves = choose_leaves_to_extend(leaf_glycans, target_composition)
    assert len(selected_leaves) == 0


def test_choose_leaves_to_extend_multiple_equal():
    """Test selection when multiple leaves are equally close to target"""
    leaf_glycans = {
        "GlcNAc(b1-3)Gal(b1-4)Glc",
        "GlcNAc(b1-6)Gal(b1-4)Glc"
    }
    target_composition = {"Hex": 2, "HexNAc": 2}
    selected_leaves = choose_leaves_to_extend(leaf_glycans, target_composition)
    # Both glycans should be selected as they're equally close to target
    assert len(selected_leaves) == 2


@pytest.fixture
def evo_test_networks():
    """Create test networks for evolutionary pruning"""
    # Create main network
    glycans = ['Gal(b1-4)Glc-ol', "GlcNAc(b1-3)[GlcNAc(b1-6)]Gal(b1-4)Glc-ol", 'Fuc(a1-2)Gal(b1-4)Glc-ol']
    abundances = [100.0, 30.0, 50.0]
    main_net = construct_network(glycans, abundances = abundances)
    # Create comparison networks
    network_dic = {
        'species1': main_net.copy(),
        'species2': main_net.copy(),
        'species3': main_net.copy()
    }
    return main_net, network_dic


@pytest.fixture
def extension_test_network():
    """Create test network for extension"""
    glycans = ['Gal(b1-4)Glc-ol', "GlcNAc(b1-3)Gal(b1-4)Glc-ol", 'Fuc(a1-2)Gal(b1-4)Glc-ol']
    network = construct_network(glycans)
    return network


@pytest.fixture
def n_glycan_network():
    """Create test network for extension"""
    glycans = ['Man(a1-2)Man(a1-3)[Man(a1-3)[Man(a1-6)]Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc',
               'GlcNAc(b1-2)Man(a1-3)[GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc',
               'Neu5Ac(a2-6)Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Gal(b1-4)GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc',
               'Neu5Ac(a2-6)Gal(b1-4)GlcNAc(b1-2)Man(a1-3)[Neu5Gc(a2-3)Gal(b1-4)GlcNAc(b1-2)Man(a1-6)]Man(b1-4)GlcNAc(b1-4)[dHex(a1-6)]GlcNAc']
    network = construct_network(glycans)
    return network


def test_evoprune_network_basic(evo_test_networks):
    """Test basic evolutionary pruning functionality"""
    main_net, network_dic = evo_test_networks
    pruned_net = evoprune_network(
        main_net,
        network_dic=network_dic,
        species_list=['species1', 'species2', 'species3'],
        threshold=0.01
    )
    # Check that network is still connected
    assert nx.is_weakly_connected(pruned_net)
    # Check that root node is preserved
    assert 'Gal(b1-4)Glc-ol' in pruned_net.nodes()
    # Check virtual node attributes
    virtual_nodes = [n for n, d in pruned_net.nodes(data=True) if d.get('virtual') == 1]
    for vnode in virtual_nodes:
        assert pruned_net.nodes[vnode].get('abundance', 0) > 0


def test_evoprune_network_threshold(evo_test_networks):
    """Test threshold-based pruning"""
    main_net, network_dic = evo_test_networks
    # Set high threshold to force pruning
    pruned_net = evoprune_network(
        main_net,
        network_dic=network_dic,
        threshold=0.5
    )
    # Check that low-abundance virtual nodes are removed
    virtual_nodes = [n for n, d in pruned_net.nodes(data=True) if d.get('virtual') == 1]
    assert len(virtual_nodes) < len([n for n, d in main_net.nodes(data=True) if d.get('virtual') == 1])


def test_extend_network_basic(extension_test_network):
    """Test basic network extension"""
    extended_net, new_glycans = extend_network(
        extension_test_network,
        steps=1,
        to_extend="all"
    )
    # Check that network grew
    assert len(extended_net.nodes()) > len(extension_test_network.nodes())
    assert len(new_glycans) > 0
    # Check that new glycans are connected to existing ones
    for new_glycan in new_glycans:
        paths = nx.single_source_shortest_path(extended_net, 'Gal(b1-4)Glc-ol')
        assert any(new_glycan in path for path in paths.values())
    extended_net, new_glycans = extend_network(
        extension_test_network,
        steps=1,
        to_extend="all",
        strict_context=True
    )


def test_extend_network_specific_target(extension_test_network):
    """Test network extension toward specific target"""
    target_composition = {"Hex": 2, "HexNAc": 2}
    extended_net, new_glycans = extend_network(
        extension_test_network,
        steps=1,
        to_extend=target_composition
    )
    # Check that at least one new glycan moves closer to target composition
    assert any(
        sum((Counter(target_composition) - Counter(glycan_to_composition(g))).values()) <
        sum((Counter(target_composition) - Counter(glycan_to_composition('Gal(b1-4)Glc-ol'))).values())
        for g in new_glycans
    )


def test_extend_network_specific_leaf(extension_test_network):
    """Test network extension from specific leaf"""
    leaf_to_extend = 'GlcNAc(b1-3)Gal(b1-4)Glc-ol'
    extended_net, new_glycans = extend_network(
        extension_test_network,
        steps=1,
        to_extend=leaf_to_extend
    )
    # Check that new glycans are extensions of specified leaf
    for new_glycan in new_glycans:
        assert subgraph_isomorphism(new_glycan, leaf_to_extend)


@patch('bokeh.io.output_notebook')
@patch('bokeh.plotting.show')
def test_plot_network_basic(mock_show, mock_enable, evo_test_networks):
    """Test basic network plotting functionality"""
    main_net, _ = evo_test_networks
    plot = plot_network(main_net, plot_format='kamada_kawai')
    # Check that plot was created
    assert plot is not None
    # Check renderer properties
    assert len(plot.renderers) > 0
    # Test without edge labels
    plot = plot_network(main_net, plot_format='kamada_kawai', edge_label_draw=False)
    assert plot is not None
    plt.close()


@patch('bokeh.io.output_notebook')
@patch('bokeh.plotting.show')
def test_plot_network_n(mock_show, mock_enable, n_glycan_network):
    """Test basic network plotting functionality for N-glycans"""
    plot = plot_network(n_glycan_network, plot_format='spring')
    # Check that plot was created
    assert plot is not None
    # Check renderer properties
    assert len(plot.renderers) > 0
    # Test without edge labels
    plot = plot_network(n_glycan_network, plot_format='spring', edge_label_draw=False)
    assert plot is not None
    plt.close()


@patch('bokeh.io.output_notebook')
@patch('bokeh.plotting.show')
def test_plot_network_with_edge_labels(mock_show, mock_enable, evo_test_networks):
    """Test network plotting with edge labels"""
    main_net, _ = evo_test_networks
    plot = plot_network(main_net, plot_format='kamada_kawai', edge_label_draw=True)
    # Check for edge labels
    assert plot is not None
    # Check for renderers (indirect check for edge labels)
    assert len(plot.renderers) > len(main_net.nodes())
    plt.close()


@patch('bokeh.io.output_notebook')
@patch('bokeh.plotting.show')
def test_plot_network_with_lfc(mock_show, mock_enable, evo_test_networks):
    """Test network plotting with log fold change data"""
    lfc_dict = {'Fuc(a1-2)': 1.5, 'GlcNAc(b1-3)': -0.5}
    main_net, _ = evo_test_networks
    plot = plot_network(main_net, plot_format='kamada_kawai', edge_label_draw=True, lfc_dict=lfc_dict)
    # Verify plot was created
    assert plot is not None
    # Check for renderers
    assert len(plot.renderers) > 0
    plt.close()


@patch('bokeh.io.output_notebook')
@patch('bokeh.plotting.show')
def test_plot_network_layouts(mock_show, mock_enable, evo_test_networks):
    """Test different layout options with fallback handling"""
    # Test kamada_kawai and spring layouts which don't require external dependencies
    safe_layouts = ['kamada_kawai', 'spring']
    main_net, _ = evo_test_networks
    for layout in safe_layouts:
        plot = plot_network(main_net, plot_format=layout)
        assert plot is not None
        plt.close()
    # Test pydot2 layout with proper error handling
    try:
        with suppress_pydot_warnings():
            plot = plot_network(main_net, plot_format='pydot2')
            assert plot is not None
    except (ImportError, FileNotFoundError):
        print("Graphviz not installed, skipping pydot2 layout test")
    plt.close()


def test_plot_network_no_notebook(evo_test_networks):
    """Test fallback behavior when not in notebook"""
    main_net, _ = evo_test_networks
    with patch('bokeh.io.output_notebook', side_effect=Exception):
        with patch('bokeh.plotting.show') as mock_show:
            plot = plot_network(main_net, plot_format='kamada_kawai')
            # Even with notebook initialization failing, should still return a plot
            assert plot is not None


def test_add_high_man_removal(evo_test_networks):
    main_net, _ = evo_test_networks
    network = add_high_man_removal(main_net)
    assert len(main_net.edges()) == len(network.edges())


def test_net_dic():
    assert len(net_dic) > 0


def test_find_shared_virtuals():
    glycan_a = "Glc-ol"
    glycan_b = "GlcNAc(b1-3)Gal(b1-4)Gal-ol"
    assert find_shared_virtuals(glycan_a, glycan_b, {}) == []


def test_create_adjacency_matrix():
    glycans = ["Glc-ol", "Gal-ol"]
    assert create_adjacency_matrix(glycans, {})[1] == []
    glycans = ["Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol", "Gal(b1-4)Glc-ol"]
    assert create_adjacency_matrix(glycans, {})[1] == ['GlcNAc(b1-3)Gal(b1-4)Glc-ol']


def test_find_ptm():
    assert find_ptm("Gal(b1-4)Glc-ol", ["Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol", "Gal(b1-4)Glc-ol"], {}, {}) == 0
    from glycowork.motif.tokenization import stem_lib
    assert find_ptm("Sug(b1-4)Gal6S(b1-4)Glc-ol", ["Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol", "Gal(b1-4)Glc-ol"], {}, stem_lib) == 0
    assert find_ptm("Gal(b1-4)Gal6S(b1-4)Glc-ol", ["Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol", "Gal(b1-4)Glc-ol"], {}, stem_lib)[1] == "6S"


@patch('bokeh.io.output_notebook')
@patch('bokeh.plotting.show')
def test_infer_network(mock_show, mock_enable):
    net = construct_network(["Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol", "Gal(b1-4)Glc-ol"])
    spec_dic = {"test": construct_network(["GlcNAc(b1-3)Gal(b1-4)Glc-ol", "Gal(b1-4)Glc-ol"]), "org": net}
    net2 = infer_network(net, "org", ["test", "org"], spec_dic)
    assert nx.get_node_attributes(net2, "virtual")["GlcNAc(b1-3)Gal(b1-4)Glc-ol"] == 2
    plot = plot_network(net2)


def test_export_network(tmp_path):
    net = construct_network(["Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol", "Gal(b1-4)Glc-ol"])
    filepath = os.path.join(tmp_path, "test_network")
    export_network(net, filepath, other_node_attributes=["extra_attr"])
    net = construct_network(["Gal(b1-4)Glc-ol"])
    export_network(net, filepath, other_node_attributes=["extra_attr"])
    # Clean up
    os.remove(f"{filepath}_edge_list.csv")
    os.remove(f"{filepath}_node_labels.csv")


@pytest.fixture
def sample_objects():
    return {
        'obj1': [1, 2, 3, 4],
        'obj2': [2, 3, 4, 5],
        'obj3': [3, 4, 5, 6]
    }


@pytest.fixture
def sample_taxonomy_data():
    return pd.DataFrame({
        'glycan': [
            'Gal(b1-4)Glc-ol',
            'Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol',
            'Fuc(a1-2)Gal(b1-4)Glc-ol',
            'Gal(b1-4)Glc-ol',
            'GalNAc(b1-3)Gal(b1-4)Glc-ol',
            'Neu5Ac(a2-3)Gal(b1-4)Glc-ol'
        ],
        'Species': ['Species1', 'Species1', 'Species1', 'Species2', 'Species2', 'Species2'],
        'Order': ['Order1', 'Order1', 'Order1', 'Order2', 'Order2', 'Order2'],
        'Class': ['Class1', 'Class1', 'Class1', 'Class2', 'Class2', 'Class2']
    }).set_index('glycan')


@pytest.fixture
def sample_networks():
    # Create network for Species1
    G1 = nx.DiGraph()
    G1.add_edge('Gal(b1-4)Glc-ol', 'GlcNAc(b1-3)Gal(b1-4)Glc-ol', diffs='GlcNAc(b1-3)')
    G1.add_edge('GlcNAc(b1-3)Gal(b1-4)Glc-ol', 'Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol', diffs='Gal(b1-4)')
    G1.add_edge('Gal(b1-4)Glc-ol', 'Fuc(a1-2)Gal(b1-4)Glc-ol', diffs='Fuc(a1-2)')
    nx.set_node_attributes(G1, {node: {'virtual': 0} for node in G1.nodes()})
    # Create network for Species2
    G2 = nx.DiGraph()
    G2.add_edge('Gal(b1-4)Glc-ol', 'GalNAc(b1-3)Gal(b1-4)Glc-ol', diffs='GalNAc(b1-3)')
    G2.add_edge('Gal(b1-4)Glc-ol', 'Neu5Ac(a2-3)Gal(b1-4)Glc-ol', diffs='Neu5Ac(a2-3)')
    nx.set_node_attributes(G2, {node: {'virtual': 0} for node in G2.nodes()})
    return [G1, G2]


@pytest.fixture
def network_dict(sample_networks):
    return {
        'Species1': sample_networks[0],
        'Species2': sample_networks[1]
    }


@pytest.fixture
def sample_embeddings():
    return pd.DataFrame({
        'dim1': [0.1, 0.2, 0.3, 0.1, 0.2, 0.3],
        'dim2': [0.4, 0.5, 0.6, 0.4, 0.5, 0.6],
        'dim3': [0.7, 0.8, 0.9, 0.7, 0.8, 0.9]
    }, index=[
        'Gal(b1-4)Glc-ol',
        'Gal(b1-4)GlcNAc(b1-3)Gal(b1-4)Glc-ol',
        'Fuc(a1-2)Gal(b1-4)Glc-ol',
        'Gal(b1-4)Glc-ol',
        'GalNAc(b1-3)Gal(b1-4)Glc-ol',
        'Neu5Ac(a2-3)Gal(b1-4)Glc-ol'
    ])


def test_calculate_distance_matrix(sample_objects):
    # Test with dictionary input
    dm = calculate_distance_matrix(sample_objects, lambda x, y: abs(sum(x) - sum(y)))
    assert isinstance(dm, pd.DataFrame)
    assert list(dm.columns) == list(sample_objects.keys())
    assert dm.shape == (3, 3)
    assert (dm.values >= 0).all()
    assert (dm.values == dm.values.T).all()  # Symmetry
    assert np.all(np.diag(dm) == 0)  # Zero diagonal


def test_calculate_distance_matrix_with_list():
    # Test with list input
    objects = [[1, 2], [2, 3], [3, 4]]
    dm = calculate_distance_matrix(objects, lambda x, y: abs(sum(x) - sum(y)))
    assert isinstance(dm, pd.DataFrame)
    assert dm.shape == (3, 3)
    assert list(dm.columns) == [0, 1, 2]


def test_jaccard():
    list1 = [1, 2, 3]
    list2 = [2, 3, 4]
    distance = jaccard(list1, list2)
    assert isinstance(distance, float)
    assert 0 <= distance <= 1
    assert jaccard(list1, list1) == 0  # Identity
    # Test with networks
    G1 = nx.Graph([(1, 2), (2, 3)])
    G2 = nx.Graph([(2, 3), (3, 4)])
    net_distance = jaccard(G1.nodes(), G2.nodes())
    assert isinstance(net_distance, float)
    assert 0 <= net_distance <= 1


def test_distance_from_metric(sample_taxonomy_data, sample_networks):
    dm = distance_from_metric(
        sample_taxonomy_data,
        sample_networks,
        metric="Jaccard",
        cut_off=1,
        rank="Species"
    )
    assert isinstance(dm, pd.DataFrame)
    assert dm.shape[0] == dm.shape[1]  # Square matrix
    assert (dm.values >= 0).all()
    assert (dm.values == dm.values.T).all()  # Symmetry
    # Test invalid metric case
    try:
        dm_invalid = distance_from_metric(
            sample_taxonomy_data,
            sample_networks,
            metric="InvalidMetric",
            cut_off=1,
            rank="Species"
            )
        assert False, "Should have raised an exception for invalid metric"
    except ValueError:
        pass  # Expected behavior
    # Test dendrogram generation
    if dm is not None:
        # Test without file saving
        dendrogram_from_distance(dm, ylabel="Test")
        # Test with file saving
        dendrogram_from_distance(dm, ylabel="Test", filepath="test.png")


def test_distance_from_embeddings(sample_taxonomy_data, sample_embeddings):
    dm = distance_from_embeddings(
        sample_taxonomy_data,
        sample_embeddings,
        cut_off=1,
        rank="Species",
        averaging="median"
    )
    assert isinstance(dm, pd.DataFrame)
    assert dm.shape[0] == dm.shape[1]
    assert (dm.values >= 0).all()
    assert (dm.values == dm.values.T).all()


def test_distance_from_embeddings_invalid_averaging(sample_taxonomy_data, sample_embeddings):
    result = distance_from_embeddings(
        sample_taxonomy_data,
        sample_embeddings,
        averaging="invalid"
    )
    assert result is None


def test_check_conservation():
    # Create sample data
    df = pd.DataFrame({
        'Species': ['Species1', 'Species1', 'Species2', 'Species2'],
        'Order': ['Order1', 'Order1', 'Order1', 'Order1'],
        'glycan': ['Gal(b1-4)Glc-ol'] * 4
    })
    # Create sample networks
    network_dic = {
        'Species1': nx.Graph([('Gal(b1-4)Glc-ol', 'GlcNAc(b1-3)Gal(b1-4)Glc-ol')]),
        'Species2': nx.Graph([('Gal(b1-4)Glc-ol', 'Fuc(a1-2)Gal(b1-4)Glc-ol')])
    }
    conservation = check_conservation(
        'Gal(b1-4)Glc-ol',
        df,
        network_dic,
        threshold=1
    )
    assert isinstance(conservation, dict)
    assert all(0 <= v <= 1 for v in conservation.values())
    # Test motif search
    conservation_motif = check_conservation(
        'Gal(b1-4)Glc',
        df,
        network_dic,
        threshold=1,
        motif=True
    )
    assert isinstance(conservation_motif, dict)
    assert all(0 <= v <= 1 for v in conservation_motif.values())
    conservation_motif = check_conservation(
        'Gal(b1-4)',
        df,
        network_dic,
        threshold=1,
        motif=True
    )
    assert isinstance(conservation_motif, dict)


def test_get_communities(sample_networks):
    communities = get_communities(sample_networks, ['net1', 'net2', 'net3'])
    assert isinstance(communities, dict)
    assert len(communities) > 0
    assert all(isinstance(v, list) for v in communities.values())
    assert all('_' in k for k in communities.keys())
    # Test without labels
    communities_no_labels = get_communities(sample_networks)
    assert isinstance(communities_no_labels, dict)
    assert len(communities_no_labels) > 0


def test_check_conservation_threshold():
    df = pd.DataFrame({
        'Species': ['Species1'] * 3,
        'Order': ['Order1'] * 3,
        'glycan': ['Gal(b1-4)Glc-ol'] * 3
    })
    network_dic = {
        'Species1': nx.Graph([('Gal(b1-4)Glc-ol', 'GlcNAc(b1-3)Gal(b1-4)Glc-ol')])
    }
    # Test with threshold higher than available glycans
    conservation = check_conservation(
        'Gal(b1-4)Glc-ol',
        df,
        network_dic,
        threshold=5
    )
    assert len(conservation) == 0


def test_get_communities_single_node():
    # Test with single-node networks
    single_node_networks = [nx.Graph([('A', 'A')]), nx.Graph([('B', 'B')])]
    communities = get_communities(single_node_networks)
    assert isinstance(communities, dict)
    assert len(communities) > 0
    assert all(isinstance(v, list) for v in communities.values())


@pytest.fixture
def sample_glycan_df():
    df = pd.DataFrame({
        'glycan': [
            'Gal(b1-4)Glc-ol',
            'Fuc(a1-3)[Gal(b1-4)]GlcNAc(b1-3)Gal(b1-4)Glc-ol',
            'Fuc(a1-2)Gal(b1-4)Glc-ol'
        ],
        'Species': [['Homo_sapiens'], ['Homo_sapiens'], ['Mus_musculus']],
        'Phylum': [['Chordata'], ['Chordata'], ['Chordata']],
        'tissue_sample': [['blood'], ['blood', 'liver'], ['brain']],
        'disease_association': [[], ['cancer'], []],
        'disease_direction': [[], ['up'], []],
        'disease_sample': [[], ['tumor'], []],
        'glytoucan_id': ['G00001', 'G00002', 'G00003']
    })
    # Add graph attribute for fast comparison
    df['graph'] = [glycan_to_nxGraph(g) for g in df['glycan']]
    return df


def test_check_presence_existing_glycan(sample_glycan_df, capsys):
    check_presence('Gal(b1-4)Glc-ol', sample_glycan_df)
    captured = capsys.readouterr()
    assert "Glycan already in dataset." in captured.out
    check_presence("WURCS=2.0/3,5,4/[a2122h-1b_1-5_2*NCC/3=O][a1122h-1b_1-5][a1122h-1a_1-5]/1-1-2-3-3/a4-b1_b4-c1_c3-d1_c6-e1", sample_glycan_df)


def test_check_presence_new_glycan(sample_glycan_df, capsys):
    check_presence('Neu5Ac(a2-3)Gal(b1-4)Glc-ol', sample_glycan_df)
    captured = capsys.readouterr()
    assert "It's your lucky day, this glycan is new!" in captured.out


def test_check_presence_with_species(sample_glycan_df, capsys):
    check_presence('Gal(b1-4)Glc-ol', sample_glycan_df, name='Mus_musculus')
    captured = capsys.readouterr()
    assert "It's your lucky day, this glycan is new!" in captured.out


def test_check_presence_invalid_species(sample_glycan_df, capsys):
    check_presence('Gal(b1-4)Glc-ol', sample_glycan_df, name='Invalid_species')
    captured = capsys.readouterr()
    assert "Invalid_species is not in dataset" in captured.out


def test_check_presence_fast_mode(sample_glycan_df, capsys):
    check_presence('Gal(b1-4)Glc-ol', sample_glycan_df, fast=True)
    captured = capsys.readouterr()
    assert "Glycan already in dataset." in captured.out


def test_get_insight():
    get_insight('Gal(b1-4)Glc-ol')
    get_insight('Man(a1-3)[Man(a1-6)]Man(b1-4)GlcNAc(b1-4)GlcNAc')


def test_glytoucan_to_glycan_forward():
    result = glytoucan_to_glycan(['G26039ES', 'G65562ZE'])
    assert len(result) == 2
    assert 'Gal(b1-4)Glc-ol' in result
    assert 'Neu5Ac(a2-3)Gal(b1-3)GalNAc' in result


def test_glytoucan_to_glycan_reverse():
    result = glytoucan_to_glycan(['Gal(b1-4)Glc-ol', 'Neu5Ac(a2-3)Gal(b1-3)GalNAc'], revert=True)
    assert len(result) == 2
    assert 'G26039ES' in result
    assert 'G65562ZE' in result


def test_glytoucan_to_glycan_missing_id(capsys):
    glytoucan_to_glycan(['G26039ES', 'MISSING'])
    captured = capsys.readouterr()
    assert 'These IDs are not in our database: ' in captured.out
    assert 'MISSING' in captured.out


def test_glytoucan_to_glycan_missing_glycan(capsys):
    glytoucan_to_glycan(['Gal(b1-4)Glc-ol', 'MISSING'], revert=True)
    captured = capsys.readouterr()
    assert 'These glycans are not in our database: ' in captured.out
    assert 'MISSING' in captured.out


def test_seed_wildcard_hierarchy_basic():
    glycans = ["GalNAc", "GlcNAc", "Man"]
    labels = [1, 2, 3]
    wildcard_list = ["Gal", "Glc"]
    wildcard_name = "Hex"
    # Set random seed for reproducibility
    np.random.seed(42)
    new_glycans, new_labels = seed_wildcard_hierarchy(
        glycans, labels, wildcard_list, wildcard_name, r=1.0
    )
    assert len(new_glycans) >= len(glycans)
    assert len(new_glycans) == len(new_labels)
    assert all(label in labels for label in new_labels)


def test_seed_wildcard_hierarchy_no_matches():
    glycans = ["XXX", "YYY", "ZZZ"]
    labels = [1, 2, 3]
    wildcard_list = ["Gal", "Glc"]
    wildcard_name = "Hex"
    new_glycans, new_labels = seed_wildcard_hierarchy(
        glycans, labels, wildcard_list, wildcard_name
    )
    assert new_glycans == glycans
    assert new_labels == labels


def test_hierarchy_filter_basic():
    data = {
        'glycan': ['GalNAc1', 'GlcNAc1', 'Man1', 'GalNAc2', 'GlcNAc2',
                   'Man2', 'GalNAc3', 'GlcNAc3', 'Man3', 'GalNAc4'],
        'Species': ['sp1']*5 + ['sp2']*5,
        'Genus': ['g1']*5 + ['g2']*5,
        'Family': ['f1']*5 + ['f2']*5,
        'Order': ['o1']*5 + ['o2']*5,
        'Class': ['c1']*5 + ['c2']*5,
        'Phylum': ['p1']*5 + ['p2']*5,
        'Kingdom': ['k1']*5 + ['k2']*5,
        'Domain': ['d1']*5 + ['d2']*5
    }
    df = pd.DataFrame(data)
    train_x, val_x, train_y, val_y, _, class_list, class_converter = hierarchy_filter(
        df, rank='Domain', min_seq=1
    )
    assert len(train_x) + len(val_x) == len(set(df['glycan']))  # Check for duplicates removal
    assert len(train_y) == len(train_x)
    assert len(val_y) == len(val_x)
    assert all(isinstance(y, int) for y in train_y + val_y)
    assert len(class_list) == len(set(df['Domain']))
    assert len(class_converter) == len(class_list)


def test_hierarchy_filter_min_seq():
    data = {
        'glycan': ['GalNAc1', 'GlcNAc1', 'Man1', 'GalNAc2', 'GlcNAc2',
                   'Man2', 'GalNAc3', 'GlcNAc3', 'Man3', 'GalNAc4',
                   'GlcNAc4', 'Man4', 'GalNAc5', 'GlcNAc5', 'Man5'],
        'Domain': ['d1']*2 + ['d2']*13,  # d1 has 2 samples, d2 has 13
        'Kingdom': ['k1']*2 + ['k2']*13,
        'Phylum': ['p1']*2 + ['p2']*13,
        'Class': ['c1']*2 + ['c2']*13,
        'Order': ['o1']*2 + ['o2']*13,
        'Family': ['f1']*2 + ['f2']*13,
        'Genus': ['g1']*2 + ['g2']*13,
        'Species': ['s1']*2 + ['s2']*13
    }
    df = pd.DataFrame(data)
    train_x, val_x, train_y, val_y, id_val, class_list, class_converter = hierarchy_filter(
        df, rank='Domain', min_seq=3
    )
    # Only d2 should remain as it's the only one with 3 or more sequences
    assert len(class_list) == 1
    assert class_list[0] == 'd2'
    assert len(train_x) + len(val_x) < len(df)  # Some samples should be filtered out
    assert all(y == 0 for y in train_y + val_y)  # All should be mapped to 0 as only one class remains


def test_general_split_basic():
    glycans = ['G1', 'G2', 'G3', 'G4', 'G5']
    labels = [1, 2, 3, 4, 5]
    train_x, test_x, train_y, test_y = general_split(glycans, labels, test_size=0.2)
    assert len(train_x) == 4  # 80% of 5
    assert len(test_x) == 1   # 20% of 5
    assert len(train_y) == len(train_x)
    assert len(test_y) == len(test_x)
    assert set(train_y + test_y) == set(labels)


def test_general_split_reproducibility():
    glycans = ['G1', 'G2', 'G3', 'G4', 'G5']
    labels = [1, 2, 3, 4, 5]
    train_x1, test_x1, train_y1, test_y1 = general_split(glycans, labels)
    train_x2, test_x2, train_y2, test_y2 = general_split(glycans, labels)
    assert train_x1 == train_x2
    assert test_x1 == test_x2
    assert train_y1 == train_y2
    assert test_y1 == test_y2


def test_prepare_multilabel_basic():
    data = {
        'glycan': ['G1', 'G1', 'G2', 'G2', 'G3'],
        'Species': ['sp1', 'sp2', 'sp1', 'sp3', 'sp2']
    }
    df = pd.DataFrame(data)
    glycans, labels = prepare_multilabel(df, rank='Species')
    assert len(glycans) == 3  # unique glycans
    assert len(labels) == len(glycans)
    assert all(len(label) == 3 for label in labels)  # 3 unique species
    assert all(isinstance(label, list) for label in labels)
    assert all(all(isinstance(v, float) for v in label) for label in labels)


def test_prepare_multilabel_single_association():
    data = {
        'glycan': ['G1', 'G2', 'G3'],
        'Species': ['sp1', 'sp2', 'sp3']
    }
    df = pd.DataFrame(data)
    glycans, labels = prepare_multilabel(df)
    assert len(glycans) == 3
    assert len(labels) == 3
    assert all(sum(label) == 1.0 for label in labels)  # each glycan has exactly one association


def test_hierarchy_filter_with_wildcards():
    data = {
        'glycan': ['GalNAc', 'GlcNAc', 'Man', 'GalNAc(b1-4)Gal', 'GlcNAc(b1-4)GlcNAc',
                   'Man(a1-2)Man', 'GalNAc(a1-3)Gal', 'GlcNAc(b1-2)Glc', 'Man(a1-3)Man', 'GalNAc(b1-4)GlcNAc',
                   'GlcNAc(b1-3)Gal', 'Man(a1-6)Man', 'GalNAc(b1-3)GlcNAc', 'GlcNAc(a1-3)Gal', 'Man(b1-4)GlcNAc'],
        'Domain': ['d1']*8 + ['d2']*7,
        'Kingdom': ['k1']*8 + ['k2']*7,
        'Phylum': ['p1']*8 + ['p2']*7,
        'Class': ['c1']*8 + ['c2']*7,
        'Order': ['o1']*8 + ['o2']*7,
        'Family': ['f1']*8 + ['f2']*7,
        'Genus': ['g1']*8 + ['g2']*7,
        'Species': ['s1']*8 + ['s2']*7
    }
    df = pd.DataFrame(data)
    train_x, val_x, train_y, val_y, id_val, class_list, class_converter = hierarchy_filter(
        df,
        rank='Domain',
        min_seq=1,
        wildcard_seed=True,
        wildcard_list=['Gal', 'Glc'],
        wildcard_name='Hex',
        r=1.0
    )
    assert len(train_x) >= len(set(df['glycan']))  # Should have additional entries due to wildcards
    assert len(train_y) == len(train_x)
    assert len(val_y) == len(val_x)
    assert all(y in [0, 1] for y in train_y + val_y)  # Check labels are properly converted


@pytest.fixture
def mock_glycan_data():
    """Create a mock PyG Data object representing a glycan"""
    data = Data(
        x=torch.tensor([[0], [1], [2], [3], [4]], dtype=torch.float),
        edge_index=torch.tensor([[0, 1], [1, 2], [2, 3], [3, 4]], dtype=torch.long).t(),
        string_labels=['Gal', 'b1-4', 'GlcNAc', 'b1-2', 'Man'],
        labels=[0, 5, 1, 8, 2]
    )
    return data


@pytest.fixture
def mock_library():
    """Create a mock glycoletter library"""
    return {
        'Gal': 0,
        'GlcNAc': 1,
        'Man': 2,
        'Hex': 3,
        'HexNAc': 4,
        'b1-4': 5,
        'a1-3': 6,
        'a1-6': 7,
        'b1-2': 8
    }

@pytest.fixture
def mock_glycan_dataset():
    """Create a list of mock glycans and labels"""
    glycans = [
        "Gal(b1-4)GlcNAc",
        "Man(a1-3)[Man(a1-6)]Man",
        "Gal(b1-4)GlcNAc(b1-2)Man"
    ]
    labels = [0, 1, 0]
    return glycans, labels


def test_augment_glycan(mock_glycan_data, mock_library):
    np.random.seed(42)
    torch.manual_seed(42)
    augmented = augment_glycan(
        mock_glycan_data,
        mock_library,
        generalization_prob=1.0  # Force augmentation
    )
    assert isinstance(augmented, Data)
    assert len(augmented.string_labels) == len(mock_glycan_data.string_labels)
    assert augmented.string_labels != mock_glycan_data.string_labels
    assert all(label in mock_library or label in ['Hex', 'HexNAc'] or '-' in label
              for label in augmented.string_labels)


def test_augmented_glycan_dataset(mock_glycan_data, mock_library):
    np.random.seed(42)
    torch.manual_seed(42)
    dataset = [mock_glycan_data]
    augmented_dataset = AugmentedGlycanDataset(
        dataset,
        mock_library,
        augment_prob=1.0,  # Force augmentation
        generalization_prob=1.0  # Force wildcarding
    )
    assert len(augmented_dataset) == len(dataset)
    item = augmented_dataset[0]
    assert isinstance(item, Data)
    assert item.string_labels != mock_glycan_data.string_labels


def test_dataset_to_graphs(mock_glycan_dataset, mock_library):
    glycans, labels = mock_glycan_dataset
    data_list = dataset_to_graphs(
        glycans,
        labels,
        libr=mock_library,
        label_type=torch.long
    )
    assert len(data_list) == len(glycans)
    assert all(isinstance(data, Data) for data in data_list)
    assert all(hasattr(data, 'y') for data in data_list)
    assert all(data.y.dtype == torch.long for data in data_list)


def test_dataset_to_dataloader(mock_glycan_dataset, mock_library):
    glycans, labels = mock_glycan_dataset
    extra_features = [0.5] * len(glycans)

    dataloader = dataset_to_dataloader(
        glycans,
        labels,
        libr=mock_library,
        batch_size=2,
        extra_feature=extra_features,
        augment_prob=0.5
    )
    assert isinstance(dataloader, torch.utils.data.DataLoader)
    assert dataloader.batch_size == 2
    # Check first batch
    batch = next(iter(dataloader))
    assert hasattr(batch, 'y')
    assert hasattr(batch, 'train_idx')


def test_split_data_to_train(mock_glycan_dataset, mock_library):
    glycans, labels = mock_glycan_dataset
    # Split data for test
    train_glycans = glycans[:2]
    val_glycans = glycans[2:]
    train_labels = labels[:2]
    val_labels = labels[2:]
    extra_train = [0.5] * len(train_glycans)
    extra_val = [0.5] * len(val_glycans)
    loaders = split_data_to_train(
        train_glycans,
        val_glycans,
        train_labels,
        val_labels,
        libr=mock_library,
        batch_size=1,
        extra_feature_train=extra_train,
        extra_feature_val=extra_val,
        augment_prob=0.5
    )
    assert 'train' in loaders
    assert 'val' in loaders
    assert isinstance(loaders['train'], torch.utils.data.DataLoader)
    assert isinstance(loaders['val'], torch.utils.data.DataLoader)
    # Check validation loader has no augmentation
    val_batch = next(iter(loaders['val']))
    assert hasattr(val_batch, 'y')
    assert hasattr(val_batch, 'train_idx')
    loaders = split_data_to_train(
        train_glycans,
        val_glycans,
        train_labels,
        val_labels,
        batch_size=1,
        extra_feature_train=extra_train,
        extra_feature_val=extra_val,
        augment_prob=0.5
    )


def test_dataset_to_graphs_caching(mock_glycan_dataset, mock_library):
    glycans, labels = mock_glycan_dataset
    # Add duplicate glycan
    glycans = glycans + [glycans[0]]
    labels = labels + [labels[0]]
    data_list = dataset_to_graphs(
        glycans,
        labels,
        libr=mock_library
    )
    assert len(data_list) == len(glycans)
    # Check that duplicate glycans have same structure but different objects
    assert torch.equal(data_list[0].labels, data_list[-1].labels)
    assert torch.equal(data_list[0].edge_index, data_list[-1].edge_index)
    assert id(data_list[0]) != id(data_list[-1])


def test_augment_glycan_no_change(mock_glycan_data, mock_library):
    augmented = augment_glycan(
        mock_glycan_data,
        mock_library,
        generalization_prob=0.0  # Prevent augmentation
    )
    assert augmented.string_labels == mock_glycan_data.string_labels
    assert augmented.labels == mock_glycan_data.labels


def test_dataset_to_dataloader_batch_size(mock_glycan_dataset, mock_library):
    glycans, labels = mock_glycan_dataset
    # Test with different batch sizes
    for batch_size in [1, 2, len(glycans)]:
        dataloader = dataset_to_dataloader(
            glycans,
            labels,
            libr=mock_library,
            batch_size=batch_size
        )
        assert dataloader.batch_size == batch_size
        _ = next(iter(dataloader))


@pytest.fixture(params=[
    'regression',
    'classification',
    'multilabel'
])
def mode(request):
    return request.param


@pytest.fixture
def expected_metrics(mode):
    metrics_map = {
        'regression': ['loss', 'mse', 'mae', 'r2'],
        'classification': ['loss', 'acc', 'mcc', 'auroc'],
        'multilabel': ['loss', 'acc', 'mcc', 'lrap', 'ndcg']
    }
    return metrics_map[mode]


@pytest.fixture
def mock_model(mode):
    """Create a mock model that adapts its output size based on the mode"""
    class SimpleModel(nn.Module):
        def __init__(self, output_size):
            super().__init__()
            self.embedding = nn.Embedding(10, 32)
            self.fc = nn.Linear(32, output_size)
            self.bn = nn.BatchNorm1d(32)

        def forward(self, x, edge_index, batch):
            x = self.embedding(x)
            x = self.bn(x)
            graph_embed = torch.zeros(batch.max().item() + 1, x.size(1),
                                    device=x.device)
            graph_embed.index_add_(0, batch, x)
            out = self.fc(graph_embed)
            return out
    output_sizes = {
        'regression': 1,
        'classification': 2,
        'multilabel': 2
    }
    return SimpleModel(output_sizes[mode]).to(device)


@pytest.fixture
def mock_dataloader(mode):
    class MockData:
        def __init__(self, mode):
            # Node features for 6 nodes across 2 graphs
            self.labels = torch.tensor([0, 1, 0, 1, 0, 1], dtype=torch.long)
            # Edge connections
            self.edge_index = torch.tensor([[0, 1, 2, 3, 4, 5],
                                          [1, 2, 0, 4, 5, 3]], dtype=torch.long)
            # Batch assignments: first 3 nodes to first graph, last 3 to second
            self.batch = torch.tensor([0, 0, 0, 1, 1, 1], dtype=torch.long)
            # Adapt y based on mode
            if mode == 'regression':
                self.y = torch.tensor([0.5, 1.5], dtype=torch.float)
            elif mode == 'classification':
                self.y = torch.tensor([0, 1], dtype=torch.long)
            else:  # multilabel
                self.y = torch.tensor([[1.0, 0.0], [0.0, 1.0]], dtype=torch.float)
    class MockLoader:
        def __init__(self, data_list):
            self.data_list = data_list
        def __iter__(self):
            return iter(self.data_list)
        def __len__(self):
            return len(self.data_list)
    mock_data_list = [MockData(mode=mode) for _ in range(3)]
    loader = MockLoader(mock_data_list)
    return {'train': loader, 'val': loader}


@pytest.fixture
def mock_xgb_data():
    X_train = pd.DataFrame({
        'feature1': [1, 2, 3],
        'feature2': [4, 5, 6]
    })
    X_test = pd.DataFrame({
        'feature1': [7, 8, 9],
        'feature2': [10, 11, 12]
    })
    y_train = [0, 1, 0]
    y_test = [1, 0, 1]
    return X_train, X_test, y_train, y_test


def test_early_stopping():
    early_stopping = EarlyStopping(patience=2, verbose=True)
    model = Mock()
    # Should not stop after first higher loss
    early_stopping(0.5, model)
    early_stopping(0.6, model)
    assert not early_stopping.early_stop
    # Should stop after patience exceeded
    early_stopping(0.7, model)
    assert early_stopping.early_stop
    # Should reset counter on improvement
    early_stopping = EarlyStopping(patience=2)
    early_stopping(0.5, model)
    early_stopping(0.4, model)  # Improvement
    early_stopping(0.45, model)  # Worse
    assert not early_stopping.early_stop  # Counter should have reset


def test_batch_norm_stats(mock_model):
    # Test disabling
    disable_running_stats(mock_model)
    assert mock_model.bn.momentum == 0
    assert hasattr(mock_model.bn, 'backup_momentum')
    # Test enabling
    enable_running_stats(mock_model)
    assert mock_model.bn.momentum == mock_model.bn.backup_momentum


@patch('torch.cuda.is_available', return_value=False)
def test_poly1_cross_entropy_loss(_):
    criterion = Poly1CrossEntropyLoss(num_classes=3, epsilon=1.0)
    logits = torch.tensor([[2.0, 1.0, 0.0], [0.0, 2.0, 1.0]])
    labels = torch.tensor([0, 1])
    loss = criterion(logits, labels)
    assert isinstance(loss, torch.Tensor)
    assert loss.ndim == 0  # Scalar tensor
    criterion = Poly1CrossEntropyLoss(num_classes=3, epsilon=1.0, reduction="sum")
    loss = criterion(logits, labels)
    labels = torch.tensor([[1, 0, 0], [0, 1, 0]])
    loss = criterion(logits, labels)


def test_sam_optimizer(mock_model, mode):
    if mode == "classification":
        # Create a dummy input and target
        x = torch.randint(0, 10, (6,)).to(device)  # 6 nodes with features in range [0, 10)
        edge_index = torch.tensor([[0, 1, 2], [1, 2, 0]], dtype=torch.long).to(device)  # Some edges
        batch = torch.tensor([0, 0, 0, 1, 1, 1]).to(device)  # Two graphs
        target = torch.tensor([0, 1]).to(device)  # Labels for the two graphs
        # Create optimizer
        sam = SAM(
            mock_model.parameters(),
            base_optimizer=torch.optim.SGD,
            rho=0.5,
            adaptive=True,
            lr=0.1
        )
        # Forward pass
        output = mock_model(x, edge_index, batch)
        loss = torch.nn.functional.cross_entropy(output, target).to(device)
        # Backward pass to create gradients
        loss.backward()
        # Now test SAM steps
        sam.first_step(zero_grad=True)
        # Another forward-backward pass
        output = mock_model(x, edge_index, batch)
        loss = torch.nn.functional.cross_entropy(output, target).to(device)
        loss.backward()
        sam.second_step(zero_grad=True)
        # Verify state
        assert hasattr(sam, 'base_optimizer')
        assert isinstance(sam.base_optimizer, torch.optim.SGD)


def test_sam_optimizer_state():
    # Create a simple model for testing
    model = torch.nn.Linear(2, 2)
    # Create dummy data
    x = torch.randn(4, 2)
    y = torch.tensor([0, 1, 0, 1])
    # Initialize SAM
    sam = SAM(
        model.parameters(),
        base_optimizer=torch.optim.SGD,
        rho=0.5,
        adaptive=True,
        lr=0.1
    )
    # Initial parameter values
    initial_params = {name: param.clone() for name, param in model.named_parameters()}
    # Forward pass
    output = model(x)
    loss = torch.nn.functional.cross_entropy(output, y)
    loss.backward()
    # First step
    sam.first_step(zero_grad=True)
    # Check parameters were updated
    for name, param in model.named_parameters():
        assert not torch.equal(param, initial_params[name]), f"Parameters {name} were not updated in first step"
    # Store parameters after first step
    params_after_first = {name: param.clone() for name, param in model.named_parameters()}
    # Another forward-backward pass
    output = model(x)
    loss = torch.nn.functional.cross_entropy(output, y)
    loss.backward()
    # Second step
    sam.second_step(zero_grad=True)
    # Check parameters were updated again
    for name, param in model.named_parameters():
        assert not torch.equal(param, params_after_first[name]), f"Parameters {name} were not updated in second step"


def test_sam_optimizer_zero_grad():
    model = torch.nn.Linear(2, 2)
    x = torch.randn(4, 2)
    y = torch.tensor([0, 1, 0, 1])
    sam = SAM(
        model.parameters(),
        base_optimizer=torch.optim.SGD,
        rho=0.5,
        adaptive=True,
        lr=0.1
    )
    # Forward and backward pass
    output = model(x)
    loss = torch.nn.functional.cross_entropy(output, y)
    loss.backward()
    # Check gradients exist
    assert all(p.grad is not None for p in model.parameters())
    # Test zero_grad
    sam.zero_grad()
    assert all(p.grad is None or torch.all(p.grad == 0) for p in model.parameters())


def test_gsam_gradient_decompose_coverage():
  """Minimal test to cover _gradient_decompose method for GSAM"""
  model = nn.Linear(2, 1).to(device)
  x = torch.randn(4, 2, device=device)
  y = torch.randn(4, 1, device=device)
  criterion = nn.MSELoss()
  optimizer = SAM(model.parameters(), torch.optim.SGD, lr=0.1, alpha=0.1)
  assert optimizer.minimize_surrogate_gap == True
  optimizer.zero_grad()
  loss1 = criterion(model(x), y)
  loss1.backward()
  optimizer.first_step(zero_grad=True)
  loss2 = criterion(model(x), y)
  loss2.backward()
  optimizer.second_step(zero_grad=True)


@patch('torch.cuda.is_available', return_value=False)
def test_training_setup(_, mock_model):
    optimizer, scheduler, criterion = training_setup(
        mock_model,
        lr=0.001,
        mode='multiclass',
        num_classes=3
    )
    assert isinstance(optimizer, SAM)
    assert isinstance(scheduler, torch.optim.lr_scheduler.ReduceLROnPlateau)
    assert isinstance(criterion, Poly1CrossEntropyLoss)
    optimizer, scheduler, criterion = training_setup(
        mock_model,
        lr=0.001,
        mode='multilabel'
    )
    assert isinstance(optimizer, SAM)
    assert isinstance(scheduler, torch.optim.lr_scheduler.ReduceLROnPlateau)
    assert isinstance(criterion, Poly1CrossEntropyLoss)
    optimizer, scheduler, criterion = training_setup(
        mock_model,
        lr=0.001,
        mode='binary'
    )
    assert not isinstance(optimizer, SAM)
    assert isinstance(scheduler, torch.optim.lr_scheduler.ReduceLROnPlateau)
    assert isinstance(criterion, Poly1CrossEntropyLoss)
    optimizer, scheduler, criterion = training_setup(
        mock_model,
        lr=0.001,
        mode='regression'
    )
    assert not isinstance(optimizer, SAM)
    assert isinstance(scheduler, torch.optim.lr_scheduler.ReduceLROnPlateau)
    assert isinstance(criterion, torch.nn.MSELoss)
    try:
        optimizer, scheduler, criterion = training_setup(
        mock_model,
        lr=0.001,
        mode='multiclass',
        num_classes=2
        )
        return False
    except ValueError:
        pass
    try:
        optimizer, scheduler, criterion = training_setup(
        mock_model,
        lr=0.001,
        mode='wrong',
        num_classes=2
        )
        return False
    except ValueError:
        pass


def test_train_ml_model(mock_xgb_data):
    X_train, X_test, y_train, y_test = mock_xgb_data
    # Test classification
    model = train_ml_model(
        X_train, X_test, y_train, y_test,
        mode='classification',
        feature_calc=False
    )
    assert isinstance(model, xgb.XGBClassifier)
    # Test regression
    model = train_ml_model(
        X_train, X_test, y_train, y_test,
        mode='regression',
        feature_calc=False
    )
    assert isinstance(model, xgb.XGBRegressor)
    # Test feature calculation
    X_train = ["Neu5Ac(a2-3)Gal(b1-4)Glc", "Fuc(a1-3)[Gal(b1-4)]GlcNAc", "Glc(a1-4)Glc"]
    X_test = ["Fuc(a1-2)Gal(b1-4)Glc", "Man(b1-4)GlcNAc(b1-4)[Fuc(a1-6)]GlcNAc", "Neu5Ac(a2-6)GalNAc"]
    model = train_ml_model(
        X_train, X_test, y_train, y_test,
        mode='classification',
        feature_calc=False
    )
    additional_train = pd.DataFrame({'abc': [0, 1, 2], 'def': [3, 4, 5]})
    additional_test = pd.DataFrame({'abc': [9, 8, 7], 'def': [6, 5, 4]})
    model, train, test = train_ml_model(
        X_train, X_test, y_train, y_test,
        mode='classification',
        feature_calc=True,
        feature_set=['terminal'],
        return_features=True,
        additional_features_train=additional_train,
        additional_features_test=additional_test
    )
    assert isinstance(model, xgb.XGBClassifier)
    assert isinstance(train, pd.DataFrame)


@patch('matplotlib.pyplot.show')
def test_analyze_ml_model(mock_show, mock_xgb_data):
    X_train, X_test, y_train, y_test = mock_xgb_data
    model = train_ml_model(X_train, X_test, y_train, y_test, mode='classification')
    analyze_ml_model(model)
    mock_show.assert_called_once()


def test_get_mismatch(mock_xgb_data):
    X_train, X_test, y_train, y_test = mock_xgb_data
    model = train_ml_model(X_train, X_test, y_train, y_test, mode='classification')
    mismatches = get_mismatch(model, X_test, y_test, n=2)
    assert isinstance(mismatches, list)
    assert all(isinstance(m, tuple) and len(m) == 2 for m in mismatches)
    assert len(mismatches) <= 2


# Tests for SweetNet
def test_sweetnet_initialization():
    model = SweetNet(lib_size=100, num_classes=2, hidden_dim=64)
    assert isinstance(model, torch.nn.Module)
    assert model.lin3.out_features == 2
    assert model.item_embedding.num_embeddings == 101  # lib_size + 1


def test_sweetnet_forward():
    model = SweetNet(lib_size=100, num_classes=1, hidden_dim=64)
    batch_size = 3
    num_nodes = 5
    # Create dummy inputs
    x = torch.randint(0, 100, (num_nodes,))
    edge_index = torch.tensor([[0, 1, 2, 3], [1, 2, 3, 4]], dtype=torch.long)
    batch = torch.tensor([0, 0, 1, 1, 2])  # Assign nodes to 3 different graphs
    # Test normal forward pass
    output = model(x, edge_index, batch)
    assert output.shape == (batch_size,)
    # Test inference mode
    output, embedding = model(x, edge_index, batch, inference=True)
    assert isinstance(output, torch.Tensor)
    assert isinstance(embedding, torch.Tensor)


# Tests for NSequonPred
def test_nsequon_pred_initialization():
    model = NSequonPred()
    assert isinstance(model, torch.nn.Module)
    assert model.fc4.out_features == 1


def test_nsequon_pred_forward():
    model = NSequonPred()
    batch_size = 2
    input_dim = 1280
    x = torch.randn(batch_size, input_dim)
    output = model(x)
    assert output.shape == (batch_size, 1)


# Tests for sigmoid_range and SigmoidRange
def test_sigmoid_range():
    x = torch.tensor([0.0, 1.0, 2.0])
    low, high = -1.0, 1.0
    result = sigmoid_range(x, low, high)
    assert torch.all(result >= low)
    assert torch.all(result <= high)


def test_sigmoid_range_module():
    module = SigmoidRange(-1.0, 1.0)
    x = torch.tensor([0.0, 1.0, 2.0])
    result = module(x)
    assert torch.all(result >= -1.0)
    assert torch.all(result <= 1.0)


# Tests for LectinOracle
def test_lectin_oracle_initialization():
    model = LectinOracle(input_size_glyco=100)
    assert isinstance(model, torch.nn.Module)
    assert model.num_classes == 1
    assert model.input_size_glyco == 100


def test_lectin_oracle_forward():
    model = LectinOracle(input_size_glyco=100)
    batch_size = 2
    num_nodes = 4
    # Create dummy inputs
    prot = torch.randn(batch_size, 960)
    nodes = torch.randint(0, 100, (num_nodes,))
    edge_index = torch.tensor([[0, 1, 2], [1, 2, 3]], dtype=torch.long)
    batch = torch.tensor([0, 0, 1, 1])
    # Test normal forward pass
    output = model(prot, nodes, edge_index, batch)
    assert output.shape == (batch_size, 1)
    # Test inference mode
    output, prot_emb, glyco_emb = model(prot, nodes, edge_index, batch, inference=True)
    assert isinstance(output, torch.Tensor)
    assert isinstance(prot_emb, torch.Tensor)
    assert isinstance(glyco_emb, torch.Tensor)


def test_lectin_oracle_flex_forward():
    model = LectinOracle_flex(input_size_glyco=100)
    batch_size = 2
    num_nodes = 4
    # Create dummy inputs
    prot = prot_to_coded(["MKVSAAALAVILIATALCAPASASPYSSDTTPCCFAYIARPLPRAHIKEYFYTSGKCSNPAVVFVTRKNRQVCANPEKKWVREYINSLEMS"])
    prot = torch.tensor(prot, dtype=torch.float)
    prot = prot.repeat(batch_size, 1)
    nodes = torch.randint(0, 100, (num_nodes,))
    edge_index = torch.tensor([[0, 1, 2], [1, 2, 3]], dtype=torch.long)
    batch = torch.tensor([0, 0, 1, 1])
    # Test normal forward pass
    output = model(prot, nodes, edge_index, batch)
    assert output.shape == (batch_size, 1)
    # Test inference mode
    output, prot_emb, glyco_emb = model(prot, nodes, edge_index, batch, inference=True)
    assert isinstance(output, torch.Tensor)
    assert isinstance(prot_emb, torch.Tensor)
    assert isinstance(glyco_emb, torch.Tensor)


# Tests for init_weights
def test_init_weights():
    model = torch.nn.Linear(10, 5)
    # Test sparse initialization
    init_weights(model, mode='sparse', sparsity=0.1)
    assert torch.sum(model.weight == 0) > 0
    # Test kaiming initialization
    init_weights(model, mode='kaiming')
    assert not torch.all(model.weight == 0)
    # Test xavier initialization
    init_weights(model, mode='xavier')
    assert not torch.all(model.weight == 0)
    try:
        init_weights(model, mode='wrong')
        return False
    except:
        pass


# Tests for prep_model
@pytest.mark.parametrize("model_type,num_classes,expected_class", [
    ("SweetNet", 1, SweetNet),
    ("LectinOracle", 2, LectinOracle),
    ("LectinOracle_flex", 1, LectinOracle_flex),
    ("NSequonPred", 1, NSequonPred),
])


def test_prep_model(model_type: str, num_classes: int, expected_class: type):
    mock_lib = {"A": 0, "B": 1, "C": 2}
    model = prep_model(model_type, num_classes, libr=mock_lib, trained=False)
    assert isinstance(model, expected_class)
    device = next(model.parameters()).device
    assert str(device).startswith("cuda" if torch.cuda.is_available() else "cpu")


def test_prep_model_trained():
    model = prep_model("LectinOracle", num_classes=1, trained=True)
    assert isinstance(model, LectinOracle)


@pytest.mark.parametrize("invalid_input", [
    {"model_type": "InvalidModel", "num_classes": 1},
    {"model_type": "SweetNet", "num_classes": 1, "hidden_dim": 64, "trained": True},
])


def test_prep_model_invalid_inputs(invalid_input):
    with pytest.raises((ValueError, Exception)):
        prep_model(**invalid_input)


@pytest.fixture
def sample_data():
    glycans = ["Gal(b1-3)[Neu5Ac(a2-6)]GalNAc", "Neu5Ac(a2-3)Gal(b1-3)GalNAc"]
    protein = "MAEGEITTFTALTEKFNLPPGNYKKPKLLYCSNGGHFLRILPDGTVDGTRDRSDQHIQLQLSAESVGEVYIKSTETGQYLAMDTDGLLYGSQTPNEECLFLERLEENHYNTYISKKHAEKNWFVGLKKNGSCKRGPRTHYGQKAILFLPLPV"
    mock_embeddings = np.random.rand(960).tolist()
    prot_dict = {protein: mock_embeddings}
    return {
        'glycans': glycans,
        'protein': protein,
        'prot_dict': prot_dict,
        'mock_embeddings': mock_embeddings
    }


@pytest.fixture
def mock_models():
    class MockSweetNet(torch.nn.Module):
        def forward(self, x, edge_index, batch, inference=False):
            batch_size = max(batch).item() + 1
            if inference:
                return torch.randn(batch_size, 2), torch.randn(batch_size, 128)
            return torch.randn(batch_size, 2)
    class MockLectinOracle(torch.nn.Module):
        def forward(self, prot, x, edge_index, batch, inference=False):
            batch_size = max(batch).item() + 1
            return torch.tensor([[float(i)] for i in range(batch_size)], dtype=torch.float64)
    class MockNSequonPred(torch.nn.Module):
        def forward(self, x):
            return torch.randn(x.shape[0], 1)
    return {
        'sweetnet': MockSweetNet().to(device).eval(),
        'lectin_oracle': MockLectinOracle().to(device).eval(),
        'nsequon_pred': MockNSequonPred().to(device).eval()
    }


def test_simple_dataset():
    x = [[1.0, 2.0], [3.0, 4.0]]
    y = [0.5, 1.5]
    dataset = SimpleDataset(x, y)
    assert len(dataset) == 2
    x_tensor, y_tensor = dataset[0]
    assert isinstance(x_tensor, torch.FloatTensor)
    assert isinstance(y_tensor, torch.FloatTensor)
    assert torch.equal(x_tensor, torch.FloatTensor([1.0, 2.0]))
    assert torch.equal(y_tensor, torch.FloatTensor([0.5]))


def test_sigmoid():
    # Test basic sigmoid functionality
    assert abs(sigmoid(0) - 0.5) < 1e-6
    assert sigmoid(100) > 0.99999
    assert sigmoid(-100) < 0.00001
    # Test with different inputs
    test_values = [-1.0, 0.0, 1.0, 2.0, -2.0]
    for x in test_values:
        y = sigmoid(x)
        assert 0 <= y <= 1


def test_glycans_to_emb(sample_data, mock_models):
    glycans = sample_data['glycans']
    model = mock_models['sweetnet']
    # Test representation mode
    rep_df = glycans_to_emb(glycans, model, rep=True)
    assert isinstance(rep_df, pd.DataFrame)
    assert len(rep_df) == len(glycans)
    # Test prediction mode
    class_list = ['class1', 'class2']
    preds = glycans_to_emb(glycans, model, rep=False, class_list=class_list)
    assert isinstance(preds, list)
    assert len(preds) == len(glycans)
    assert all(p in class_list for p in preds)


def test_get_multi_pred(sample_data, mock_models):
    model = mock_models['lectin_oracle']
    results = get_multi_pred(
        prot=sample_data['protein'],
        glycans=sample_data['glycans'],
        model=model,
        prot_dic=sample_data['prot_dict']
    )
    assert isinstance(results, list)
    assert len(results) == len(sample_data['glycans'])
    # Test with background correction
    correction_df = pd.DataFrame({
        'motif': sample_data['glycans'],
        'pred': pd.Series([0.1, 0.2], dtype=np.float64)
    })
    results_corrected = get_multi_pred(
        prot=sample_data['protein'],
        glycans=sample_data['glycans'],
        model=model,
        prot_dic=sample_data['prot_dict'],
        background_correction=True,
        correction_df=correction_df
    )
    assert len(results_corrected) == len(sample_data['glycans'])
    results_corrected = get_multi_pred(
        prot=sample_data['protein'],
        glycans=sample_data['glycans'],
        model=model,
        prot_dic=sample_data['prot_dict'],
        background_correction=True,
        correction_df=correction_df,
        flex=True
    )
    correction_df = pd.DataFrame({
        'motif': sample_data['glycans'][:-1],
        'pred': pd.Series([0.1], dtype=np.float64)
    })
    results_corrected = get_multi_pred(
        prot=sample_data['protein'],
        glycans=sample_data['glycans'],
        model=model,
        prot_dic=sample_data['prot_dict'],
        background_correction=True,
        correction_df=correction_df
    )


def test_get_lectin_preds(sample_data, mock_models):
    model = mock_models['lectin_oracle']
    # Test basic functionality
    df = get_lectin_preds(
        prot=sample_data['protein'],
        glycans=sample_data['glycans'],
        model=model,
        prot_dic=sample_data['prot_dict']
    )
    assert isinstance(df, pd.DataFrame)
    assert 'motif' in df.columns
    assert 'pred' in df.columns
    assert len(df) == len(sample_data['glycans'])
    try:
        df = get_lectin_preds(
        prot=sample_data['protein'],
        glycans=sample_data['glycans'],
        model=model,
        prot_dic=None
        )
        return False
    except:
        pass
    # Test with background correction
    correction_df = pd.DataFrame({
        'motif': sample_data['glycans'],
        'pred': pd.Series([0.1, 0.2], dtype=np.float64)
    })
    df_corrected = get_lectin_preds(
        prot=sample_data['protein'],
        glycans=sample_data['glycans'],
        model=model,
        prot_dic=sample_data['prot_dict'],
        background_correction=True,
        correction_df=correction_df
    )
    assert len(df_corrected) == len(sample_data['glycans'])


def test_get_esmc_representations(sample_data):
    # It's in fair-esm, nothing we can do about it on our end
    warnings.filterwarnings("ignore", category=FutureWarning,
                         message="You are using `torch.load` with `weights_only=False`.*")
    class MockESMC(torch.nn.Module):
      def encode(self, protein):
        return torch.zeros(1, 100, 960)
      def logits(self, protein_tensor, config):
        class MockLogitsOutput:
          def __init__(self):
            self.embeddings = torch.randn(1, 100, 960)
        return MockLogitsOutput()
    model = MockESMC()
    proteins = [sample_data['protein']]
    with patch.dict(sys.modules, {'esm.sdk.api': None}):
        result = get_esmc_representations(proteins, model)
    assert isinstance(result, dict)
    assert len(result) == len(set(proteins))
    assert all(isinstance(v, list) for v in result.values())
    assert all(len(v) == 960 for v in result.values())


def test_get_Nsequon_preds(sample_data, mock_models):
    model = mock_models['nsequon_pred']
    sequences = [
        "zzzzzzzzzzzzzzzzzzzNGSzzzzzzzzzzzzzzzzzz",
        "zzzzzzzzzzzzzzzzzzzNATzzzzzzzzzzzzzzzzzz"
    ]
    df = get_Nsequon_preds(
        prots=sequences,
        model=model,
        prot_dic={seq: sample_data['mock_embeddings'] for seq in sequences}
    )
    assert isinstance(df, pd.DataFrame)
    assert 'seq' in df.columns
    assert 'glycosylated' in df.columns
    assert len(df) == len(sequences)
    assert all(0 <= p <= 1 for p in df['glycosylated'])


def test_mock_data(mock_dataloader):
    """Helper function to verify mock data structure"""
    data = next(iter(mock_dataloader['train']))
    assert hasattr(data, 'labels')
    assert hasattr(data, 'y')
    assert hasattr(data, 'edge_index')
    assert hasattr(data, 'batch')
    assert isinstance(data.labels, torch.Tensor)
    assert isinstance(data.y, torch.Tensor)
    assert isinstance(data.edge_index, torch.Tensor)
    assert isinstance(data.batch, torch.Tensor)
    assert data.batch.max() == 1  # Should have exactly 2 graphs
    assert len(data.y) == 2  # Should have 2 labels (one per graph)
    assert len(data.labels) == 6  # Should have 6 nodes
    assert (data.batch == 0).sum() == 3  # 3 nodes in first graph
    assert (data.batch == 1).sum() == 3  # 3 nodes in second graph


def test_train_model_all_modes(mode, expected_metrics, mock_model, mock_dataloader):
    # Configure based on mode
    if mode == 'regression':
        criterion = nn.MSELoss().to(device)
        optimizer = torch.optim.Adam(mock_model.parameters(), lr=0.1)
    elif mode == 'classification':
        criterion = nn.CrossEntropyLoss().to(device)
        optimizer = SAM(mock_model.parameters(), torch.optim.SGD, lr=0.1)
    else:  # multilabel
        criterion = nn.BCEWithLogitsLoss().to(device)
        optimizer = SAM(mock_model.parameters(), torch.optim.SGD, lr=0.1)
    scheduler = torch.optim.lr_scheduler.StepLR(
        optimizer.base_optimizer if isinstance(optimizer, SAM) else optimizer,
        step_size=1
    )
    # Run training
    _, metrics = train_model(
        mock_model,
        mock_dataloader,
        criterion,
        optimizer,
        scheduler,
        num_epochs=2,
        mode=mode,
        mode2='multi' if mode != 'regression' else 'binary',
        return_metrics=True
    )
    # Verify metrics structure
    assert isinstance(metrics, dict)
    assert 'train' in metrics and 'val' in metrics
    for phase in ['train', 'val']:
        assert all(key in metrics[phase] for key in expected_metrics)
        for metric in metrics[phase].values():
            assert len(metric) == 2  # Two epochs
            assert all(isinstance(v, float) for v in metric)
    _ = train_model(
        mock_model,
        mock_dataloader,
        criterion,
        optimizer,
        scheduler,
        num_epochs=2,
        mode=mode,
        mode2='multi' if mode != 'regression' else 'binary',
        return_metrics=False
    )


def test_train_model_plotting(mock_model, mock_dataloader):
    if mode == "classification":
        criterion = nn.CrossEntropyLoss()
        optimizer = SAM(mock_model.parameters(), torch.optim.SGD, lr=0.1, alpha=0.1)
        scheduler = torch.optim.lr_scheduler.StepLR(
            optimizer.base_optimizer,
            step_size=1
        )
        # Run training
        _ = train_model(
            mock_model,
            mock_dataloader,
            criterion,
            optimizer,
            scheduler,
            num_epochs=2,
            mode=mode,
            mode2='multi',
            return_metrics=False
        )
