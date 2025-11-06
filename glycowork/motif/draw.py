from pathlib import Path
from glycowork.glycan_data.loader import unwrap, motif_list, lib
from glycowork.motif.regex import get_match
from glycowork.motif.graph import glycan_to_nxGraph, subgraph_isomorphism, compare_glycans, graph_to_string
from glycowork.motif.tokenization import get_core, get_modification
from glycowork.motif.processing import min_process_glycans, rescue_glycans, in_lib, expand_lib, get_matching_indices
import matplotlib.pyplot as plt
from io import BytesIO
from typing import Dict, List, Tuple, Optional, Union, Any
import networkx as nx
import drawsvg as draw
from glycorender.render import convert_svg_to_pdf, convert_svg_to_png
import numpy as np
import pandas as pd
import re
from math import sin, cos, radians, sqrt, atan, degrees


# Adjusted SNFG color palette
col_dict_base = {
    'snfg_white': '#FFFFFF', 'snfg_alt_blue': '#0385AE', 'snfg_green': '#058F60', 'snfg_yellow': '#FCC326',
    'snfg_light_blue': '#91D3E3', 'snfg_pink': '#F39EA0', 'snfg_purple': '#A15989', 'snfg_brown': '#9F6D55',
    'snfg_orange': '#EF6130', 'snfg_red': '#C23537', 'black': '#000000', 'grey': '#7F7F7F'
}

col_dict_transparent = {
    'snfg_white': '#FFFFFF', 'snfg_alt_blue': '#CDE7EF', 'snfg_green': '#CDE9DF', 'snfg_yellow': '#FFF6DE',
    'snfg_light_blue': '#EEF8FB', 'snfg_pink': '#FDF0F1', 'snfg_purple': '#F1E6ED', 'snfg_brown': '#F1E9E5',
    'snfg_orange': '#FDE7E0', 'snfg_red': '#F7E0E0', 'black': '#D9D9D9', 'grey': '#ECECEC'
}

# Shape-color mapping
sugar_dict = {
  "Hex": ['Hex', 'snfg_white', False], "Glc": ['Hex', 'snfg_alt_blue', False],
  "Glcf": ['Hex', 'snfg_alt_blue', True], "Man": ['Hex', 'snfg_green', False],
  "Manf": ['Hex', 'snfg_green', True], "Gal": ['Hex', 'snfg_yellow', False],
  "Galf": ['Hex', 'snfg_yellow', True], "Gul": ['Hex', 'snfg_orange', False],
  "Alt": ['Hex', 'snfg_pink', False], "All": ['Hex', 'snfg_purple', False],
  "Tal": ['Hex', 'snfg_light_blue', False], "Ido": ['Hex', 'snfg_brown', False],

  "HexNAc": ['HexNAc', 'snfg_white', False], "GlcNAc": ['HexNAc', 'snfg_alt_blue', False],
  "GlcfNAc": ['HexNAc', 'snfg_alt_blue', True], "ManNAc": ['HexNAc', 'snfg_green', False],
  "ManfNAc": ['HexNAc', 'snfg_green', True], "GalNAc": ['HexNAc', 'snfg_yellow', False],
  "GalfNAc": ['HexNAc', 'snfg_yellow', True], "GulNAc": ['HexNAc', 'snfg_orange', False],
  "AltNAc": ['HexNAc', 'snfg_pink', False], "AllNAc": ['HexNAc', 'snfg_purple', False],
  "TalNAc": ['HexNAc', 'snfg_light_blue', False], "IdoNAc": ['HexNAc', 'snfg_brown', False],

  "HexN": ['HexN', 'snfg_white', False], "GlcN": ['HexN', 'snfg_alt_blue', False],
  "ManN": ['HexN', 'snfg_green', False], "GalN": ['HexN', 'snfg_yellow', False],
  "GulN": ['HexN', 'snfg_orange', False], "AltN": ['HexN', 'snfg_pink', False],
  "AllN": ['HexN', 'snfg_purple', False], "TalN": ['HexN', 'snfg_light_blue', False],
  "IdoN": ['HexN', 'snfg_brown', False],

  "HexA": ['HexA', 'snfg_white', False], "GlcA": ['HexA', 'snfg_alt_blue', False],
  "ManA": ['HexA', 'snfg_green', False], "GalA": ['HexA', 'snfg_yellow', False],
  "GulA": ['HexA', 'snfg_orange', False], "AltA": ['HexA_2', 'snfg_pink', False],
  "AllA": ['HexA', 'snfg_purple', False], "TalA": ['HexA', 'snfg_light_blue', False],
  "IdoA": ['HexA_2', 'snfg_brown', False],

  "dHex": ['dHex', 'snfg_white', False], "Qui": ['dHex', 'snfg_alt_blue', False],
  "Rha": ['dHex', 'snfg_green', False], "6dGul": ['dHex', 'snfg_orange', False],
  "6dAlt": ['dHex', 'snfg_pink', False], "6dAltf": ['dHex', 'snfg_pink', True],
  "6dTal": ['dHex', 'snfg_light_blue', False], "Fuc": ['dHex', 'snfg_red', False],
  "Fucf": ['dHex', 'snfg_red', True],

  "dHexNAc": ['dHexNAc', 'snfg_white', False], "QuiNAc": ['dHexNAc', 'snfg_alt_blue', False],
  "RhaNAc": ['dHexNAc', 'snfg_green', False], "6dAltNAc": ['dHexNAc', 'snfg_pink', False],
  "6dTalNAc": ['dHexNAc', 'snfg_light_blue', False], "FucNAc": ['dHexNAc', 'snfg_red', False],
  "FucfNAc": ['dHexNAc', 'snfg_red', True],

  "ddHex": ['ddHex', 'snfg_white', False], "Oli": ['ddHex', 'snfg_alt_blue', False],
  "Tyv": ['ddHex', 'snfg_green', False], "Abe": ['ddHex', 'snfg_orange', False],
  "Par": ['ddHex', 'snfg_pink', False], "Dig": ['ddHex', 'snfg_purple', False],
  "Col": ['ddHex', 'snfg_light_blue', False],

  "Pen": ['Pen', 'snfg_white', False], "Ara": ['Pen', 'snfg_green', False],
  "Araf": ['Pen', 'snfg_green', True], "Lyx": ['Pen', 'snfg_yellow', False],
  "Lyxf": ['Pen', 'snfg_yellow', True], "Xyl": ['Pen', 'snfg_orange', False],
  "Xylf": ['Pen', 'snfg_orange', True], "Rib": ['Pen', 'snfg_pink', False],
  "Ribf": ['Pen', 'snfg_pink', True],

  "dNon": ['dNon', 'snfg_white', False], "Kdn": ['dNon', 'snfg_green', False],
  "Neu5Ac": ['dNon', 'snfg_purple', False], "Neu5Gc": ['dNon', 'snfg_light_blue', False],
  "Neu": ['dNon', 'snfg_brown', False], "Sia": ['dNon', 'snfg_red', False],

  "ddNon": ['ddNon', 'snfg_white', False], "Pse": ['ddNon', 'snfg_green', False],
  "Leg": ['ddNon', 'snfg_yellow', False], "Aci": ['ddNon', 'snfg_pink', False],
  "4eLeg": ['ddNon', 'snfg_light_blue', False],

  "Unknown": ['Unknown', 'snfg_white', False], "Bac": ['Unknown', 'snfg_alt_blue', False],
  "LDManHep": ['Unknown', 'snfg_green', False], "Kdo": ['Unknown', 'snfg_yellow', False],
  "Kdof": ['Unknown', 'snfg_yellow', True], "Dha": ['Unknown', 'snfg_orange', False],
  "DDManHep": ['Unknown', 'snfg_pink', False], "MurNAc": ['Unknown', 'snfg_purple', False],
  "MurNGc": ['Unknown', 'snfg_light_blue', False], "Mur": ['Unknown', 'snfg_brown', False],

  "Assigned": ['Assigned', 'snfg_white', False], "Api": ['Assigned', 'snfg_alt_blue', False],
  "Apif": ['Assigned', 'snfg_alt_blue', True], "Fru": ['Assigned', 'snfg_green', False],
  "Fruf": ['Assigned', 'snfg_green', True], "Tag": ['Assigned', 'snfg_yellow', False],
  "Sor": ['Assigned', 'snfg_orange', False], "Psi": ['Assigned', 'snfg_pink', False],
  "non_glycan": ['Assigned', 'black', False],

  "blank": ['empty', 'snfg_white', False], "text": ['text', None, None], "-": ['empty', None, None],
  "red_end": ['red_end', None, None], "free": ['free', None, None], "show": ['empty', None, None], "hide": ['empty', None, None],
  "04X": ['04X', None, None], "15A": ['15A', None, None], "02A": ['02A', None, None], "13X": ['13X', None, None],
  "24X": ['24X', None, None], "35X": ['35X', None, None], "04A": ['04A', None, None], "15X": ['15X', None, None],
  "02X": ['02X', None, None], "13A": ['13A', None, None], "24A": ['24A', None, None], "35A": ['35A', None, None],
  "25A": ['25A', None, None], "03A": ['03A', None, None], "14X": ['14X', None, None], "25X": ['25X', None, None],
  "03X": ['03X', None, None], "14A": ['14A', None, None], "Z": ['Z', None, None], "Y": ['Y', None, None],
  "B": ['B', None, None], "C": ['C', None, None]
}


domon_costello = {'B', 'C', 'Z', 'Y', '04X', '15A', '02A', '13X', '24X', '35X', '04A', '15X', '02X', '13A', '24A', '35A', '25A', '03A', '14X', '25X', '03X', '14A'}


def draw_hex(
    x_pos: float, # X coordinate of hexagon center
    y_pos: float, # Y coordinate of hexagon center
    dim: float, # Base dimension for scaling
    col_dict: Dict[str, str], # Color mapping dictionary
    drawing: draw.Drawing, # Glycan drawing to be modified
    color: str = 'white', # Fill color
    outline_only: bool = False # Whether to draw only the circumference
    ) -> None:
  "Draws filled hexagon shape with border at specified position and scale"
  x_base = -x_pos * dim
  y_base = y_pos * dim
  half_dim = 0.5 * dim
  stroke_width = 0.04 * dim
  points = []
  for angle in (0, 60, 120, 180, 240, 300):
    rad = radians(angle)
    points.extend([x_base + half_dim * cos(rad), y_base + half_dim * sin(rad)])
  if outline_only:
    p = draw.Path(stroke_width = stroke_width, stroke = col_dict['black'], fill = 'none')
    p.M(points[0], points[1])  # Move to first point
    for i in range(2, len(points), 2):  # Line to subsequent points
      p.L(points[i], points[i+1])
    p.Z()  # Close path
    drawing.append(p)
  else:
    # Draw filled hexagon with border
    drawing.append(draw.Lines(*points, close = True, fill = color, stroke = col_dict['black'], stroke_width = stroke_width))


def add_customization(
    drawing, # Drawing object to modify
    x_base: float, # X coordinate of base position
    y_base: float, # Y coordinate of base position
    dim: float, # Base dimension for scaling
    modification: str, # Text annotation for modifications
    col_dict: Dict[str, str], # Color mapping dictionary
    conf: str = None, # Ring configuration text
    furanose: bool = False, # Draw furanose indicator
    text_anchor: str = 'middle' # Text alignment
    ) -> None:
  "Adds text annotations and indicators to glycan symbol"
  half_dim = dim / 2
  # Text annotation
  p = draw.Path(stroke_width = 0)
  p.M(x_base-dim, y_base+half_dim)
  p.L(x_base+dim, y_base+half_dim)
  drawing.append(p)
  drawing.append(draw.Text(modification, dim*0.35, path = p, fill = col_dict['black'], text_anchor = text_anchor, line_offset = -3.15))
  if furanose or conf:
    conf_text = ""
    if conf:
      conf_dict = {'L-': 'L', 'D-': 'D', '1,7lactone': 'on'}
      conf_text = conf_dict[conf] if isinstance(conf, str) and conf in conf_dict else conf
    if furanose:
      conf_text += "f"
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base)
    p.L(x_base+dim, y_base)
    drawing.append(p)
    drawing.append(draw.Text(conf_text, dim*0.3, path = p, fill = col_dict['black'], text_anchor = text_anchor, center = True))


def draw_shape(
    shape: str, # SNFG shape designation
    color: str, # SNFG color designation
    x_pos: float, # X coordinate of shape center
    y_pos: float, # Y coordinate of shape center
    col_dict: Dict[str, str], # Color mapping dictionary
    drawing: draw.Drawing, # Glycan drawing to be modified
    modification: str = '', # Text annotation for modifications
    dim: float = 50, # Base dimension for scaling
    furanose: bool = False, # Draw furanose indicator
    conf: str = '', # Ring configuration text
    deg: float = 0, # Rotation angle in degrees
    text_anchor: str = 'middle', # Text alignment for postbiosynthetic modifications
    scalar: float = 0 # Intensity scaling factor for drawsvg output
    ) -> None:
  "Draws SNFG glycan symbol with specified shape, color, position, and optional annotations"
  x_base = -x_pos * dim
  y_base = y_pos * dim
  stroke_w = 0.04 * dim
  half_dim = dim / 2
  inside_hex_dim = ((sqrt(3))/2) * half_dim
  if scalar:
    col_dict_scalar = {
    'snfg_white': '#FFFFFF', 'snfg_alt_blue': 'darkblue', 'snfg_green': 'green', 'snfg_yellow': 'darkgoldenrod',
    'snfg_light_blue': 'skyblue', 'snfg_pink': 'orchid', 'snfg_purple': 'purple', 'snfg_brown': 'saddlebrown',
    'snfg_orange': 'orangered', 'snfg_red': 'firebrick', 'black': '#000000', 'grey': '#7F7F7F'}
    radius = 2.35 if shape == "HexNAc" else 2.2
    gradient = draw.RadialGradient(x_base, y_base, half_dim * radius)
    opacity = max(0, min(1, scalar))*0.8  # Normalize opacity to [0, 1]
    opacity = opacity * 1.3 if color in ['snfg_yellow', 'snfg_light_blue', 'snfg_white'] else opacity
    gradient.add_stop(0, col_dict_scalar[color], opacity = opacity)
    gradient.add_stop(0.6, col_dict_scalar[color], opacity = opacity * 0.4)
    gradient.add_stop(1, 'white', opacity = 0)
    drawing.append(draw.Circle(x_base, y_base, half_dim * radius, fill = gradient))

  if shape == 'Hex':
    # Hexose - circle
    drawing.append(draw.Circle(x_base, y_base, half_dim, fill = col_dict[color], stroke_width = stroke_w, stroke = col_dict['black']))
  elif shape == 'HexNAc':
    # HexNAc - square
    drawing.append(draw.Rectangle(x_base-half_dim, y_base-half_dim, dim, dim, fill = col_dict[color], stroke_width = stroke_w, stroke = col_dict['black']))
  elif shape == 'HexN':
    # Hexosamine - crossed square
    drawing.append(draw.Rectangle(x_base-half_dim, y_base-half_dim, dim, dim, fill = 'white', stroke_width = stroke_w, stroke = col_dict['black']))
    drawing.append(draw.Lines(x_base-half_dim, y_base-half_dim,
                        x_base+half_dim, y_base-half_dim,
                        x_base+half_dim, y_base+half_dim,
                        x_base-half_dim, y_base-half_dim,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = 0))
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base-half_dim, y_base-half_dim)
    p.L(x_base+half_dim, y_base-half_dim)
    p.M(x_base+half_dim, y_base-half_dim)
    p.L(x_base+half_dim, y_base+half_dim)
    p.M(x_base+half_dim, y_base+half_dim)
    p.L(x_base-half_dim, y_base-half_dim)
    drawing.append(p)
  elif shape in ['HexA_2', 'HexA']:
    # Hexuronate - divided diamond;  AltA / IdoA for HexA_2 and flipped colors for HexA
    drawing.append(draw.Lines(x_base,         y_base+half_dim,
                        x_base+half_dim, y_base,
                        x_base,         y_base-half_dim,
                        x_base-half_dim, y_base,
                        close = True, fill = 'white', stroke = col_dict['black'], stroke_width = stroke_w))
    y_direction = half_dim if shape == 'HexA_2' else -half_dim
    drawing.append(draw.Lines(x_base-half_dim, y_base,
                        x_base, y_base+y_direction,
                        x_base+half_dim, y_base,
                        x_base-half_dim, y_base,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = 0))
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'], fill = 'none')
    p.M(x_base-half_dim, y_base)
    p.L(x_base, y_base+y_direction)
    p.L(x_base+half_dim, y_base)
    drawing.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'], fill = 'none')
    p.M(x_base-half_dim, y_base)
    p.L(x_base+half_dim, y_base)
    drawing.append(p)
  elif shape == 'dHex':
    # Deoxyhexose - triangle
    drawing.append(draw.Lines(x_base- half_dim, y_base+inside_hex_dim,  # -(dim*1/3)
                        x_base, y_base-inside_hex_dim,
                        x_base+half_dim, y_base+inside_hex_dim,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = stroke_w))
  elif shape == 'dHexNAc':
    # Deoxyhexnac - divided triangle
    drawing.append(draw.Lines(x_base-half_dim, y_base+inside_hex_dim,  # -(dim*1/3) for center of triangle
                        x_base, y_base-inside_hex_dim,  # -(dim*1/3) for bottom alignment
                        x_base+half_dim, y_base+inside_hex_dim,  # -(((3**0.5)/2)*dim*0.5) for half of triangle height
                        close = True, fill = 'white', stroke = col_dict['black'], stroke_width = stroke_w))
    drawing.append(draw.Lines(x_base, y_base+inside_hex_dim,  # -(dim*1/3)
                        x_base, y_base-inside_hex_dim,
                        x_base+half_dim, y_base+inside_hex_dim,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = 0))
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'], fill = 'none')
    p.M(x_base, y_base+inside_hex_dim)
    p.L(x_base, y_base-inside_hex_dim)
    drawing.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'], fill = 'none')
    p.M(x_base, y_base+inside_hex_dim)
    p.L(x_base+half_dim, y_base+inside_hex_dim)
    drawing.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'], fill = 'none')
    p.M(x_base, y_base-inside_hex_dim)
    p.L(x_base+half_dim, y_base+inside_hex_dim)
    drawing.append(p)
  elif shape == 'ddHex':
    # Dideoxyhexose - flat rectangle
    drawing.append(draw.Lines(x_base-half_dim,         y_base+(dim*7/12*0.5),  # -(dim*0.5/12)
                        x_base+half_dim,         y_base+(dim*7/12*0.5),
                        x_base+half_dim,         y_base-(dim*7/12*0.5),
                        x_base-half_dim,         y_base-(dim*7/12*0.5),
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = stroke_w))
  elif shape == 'Pen':
    # Pentose - star
    cos18 = cos(radians(18))
    cos54 = cos(radians(54))
    sin18 = sin(radians(18))
    sin54 = sin(radians(54))
    base_r = half_dim/cos18
    small_r = (0.25*dim)/cos18
    drawing.append(draw.Lines(x_base, y_base-base_r,
                    x_base+small_r*cos54, y_base-small_r*sin54,
                    x_base+base_r*cos18, y_base-base_r*sin18,
                    x_base+small_r*cos18, y_base+small_r*sin18,
                    x_base+base_r*cos54, y_base+base_r*sin54,
                    x_base, y_base+small_r,
                    x_base-base_r*cos54, y_base+base_r*sin54,
                    x_base-small_r*cos18, y_base+small_r*sin18,
                    x_base-base_r*cos18, y_base-base_r*sin18,
                    x_base-small_r*cos54, y_base-small_r*sin54,
                   close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = stroke_w))
  elif shape in ['dNon', 'ddNon']:
    # Deoxynonulosonate - diamond or Dideoxynonulosonate - flat diamond
    diamond_adjust = 0 if shape == 'dNon' else dim*1/8
    drawing.append(draw.Lines(x_base,         y_base+half_dim-diamond_adjust,
                        x_base+half_dim+diamond_adjust, y_base,
                        x_base,         y_base-half_dim+diamond_adjust,
                        x_base-half_dim-diamond_adjust, y_base,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = stroke_w))
  elif shape == 'Unknown':
    # Unknown - flat hexagon
    flat_adjust = dim*1/8
    extra = dim*0.2
    drawing.append(draw.Lines(x_base-half_dim+flat_adjust, y_base+half_dim-flat_adjust,
                        x_base+half_dim-flat_adjust, y_base+half_dim-flat_adjust,
                        x_base+half_dim-flat_adjust+extra, y_base,
                        x_base+half_dim-flat_adjust, y_base-half_dim+flat_adjust,
                        x_base-half_dim+flat_adjust, y_base-half_dim+flat_adjust,
                        x_base-half_dim+flat_adjust-extra, y_base,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = stroke_w))
  elif shape == 'Assigned':
    # Assigned - pentagon
    cos18 = cos(radians(18))
    cos54 = cos(radians(54))
    sin18 = sin(radians(18))
    sin54 = sin(radians(54))
    base_r = half_dim/cos18
    drawing.append(draw.Lines(x_base, y_base-base_r,
                        x_base+base_r*cos18, y_base-base_r*sin18,
                        x_base+base_r*cos54, y_base+base_r*sin54,
                        x_base-base_r*cos54, y_base+base_r*sin54,
                        x_base-base_r*cos18, y_base-base_r*sin18,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = stroke_w))
  elif shape == 'empty':
    drawing.append(draw.Circle(x_base, y_base, dim/2, fill = 'none', stroke_width = stroke_w, stroke = 'none'))
  elif shape == 'text':
    drawing.append(draw.Text(modification, dim*0.35, x_base, y_base, text_anchor = text_anchor, fill = col_dict['black']))
  elif shape in {'red_end', 'free'}:
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'], fill = 'none')
    p.M((x_base+0.1*dim), (y_base-0.4*dim))  # Start path at point (-10, 20)
    p.C((x_base-0.3*dim), (y_base-0.1*dim),
        (x_base+0.3*dim), (y_base+0.1*dim),
        (x_base-0.1*dim), (y_base+0.4*dim))
    drawing.append(p)
    if shape == 'red_end':
      drawing.append(draw.Circle(x_base, y_base, 0.15 * dim, fill = 'white',
                                 stroke_width = stroke_w, stroke = col_dict['black']))
  # Handle segmented Hex shapes (04X, 15A, etc.)
  elif any(x in shape for x in {'04', '15', '02', '13', '24', '35', '25', '03', '14'}):
    use_grey_base = shape in {'04A', '15X', '02X', '13A', '24A', '35A', '14A'}
    segment_fill = 'white' if use_grey_base else col_dict['grey']
    # Define angle pairs for each shape type
    angles = {
      '04': (30, 150, [60, 120]), '15': (90, 330, [60, 0]),
      '02': (30, 270, [0, 300]), '13': (330, 210, [300, 240]),
      '24': (270, 150, [240, 180]), '35': (210, 90, [180, 120]),
      '25': (90, 270, [60, 0, 300]), '03': (30, 210, [0, 300, 240]),
      '14': (330, 150, [300, 240, 180])
    }
    start_angle, end_angle, mid_angles = angles[shape[:2]]
    draw_hex(x_pos, y_pos, dim, col_dict, drawing, color = col_dict['grey'] if use_grey_base else 'white')
    # Draw the segment
    points = [x_base, y_base]
    points.extend([x_base+inside_hex_dim*cos(radians(start_angle)), y_base-inside_hex_dim*sin(radians(start_angle))])
    for angle in mid_angles:
      points.extend([x_base+half_dim*cos(radians(angle)), y_base-half_dim*sin(radians(angle))])
    points.extend([x_base+inside_hex_dim*cos(radians(end_angle)), y_base-inside_hex_dim*sin(radians(end_angle))])
    drawing.append(draw.Lines(*points, close = True, fill = segment_fill, stroke = col_dict['black'], stroke_width = 0))
    # Draw the dividing line - either center-to-edge or edge-to-edge
    if shape[:2] in {'25', '03', '14'}:
      p = draw.Path(stroke_width=stroke_w, stroke=col_dict['black'])
      p.M(x_base+inside_hex_dim*cos(radians(start_angle)), y_base-inside_hex_dim*sin(radians(start_angle)))
      p.L(x_base+inside_hex_dim*cos(radians(end_angle)), y_base-inside_hex_dim*sin(radians(end_angle)))
      drawing.append(p)
    else:
      for angle in [start_angle, end_angle]:
        p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
        p.M(x_base, y_base)
        p.L(x_base+inside_hex_dim*cos(radians(angle)), y_base-inside_hex_dim*sin(radians(angle)))
        drawing.append(p)
    # Draw outline
    draw_hex(x_pos, y_pos, dim, col_dict, drawing, outline_only = True)
  elif shape in {'Z', 'Y'}:
    rot = f'rotate({deg} {-abs(x_pos)*dim} {-abs(y_pos)*dim})'
    g = draw.Group(transform = rot)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base, y_base-half_dim)
    p.L(x_base, y_base+half_dim)
    p.M(x_base-0.02*dim, y_base-half_dim)
    p.L(x_base+0.4*dim, y_base-half_dim)
    g.append(p)
    if shape == 'Y':
      g.append(draw.Circle(x_base + 0.4 * dim, y_base, 0.15 * dim, fill = 'none',
                           stroke_width = stroke_w, stroke = col_dict['black']))
    drawing.append(g)
  elif shape in {'B', 'C'}:
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base, y_base-half_dim)
    p.L(x_base, y_base+half_dim)
    p.M(x_base+0.02*dim, y_base+half_dim)
    p.L(x_base-0.4*dim, y_base+half_dim)
    drawing.append(p)
    if shape == 'C':
      drawing.append(draw.Circle(x_base - 0.4 * dim, y_base, 0.15 * dim, fill = 'none',
                                 stroke_width = stroke_w, stroke = col_dict['black']))
  if shape not in {'empty', 'text', 'red_end', 'free', 'Z', 'Y', 'B', 'C'} and not any(x in shape for x in {'04', '15', '02', '13', '24', '35', '25', '03', '14'}):
    add_customization(drawing, x_base, y_base, dim, modification, col_dict, conf, furanose, text_anchor)


def add_bond(
    x_start: float, # Starting X coordinate
    x_stop: float, # Ending X coordinate
    y_start: float, # Starting Y coordinate
    y_stop: float, # Ending Y coordinate
    drawing: draw.Drawing, # Glycan drawing to be modified
    label: str = '', # Bond label text
    dim: float = 50, # Base dimension for scaling
    compact: bool = False, # Use compact drawing style
    highlight: str = 'show', # Highlight state: 'show' or 'hide'
    color_highlight:  bool = False # Whether to highlight this linkage in red
    ) -> None:
  "Draws glycosidic bond line with optional label between specified coordinates"
  col_dict = col_dict_transparent if highlight == 'hide' else col_dict_base
  scaling_factor = 1.2 if compact else 2
  y_scaling = 0.6 if compact else 1
  x_start, x_stop = [-x * scaling_factor * dim for x in (x_start, x_stop)]
  y_start, y_stop = [y * y_scaling * dim for y in (y_start, y_stop)]
  final_width = 0.12*dim if color_highlight else 0.08*dim
  p = draw.Path(stroke_width = final_width, stroke = col_dict['snfg_red'] if color_highlight else col_dict['black'],)
  p.M(x_start, y_start).L(x_stop, y_stop)
  drawing.append(p)
  if label and label != '-':
    drawing.append(draw.Text(label, dim*0.4, path = p, text_anchor = 'middle',
                             fill = col_dict['black'], valign = 'middle', line_offset = -0.5))


def add_sugar(
    monosaccharide: str, # IUPAC monosaccharide name
    drawing: draw.Drawing, # Glycan drawing to be modified
    x_pos: float = 0, # X coordinate of sugar center
    y_pos: float = 0, # Y coordinate of sugar center
    modification: str = '', # Text annotation for modifications
    dim: float = 50, # Base dimension for scaling
    compact: bool = False, # Use compact drawing style
    conf: str = '', # Ring configuration text
    deg: float = 0, # Rotation angle in degrees
    text_anchor: str = 'middle', # Text alignment for postbiosynthetic modifications
    highlight: str = 'show', # Highlight state: 'show' or 'hide'
    scalar: float = 0 # Intensity scaling factor
    ) -> None:
  "Draws SNFG monosaccharide symbol with specified parameters at given position"
  col_dict = col_dict_transparent if highlight == 'hide' else col_dict_base
  x_pos = x_pos * (1.2 if compact else 2)
  y_pos = y_pos * (0.6 if compact else 1)
  if monosaccharide in sugar_dict:
    shape, color, furanose = sugar_dict[monosaccharide]
    draw_shape(shape = shape, color = color, x_pos = x_pos, y_pos = y_pos, drawing = drawing, modification = modification,
               conf = conf, furanose = furanose, dim = dim, deg = deg, text_anchor = text_anchor, col_dict = col_dict, scalar = scalar)
  else:
    x_base = -x_pos * dim
    y_base = y_pos * dim
    half_dim = dim / 2
    p = draw.Path(stroke_width = 0.04 * dim, stroke = 'black')
    p.M(x_base-half_dim, y_base+half_dim)
    p.L(x_base+half_dim, y_base-half_dim)
    p.M(x_base+half_dim, y_base+half_dim)
    p.L(x_base-half_dim, y_base-half_dim)
    drawing.append(p)


def process_bonds(
    linkage_list: Union[List[str], List[List[str]]] # Glycosidic linkages
    ) -> Union[List[str], List[List[str]]]: # Formatted linkage text
  "Formats glycosidic linkage text for visualization"
  ALPHA_PATTERN = re.compile(r"^a\d")
  BETA_PATTERN = re.compile(r"^b\d")
  DIGIT_PATTERN = re.compile(r"^\d-\d")
  def process_single_linkage(linkage: str) -> str:
    if '-' in linkage:
      first, last = linkage[0], re.search(r"-(.*)", linkage).group(1)
    else:
      first, last = linkage[0], linkage[-1]
    if '?' in first and '?' in last: return '?'
    if '?' in first: return f' {last}'
    if '?' in last:
      if ALPHA_PATTERN.match(linkage): return '\u03B1'
      if BETA_PATTERN.match(linkage): return '\u03B2'
      return '-'
    if ALPHA_PATTERN.match(linkage): return f'\u03B1 {last}'
    if BETA_PATTERN.match(linkage): return f'\u03B2 {last}'
    if DIGIT_PATTERN.match(linkage): return f'{first} - {last}'
    return '-'
  if any(isinstance(el, list) for el in linkage_list):
    return [[process_single_linkage(linkage) for linkage in sub_list] for sub_list in linkage_list]
  return [process_single_linkage(linkage) for linkage in linkage_list]


def get_highlight_attribute(
    glycan_graph: nx.DiGraph, # NetworkX glycan graph
    motif_string: str, # Motif to highlight
    termini_list: List = [], # Terminal position specifications
    reverse_highlight: bool = False # Whether to highlight everything EXCEPT highlight_motif
    ) -> nx.DiGraph: # Graph with highlight attributes
  "Labels nodes in glycan graph based on presence in specified motif"
  if motif_string:
    motif = glycan_to_nxGraph(motif_string, termini = 'provided' if termini_list else None, termini_list = termini_list)
    _, mappings = subgraph_isomorphism(glycan_graph, motif, termini_list = termini_list, return_matches = True)
    if reverse_highlight:
      mapping_show = {node: 'hide' if node in set(unwrap(mappings)) else 'show' for node in glycan_graph.nodes()}
    else:
      mapping_show = {node: 'show' if node in set(unwrap(mappings)) else 'hide' for node in glycan_graph.nodes()}
  else:
    mapping_show = {node: 'show' for node in glycan_graph.nodes()}
  nx.set_node_attributes(glycan_graph, dict(sorted(mapping_show.items())), 'highlight_labels')
  return glycan_graph


def process_repeat(
    repeat: str # Repeating unit specification, terminating either in linkage connecting units or first monosaccharide of the unit
    ) -> str: # Processed repeat unit
  "Formats repeating unit glycan sequence for drawing"
  backbone = re.findall(r'.*\((?!.*\()', repeat)[0]
  repeat_connection = re.sub(r'\)(.*)', '', re.sub(r'.*\((?!.*\()', '', repeat))
  return f'blank(?1-{repeat_connection[-1]}){backbone}{repeat_connection[:2]}-?)'


def get_branches_from_graph(graph, main_chain, main_chain_sugars):
  """Extract branch structures based on paths to lowest node index tips"""
  main_chain_set = set(main_chain)
  all_nodes = main_chain_set.copy()
  main_chain_sugars = sorted(main_chain_sugars, reverse = True)

  # Find branch by always following lowest node index
  def follow_lowest_index_path(start_node):
    path = [start_node]
    current = start_node
    while True:
      successors = [n for n in graph.successors(current) if n not in all_nodes]
      if not successors: break
      # Always take the lowest node index
      next_node = min(successors)
      path.append(next_node)
      current = next_node
    # Extract sugars and bonds
    sugar_nodes = sorted([n for n in path if n % 2 == 0], reverse = True)
    bond_nodes = sorted([n for n in path if n % 2 == 1], reverse = True)
    return path, sugar_nodes, bond_nodes

  def add_branch(branch_list, start_node, connection):
    path, sugar_nodes, bond_nodes = follow_lowest_index_path(start_node)
    branch_list.append({
      'nodes': path, 'sugar_nodes': sugar_nodes,
      'bond_nodes': bond_nodes, 'connection': connection
    })
    all_nodes.update(path)

  def process_level(parent_level, parent_idx, connect_idx_getter):
    level = []
    for i, branch in enumerate(parent_level):
      for j, node in enumerate(branch['sugar_nodes']):
        if graph.out_degree(node) > 1:
          for succ in sorted(graph.successors(node)):
            if succ not in all_nodes:
              add_branch(level, succ, (i, j))
    return level

  first_level = []
  for node in main_chain:
    if node % 2 == 0 and graph.out_degree(node) > 1:  # Sugar with branches
      for succ in sorted(graph.successors(node)):
        if succ not in all_nodes:
          add_branch(first_level, succ, (0, main_chain_sugars.index(node)))
  second_level = process_level(first_level, 0, lambda i: i)
  third_level = process_level(second_level, 1, lambda i: i)
  return first_level, second_level, third_level


def get_coordinates_and_labels(
    draw_this: str, # IUPAC-condensed glycan sequence
    highlight_motif: Optional[str], # Motif to highlight
    show_linkage: bool = True, # Show linkage labels
    termini_list: List = [], # Terminal position specifications (from 'terminal', 'internal', and 'flexible')
    reverse_highlight: bool = False # Whether to highlight everything EXCEPT highlight_motif
) -> List[List]: # Drawing coordinates and labels (monosaccharide label, x position, y position, modification, bond, conformation)
  "Calculates drawing coordinates and formats labels for glycan visualization"
  graph = glycan_to_nxGraph(draw_this, termini = 'calc' if termini_list else 'ignore')
  graph = get_highlight_attribute(graph, highlight_motif, termini_list = termini_list, reverse_highlight = reverse_highlight)
  node_values = list(nx.get_node_attributes(graph, 'string_labels').values())
  highlight_values = list(nx.get_node_attributes(graph, 'highlight_labels').values())
  root = max(graph.nodes())
  leaves = [n for n in graph.nodes() if graph.out_degree(n) == 0 and n != root] if len(graph) > 1 else [0]
  main_chain = nx.shortest_path(graph.reverse(), leaves[0], root) if leaves else []
  main_label_sugar = [node for node in main_chain if node % 2 == 0]
  main_sugar = [node_values[node] for node in main_chain if node % 2 == 0]  # Even indices are sugars
  main_sugar = [get_core(node) if node not in domon_costello else node for node in main_sugar][::-1]
  main_sugar_modification = [get_modification(node_values[node]).replace('O', '').replace('-ol', '') for node in main_chain if node % 2 == 0][::-1]
  main_bond = [node_values[node] for node in main_chain if node % 2 == 1][::-1]  # Odd indices are bonds
  main_sugar_highlight = [highlight_values[node] for node in main_chain if node % 2 == 0][::-1]
  main_bond_highlight = [highlight_values[node] for node in main_chain if node % 2 == 1][::-1]
  main_sugar_x_pos = list(range(len(main_sugar)))

  branch_level1, branch_level2, branch_level3 = get_branches_from_graph(graph, main_chain, main_label_sugar)

  def process_branch_data(branches):
    sugar, sugar_mod, bond, connection, sugar_label, bond_label = [], [], [], [], [], []
    for branch in branches:
      # Extract sugar and bond labels
      sugar_nodes = branch['sugar_nodes']
      bond_nodes = [m for m in branch['nodes'] if m % 2 == 1]
      sugar.append([get_core(node_values[n]) if node_values[n] not in domon_costello else node_values[n] for n in sugar_nodes])
      sugar_mod.append([get_modification(node_values[n]).replace('O', '') for n in sugar_nodes])
      bond.append([node_values[n] for n in bond_nodes])
      connection.append(branch['connection'])
      sugar_label.append([highlight_values[n] for n in sugar_nodes])
      bond_label.append([highlight_values[n] for n in bond_nodes])
    return sugar, sugar_mod, bond, connection, sugar_label, bond_label

  # Get branch data for all levels
  l1_sugar, l1_sugar_modification, l1_bond, l1_connection, l1_sugar_label, l1_bond_label = process_branch_data(branch_level1)
  l2_sugar, l2_sugar_modification, l2_bond, l2_connection, l2_sugar_label, l2_bond_label = process_branch_data(branch_level2)
  l3_sugar, l3_sugar_modification, l3_bond, l3_connection, l3_sugar_label, l3_bond_label = process_branch_data(branch_level3)

  # Process linkages
  main_bond = process_bonds(main_bond)
  l1_bond = process_bonds(l1_bond)
  l2_bond = process_bonds(l2_bond)
  l3_bond = process_bonds(l3_bond)

  # Main chain x
  if (main_sugar[-1]  == 'Fuc' and len(main_bond) > 1) or (main_sugar[-1] == 'Xyl' and len(main_bond) > 1 and main_bond[-1] == 'β 2'):
    main_sugar_x_pos[-1] -= 1

  # Calculate x positions for branches
  def calculate_x_positions(sugars, connections, parent_x_positions, level = 1):
    x_positions = []
    for j, branch in enumerate(sugars):
      if level == 1:
        start_x = parent_x_positions[connections[j][1]]
      else:
        parent_branch, parent_idx = connections[j]
        start_x = parent_x_positions[parent_branch][parent_idx]
      tmp = [start_x + 1 + k for k in range(len(branch))]
      if branch[-1] in {'Fuc', 'Xyl'}:
        tmp[-1] -= 1
      x_positions.append(tmp)
    return x_positions

  # Calculate x positions for all branch levels
  l1_x_pos = calculate_x_positions(l1_sugar, l1_connection, main_sugar_x_pos)
  l2_x_pos = calculate_x_positions(l2_sugar, l2_connection, l1_x_pos, level = 2)
  l3_x_pos = calculate_x_positions(l3_sugar, l3_connection, l2_x_pos, level = 3)

  # Initialize y positions - ALL START AT Y=0 (except Fuc)
  main_sugar_y_pos = [2 if s == "Fuc" and i == len(main_sugar)-1 and draw_this.count('(') > 1 else 0 for i, s in enumerate(main_sugar)]
  l1_y_pos = [[2 if s == "Fuc" else 0 for s in sugars] for sugars in l1_sugar]
  l2_y_pos = [[2 if s == "Fuc" else 0 for s in sugars] for sugars in l2_sugar]
  l3_y_pos = [[2 if s == "Fuc" else 0 for s in sugars] for sugars in l3_sugar]

  SPACING = 1
  # Main chain goes down, branches go up
  branch_points = {conn[1] for conn in l1_connection}
  # For each branch point, main chain beyond it goes down
  for parent_idx in sorted(branch_points):
    # Core fucose special case (don't push down main chain)
    branch_indices = [j for j, conn in enumerate(l1_connection) if conn[1] == parent_idx]
    core_branches = [l1_sugar[j] for j in branch_indices]
    is_core_fuc = all(j in [['Fuc'], ['Xyl']] for j in core_branches)
    is_fuc_partner = main_sugar[parent_idx + 1] == 'Fuc'
    if not is_core_fuc and not is_fuc_partner:
      branch_sugar = max(core_branches, key = len)
      l2_connected_branches = [l2_sugar[k] for k, conn in enumerate(l2_connection) if conn[0] in branch_indices]
      l1_main_chain_branches = [l1_sugar[k] for k, conn in enumerate(l1_connection) if conn[1] > parent_idx and l1_sugar[k] not in [['Fuc'], ['Gal']]]
      has_fuc = ('Fuc' in branch_sugar) or ('Fuc' in unwrap(l2_connected_branches))
      has_own_fuc = 'Fuc' in unwrap(l1_main_chain_branches)
      has_bisecting = any('GlcNAc' in b[0] for b in core_branches) and main_sugar[parent_idx] == 'Man' and main_bond[parent_idx-1] == 'β 4'
      has_triple_branch = len(branch_indices) == 2 and not 'Xyl' in unwrap(core_branches)
      l2_connected_branches = [k for k in l2_connected_branches if k not in [['Fuc'], ['Gal']]]
      max_l2_len = max((len(b) for b in l2_connected_branches), default = 0)
      max_l1_len = max((len(b) for b in l1_main_chain_branches if b not in [['Fuc'], ['Gal']]), default = 0)
      spacing_spec = SPACING + has_fuc*SPACING + has_own_fuc*SPACING + (max_l1_len > 0)*0.25*SPACING
      if (max_l2_len > 0) and (spacing_spec < 2.5 or (has_fuc and has_own_fuc)):
        spacing_spec += 0.5*SPACING
      if (has_bisecting or has_triple_branch) and spacing_spec < 1.1:
        spacing_spec += SPACING
      # Push main chain down after branch point
      for i in range(parent_idx + 1, len(main_sugar)):
        main_sugar_y_pos[i] += spacing_spec

  # All branches go up
  for j, conn in enumerate(l1_connection):
    parent_idx = conn[1]
    branch_sugar = l1_sugar[j]
    # Special case for core fucose
    if parent_idx == 0 and branch_sugar == ['Fuc'] and l1_bond[j] == ['α 6']:
      # Core fucose goes up
      l1_y_pos[j] = [-2*SPACING] * len(branch_sugar)
    else:
      is_bisecting = branch_sugar[0] in ['GlcNAc'] and main_sugar[parent_idx] == 'Man' and main_bond[parent_idx-1] == 'β 4'
      is_leading_xyl = main_sugar[-1] == 'Xyl'
      is_fuc_partner = main_sugar[parent_idx+1] == 'Fuc'
      parent_branches = [(k, c) for k, c in enumerate(l1_connection) if c[1] == parent_idx]  # + 1 from main chain
      is_triple_branch = len(parent_branches) == 2 and j == parent_branches[0][0]
      # All other branches go up by spacing amount
      if len(branch_sugar) == 1 and branch_sugar[0] in ['Fuc', 'Xyl']:
        l1_y_pos[j][0] = main_sugar_y_pos[parent_idx] + 2*SPACING
      elif is_leading_xyl and j == 0:
        l1_y_pos[j][0] = main_sugar_y_pos[parent_idx] + SPACING
      elif len(branch_sugar) == 1 and (is_bisecting or is_fuc_partner or is_triple_branch):
        l1_y_pos[j][0] = main_sugar_y_pos[parent_idx]
      elif len(branch_sugar) == 1:
        l1_y_pos[j][0] = main_sugar_y_pos[parent_idx] - SPACING
      else:
        offset = main_sugar_y_pos[parent_idx + 1] - main_sugar_y_pos[parent_idx]
        shift_amount = main_sugar_y_pos[parent_idx] - offset
        l1_y_pos[j] = [p + shift_amount for p in l1_y_pos[j]]

  def process_branch_level(level_sugar, level_y_pos, level_connection, next_level_sugar, next_level_connection, parent_level_y_pos, parent_level_sugar):
    # At each branch point, push remaining sugars down
    for idx, (parent_branch, parent_idx) in enumerate(level_connection):
      is_fuc_partner = parent_level_sugar[parent_branch][parent_idx+1] == 'Fuc' if parent_idx+1 < len(parent_level_sugar[parent_branch]) else False
      if not is_fuc_partner and level_sugar[idx][0] != 'Fuc':
        branch_indices = [j for j, conn in enumerate(level_connection) if conn[0] == parent_branch and conn[1] == parent_idx]
        core_branches = [level_sugar[j] for j in branch_indices]
        branch_sugar = max(core_branches, key = len)
        has_next_level = len(next_level_sugar) > 0
        next_level_connected_branches = [next_level_sugar[k] for k, conn in enumerate(next_level_connection) if conn[0] in branch_indices] if has_next_level else []
        has_fuc = ('Fuc' in branch_sugar) or ('Fuc' in unwrap(next_level_connected_branches) if has_next_level else False)
        spacing_spec = SPACING + has_fuc*SPACING
        for i in range(parent_idx + 1, len(parent_level_y_pos[parent_branch])):
          parent_level_y_pos[parent_branch][i] += spacing_spec
    # Branches go up
    for j, (parent_branch, parent_idx) in enumerate(level_connection):
      parent_y = parent_level_y_pos[parent_branch][parent_idx]
      is_fuc_partner = parent_level_sugar[parent_branch][parent_idx+1] == 'Fuc' if parent_idx+1 < len(parent_level_sugar[parent_branch]) else False
      if len(level_sugar[j]) == 1 and level_sugar[j][0] in ['Fuc', 'Xyl']:
        level_y_pos[j][0] = parent_y + 2*SPACING
      elif len(level_sugar[j]) == 1 and is_fuc_partner:
        level_y_pos[j][0] = parent_y
      else:
        offset = parent_level_y_pos[parent_branch][parent_idx+1] - parent_y if parent_idx+1 < len(parent_level_y_pos[parent_branch]) else 0
        shift_amount = parent_y - offset
        level_y_pos[j] = [p + shift_amount for p in level_y_pos[j]]
    return level_y_pos

  l2_y_pos = process_branch_level(l2_sugar, l2_y_pos, l2_connection, l3_sugar, l3_connection, l1_y_pos, l1_sugar)
  l3_y_pos = process_branch_level(l3_sugar, l3_y_pos, l3_connection, [], [], l2_y_pos, l2_sugar)

  def extract_conformation(sugar_modifications):
    conf_pattern = r'^L-|^D-|(\d,\d+lactone)'
    if sugar_modifications and isinstance(sugar_modifications[0], list):
      return [[k.group() if k is not None else '' for k in j] for j in [[re.search(conf_pattern, k) for k in j] for j in sugar_modifications]], \
             [[re.sub(conf_pattern, '', k) for k in j] for j in sugar_modifications]
    else:
      return [k.group() if k is not None else '' for k in [re.search(conf_pattern, k) for k in sugar_modifications]], \
             [re.sub(conf_pattern, '', k) for k in sugar_modifications]

  main_conf, main_sugar_modification = extract_conformation(main_sugar_modification)
  l1_conf, l1_sugar_modification = extract_conformation(l1_sugar_modification)
  l2_conf, l2_sugar_modification = extract_conformation(l2_sugar_modification)
  l3_conf, l3_sugar_modification = extract_conformation(l3_sugar_modification)

  data_combined = [
      [main_sugar, main_sugar_x_pos, main_sugar_y_pos, main_sugar_modification, main_bond, main_conf, main_sugar_highlight, main_bond_highlight],
      [l1_sugar, l1_x_pos, l1_y_pos, l1_sugar_modification, l1_bond, l1_connection, l1_conf, l1_sugar_label, l1_bond_label],
      [l2_sugar, l2_x_pos, l2_y_pos, l2_sugar_modification, l2_bond, l2_connection, l2_conf, l2_sugar_label, l2_bond_label],
      [l3_sugar, l3_x_pos, l3_y_pos, l3_sugar_modification, l3_bond, l3_connection, l3_conf, l3_sugar_label, l3_bond_label]
  ]
  return data_combined


def draw_bracket(
    x: float, # X coordinate
    y_min_max: List[float], # [Min Y, Max Y] coordinates
    drawing: draw.Drawing, # Glycan drawing to be modified
    direction: str = 'right', # Bracket direction ("left", "right")
    dim: float = 50, # Base dimension for scaling
    highlight: str = 'show', # Highlight state
    deg: float = 0 # Rotation angle in degrees
    ) -> None:
  "Draws bracket shape at specified position and dimensions"
  col_dict = col_dict_transparent if highlight == 'hide' else col_dict_base
  stroke_opts = {'stroke_width': 0.04 * dim, 'stroke': col_dict['black']}
  x_common = -x * dim
  y_min = y_min_max[0] * dim - 0.75 * dim
  y_max = y_min_max[1] * dim + 0.75 * dim
  # Vertical
  g = draw.Group(transform = f'rotate({deg} {x_common} {np.mean(y_min_max)})')
  p = draw.Path(**stroke_opts)
  p.M(x_common, y_max)
  p.L(x_common, y_min)
  g.append(p)
  offset = 0.25 * dim * (1 if direction == 'right' else -1)
  for y in [y_min, y_max]:
    p = draw.Path(**stroke_opts)
    p.M(x_common - offset/12.5, y)
    p.L(x_common + offset, y)
    g.append(p)
  drawing.append(g)


def is_jupyter() -> bool:
  "Detects if code is running in Jupyter notebook environment"
  try:
    from IPython import get_ipython
    return 'IPKernelApp' in get_ipython().config  # Check if in IPython kernel
  except AttributeError:
    return False


def display_svg_with_matplotlib(
    svg_data: Any, # SVG drawing object
    chem: bool = False # Whether svg_data comes from RDKit chemical
    ) -> None:
  "Renders SVG using matplotlib for non-Jupyter environments"
  from PIL import Image
  svg_data = svg_data if isinstance(svg_data, str) else svg_data.as_svg()
  # Get original SVG dimensions and scale them up
  size_multiplier = 4  # Make everything 4x bigger
  width = svg_data.width if hasattr(svg_data, 'width') else 800
  height = svg_data.height if hasattr(svg_data, 'height') else 800
  # Convert to PNG with larger dimensions
  png_output = convert_svg_to_png(svg_data, output_width = width * size_multiplier,
                                  output_height = height * size_multiplier, scale = 2.0, return_bytes = True,
                                  chem = chem)
  # Use PIL to crop aggressively
  img = Image.open(BytesIO(png_output))
  bbox = img.convert('RGBA').getbbox()
  if bbox:
    # Add minimal padding - just enough to not cut off edges
    padding = int(10 * size_multiplier)
    bbox = (max(0, bbox[0] - padding), max(0, bbox[1] - padding),
            min(img.width, bbox[2] + padding), min(img.height, bbox[3] + padding))
  img_cropped = img.crop(bbox).convert('RGBA')
  # Display with appropriate figure size
  dpi = plt.rcParams['figure.dpi']
  figsize = (img_cropped.width / dpi, img_cropped.height / dpi)
  plt.figure(figsize = figsize)
  plt.imshow(img_cropped)
  plt.axis('off')
  plt.show()


def process_per_residue(
    draw_this: str, # reordered IUPAC-condensed glycan sequence
    per_residue: List[float], # Scalar values per residue
    glycan: str, # original IUPAC-condensed glycan sequence
    ) -> Tuple[List[float], List[List[float]], List[List[float]]]: # (main chain values, side chain values, branched side chain values)
  "Maps per-residue scalar values to main chain, side chains, and branched side chains"
  if glycan != draw_this:
    g1 = glycan_to_nxGraph(glycan)
    g2 = glycan_to_nxGraph(draw_this)
    _, mappy = compare_glycans(g2, g1, return_matches = True)
    per_residue = [per_residue[mappy[i*2]//2] for i in range(len(per_residue))]
  temp = re.sub(r'\([^)]*\)', 'x', draw_this) + 'x'
  temp = re.sub(r'[^x\[\]]', '', temp)
  main_chain_indices, l1_indices = [], []
  l2_indices, l1_stack = [], []
  idx = 0
  for char in temp:
    if char == '[':
      l1_stack.append([])
    elif char == ']':
      if len(l1_stack) == 1:
        l1_indices.append(l1_stack.pop())
      else:
        l2_indices.append(l1_stack.pop())
    elif char == 'x':
      if l1_stack:
        l1_stack[-1].append(per_residue[idx])
      else:
        main_chain_indices.append(per_residue[idx])
      idx += 1
  l1_indices = [k[::-1] for k in l1_indices if k]
  l2_indices = [k[::-1] for k in l2_indices if k]
  return main_chain_indices[::-1], l1_indices, l2_indices


def process_per_linkage(
    draw_this: str, # reordered IUPAC-condensed glycan sequence
    highlight_linkages: List[int], # Which linkages to highlight
    glycan: str, # original IUPAC-condensed glycan sequence
    ) -> Tuple[List[bool], List[List[bool]], List[List[bool]]]: # (main chain values, side chain values, branched side chain values)
  "Maps which linkages to highlight to main chain, side chains, and branched side chains"
  per_linkage = [i in highlight_linkages for i in range(glycan.count('('))]
  if glycan != draw_this:
    g1 = glycan_to_nxGraph(glycan)
    g2 = glycan_to_nxGraph(draw_this)
    _, mappy = compare_glycans(g2, g1, return_matches = True)
    per_linkage = [per_linkage[mappy[i*2]//2] for i in range(len(per_linkage))]
  temp = re.sub(r'\([^)]*\)', 'x', draw_this) + 'x'
  temp = re.sub(r'[^x\[\]]', '', temp)
  main_chain_indices, l1_indices = [], []
  l2_indices, l1_stack = [], []
  idx = 0
  for char in temp[:-1]:
    if char == '[':
      l1_stack.append([])
    elif char == ']':
      if len(l1_stack) == 1:
        l1_indices.append(l1_stack.pop())
      else:
        l2_indices.append(l1_stack.pop())
    elif char == 'x':
      if l1_stack:
        l1_stack[-1].append(per_linkage[idx])
      else:
        main_chain_indices.append(per_linkage[idx])
      idx += 1
  l1_indices = [k[::-1] for k in l1_indices if k]
  l2_indices = [k[::-1] for k in l2_indices if k]
  return main_chain_indices[::-1], l1_indices, l2_indices


mono_list = ['Glc', 'GlcNAc', 'GlcA', 'Man', 'ManNAc', 'Gal', 'GalNAc', 'Gul', 'GulNAc',
                 'Alt', 'AltNAc', 'All', 'AllNAc', 'Neu5Ac', 'Tal', 'TalNAc', 'Neu5Gc', 'Ido', 'IdoNAc', 'IdoA', 'Fuc']

chem_cols = ['#CDE7EF', '#CDE7EF', '#CDE7EF',     # blue
        '#CDE9DF', '#CDE9DF',                # green
        '#FFF6DE', '#FFF6DE',                # yellow
        '#FDE7E0', '#FDE7E0',                # orange
        '#FDF0F1', '#FDF0F1',                # pink
        '#F1E6ED', '#F1E6ED', '#F1E6ED',     # purple
        '#EEF8FB', '#EEF8FB', '#EEF8FB',     # light blue
        '#F1E9E5', '#F1E9E5', '#F1E9E5',     # brown
        '#F7E0E0', '#F7E0E0']                # red

chem_cols_alpha = ['#0385AE', '#0385AE', '#0385AE',     # blue
        '#058F60', '#058F60',                # green
        '#FCC326', '#FCC326',                # yellow
        '#EF6130', '#EF6130',                # orange
        '#F39EA0', '#F39EA0',                # pink
        '#A15989', '#A15989', '#A15989',     # purple
        '#91D3E3', '#91D3E3', '#91D3E3',     # light blue
        '#9F6D55', '#9F6D55', '#9F6D55',     # brown
        '#C23537']                           # red


def get_hit_atoms_and_bonds(
    mol: Any, # RDKit molecule object
    smt: str # SMARTS pattern string
    ) -> Tuple[List[int], List[int]]: # (matching atom indices, matching bond indices)
  "Identifies atoms and bonds matching SMARTS pattern in molecule"
  # Adapted from https://github.com/rdkit/rdkit/blob/master/Docs/Book/data/test_multi_colours.py
  try:
    from rdkit.Chem import MolFromSmarts
  except ImportError:
    raise ImportError("You must install the 'chem' dependencies to use this feature. Try 'pip install glycowork[chem]'.")
  bonds = []
  q = MolFromSmarts(smt)
  atoms = [atom for match in mol.GetSubstructMatches(q, useChirality = True) for atom in match]
  for ha1 in atoms:
    for ha2 in atoms:
      if ha1 > ha2:
        b = mol.GetBondBetweenAtoms(ha1, ha2)
        if b:
          bonds.append(b.GetIdx())
  return atoms, bonds


def add_colours_to_map(
    els: List[int], # Element indices
    cols: Dict[int, List], # Color map dictionary
    col_num: int, # Color index
    alpha: bool = True, # Use alpha-adjusted colors
    hex_codes: bool = True # Return hex color codes
    ) -> None:
  "Adds color assignments to mapping dictionary for chemical structure visualization"
  from matplotlib.colors import ColorConverter
  color = chem_cols_alpha[col_num] if alpha else chem_cols[col_num]
  color = color if hex_codes else ColorConverter().to_rgb(color)
  for el in els:
    cols.setdefault(el, [])
    if color not in cols[el]: cols[el].append(color)


def draw_chem2d(
    draw_this: str, # IUPAC-condensed glycan sequence
    mono_list: List[str], # List of monosaccharides to highlight
    filepath: Optional[Union[str, Path]] = None # Output file path
    ) -> Any: # IPython SVG display object
  "Creates 2D chemical structure drawing with highlighted monosaccharides using RDKit"
  # Adapted from https://github.com/rdkit/rdkit/blob/master/Docs/Book/data/test_multi_colours.py
  try:
    from glycowork.motif.processing import IUPAC_to_SMILES
    from rdkit.Chem import MolFromSmiles
    from rdkit.Chem.Draw import PrepareMolForDrawing
    from rdkit.Chem.Draw.rdMolDraw2D import MolDraw2DSVG
    from IPython.display import SVG
  except ImportError:
    raise ImportError("You must install the 'chem' dependencies to use this feature. Try 'pip install glycowork[chem]'.")

  mol = MolFromSmiles(IUPAC_to_SMILES([draw_this])[0])
  mol = PrepareMolForDrawing(mol)

  atom_colors, bond_colors = {}, {}
  for i, smarts in enumerate(IUPAC_to_SMILES(mono_list)):
    atoms, bonds = get_hit_atoms_and_bonds(mol, smarts)
    add_colours_to_map(atoms, atom_colors, i, hex_codes = False)
    add_colours_to_map(bonds, bond_colors, i, hex_codes = False)
  atom_colors = {k: v for k, v in atom_colors.items() if len(v) == 1}
  bond_colors = {k: [v[0]] for k, v in bond_colors.items() if len(v) == 1}

  d = MolDraw2DSVG(250, 250)
  d.drawOptions().fillHighlights = True
  d.drawOptions().useBWAtomPalette()
  d.drawOptions().rotate = 180
  d.DrawMoleculeWithHighlights(mol, '', atom_colors, bond_colors, {}, {}, -1)
  d.FinishDrawing()
  svg_data = d.GetDrawingText()

  if filepath:
    filepath = Path(filepath)
    filepath = filepath.with_name(filepath.name.replace('?', '_'))
    if filepath.suffix.lower() == '.svg':
        with open(filepath, 'w') as f:
          f.write(svg_data)
    elif filepath.suffix.lower() == '.pdf':
      convert_svg_to_pdf(svg_data, str(filepath), chem = True)
  return SVG(svg_data) if is_jupyter() else display_svg_with_matplotlib(svg_data, chem = True)


def draw_chem3d(
    draw_this: str, # IUPAC-condensed glycan sequence
    mono_list: List[str], # List of monosaccharides to highlight
    filepath: Optional[Union[str, Path]] = None, # Output file path for PDB
    pdb_file: Optional[Union[str, Path]] = None  # already existing glycan structure
    ) -> None:
  "Generates 3D chemical structure model with highlighted monosaccharides using RDKit and py3Dmol"
  # Adapted from https://github.com/rdkit/rdkit/blob/master/Docs/Book/data/test_multi_colours.py and https://github.com/rdkit/rdkit/blob/master/Docs/Book/GettingStartedInPython.rst
  try:
    from glycowork.motif.processing import IUPAC_to_SMILES
    from rdkit.Chem import MolFromSmiles, AddHs, RemoveHs, MolToPDBFile, MolFromPDBFile
    from rdkit.Chem.AllChem import EmbedMolecule, MMFFOptimizeMolecule
    if is_jupyter():
      from rdkit.Chem.Draw import IPythonConsole
      import py3Dmol
    else:
      from rdkit.Chem.Draw import rdDepictor
      from rdkit.Chem.Draw.rdMolDraw2D import MolDraw2DSVG
  except ImportError:
    raise ImportError("You must install the 'chem' dependencies to use this feature. Try 'pip install glycowork[chem]'.")

  mol = MolFromSmiles(IUPAC_to_SMILES([draw_this])[0])
  atom_colors, bond_colors = {}, {}
  for i, smarts in enumerate(IUPAC_to_SMILES(mono_list)):
    atoms, bonds = get_hit_atoms_and_bonds(mol, smarts)
    add_colours_to_map(atoms, atom_colors, i, alpha = False)
    add_colours_to_map(bonds, bond_colors, i, alpha = False)
  atom_colors = {k: ['#ECECEC'] if len(v) > 1 else v for k, v in atom_colors.items()}

  if pdb_file:
    mol = MolFromPDBFile(str(pdb_file))
  else:
    mol = AddHs(mol)
    EmbedMolecule(mol)
    MMFFOptimizeMolecule(mol)
    mol = RemoveHs(mol)
    print("Disclaimer: The conformer generated using RDKit and MMFFOptimizeMolecule is not intended to be a replacement for a 'real' conformer analysis tool.")

  if filepath:
    filepath = Path(filepath)
    if filepath.suffix.lower() == '.pdb':
      MolToPDBFile(mol, filepath)
    else:
      print("3D structure can only be saved as .pdb file.")

  if is_jupyter():
    v = py3Dmol.view(width = 500, height = 300)
    v.removeAllModels()
    IPythonConsole.addMolToView(mol, v)
    for atom_idx, colors in atom_colors.items():
      v.setStyle({'serial': atom_idx}, {'stick': {'color': colors[0]}})
    v.zoomTo()
    v.show()
  else:
    rdDepictor.Compute2DCoords(mol, clearConfs = False)
    drawer = MolDraw2DSVG(500, 500)
    drawer.drawOptions().addStereoAnnotation = True
    drawer.drawOptions().addAtomIndices = False
    drawer.drawOptions().bondLineWidth = 2
    drawer.DrawMolecule(mol, highlightAtoms = list(atom_colors.keys()),
                       highlightAtomColors = {k: tuple(int(v[0].lstrip('#')[i:i+2], 16)/255
                                          for i in (0, 2, 4)) for k, v in atom_colors.items()})
    drawer.FinishDrawing()
    display_svg_with_matplotlib(drawer.GetDrawingText(), chem = True)


@rescue_glycans
def GlycoDraw(
    glycan: str, # IUPAC-condensed glycan sequence
    vertical: bool = False, # Draw vertically
    compact: bool = False, # Use compact style
    show_linkage: bool = True, # Show linkage labels
    dim: float = 50, # Base dimension for scaling
    highlight_motif: Optional[str] = None, # Motif to highlight
    highlight_termini_list: List = [], # Terminal positions (from 'terminal', 'internal', and 'flexible')
    highlight_linkages: Optional[List[int]] = None, # Which linkages to highlight in a different color; indices, starting from 0, in glycan
    reverse_highlight: bool = False, # Whether to highlight everything EXCEPT highlight_motif
    repeat: Optional[Union[bool, int, str]] = None, # Repeat unit specification (True: n units, int: # of units, str: range of units)
    repeat_range: Optional[List[int]] = None, # Repeat unit range
    draw_method: Optional[str] = None, # Drawing method: None, 'chem2d', 'chem3d'
    filepath: Optional[Union[str, Path]] = None, # Output file path
    suppress: bool = False, # Suppress display
    per_residue: List = [], # Per-residue intensity values (order should be the same as the monosaccharides in glycan string)
    pdb_file: Optional[Union[str, Path]] = None,  # only used when draw_method='chem3d'; already existing glycan structure
    alt_text: Optional[str] = None,  # Custom ALT text for accessibility
    libr: dict = None,  # Can be modified for drawing too exotic monosaccharides
    reducing_end_label: Optional[str] = None  # Label to be drawn connected to the reducing end
    ) -> Any: # Drawing object
  "Renders glycan structure using SNFG symbols or chemical structure representation"
  if any(k in glycan for k in [';', 'β', 'α', 'RES', '=']):
    raise Exception
  if libr is None:
    libr = lib
  if glycan.startswith('Terminal') and glycan not in motif_list.motif_name.values.tolist():
    glycan = glycan.split('_')[-1]
  if glycan in motif_list.motif_name.values.tolist():
    glycan = motif_list.loc[motif_list.motif_name == glycan].motif.values[0]
  if repeat and not repeat_range:
    glycan = process_repeat(glycan)
  if glycan.endswith(')'):
    glycan += 'blank'
  draw_this = graph_to_string(glycan_to_nxGraph(glycan), order_by = "linkage") if not glycan.startswith('[') else glycan
  if per_residue:
    main_per_residue, side_per_residue, branched_side_per_residue = process_per_residue(draw_this, per_residue, glycan)
  if highlight_linkages:
    main_per_linkage, side_per_linkage, branched_side_per_linkage = process_per_linkage(draw_this, highlight_linkages, glycan)
  if compact:
    show_linkage = False
  if isinstance(highlight_motif, str) and highlight_motif.startswith('r'):
    temp = get_match(highlight_motif[1:], draw_this)
    highlight_motif = temp[0] if temp else None

  # toggle SNFG vs 2D/3D chem
  if draw_method:
    if draw_method == 'chem2d':
      return draw_chem2d(draw_this = draw_this, mono_list = mono_list, filepath = filepath)
    elif draw_method == 'chem3d':
      return draw_chem3d(draw_this = draw_this, mono_list = mono_list, filepath = filepath, pdb_file = pdb_file)
    else:
      raise ValueError('Method not supported. Please choose between "chem2d" and "chem3d".')

  # Handle floaty bits if present
  floaty_bits = []
  for openpos, closepos, _ in get_matching_indices(draw_this, opendelim = '{', closedelim = '}'):
    floaty_bits.append(f"{draw_this[openpos:closepos]}blank")
    draw_this = draw_this[:openpos-1] + len(draw_this[openpos-1:closepos+1])*'*' + draw_this[closepos+1:]
  draw_this = draw_this.replace('*', '')

  if not in_lib(draw_this, expand_lib(libr, list(sugar_dict.keys()) + [k for k in min_process_glycans([draw_this])[0] if '/' in k])): # support for super-narrow wildcard linkages
    raise Exception('Did you enter a real glycan or motif?')

  data = get_coordinates_and_labels(draw_this, show_linkage = show_linkage, highlight_motif = highlight_motif, termini_list = highlight_termini_list, reverse_highlight  = reverse_highlight)

  main_sugar, main_sugar_x_pos, main_sugar_y_pos, main_sugar_modification, main_bond, main_conf, main_sugar_label, main_bond_label = data[0]
  l1_sugar, l1_x_pos, l1_y_pos, l1_sugar_modification, l1_bond, l1_connection, l1_conf, l1_sugar_label, l1_bond_label = data[1]
  l2_sugar, l2_x_pos, l2_y_pos, l2_sugar_modification, l2_bond, l2_connection, l2_conf, l2_sugar_label, l2_bond_label = data[2]
  l3_sugar, l3_x_pos, l3_y_pos, l3_sugar_modification, l3_bond, l3_connection, l3_conf, l3_sugar_label, l3_bond_label = data[3]

  if not show_linkage:
    main_bond = ['-'] * len(main_bond)
    l1_bond = [['-' for _ in y] for y in l1_bond]
    l2_bond = [['-' for _ in y] for y in l2_bond]
    l3_bond = [['-' for _ in y] for y in l3_bond]

  # Calculate angles for main chain Y, Z fragments
  def calculate_degree(y1, y2, x1, x2):
    slope = -1 * (y2 - y1) / ((x2 * 2) - (x1 * 2))
    return degrees(atan(slope))

  main_deg = [calculate_degree(main_sugar_y_pos[k], main_sugar_y_pos[k-1], main_sugar_x_pos[k], main_sugar_x_pos[k-1])
              if sugar in {'Z', 'Y'} else 0 for k, sugar in enumerate(main_sugar)]

  # Calculate angles for branch Y, Z fragments
  l1_deg = []
  for k, sugars in enumerate(l1_sugar):
    l1_deg.append([
      calculate_degree(l1_y_pos[k][j], main_sugar_y_pos[l1_connection[k][1]], l1_x_pos[k][j], main_sugar_x_pos[l1_connection[k][1]])
      if sugar in {'Z', 'Y'} and len(sugars) == 1 else
      calculate_degree(l1_y_pos[k][j], l1_y_pos[k][j-1], l1_x_pos[k][j], l1_x_pos[k][j-1])
      if sugar in {'Z', 'Y'} else 0 for j, sugar in enumerate(sugars)
      ])

  # Calculate angles for branch_branch Y, Z fragments
  l2_deg = []
  for k, sugars in enumerate(l2_sugar):
    l2_deg.append([
      calculate_degree(l2_y_pos[k][j], l1_y_pos[l2_connection[k][0]][l2_connection[k][1]], l2_x_pos[k][j], l1_x_pos[l2_connection[k][0]][l2_connection[k][1]])
      if sugar in {'Z', 'Y'} and len(sugars) == 1 else
      calculate_degree(l2_y_pos[k][j], l2_y_pos[k][j-1], l2_x_pos[k][j], l2_x_pos[k][j-1])
      if sugar in {'Z', 'Y'} else 0 for j, sugar in enumerate(sugars)
      ])

  # Adjust drawing dimensions
  max_y = max(unwrap(l3_y_pos)+unwrap(l2_y_pos)+unwrap(l1_y_pos)+main_sugar_y_pos)
  min_y = min(unwrap(l3_y_pos)+unwrap(l2_y_pos)+unwrap(l1_y_pos)+main_sugar_y_pos)
  max_x = max(unwrap(l3_x_pos)+unwrap(l2_x_pos)+unwrap(l1_x_pos)+main_sugar_x_pos)
  min_x = min(unwrap(l3_x_pos)+unwrap(l2_x_pos)+unwrap(l1_x_pos)+main_sugar_x_pos)
  if reducing_end_label:
    min_x = min(min_x, main_sugar_x_pos[0] - 1)
  x_span = max_x - min_x
  y_span = max_y - min_y

  # Canvas size
  width = ((((x_span+1)*2)-1)*dim)+dim
  if floaty_bits:
    len_one_gw = ((max([len(j) for k in min_process_glycans(floaty_bits) for j in k]) / 6) + 1) * dim
    len_multiple_gw = (max([len(k) for k in min_process_glycans(floaty_bits)], default = 0)+1) * dim
    width += max(len_one_gw, len_multiple_gw)
  if len(floaty_bits) > len(set(floaty_bits)):
    width += dim
  if len(floaty_bits) > y_span:
    y_span += 1.0
    max_y += 0.5
    min_y -= 0.5
  height = ((((max(abs(min_y), max_y)+1)*2)-1)*dim)+60
  height = max(height, width) if vertical else height
  x_offset = abs(min_x) * dim * (1.2 if compact else 2) if reducing_end_label else 0
  x_ori = -width+(dim/2)+0.5*dim+x_offset
  y_ori = (-height/2)+(((max_y-abs(min_y))/2)*dim)

  # Generate default ALT text if not provided
  if alt_text is None:
    orientation = "vertical" if vertical else "horizontal"
    style = "compact" if compact else "standard"
    linkage_info = "with" if show_linkage else "without"
    alt_text = f"SNFG diagram of {glycan} drawn in {orientation} {style} style {linkage_info} linkage labels."
    if highlight_motif:
      alt_text += f" The motif {highlight_motif} is highlighted."
    if repeat:
      alt_text += f" Contains repeat unit (n={repeat if isinstance(repeat, (str, int)) and repeat != True else ''})."

  # Draw
  d2 = draw.Drawing(width, height, origin = (x_ori, y_ori))
  deg = 90 if vertical else 0
  d = draw.Group(transform = f'rotate({deg} {x_ori+0.5*width} {y_ori+0.5*height})')

  if reducing_end_label:
    bond_start_x = main_sugar_x_pos[0] - 0.5
    label_x = main_sugar_x_pos[0] - 0.55 - (len(reducing_end_label) * 0.1)
    label_y = main_sugar_y_pos[0]
    add_bond(bond_start_x, main_sugar_x_pos[0], label_y, main_sugar_y_pos[0], d, '-', dim = dim, compact = compact, highlight = main_sugar_label[0])
    col_dict = col_dict_transparent if main_sugar_label[0] == 'hide' else col_dict_base
    x_base = -label_x * dim * (1.2 if compact else 2)
    y_base = label_y * dim * (0.6 if compact else 1)
    d.append(draw.Text(reducing_end_label, dim*0.35, x_base, y_base, text_anchor = 'end', fill = col_dict['black'], dominant_baseline = 'middle'))
  # Bond main chain
  [add_bond(main_sugar_x_pos[k+1], main_sugar_x_pos[k], main_sugar_y_pos[k+1], main_sugar_y_pos[k], d, main_bond[k], dim = dim, compact = compact, highlight = main_bond_label[k], color_highlight = main_per_linkage[k] if highlight_linkages else False) for k in range(len(main_sugar)-1)]
  # Bond branch
  [add_bond(l1_x_pos[b_idx][s_idx+1], l1_x_pos[b_idx][s_idx], l1_y_pos[b_idx][s_idx+1], l1_y_pos[b_idx][s_idx], d, l1_bond[b_idx][s_idx+1], dim = dim, compact = compact, highlight = l1_bond_label[b_idx][s_idx+1], color_highlight = side_per_linkage[b_idx][s_idx+1] if highlight_linkages else False) for b_idx in range(len(l1_sugar)) for s_idx in range(len(l1_sugar[b_idx])-1) if len(l1_sugar[b_idx]) > 1]
  # Bond branch to main chain
  [add_bond(l1_x_pos[k][0], main_sugar_x_pos[l1_connection[k][1]], l1_y_pos[k][0], main_sugar_y_pos[l1_connection[k][1]], d, l1_bond[k][0], dim = dim, compact = compact, highlight = l1_bond_label[k][0], color_highlight = side_per_linkage[k][0] if highlight_linkages else False) for k in range(len(l1_sugar))]
  # Bond branch branch
  [add_bond(l2_x_pos[b_idx][s_idx+1], l2_x_pos[b_idx][s_idx], l2_y_pos[b_idx][s_idx+1], l2_y_pos[b_idx][s_idx], d, l2_bond[b_idx][s_idx+1], dim = dim, compact = compact, highlight = l2_bond_label[b_idx][s_idx+1], color_highlight = branched_side_per_linkage[b_idx][s_idx+1] if highlight_linkages else False) for b_idx in range(len(l2_sugar)) for s_idx in range(len(l2_sugar[b_idx])-1) if len(l2_sugar[b_idx]) > 1]
  # Bond branch branch branch
  [add_bond(l3_x_pos[b_idx][s_idx+1], l3_x_pos[b_idx][s_idx], l3_y_pos[b_idx][s_idx+1], l3_y_pos[b_idx][s_idx], d, l3_bond[b_idx][s_idx+1], dim = dim, compact = compact, highlight = l3_bond_label[b_idx][s_idx+1]) for b_idx in range(len(l3_sugar)) for s_idx in range(len(l3_sugar[b_idx])-1) if len(l3_sugar[b_idx]) > 1]
  # Bond branch_branch to branch
  [add_bond(l2_x_pos[k][0], l1_x_pos[l2_connection[k][0]][l2_connection[k][1]], l2_y_pos[k][0], l1_y_pos[l2_connection[k][0]][l2_connection[k][1]], d, l2_bond[k][0], dim = dim, compact = compact, highlight = l2_bond_label[k][0], color_highlight = branched_side_per_linkage[k][0] if highlight_linkages else False) for k in range(len(l2_sugar))]
  # Bond branch_branch_branch to branch_branch
  [add_bond(l3_x_pos[k][0], l2_x_pos[l3_connection[k][0]][l3_connection[k][1]], l3_y_pos[k][0], l2_y_pos[l3_connection[k][0]][l3_connection[k][1]], d, l3_bond[k][0], dim = dim, compact = compact, highlight = l3_bond_label[k][0]) for k in range(len(l3_sugar))]

  # Sugar main chain
  [add_sugar(main_sugar[k], d, main_sugar_x_pos[k], main_sugar_y_pos[k], modification = main_sugar_modification[k], conf = main_conf[k], compact = compact, dim = dim, deg = main_deg[k], highlight = main_sugar_label[k], scalar = main_per_residue[k] if per_residue else 0) for k in range(len(main_sugar))]
  # Sugar branch
  [add_sugar(l1_sugar[b_idx][s_idx], d, l1_x_pos[b_idx][s_idx], l1_y_pos[b_idx][s_idx], modification = l1_sugar_modification[b_idx][s_idx], conf = l1_conf[b_idx][s_idx], compact = compact, dim = dim, deg = l1_deg[b_idx][s_idx], highlight = l1_sugar_label[b_idx][s_idx], scalar = side_per_residue[b_idx][s_idx] if per_residue else 0) for b_idx in range(len(l1_sugar)) for s_idx in range(len(l1_sugar[b_idx]))]
  # Sugar branch_branch
  [add_sugar(l2_sugar[b_idx][s_idx], d, l2_x_pos[b_idx][s_idx], l2_y_pos[b_idx][s_idx], modification = l2_sugar_modification[b_idx][s_idx], conf = l2_conf[b_idx][s_idx], compact = compact, dim = dim, deg = l2_deg[b_idx][s_idx], highlight = l2_sugar_label[b_idx][s_idx], scalar = branched_side_per_residue[b_idx][s_idx] if per_residue else 0) for b_idx in range(len(l2_sugar)) for s_idx in range(len(l2_sugar[b_idx]))]
  # Sugar branch branch branch
  [add_sugar(l3_sugar[b_idx][s_idx], d, l3_x_pos[b_idx][s_idx], l3_y_pos[b_idx][s_idx], modification = l3_sugar_modification[b_idx][s_idx], conf = l3_conf[b_idx][s_idx], compact = compact, dim = dim, highlight = l3_sugar_label[b_idx][s_idx]) for b_idx in range(len(l3_sugar)) for s_idx in range(len(l3_sugar[b_idx]))]

  highlight = 'show' if highlight_motif == None else 'hide'
  if floaty_bits != []:
    fb_count = {i: floaty_bits.count(i) for i in floaty_bits}
    floaty_bits = list(set(floaty_bits))
    floaty_data = []
    for k, k_val in enumerate(floaty_bits):
      if in_lib(min_process_glycans([k_val])[0][0], libr):
        floaty_data.append(get_coordinates_and_labels(k_val, show_linkage = show_linkage, highlight_motif = None))
      else:
        floaty_data.append(get_coordinates_and_labels('blank(-)blank', show_linkage = show_linkage, highlight_motif = None))
    n_floats = len(floaty_bits)
    y_spacing = (y_span / (n_floats - 1)) if n_floats > 1 else 0
    for j, j_val in enumerate(floaty_data):
      floaty_sugar, floaty_sugar_x_pos, floaty_sugar_y_pos, floaty_sugar_modification, floaty_bond, floaty_conf, _, _ = j_val[0]
      floaty_sugar_label = ['show' if highlight_motif == None else 'hide' for k in floaty_sugar]
      floaty_bond_label = ['show' if highlight_motif == None else 'hide' for k in floaty_bond]
      floaty_sugar_x_pos = [floaty_sugar_x_pos[k] + max_x + 1 for k in floaty_sugar_x_pos]
      current_y = (min_y + (j * y_spacing)) if n_floats > 1 else ((min_y + max_y) / 2)
      floaty_sugar_y_pos = [current_y for _ in range(len(floaty_sugar_y_pos))]
      if floaty_sugar != ['blank', 'blank']:
        [add_bond(floaty_sugar_x_pos[k+1], floaty_sugar_x_pos[k], floaty_sugar_y_pos[k+1], floaty_sugar_y_pos[k], d, floaty_bond[k], dim = dim, compact = compact, highlight = floaty_bond_label[k]) for k in range(len(floaty_sugar)-1)]
        [add_sugar(floaty_sugar[k], d, floaty_sugar_x_pos[k], floaty_sugar_y_pos[k], modification = floaty_sugar_modification[k], conf = floaty_conf, compact = compact, dim = dim, highlight = floaty_sugar_label[k]) for k in range(len(floaty_sugar))]
      else:
        add_sugar('text', d, min(floaty_sugar_x_pos)-0.3, floaty_sugar_y_pos[-1], modification = floaty_bits[j].translate(str.maketrans("123456789", "\u2081\u2082\u2083\u2084\u2085\u2086\u2087\u2088\u2089")).replace('blank', ''), compact = compact, dim = dim, text_anchor = 'end', highlight = highlight)

      if fb_count[floaty_bits[j]] > 1:
        x_offset = 0.5 if not compact else 0.75
        add_sugar('text', d, max(floaty_sugar_x_pos)+x_offset, floaty_sugar_y_pos[-1], modification = f"{fb_count[floaty_bits[j]]}x", compact = compact, dim = dim, highlight = highlight)

    bracket_x = max_x * (2 if not compact else 1.2) + 1
    bracket_y = (min_y, max_y) if not compact else ((min_y * 0.5) * 1.2, (max_y * 0.5) * 1.2)
    draw_bracket(bracket_x, bracket_y, d, direction = 'right', dim = dim, highlight = highlight)

  # add brackets around repeating unit
  if repeat:
    # process annotation
    repeat_annot = 'n' + (' = ' + str(repeat) if isinstance(repeat, (str, int)) and repeat != True else '')
    # repeat range code block
    if repeat_range:
      bracket_open = (main_sugar_x_pos[repeat_range[1]]*2)+1 if not compact else (main_sugar_x_pos[repeat_range[1]]*1.2)+0.6
      bracket_close = (main_sugar_x_pos[repeat_range[0]]*2)-1 if not compact else (main_sugar_x_pos[repeat_range[0]]*1.2)-0.6
      bracket_y_open =  (main_sugar_y_pos[repeat_range[1]], main_sugar_y_pos[repeat_range[1]]) if not compact else (((np.mean(main_sugar_y_pos[repeat_range[1]]) * 0.5) * 1.2)+0.0, ((np.mean(main_sugar_y_pos[repeat_range[1]]) * 0.5) * 1.2)-0.0)
      bracket_y_close = (main_sugar_y_pos[repeat_range[0]], main_sugar_y_pos[repeat_range[0]]) if not compact else (((np.mean(main_sugar_y_pos[repeat_range[0]]) * 0.5) * 1.2)+0.0, ((np.mean(main_sugar_y_pos[repeat_range[0]]) * 0.5) * 1.2)-0.0)
      text_x = main_sugar_x_pos[repeat_range[0]]-0.5
      text_y = main_sugar_y_pos[0]+1.05 if not compact else (main_sugar_y_pos[0]+1.03)/0.6
      draw_bracket(bracket_close, bracket_y_close, d, direction = 'left', dim = dim, highlight = highlight, deg = 0)
      draw_bracket(bracket_open, bracket_y_open, d, direction = 'right', dim = dim, highlight = highlight, deg = 0)
      add_sugar('text', d, text_x, text_y, modification = repeat_annot, compact = compact, dim = dim, text_anchor = 'start', highlight = highlight)
    # repeat unit code block
    else:
      open_deg = calculate_degree(main_sugar_y_pos[-1], main_sugar_y_pos[-2], main_sugar_x_pos[-1], main_sugar_x_pos[-2])
      if open_deg == 0:
        bracket_open = np.mean([k*2 for k in main_sugar_x_pos][-2:])+0.2 if not compact else np.mean([k*1.2 for k in main_sugar_x_pos][-2:])+0.15
        bracket_y_open = (np.mean(main_sugar_y_pos[-2:]), np.mean(main_sugar_y_pos[-2:])) if not compact else (((np.mean(main_sugar_y_pos[-2:]) * 0.5) * 1.2)+0.0, ((np.mean(main_sugar_y_pos[-2:]) * 0.5) * 1.2)-0.0)
        bracket_y_close = (main_sugar_y_pos[0], main_sugar_y_pos[0]) if not compact else (((np.mean(main_sugar_y_pos[0]) * 0.5) * 1.2)+0.0, ((np.mean(main_sugar_y_pos[0]) * 0.5) * 1.2)-0.0)
      else:
        bracket_open = np.mean([k*2 for k in main_sugar_x_pos][-2:])+0.0 if not compact else np.mean([k*1.2 for k in main_sugar_x_pos][-2:])+0
        bracket_y_open = (np.mean(main_sugar_y_pos[-2:]), np.mean(main_sugar_y_pos[-2:])) if not compact else (((np.mean(main_sugar_y_pos[-2:]) * 0.5) * 1.2)+0.3, ((np.mean(main_sugar_y_pos[-2:]) * 0.5) * 1.2)-0.3)
        bracket_y_close = (main_sugar_y_pos[0], main_sugar_y_pos[0]) if not compact else (((np.mean(main_sugar_y_pos[0]) * 0.5) * 1.2)+0.3, ((np.mean(main_sugar_y_pos[0]) * 0.5) * 1.2)-0.3)
      bracket_close = np.mean([k*2 for k in main_sugar_x_pos][:2])-0.2 if not compact else np.mean([k*1.2 for k in main_sugar_x_pos][:2])-0.15
      text_x = bracket_close - (0.42) if not compact else bracket_close - (0.13)
      text_y = main_sugar_y_pos[0] + 1.05 if not compact else (main_sugar_y_pos[0] + 1.03)/0.6
      draw_bracket(bracket_open, bracket_y_open, d, direction = 'right', dim = dim, highlight = highlight, deg = open_deg)
      draw_bracket(bracket_close, bracket_y_close, d, direction = 'left', dim = dim, highlight = highlight, deg = 0)
      add_sugar('text', d, text_x, text_y, modification = repeat_annot, compact = compact, dim = dim, text_anchor = 'start', highlight = highlight)

  d2.append(d)

  if filepath:
    filepath = Path(filepath)
    filepath = filepath.with_name(filepath.name.replace('?', '_'))
    data = d2.as_svg()
    data = data.replace('<svg ', f'<svg aria-label="{alt_text}" role="img" ', 1)
    if filepath.suffix.lower() == '.svg':
      with open(filepath, 'w', encoding = "utf-8") as f:
        f.write(data)
    elif filepath.suffix.lower() == '.pdf':
      convert_svg_to_pdf(data, str(filepath))
    elif filepath.suffix.lower() == '.png':
      convert_svg_to_png(data, str(filepath))
  return d2 if is_jupyter() or suppress or filepath else display_svg_with_matplotlib(d2)


def scale_in_range(
    listy: List[float], # Numbers to normalize
    a: float, # Target minimum
    b: float # Target maximum
    ) -> List[float]: # Normalized numbers
  "Normalizes list of numbers to specified range"
  min_val = min(listy)
  max_val = max(listy)
  range_val = max(max_val - min_val, 1e-6)
  return [(b - a) * ((x - min_val) / range_val) + a for x in listy]


def annotate_figure(
    svg_input: str, # Input SVG file path
    scale_range: Tuple[int, int] = (25, 80), # Min/max glycan dimensions
    compact: bool = False, # Use compact style
    glycan_size: str = 'medium', # Glycan size preset ('small', 'medium', 'large')
    filepath: Union[str, Path] = '', # Output file path
    scale_by_DE_res: Optional[pd.DataFrame] = None, # Differential expression results (motif_analysis.get_differential_expression)
    x_thresh: float = 1, # X metric threshold
    y_thresh: float = 0.05, # P-value threshold
    x_metric: str = 'Log2FC' # X axis metric ('Log2FC', 'Effect size')
    ) -> Optional[str]: # Modified SVG code
  "Replaces text labels with glycan drawings in SVG figure"
  glycan_size_dict = {
      'small': 'scale(0.1 0.1)  translate(0, -74)',
      'medium': 'scale(0.2 0.2)  translate(0, -55)',
      'large': 'scale(0.3 0.3)  translate(0, -49)'
      }
  glycan_scale = ''

  if scale_by_DE_res is not None:
    res_df = scale_by_DE_res.loc[(abs(scale_by_DE_res[x_metric]) > x_thresh) & (scale_by_DE_res['corr p-val'] < y_thresh)]
    y = -np.log10(res_df['corr p-val'].values.tolist())
    labels = res_df['Glycan'].values.tolist()
    glycan_scale = [y, labels]

  # Get svg code
  svg_tmp = open(svg_input, "r").read() if '?xml' not in svg_input else svg_input
  # Get all text labels
  label_pattern = re.compile(r'<!--\s*(.*?)\s*-->')
  transform_pattern = re.compile(r'<g transform\s*(.*?)\s*">')
  matches = re.findall(r"<!--.*-->[\s\S]*?<\/g>", svg_tmp)
  # Prepare for appending
  svg_tmp = svg_tmp.replace('</svg>', '')
  element_id = 0
  edit_svg = False
  motifs = motif_list.motif_name.values.tolist()

  for match in matches:
    # Keep track of current label and position in figure
    current_label = label_pattern.findall(match)[0]
    if current_label.startswith('Terminal') and current_label not in motifs:
      if in_lib(current_label.split('_')[-1], lib):
        edit_svg = True
    # Check if label is glycan
    if in_lib(current_label, lib):
      edit_svg = True
    else:
      pass
    try:
      if in_lib(motif_list.loc[motif_list.motif_name == current_label].motif.values.tolist()[0], lib):
        edit_svg = True
      else:
        pass
    except Exception:
      pass
    # Delete text label, append glycan figure
    if edit_svg:
      current_pos = '<g transform' + transform_pattern.findall(match)[0] + '">'
      current_pos = current_pos.replace('scale(0.1 -0.1)', glycan_size_dict[glycan_size])
      svg_tmp = svg_tmp.replace(match, '')
      if glycan_scale == '':
        d = GlycoDraw(current_label, compact = compact, suppress = True)
      else:
        d = GlycoDraw(current_label, compact = compact, dim = scale_in_range(glycan_scale[0], scale_range[0], scale_range[1])[glycan_scale[1].index(current_label)], suppress = True)
      data = d.as_svg().replace('<?xml version="1.0" encoding="UTF-8"?>\n', '')
      id_matches = re.findall(r'd\d+', data)
      # Reassign element ids to avoid duplicates
      for idx in id_matches:
        data = data.replace(idx, 'd' + str(element_id))
        element_id += 1
      svg_tmp += '\n' + current_pos + '\n' + data + '\n</g>'
      edit_svg = False
  svg_tmp += '</svg>'

  if filepath:
    if filepath.endswith('.pdf'):
      convert_svg_to_pdf(svg_tmp, str(filepath))
    elif filepath.endswith('.svg'):
      with open(filepath, 'w', encoding = "utf-8") as f:
        f.write(svg_tmp)
    elif filepath.endswith('.png'):
      convert_svg_to_png(svg_tmp, str(filepath))
  else:
    return svg_tmp


def plot_glycans_excel(
    df: Union[pd.DataFrame, str, Path], # DataFrame or filepath with glycans
    folder_filepath: Union[str, Path], # Output folder path
    glycan_col_num: int = 0, # Glycan column index
    scaling_factor: float = 0.2, # Image scaling
    compact: bool = False # Use compact style
    ) -> None:
  "Creates Excel file with SNFG glycan images in a new column"
  from openpyxl.drawing.image import Image as OpenpyxlImage
  from openpyxl.utils import get_column_letter
  from PIL import Image
  if isinstance(df, (str, Path)):
    df = pd.read_csv(df) if Path(df).suffix.lower() == ".csv" else pd.read_csv(df, sep = "\t") if Path(df).suffix.lower() == ".tsv" else pd.read_excel(df)
  df["SNFG"] = [np.nan for k in range(len(df))]
  image_column_number = df.columns.tolist().index("SNFG") + 1
  # Convert df_out to Excel
  writer = pd.ExcelWriter(Path(folder_filepath) / "output.xlsx", engine = "openpyxl")
  df.to_excel(writer, index = False)
  # Load the workbook and get the active sheet
  workbook = writer.book
  sheet = writer.sheets["Sheet1"]
  min_padding = 5  # Minimum padding in pixels
  for i, glycan_structure in enumerate(df.iloc[:, glycan_col_num]):
    if glycan_structure and glycan_structure[0]:
      if not isinstance(glycan_structure[0], str):
        glycan_structure = glycan_structure[0][0]
      # Generate glycan image using GlycoDraw
      svg_data = GlycoDraw(glycan_structure, compact = compact, suppress = True).as_svg()
      # Get SVG dimensions and scale them
      width = svg_data.width if hasattr(svg_data, 'width') else 800
      height = svg_data.height if hasattr(svg_data, 'height') else 800
      # Convert SVG data to image
      temp_bytes = BytesIO(convert_svg_to_png(svg_data.encode('utf-8').decode('utf-8'), output_width = width,
                         output_height = height, scale = 2.0, return_bytes = True))
      # Load and crop image
      img = Image.open(temp_bytes)
      bbox = img.convert('RGBA').getbbox()
      if bbox:
        # Add minimal padding
        bbox = (max(0, bbox[0] - min_padding), max(0, bbox[1] - min_padding),
               min(img.width, bbox[2] + min_padding), min(img.height, bbox[3] + min_padding))
        img = img.crop(bbox).convert('RGBA')
      # Apply user scaling factor
      img_width, img_height = img.size
      img = img.resize((int(img_width * scaling_factor), int(img_height * scaling_factor)), Image.BICUBIC)
      # Save the image to a BytesIO object
      img_stream = BytesIO()
      img.save(img_stream, format = 'PNG')
      img_stream.seek(0)
      # Create an image
      img_for_excel = OpenpyxlImage(img_stream)
      img_for_excel.width, img_for_excel.height = img.width, img.height  # Set width and height
      # Find the cell to insert the image
      cell = sheet.cell(row = i + 2, column = image_column_number)  # +2 because Excel is 1-indexed and there's a header row
      # Insert the image into the cell
      sheet.add_image(img_for_excel, cell.coordinate)
      # Resize the cell to fit the image
      column_letter = get_column_letter(image_column_number)
      sheet.column_dimensions[column_letter].width = img.width * 0.75 * 0.15
      sheet.row_dimensions[cell.row].height = img.height * 0.75
  # Save the workbook
  workbook.save(filename = Path(folder_filepath) / "output.xlsx")
