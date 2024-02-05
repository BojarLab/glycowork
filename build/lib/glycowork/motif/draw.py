from glycowork.glycan_data.loader import lib, unwrap, motif_list, multireplace
from glycowork.motif.regex import get_match
from glycowork.motif.graph import glycan_to_nxGraph, categorical_node_match_wildcard
from glycowork.motif.tokenization import get_core, get_modification
from glycowork.motif.processing import expand_lib, min_process_glycans, get_possible_linkages, get_possible_monosaccharides, rescue_glycans
from io import BytesIO
import networkx as nx
import copy

try:
  import drawsvg as draw
  from openpyxl.drawing.image import Image as OpenpyxlImage
  from openpyxl.utils import get_column_letter
  from PIL import Image
except ImportError:
  raise ImportError("<draw dependencies missing; did you do 'pip install glycowork[draw]'?>")
import numpy as np
import pandas as pd
import sys
import re
from math import sin, cos, radians, sqrt, atan, degrees


def matches(line, opendelim = '(', closedelim = ')'):
  """
  Find matching pairs of delimiters in a given string.\n
  | Arguments:
  | :-
  | line (str): The string to search for delimiter pairs.
  | opendelim (str, optional): The character to use as the opening delimiter. Default: '('.
  | closedelim (str, optional): The character to use as the closing delimiter. Default: ')'.\n
  | Returns:
  | :-
  | tuple: A tuple containing the start and end positions of the matched delimiter pair, and the current depth of the stack.\n
  ref: https://stackoverflow.com/questions/5454322/python-how-to-match-nested-parentheses-with-regex
  """
  stack = []
  for m in re.finditer(r'[{}{}]'.format(opendelim, closedelim), line):
      pos = m.start()
      if line[pos-1] == '\\':
          # Skip escape sequence
          continue

      c = line[pos]
      if c == opendelim:
          stack.append(pos+1)
      elif c == closedelim:
          if stack:
              # print("matched", prevpos, pos, line[prevpos:pos])
              yield (stack.pop(), pos, len(stack))
          else:
              # Error
              print(f"Encountered extraneous closing quote at pos {pos}: '{line[pos:]}'")
              pass

  if stack:
      for pos in stack:
          print(f"Expecting closing quote to match open quote starting at: '{line[pos - 1:]}'")


# Adjusted SNFG color palette
col_dict_base = {
    'snfg_white'     : '#FFFFFF', 
    'snfg_alt_blue'  : '#0385AE', 
    'snfg_green'     : '#058F60',
    'snfg_yellow'    : '#FCC326',
    'snfg_light_blue': '#91D3E3',
    'snfg_pink'      : '#F39EA0',
    'snfg_purple'    : '#A15989',
    'snfg_brown'     : '#9F6D55',
    'snfg_orange'    : '#EF6130',
    'snfg_red'       : '#C23537',
    'black'          : '#000000',
    'grey'           : '#7F7F7F'
}

col_dict_transparent = {
    'snfg_white'     : '#FFFFFF', 
    'snfg_alt_blue'  : '#CDE7EF', 
    'snfg_green'     : '#CDE9DF',
    'snfg_yellow'    : '#FFF6DE',
    'snfg_light_blue': '#EEF8FB',
    'snfg_pink'      : '#FDF0F1',
    'snfg_purple'    : '#F1E6ED',
    'snfg_brown'     : '#F1E9E5',
    'snfg_orange'    : '#FDE7E0',
    'snfg_red'       : '#F7E0E0',
    'black'          : '#D9D9D9',
    'grey'           : '#ECECEC'
}

# col_dict_base['snfg_white']

# Extensions for draw_lib
additions = ['-', 'blank', 'red_end', 'free',
             '04X', '15A', '02A', '13X',
             '24X', '35X', '04A', '15X',
             '02X', '13A', '24A', '35A',
             '25A', '03A', '14X', '25X',
             '03X', '14A', 'Z', 'Y',
             'B', 'C', 'text', 'non_glycan',
             'show', 'hide']

# Shape-color mapping
sugar_dict = {
  "Hex": ['Hex', 'snfg_white', False], "Glc": ['Hex', 'snfg_alt_blue', False],
  "Glcf": ['Hex', 'snfg_alt_blue', True], "Man": ['Hex', 'snfg_green', False],
  "Manf": ['Hex', 'snfg_green', True], "Gal": ['Hex', 'snfg_yellow', False],
  "Galf": ['Hex', 'snfg_yellow', True], "Gul": ['Hex', 'snfg_orange', False],
  "Alt": ['Hex', 'snfg_pink', False], "All": ['Hex', 'snfg_purple', False],
  "Tal": ['Hex', 'snfg_light_blue', False], "Ido": ['Hex', 'snfg_brown', False],

  "HexNAc": ['HexNAc', 'snfg_white', False], "GlcNAc": ['HexNAc', 'snfg_alt_blue', False],
  "GlcfNAc": ['HexNAc', 'snfg_alt_blue', True], "ManNAc": ['HexNAc', 'snfg_green', False],
  "ManfNAc": ['HexNAc', 'snfg_green', True], "GalNAc": ['HexNAc', 'snfg_yellow', False],
  "GalfNAc": ['HexNAc', 'snfg_yellow', True], "GulNAc": ['HexNAc', 'snfg_orange', False],
  "AltNAc": ['HexNAc', 'snfg_pink', False], "AllNAc": ['HexNAc', 'snfg_purple', False],
  "TalNAc": ['HexNAc', 'snfg_light_blue', False], "IdoNAc": ['HexNAc', 'snfg_brown', False],

  "HexN": ['HexN', 'snfg_white', False], "GlcN": ['HexN', 'snfg_alt_blue', False],
  "ManN": ['HexN', 'snfg_green', False], "GalN": ['HexN', 'snfg_yellow', False],
  "GulN": ['HexN', 'snfg_orange', False], "AltN": ['HexN', 'snfg_pink', False],
  "AllN": ['HexN', 'snfg_purple', False], "TalN": ['HexN', 'snfg_light_blue', False],
  "IdoN": ['HexN', 'snfg_brown', False],

  "HexA": ['HexA', 'snfg_white', False], "GlcA": ['HexA', 'snfg_alt_blue', False],
  "ManA": ['HexA', 'snfg_green', False], "GalA": ['HexA', 'snfg_yellow', False],
  "GulA": ['HexA', 'snfg_orange', False], "AltA": ['HexA_2', 'snfg_pink', False],
  "AllA": ['HexA', 'snfg_purple', False], "TalA": ['HexA', 'snfg_light_blue', False],
  "IdoA": ['HexA_2', 'snfg_brown', False],

  "dHex": ['dHex', 'snfg_white', False], "Qui": ['dHex', 'snfg_alt_blue', False],
  "Rha": ['dHex', 'snfg_green', False], "6dGul": ['dHex', 'snfg_orange', False],
  "6dAlt": ['dHex', 'snfg_pink', False], "6dAltf": ['dHex', 'snfg_pink', True],
  "6dTal": ['dHex', 'snfg_light_blue', False], "Fuc": ['dHex', 'snfg_red', False],
  "Fucf": ['dHex', 'snfg_red', True],

  "dHexNAc": ['dHexNAc', 'snfg_white', False], "QuiNAc": ['dHexNAc', 'snfg_alt_blue', False],
  "RhaNAc": ['dHexNAc', 'snfg_green', False], "6dAltNAc": ['dHexNAc', 'snfg_pink', False],
  "6dTalNAc": ['dHexNAc', 'snfg_light_blue', False], "FucNAc": ['dHexNAc', 'snfg_red', False],
  "FucfNAc": ['dHexNAc', 'snfg_red', True],

  "ddHex": ['ddHex', 'snfg_white', False], "Oli": ['ddHex', 'snfg_alt_blue', False],
  "Tyv": ['ddHex', 'snfg_green', False], "Abe": ['ddHex', 'snfg_orange', False],
  "Par": ['ddHex', 'snfg_pink', False], "Dig": ['ddHex', 'snfg_purple', False],
  "Col": ['ddHex', 'snfg_light_blue', False],

  "Pen": ['Pen', 'snfg_white', False], "Ara": ['Pen', 'snfg_green', False],
  "Araf": ['Pen', 'snfg_green', True], "Lyx": ['Pen', 'snfg_yellow', False],
  "Lyxf": ['Pen', 'snfg_yellow', True], "Xyl": ['Pen', 'snfg_orange', False],
  "Xylf": ['Pen', 'snfg_orange', True], "Rib": ['Pen', 'snfg_pink', False],
  "Ribf": ['Pen', 'snfg_pink', True],

  "dNon": ['dNon', 'snfg_white', False], "Kdn": ['dNon', 'snfg_green', False],
  "Neu5Ac": ['dNon', 'snfg_purple', False], "Neu5Gc": ['dNon', 'snfg_light_blue', False],
  "Neu": ['dNon', 'snfg_brown', False], "Sia": ['dNon', 'snfg_red', False],

  "ddNon": ['ddNon', 'snfg_white', False], "Pse": ['ddNon', 'snfg_green', False],
  "Leg": ['ddNon', 'snfg_yellow', False], "Aci": ['ddNon', 'snfg_pink', False],
  "4eLeg": ['ddNon', 'snfg_light_blue', False],

  "Unknown": ['Unknown', 'snfg_white', False], "Bac": ['Unknown', 'snfg_alt_blue', False],
  "LDManHep": ['Unknown', 'snfg_green', False], "Kdo": ['Unknown', 'snfg_yellow', False],
  "Kdof": ['Unknown', 'snfg_yellow', True], "Dha": ['Unknown', 'snfg_orange', False],
  "DDManHep": ['Unknown', 'snfg_pink', False], "MurNAc": ['Unknown', 'snfg_purple', False],
  "MurNGc": ['Unknown', 'snfg_light_blue', False], "Mur": ['Unknown', 'snfg_brown', False],

  "Assigned": ['Assigned', 'snfg_white', False], "Api": ['Assigned', 'snfg_alt_blue', False],
  "Apif": ['Assigned', 'snfg_alt_blue', True], "Fru": ['Assigned', 'snfg_green', False],
  "Fruf": ['Assigned', 'snfg_green', True], "Tag": ['Assigned', 'snfg_yellow', False],
  "Sor": ['Assigned', 'snfg_orange', False], "Psi": ['Assigned', 'snfg_pink', False],
  "non_glycan": ['Assigned', 'black', False],

  "blank": ['empty', 'snfg_white', False], "text": ['text', None, None],
  "red_end": ['red_end', None, None], "free": ['free', None, None],
  "04X": ['04X', None, None], "15A": ['15A', None, None],
  "02A": ['02A', None, None], "13X": ['13X', None, None],
  "24X": ['24X', None, None], "35X": ['35X', None, None],
  "04A": ['04A', None, None], "15X": ['15X', None, None],
  "02X": ['02X', None, None], "13A": ['13A', None, None],
  "24A": ['24A', None, None], "35A": ['35A', None, None],
  "25A": ['25A', None, None], "03A": ['03A', None, None],
  "14X": ['14X', None, None], "25X": ['25X', None, None],
  "03X": ['03X', None, None], "14A": ['14A', None, None],
  "Z": ['Z', None, None], "Y": ['Y', None, None],
  "B": ['B', None, None], "C": ['C', None, None]
}


# Build draw_lib with glycoletter additions
draw_lib = expand_lib(lib, ['-'] + list(sugar_dict.keys()))


def hex_circumference(x_pos, y_pos, dim, col_dict):
  """Draw a hexagoncircumference at the specified position and dimensions.\n
  | Arguments:
  | :-
  | x_pos (int): X coordinate of the hexagon's center on the drawing canvas.
  | y_pos (int): Y coordinate of the hexagon's center on the drawing canvas.
  | dim (int): Arbitrary dimension unit used for scaling the hexagon's size.\n
  | Returns:
  | :-
  | None
  """
  for i in range(6):
    angle1 = radians(60 * i)
    angle2 = radians(60 * (i + 1))    
    x1 = (0 - x_pos * dim) + (0.5 * dim) * cos(angle1)
    y1 = (0 + y_pos * dim) + (0.5 * dim) * sin(angle1)
    x2 = (0 - x_pos * dim) + (0.5 * dim) * cos(angle2)
    y2 = (0 + y_pos * dim) + (0.5 * dim) * sin(angle2)
    p = draw.Path(stroke_width = 0.04 * dim, stroke = col_dict['black'])
    p.M(x1, y1)
    p.L(x2, y2)
    d.append(p)


def hex(x_pos, y_pos, dim, col_dict, color = 'white'):
  """Draw a hexagon shape at the specified position and dimensions.\n
  | Arguments:
  | :-
  | x_pos (int): X coordinate of the hexagon's center on the drawing canvas.
  | y_pos (int): Y coordinate of the hexagon's center on the drawing canvas.
  | dim (int): Arbitrary dimension unit used for scaling the hexagon's size.
  | color (str): Color of the hexagon. Default is 'white'.\n
  | Returns:
  | :-
  | None
  """  
  x_base = 0 - x_pos * dim
  y_base = 0 + y_pos * dim
  half_dim = 0.5 * dim
  points = []
  for angle in range(0, 360, 60):
    x = x_base + half_dim * cos(radians(angle))
    y = y_base + half_dim * sin(radians(angle))
    points.extend([x, y])
  d.append(draw.Lines(*points, close = True, fill = color, stroke = col_dict['black'], stroke_width = 0.04 * dim))


def draw_shape(shape, color, x_pos, y_pos, col_dict, modification = '', dim = 50, furanose = False, conf = '', deg = 0, text_anchor = 'middle'):
  """draw individual monosaccharides in shapes & colors according to the SNFG nomenclature\n
  | Arguments:
  | :-
  | shape (string): monosaccharide class; shape of icon
  | color (string): monosaccharide identity; color of icon
  | x_pos (int): x coordinate of icon on drawing canvas
  | y_pos (int): y coordinate of icon on drawing canvas
  | modification (string): icon text annotation; used for post-biosynthetic modifications
  | dim (int): arbitrary dimention unit; necessary for drawsvg; inconsequential when drawing is exported as svg graphics\n
  | Returns:
  | :-
  |
  """
  x_base = 0 - x_pos * dim
  y_base = 0 + y_pos * dim
  stroke_w = 0.04 * dim
  half_dim = dim / 2
  inside_hex_dim = ((sqrt(3))/2)*half_dim

  if shape == 'Hex':
    # Hexose - circle
    d.append(draw.Circle(x_base, y_base, half_dim, fill = col_dict[color], stroke_width = stroke_w, stroke = col_dict['black']))
    # Text annotation
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base+half_dim)
    p.L(x_base+dim, y_base+half_dim)
    d.append(p)
    d.append(draw.Text(modification, dim*0.35, path = p, fill = col_dict['black'], text_anchor = 'middle', line_offset = -3.15))
    if furanose:
      p = draw.Path(stroke_width = 0)
      p.M(x_base-dim, y_base)
      p.L(x_base+dim, y_base)
      d.append(p)
      d.append(draw.Text('f', dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))
    # Ring configuration
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base)
    p.L(x_base+dim, y_base)
    d.append(p)
    d.append(draw.Text(conf, dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))

  if shape == 'HexNAc':
    # HexNAc - square
    d.append(draw.Rectangle(x_base-half_dim, y_base-half_dim, dim, dim, fill = col_dict[color], stroke_width = stroke_w, stroke = col_dict['black']))
    # Text annotation
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base+half_dim)
    p.L(x_base+dim, y_base+half_dim)
    d.append(p)
    d.append(draw.Text(modification, dim*0.35, path = p, fill = col_dict['black'], text_anchor = 'middle', line_offset = -3.15))
    if furanose:
      p = draw.Path(stroke_width = 0)
      p.M(x_base-dim, y_base)
      p.L(x_base+dim, y_base) 
      d.append(p)
      d.append(draw.Text('f', dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))
    # Ring configuration
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base)
    p.L(x_base+dim, y_base)
    d.append(p)
    d.append(draw.Text(conf, dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))

  if shape == 'HexN':
    # Hexosamine - crossed square
    d.append(draw.Rectangle(x_base-half_dim, y_base-half_dim, dim, dim, fill = 'white', stroke_width = stroke_w, stroke = col_dict['black']))
    d.append(draw.Lines(x_base-half_dim, y_base-half_dim,
                        x_base+half_dim, y_base-half_dim,
                        x_base+half_dim, y_base+half_dim,
                        x_base-half_dim, y_base-half_dim,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = 0))
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'],)
    p.M(x_base-half_dim, y_base-half_dim)
    p.L(x_base+half_dim, y_base-half_dim)
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'],)
    p.M(x_base+half_dim, y_base-half_dim)
    p.L(x_base+half_dim, y_base+half_dim)
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'],)
    p.M(x_base+half_dim, y_base+half_dim)
    p.L(x_base-half_dim, y_base-half_dim)
    d.append(p)
    # Text annotation
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base+half_dim)
    p.L(x_base+dim, y_base+half_dim)
    d.append(p)
    d.append(draw.Text(modification, dim*0.35, path = p, fill = col_dict['black'], text_anchor = 'middle', line_offset = -3.15))
    # Ring configuration
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base)
    p.L(x_base+dim, y_base)
    d.append(p)
    d.append(draw.Text(conf, dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))

  if shape == 'HexA_2':
    # Hexuronate - divided diamond
    # AltA / IdoA
    d.append(draw.Lines(x_base,         y_base+half_dim,
                        x_base+half_dim, y_base,
                        x_base,         y_base-half_dim,
                        x_base-half_dim, y_base,
                        close = True, fill = 'white', stroke = col_dict['black'], stroke_width =stroke_w))
    d.append(draw.Lines(x_base-half_dim, y_base,
                        x_base, y_base+half_dim,
                        x_base+half_dim, y_base,
                        x_base-half_dim, y_base,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = 0))
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'],)
    p.M(x_base-half_dim, y_base)
    p.L(x_base, y_base+half_dim)
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'],)
    p.M(x_base, y_base+half_dim)
    p.L(x_base+half_dim, y_base)
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'],)
    p.M(x_base+half_dim, y_base)
    p.L(x_base-half_dim, y_base)
    d.append(p)
    # Text annotation
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base+half_dim)
    p.L(x_base+dim, y_base+half_dim)
    d.append(p)
    d.append(draw.Text(modification, dim*0.35, path = p, fill = col_dict['black'], text_anchor = 'middle', line_offset = -3.15))
    # Ring configuration
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base)
    p.L(x_base+dim, y_base)
    d.append(p)
    d.append(draw.Text(conf, dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))  

  if shape == 'HexA':
    # Hexuronate - divided diamond (col_dict[color]s flipped)
    d.append(draw.Lines(x_base,         y_base+half_dim,
                        x_base+half_dim, y_base,
                        x_base,         y_base-half_dim,
                        x_base-half_dim, y_base,
                        close = True, fill = 'white', stroke = col_dict['black'], stroke_width = stroke_w))
    d.append(draw.Lines(x_base-half_dim, y_base,
                        x_base, y_base-half_dim,
                        x_base+half_dim, y_base,
                        x_base-half_dim, y_base,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = 0))
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'],)
    p.M(x_base-half_dim, y_base)
    p.L(x_base, y_base-half_dim)
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'],)
    p.M(x_base, y_base-half_dim)
    p.L(x_base+half_dim, y_base)
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'],)
    p.M(x_base+half_dim, y_base)
    p.L(x_base-half_dim, y_base)
    d.append(p)
    # Text annotation
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base+half_dim)
    p.L(x_base+dim, y_base+half_dim)
    d.append(p)
    d.append(draw.Text(modification, dim*0.35, path = p, fill = col_dict['black'], text_anchor = 'middle', line_offset = -3.15))
    # Ring configuration
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base)
    p.L(x_base+dim, y_base)
    d.append(p)
    d.append(draw.Text(conf, dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))  

  if shape == 'dHex':
    # Deoxyhexose - triangle
    d.append(draw.Lines(x_base- half_dim, y_base+inside_hex_dim,  # -(dim*1/3)
                        x_base, y_base-inside_hex_dim,
                        x_base+half_dim, y_base+inside_hex_dim,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = stroke_w))
    # Text annotation
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base+half_dim)
    p.L(x_base+dim, y_base+half_dim)
    d.append(p)
    d.append(draw.Text(modification, dim*0.35, path = p, text_anchor = 'middle', line_offset = -3.15))
    if furanose:
      p = draw.Path(stroke_width = 0)
      p.M(x_base-dim, y_base+half_dim)
      p.L(x_base+dim, y_base+half_dim)
      d.append(p)
      d.append(draw.Text('f', dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))
    # Ring configuration
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base)
    p.L(x_base+dim, y_base)
    d.append(p)
    d.append(draw.Text(conf, dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))

  if shape == 'dHexNAc':
    # Deoxyhexnac - divided triangle
    d.append(draw.Lines(x_base-half_dim, y_base+inside_hex_dim,  # -(dim*1/3) for center of triangle
                        x_base, y_base-inside_hex_dim,  # -(dim*1/3) for bottom alignment
                        x_base+half_dim, y_base+inside_hex_dim,  # -(((3**0.5)/2)*dim*0.5) for half of triangle height
                        close = True, fill = 'white', stroke = col_dict['black'], stroke_width = stroke_w))
    d.append(draw.Lines(x_base, y_base+inside_hex_dim,  # -(dim*1/3)
                        x_base, y_base-inside_hex_dim,
                        x_base+half_dim, y_base+inside_hex_dim,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = 0))
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'],)
    p.M(x_base, y_base+inside_hex_dim)
    p.L(x_base, y_base-inside_hex_dim)
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'],)
    p.M(x_base, y_base-inside_hex_dim)
    p.L(x_base+half_dim, y_base+inside_hex_dim)
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'],)
    p.M(x_base+half_dim, y_base+inside_hex_dim)
    p.L(x_base, y_base+inside_hex_dim)
    d.append(p)
    # Text annotation
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base+half_dim)
    p.L(x_base+dim, y_base+half_dim)
    d.append(p)
    d.append(draw.Text(modification, dim*0.35, path = p, fill = col_dict['black'], text_anchor = 'middle', line_offset = -3.15))
    # Ring configuration
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base)
    p.L(x_base+dim, y_base)
    d.append(p)
    d.append(draw.Text(conf, dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))   

  if shape == 'ddHex':
    # Dideoxyhexose - flat rectangle
    d.append(draw.Lines(x_base-half_dim,         y_base+(dim*7/12*0.5),  # -(dim*0.5/12)
                        x_base+half_dim,         y_base+(dim*7/12*0.5),
                        x_base+half_dim,         y_base-(dim*7/12*0.5),
                        x_base-half_dim,         y_base-(dim*7/12*0.5),
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = stroke_w))
    # Text annotation
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base+half_dim)
    p.L(x_base+dim, y_base+half_dim)
    d.append(p)
    d.append(draw.Text(modification, dim*0.35, path = p, fill = col_dict['black'], text_anchor = 'middle', line_offset = -3.15))
    # Ring configuration
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base)
    p.L(x_base+dim, y_base)
    d.append(p)
    d.append(draw.Text(conf, dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))  

  if shape == 'Pen':
    # Pentose - star
    d.append(draw.Lines(x_base,         y_base-half_dim/cos(radians(18)),
                        x_base+((0.25*dim)/cos(radians(18)))*cos(radians(54)),         y_base-((0.25*dim)/cos(radians(18)))*sin(radians(54)),
                        x_base+(half_dim/cos(radians(18)))*cos(radians(18)),         y_base-(half_dim/cos(radians(18)))*sin(radians(18)),
                        x_base+((0.25*dim)/cos(radians(18)))*cos(radians(18)),         y_base+((0.25*dim)/cos(radians(18)))*sin(radians(18)),
                        x_base+(half_dim/cos(radians(18)))*cos(radians(54)),         y_base+(half_dim/cos(radians(18)))*sin(radians(54)),
                        x_base,         y_base+(0.25*dim)/cos(radians(18)),
                        x_base-(half_dim/cos(radians(18)))*cos(radians(54)),         y_base+(half_dim/cos(radians(18)))*sin(radians(54)),
                        x_base-((0.25*dim)/cos(radians(18)))*cos(radians(18)),         y_base+((0.25*dim)/cos(radians(18)))*sin(radians(18)),
                        x_base-(half_dim/cos(radians(18)))*cos(radians(18)),         y_base-(half_dim/cos(radians(18)))*sin(radians(18)),
                        x_base-((0.25*dim)/cos(radians(18)))*cos(radians(54)),         y_base-((0.25*dim)/cos(radians(18)))*sin(radians(54)),
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = stroke_w))
    # Text annotation
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base+half_dim)
    p.L(x_base+dim, y_base+half_dim)
    d.append(p)
    d.append(draw.Text(modification, dim*0.35, path = p, fill = col_dict['black'], text_anchor = 'middle', line_offset = -3.15))
    if furanose:
      p = draw.Path(stroke_width = 0)
      p.M(x_base-dim, y_base)
      p.L(x_base+dim, y_base)
      d.append(p)
      d.append(draw.Text('f', dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))
    # Ring configuration
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base)
    p.L(x_base+dim, y_base)
    d.append(p)
    d.append(draw.Text(conf, dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True)) 

  if shape == 'dNon':
    # Deoxynonulosonate - diamond
    d.append(draw.Lines(x_base,         y_base+half_dim,
                        x_base+half_dim, y_base,
                        x_base,         y_base-half_dim,
                        x_base-half_dim, y_base,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = stroke_w))
    # Text annotation
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base+half_dim)
    p.L(x_base+dim, y_base+half_dim)
    d.append(p)
    d.append(draw.Text(modification, dim*0.35, path = p, fill = col_dict['black'], text_anchor = 'middle', line_offset = -3.15))
    # Ring configuration
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base)
    p.L(x_base+dim, y_base)
    d.append(p)
    d.append(draw.Text(conf, dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))

  if shape == 'ddNon':
    # Dideoxynonulosonate - flat diamond
    d.append(draw.Lines(x_base,         y_base+half_dim-(dim*1/8),
                        x_base+half_dim+(dim*1/8), y_base,
                        x_base,         y_base-half_dim+(dim*1/8),
                        x_base-half_dim-(dim*1/8), y_base,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = stroke_w))
    # Text annotation
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base+half_dim)
    p.L(x_base+dim, y_base+half_dim)
    d.append(p)
    d.append(draw.Text(modification, dim*0.35, path = p, fill = col_dict['black'], text_anchor = 'middle', line_offset = -3.15))
    # Ring configuration
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base)
    p.L(x_base+dim, y_base)
    d.append(p)
    d.append(draw.Text(conf, dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))   

  if shape == 'Unknown':
    # Unknown - flat hexagon
    d.append(draw.Lines(x_base-half_dim+(dim*1/8),         y_base+half_dim-(dim*1/8),
                        x_base+half_dim-(dim*1/8),         y_base+half_dim-(dim*1/8),
                        x_base+half_dim-(dim*1/8)+(dim*0.2),         y_base,
                        x_base+half_dim-(dim*1/8),         y_base-half_dim+(dim*1/8),
                        x_base-half_dim+(dim*1/8),         y_base-half_dim+(dim*1/8),
                        x_base-half_dim+(dim*1/8)-(dim*0.2),         y_base,
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = stroke_w))
    # Text annotation
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base+half_dim)
    p.L(x_base+dim, y_base+half_dim)
    d.append(p)
    d.append(draw.Text(modification, dim*0.35, path = p, fill = col_dict['black'], text_anchor = 'middle', line_offset = -3.15))
    # Ring configuration
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base)
    p.L(x_base+dim, y_base)
    d.append(p)
    d.append(draw.Text(conf, dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True))   
  
  if shape == 'Assigned':
    # Assigned - pentagon
    d.append(draw.Lines(x_base,         y_base-half_dim/cos(radians(18)),
                        x_base+(half_dim/cos(radians(18)))*cos(radians(18)),         y_base-(half_dim/cos(radians(18)))*sin(radians(18)),
                        x_base+(half_dim/cos(radians(18)))*cos(radians(54)),         y_base+(half_dim/cos(radians(18)))*sin(radians(54)),
                        x_base-(half_dim/cos(radians(18)))*cos(radians(54)),         y_base+(half_dim/cos(radians(18)))*sin(radians(54)),
                        x_base-(half_dim/cos(radians(18)))*cos(radians(18)),         y_base-(half_dim/cos(radians(18)))*sin(radians(18)),
                        close = True, fill = col_dict[color], stroke = col_dict['black'], stroke_width = stroke_w))
    # Text annotation
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base+half_dim)
    p.L(x_base+dim, y_base+half_dim)
    d.append(p)
    d.append(draw.Text(modification, dim*0.35, path = p, fill = col_dict['black'], text_anchor = 'middle', line_offset = -3.15))
    # Ring configuration
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base)
    p.L(x_base+dim, y_base)
    d.append(p)
    d.append(draw.Text(conf, dim*0.3, path = p, fill = col_dict['black'], text_anchor = 'middle', center = True)) 

  if shape == 'empty':
    d.append(draw.Circle(x_base, y_base, dim/2, fill = 'none', stroke_width = stroke_w, stroke = 'none'))
    # Text annotation
    p = draw.Path(stroke_width = 0)
    p.M(x_base-dim, y_base+half_dim)
    p.L(x_base+dim, y_base+half_dim)
    d.append(p)
    d.append(draw.Text(modification, dim*0.35, path = p, fill = col_dict['black'], text_anchor = text_anchor, line_offset = -3.15))

  if shape == 'text':
    d.append(draw.Text(modification, dim*0.35, x_base, y_base, text_anchor = text_anchor, fill = col_dict['black']))

  if shape == 'red_end':
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'], fill = 'none')
    p.M((x_base+0.1*dim), (y_base-0.4*dim))  # Start path at point (-10, 20)
    p.C((x_base-0.3*dim), (y_base-0.1*dim),
        (x_base+0.3*dim), (y_base+0.1*dim),
        (x_base-0.1*dim), (y_base+0.4*dim))
    d.append(p)
    d.append(draw.Circle(x_base, y_base, 0.15*dim, fill = 'white', stroke_width = stroke_w, stroke = col_dict['black']))

  if shape == 'free':
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'], fill = 'none')
    p.M((x_base+0.1*dim), (y_base-0.4*dim))  # Start path at point (-10, 20)
    p.C((x_base-0.3*dim), (y_base-0.1*dim),
        (x_base+0.3*dim), (y_base+0.1*dim),
        (x_base-0.1*dim), (y_base+0.4*dim))
    d.append(p)

  if shape == '04X':
    hex(x_pos, y_pos, dim, col_dict)
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(30)),            y_base-inside_hex_dim*sin(radians(30)),
                        x_base+half_dim*cos(radians(60)),                       y_base-half_dim*sin(radians(60)),
                        x_base+half_dim*cos(radians(120)),                      y_base-half_dim*sin(radians(120)),
                        x_base+inside_hex_dim*cos(radians(150)),           y_base-inside_hex_dim*sin(radians(150)),
                        close = True, fill = col_dict['grey'], stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(30)),            y_base-inside_hex_dim*sin(radians(30)))
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(150)),           y_base-inside_hex_dim*sin(radians(150)))
    d.append(p)

  if shape == '15A':
    hex(x_pos, y_pos, dim, col_dict)
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(90)),            y_base-inside_hex_dim*sin(radians(90)),
                        x_base+half_dim*cos(radians(60)),                       y_base-half_dim*sin(radians(60)),
                        x_base+half_dim*cos(radians(0)),                      y_base-half_dim*sin(radians(0)),
                        x_base+inside_hex_dim*cos(radians(330)),           y_base-inside_hex_dim*sin(radians(330)),
                        close = True, fill = col_dict['grey'], stroke =col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(90)),            y_base-inside_hex_dim*sin(radians(90)))
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(330)),           y_base-inside_hex_dim*sin(radians(330)))
    d.append(p)

  if shape == '02A':
    hex(x_pos, y_pos, dim, col_dict)
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(30)),            y_base-inside_hex_dim*sin(radians(30)),
                        x_base+half_dim*cos(radians(0)),                      y_base-half_dim*sin(radians(0)),
                        x_base+half_dim*cos(radians(300)),                        y_base-half_dim*sin(radians(300)),
                        x_base+inside_hex_dim*cos(radians(270)),           y_base-inside_hex_dim*sin(radians(270)),
                        close = True, fill = col_dict['grey'], stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(30)),            y_base-inside_hex_dim*sin(radians(30)))
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(270)),           y_base-inside_hex_dim*sin(radians(270)))
    d.append(p)

  if shape == '13X':
    hex(x_pos, y_pos, dim, col_dict)
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(330)),            y_base-inside_hex_dim*sin(radians(330)),
                        x_base+half_dim*cos(radians(300)),                       y_base-half_dim*sin(radians(300)),
                        x_base+half_dim*cos(radians(240)),                        y_base-half_dim*sin(radians(240)),
                        x_base+inside_hex_dim*cos(radians(210)),           y_base-inside_hex_dim*sin(radians(210)),
                        close = True, fill = col_dict['grey'], stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(330)),            y_base-inside_hex_dim*sin(radians(330)))
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(210)),           y_base-inside_hex_dim*sin(radians(210)))
    d.append(p)

  if shape == '24X':
    hex(x_pos, y_pos, dim, col_dict)
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(270)),            y_base-inside_hex_dim*sin(radians(270)),
                        x_base+half_dim*cos(radians(240)),                       y_base-half_dim*sin(radians(240)),
                        x_base+half_dim*cos(radians(180)),                        y_base-half_dim*sin(radians(180)),
                        x_base+inside_hex_dim*cos(radians(150)),           y_base-inside_hex_dim*sin(radians(150)),
                        close = True, fill = col_dict['grey'], stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(270)),            y_base-inside_hex_dim*sin(radians(270)))
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(150)),           y_base-inside_hex_dim*sin(radians(150)))
    d.append(p)

  if shape == '35X':
    hex(x_pos, y_pos, dim, col_dict)
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(210)),            y_base-inside_hex_dim*sin(radians(210)),
                        x_base+half_dim*cos(radians(180)),                       y_base-half_dim*sin(radians(180)),
                        x_base+half_dim*cos(radians(120)),                        y_base-half_dim*sin(radians(120)),
                        x_base+inside_hex_dim*cos(radians(90)),           y_base-inside_hex_dim*sin(radians(90)),
                        close = True, fill = col_dict['grey'], stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(210)),            y_base-inside_hex_dim*sin(radians(210)))
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(90)),            y_base-inside_hex_dim*sin(radians(90)))
    d.append(p)

  if shape == '04A':
    hex(x_pos, y_pos, dim, color = col_dict['grey'])
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(30)),            y_base-inside_hex_dim*sin(radians(30)),
                        x_base+half_dim*cos(radians(60)),                       y_base-half_dim*sin(radians(60)),
                        x_base+half_dim*cos(radians(120)),                      y_base-half_dim*sin(radians(120)),
                        x_base+inside_hex_dim*cos(radians(150)),           y_base-inside_hex_dim*sin(radians(150)),
                        close = True, fill = 'white', stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(30)),            y_base-inside_hex_dim*sin(radians(30)))
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(150)),           y_base-inside_hex_dim*sin(radians(150)))
    d.append(p)

  if shape == '15X':
    hex(x_pos, y_pos, dim, color = col_dict['grey'])
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(90)),            y_base-inside_hex_dim*sin(radians(90)),
                        x_base+half_dim*cos(radians(60)),                       y_base-half_dim*sin(radians(60)),
                        x_base+half_dim*cos(radians(0)),                      y_base-half_dim*sin(radians(0)),
                        x_base+inside_hex_dim*cos(radians(330)),           y_base-inside_hex_dim*sin(radians(330)),
                        close = True, fill = 'white', stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(90)),            y_base-inside_hex_dim*sin(radians(90)))
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(330)),           y_base-inside_hex_dim*sin(radians(330)))
    d.append(p)

  if shape == '02X':
    hex(x_pos, y_pos, dim, color = col_dict['grey'])
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(30)),            y_base-inside_hex_dim*sin(radians(30)),
                        x_base+half_dim*cos(radians(0)),                       y_base-half_dim*sin(radians(0)),
                        x_base+half_dim*cos(radians(300)),                        y_base-half_dim*sin(radians(300)),
                        x_base+inside_hex_dim*cos(radians(270)),           y_base-inside_hex_dim*sin(radians(270)),
                        close = True, fill = 'white', stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(30)),            y_base-inside_hex_dim*sin(radians(30)))
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(270)),           y_base-inside_hex_dim*sin(radians(270)))
    d.append(p)

  if shape == '13A':
    hex(x_pos, y_pos, dim, color = col_dict['grey'])
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(330)),            y_base-inside_hex_dim*sin(radians(330)),
                        x_base+half_dim*cos(radians(300)),                       y_base-half_dim*sin(radians(300)),
                        x_base+half_dim*cos(radians(240)),                        y_base-half_dim*sin(radians(240)),
                        x_base+inside_hex_dim*cos(radians(210)),           y_base-inside_hex_dim*sin(radians(210)),
                        close = True, fill = 'white', stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(330)),            y_base-inside_hex_dim*sin(radians(330)))
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(210)),           y_base-inside_hex_dim*sin(radians(210)))
    d.append(p)

  if shape == '24A':
    hex(x_pos, y_pos, dim, color = col_dict['grey'])
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(270)),            y_base-inside_hex_dim*sin(radians(270)),
                        x_base+half_dim*cos(radians(240)),                       y_base-half_dim*sin(radians(240)),
                        x_base+half_dim*cos(radians(180)),                        y_base-half_dim*sin(radians(180)),
                        x_base+inside_hex_dim*cos(radians(150)),           y_base-inside_hex_dim*sin(radians(150)),
                        close = True, fill = 'white', stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(270)),            y_base-inside_hex_dim*sin(radians(270)))
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(150)),           y_base-inside_hex_dim*sin(radians(150)))
    d.append(p)

  if shape == '35A':
    hex(x_pos, y_pos, dim, color = col_dict['grey'])
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(210)),            y_base-inside_hex_dim*sin(radians(210)),
                        x_base+half_dim*cos(radians(180)),                       y_base-half_dim*sin(radians(180)),
                        x_base+half_dim*cos(radians(120)),                        y_base-half_dim*sin(radians(120)),
                        x_base+inside_hex_dim*cos(radians(90)),          y_base-inside_hex_dim*sin(radians(90)),
                        close = True, fill = 'white', stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(210)),            y_base-inside_hex_dim*sin(radians(210)))
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,                                                y_base)
    p.L(x_base+inside_hex_dim*cos(radians(90)),            y_base-inside_hex_dim*sin(radians(90)))
    d.append(p) 
  
  if shape == '25A':
    hex(x_pos, y_pos, dim, col_dict)
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(90)),            y_base-inside_hex_dim*sin(radians(90)),
                        x_base+half_dim*cos(radians(60)),                       y_base-half_dim*sin(radians(60)),
                        x_base+half_dim*cos(radians(0)),                        y_base-half_dim*sin(radians(0)),
                        x_base+half_dim*cos(radians(300)),                      y_base-half_dim*sin(radians(300)),
                        x_base+inside_hex_dim*cos(radians(270)),            y_base-inside_hex_dim*sin(radians(270)),
                        close = True, fill = col_dict['grey'], stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base+inside_hex_dim*cos(radians(90)),            y_base-inside_hex_dim*sin(radians(90)))
    p.L(x_base+inside_hex_dim*cos(radians(270)),            y_base-inside_hex_dim*sin(radians(270)))
    d.append(p)

  if shape == '03A':
    hex(x_pos, y_pos, dim, col_dict)
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(30)),            y_base-inside_hex_dim*sin(radians(30)),
                        x_base+half_dim*cos(radians(0)),                       y_base-half_dim*sin(radians(0)),
                        x_base+half_dim*cos(radians(300)),                        y_base-half_dim*sin(radians(300)),
                        x_base+half_dim*cos(radians(240)),                      y_base-half_dim*sin(radians(240)),
                        x_base+inside_hex_dim*cos(radians(210)),            y_base-inside_hex_dim*sin(radians(210)),
                        close = True, fill = col_dict['grey'], stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base+inside_hex_dim*cos(radians(30)),            y_base-inside_hex_dim*sin(radians(30)))
    p.L(x_base+inside_hex_dim*cos(radians(210)),            y_base-inside_hex_dim*sin(radians(210)))
    d.append(p)

  if shape == '14X':
    hex(x_pos, y_pos, dim, col_dict)
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(330)),            y_base-inside_hex_dim*sin(radians(330)),
                        x_base+half_dim*cos(radians(300)),                       y_base-half_dim*sin(radians(300)),
                        x_base+half_dim*cos(radians(240)),                        y_base-half_dim*sin(radians(240)),
                        x_base+half_dim*cos(radians(180)),                      y_base-half_dim*sin(radians(180)),
                        x_base+inside_hex_dim*cos(radians(150)),            y_base-inside_hex_dim*sin(radians(150)),
                        close = True, fill = col_dict['grey'], stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base+inside_hex_dim*cos(radians(330)),            y_base-inside_hex_dim*sin(radians(330)))
    p.L(x_base+inside_hex_dim*cos(radians(150)),            y_base-inside_hex_dim*sin(radians(150)))
    d.append(p)

  if shape == '25X':
    hex(x_pos, y_pos, dim, color = col_dict['grey'])
    d.append(draw.Lines(x_base,                                               y_base,
                        x_base+inside_hex_dim*cos(radians(90)),            y_base-inside_hex_dim*sin(radians(90)),
                        x_base+half_dim*cos(radians(60)),                       y_base-half_dim*sin(radians(60)),
                        x_base+half_dim*cos(radians(0)),                        y_base-half_dim*sin(radians(0)),
                        x_base+half_dim*cos(radians(300)),                      y_base-half_dim*sin(radians(300)),
                        x_base+inside_hex_dim*cos(radians(270)),            y_base-inside_hex_dim*sin(radians(270)),
                        close = True, fill = 'white', stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base+inside_hex_dim*cos(radians(90)),            y_base-inside_hex_dim*sin(radians(90)))
    p.L(x_base+inside_hex_dim*cos(radians(270)),            y_base-inside_hex_dim*sin(radians(270)))
    d.append(p)

  if shape == '03X':
    hex(x_pos, y_pos, dim, color = col_dict['grey'])
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(30)),            y_base-inside_hex_dim*sin(radians(30)),
                        x_base+half_dim*cos(radians(0)),                       y_base-half_dim*sin(radians(0)),
                        x_base+half_dim*cos(radians(300)),                        y_base-half_dim*sin(radians(300)),
                        x_base+half_dim*cos(radians(240)),                     y_base-half_dim*sin(radians(240)),
                        x_base+inside_hex_dim*cos(radians(210)),            y_base-inside_hex_dim*sin(radians(210)),
                        close = True, fill = 'white', stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base+inside_hex_dim*cos(radians(30)),            y_base-inside_hex_dim*sin(radians(30)))
    p.L(x_base+inside_hex_dim*cos(radians(210)),            y_base-inside_hex_dim*sin(radians(210)))
    d.append(p)

  if shape == '14A':
    hex(x_pos, y_pos, dim, color = col_dict['grey'])
    d.append(draw.Lines(x_base,                                                y_base,
                        x_base+inside_hex_dim*cos(radians(330)),            y_base-inside_hex_dim*sin(radians(330)),
                        x_base+half_dim*cos(radians(300)),                       y_base-half_dim*sin(radians(300)),
                        x_base+half_dim*cos(radians(240)),                        y_base-half_dim*sin(radians(240)),
                        x_base+half_dim*cos(radians(180)),                      y_base-half_dim*sin(radians(180)),
                        x_base+inside_hex_dim*cos(radians(150)),           y_base-inside_hex_dim*sin(radians(150)),
                        close = True, fill = 'white', stroke = col_dict['black'], stroke_width = 0))
    hex_circumference(x_pos, y_pos, dim, col_dict)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base+inside_hex_dim*cos(radians(330)),            y_base-inside_hex_dim*sin(radians(330)))
    p.L(x_base+inside_hex_dim*cos(radians(150)),            y_base-inside_hex_dim*sin(radians(150)))
    d.append(p)

  if shape == 'Z':
    # deg = 0
    rot = 'rotate(' + str(deg) + ' ' + str(0-abs(x_pos)*(dim)) + ' ' + str(0-abs(y_pos)*(dim)) + ')'
    g = draw.Group(transform = rot)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,            y_base-half_dim)
    p.L(x_base,            y_base+half_dim)
    g.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base-0.02*dim,            y_base-half_dim)
    p.L(x_base+0.4*dim,            y_base-half_dim)
    g.append(p)
    d.append(g)

  if shape == 'Y':
    # deg = 0
    rot = 'rotate(' + str(deg) + ' ' + str(0-abs(x_pos)*(dim)) + ' ' + str(0-abs(y_pos)*(dim)) + ')'
    g = draw.Group(transform = rot)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,            y_base-half_dim)
    p.L(x_base,            y_base+half_dim)
    g.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base-0.02*dim,            y_base-half_dim)
    p.L(x_base+0.4*dim,           y_base-half_dim)
    g.append(p)
    g.append(draw.Circle(x_base+0.4*dim, y_base, 0.15*dim, fill = 'none', stroke_width = stroke_w, stroke = col_dict['black']))
    d.append(g)

  if shape == 'B':
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,            y_base-half_dim)
    p.L(x_base,            y_base+half_dim)
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base+0.02*dim,            y_base+half_dim)
    p.L(x_base-0.4*dim,            y_base+half_dim)
    d.append(p)

  if shape == 'C':
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base,            y_base-half_dim)
    p.L(x_base,            y_base+half_dim)
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = col_dict['black'])
    p.M(x_base+0.02*dim,            y_base+half_dim)
    p.L(x_base-0.4*dim,            y_base+half_dim)
    d.append(p)
    d.append(draw.Circle(x_base-0.4*dim, y_base, 0.15*dim, fill = 'none', stroke_width = stroke_w, stroke = col_dict['black']))


def add_bond(x_start, x_stop, y_start, y_stop, label = '', dim = 50, compact = False, highlight = 'show'):
  """drawing lines (bonds) between start/stop coordinates\n
  | Arguments:
  | :-
  | x_start (int): x1 coordinate
  | x_stop (int): x2 coordinate
  | y_start (int): y1 coordinate
  | y_stop (int): y2 coordinate
  | label (string): bond text annotation to specify linkage
  | dim (int): arbitrary dimention unit; necessary for drawsvg; inconsequential when drawing is exported as svg graphics
  | compact (bool): drawing style; normal or compact\n
  | Returns:
  | :-
  |
  """
  if highlight == 'hide':
    col_dict = col_dict_transparent
  else:
    col_dict = col_dict_base

  label = label if label != '-' else ''
  scaling_factor = 1.2 if compact else 2
  x_start = x_start * scaling_factor
  x_stop = x_stop * scaling_factor
  if compact:
    y_start = y_start * 0.6
    y_stop = y_stop * 0.6
  p = draw.Path(stroke_width = 0.08*dim, stroke = col_dict['black'],)
  p.M(0-x_start*dim, 0+y_start*dim)
  p.L(0-x_stop*dim, 0+y_stop*dim)
  d.append(p)
  d.append(draw.Text(label, dim*0.4, path = p, text_anchor = 'middle', fill = col_dict['black'], valign = 'middle', line_offset = -0.5))


def add_sugar(monosaccharide, x_pos = 0, y_pos = 0, modification = '', dim = 50, compact = False, conf = '', deg = 0, text_anchor = 'middle', highlight = 'show'):
  """wrapper function for drawing monosaccharide at specified position\n
  | Arguments:
  | :-
  | monosaccharide (string): simplified IUPAC nomenclature
  | x_pos (int): x coordinate of icon on drawing canvas
  | y_pos (int): y coordinate of icon on drawing canvas
  | modification (string): icon text annotation; used for post-biosynthetic modifications
  | dim (int): arbitrary dimention unit; necessary for drawsvg; inconsequential when drawing is exported as svg graphics
  | compact (bool): drawing style; normal or compact\n
  | Returns:
  | :-
  |
  """
  if highlight == 'hide':
    col_dict = col_dict_transparent
  else:
    col_dict = col_dict_base
  
  x_pos = x_pos * (1.2 if compact else 2)
  y_pos = y_pos * (0.6 if compact else 1)
  if monosaccharide in sugar_dict:
    draw_shape(shape = sugar_dict[monosaccharide][0], color = sugar_dict[monosaccharide][1], x_pos = x_pos, y_pos = y_pos,
               modification = modification, conf = conf, furanose = sugar_dict[monosaccharide][2], dim = dim, deg = deg, text_anchor = text_anchor, col_dict = col_dict)
  else:
    x_base = 0 - x_pos * dim
    y_base = 0 + y_pos * dim
    stroke_w = 0.04 * dim
    half_dim = dim / 2
    p = draw.Path(stroke_width = stroke_w, stroke = 'black')
    p.M(x_base-half_dim, y_base+half_dim)
    p.L(x_base+half_dim, y_base-half_dim)
    d.append(p)
    p = draw.Path(stroke_width = stroke_w, stroke = 'black')
    p.M(x_base+half_dim, y_base+half_dim)
    p.L(x_base-half_dim, y_base-half_dim)
    d.append(p)


def multiple_branches(glycan):
  """Reorder multiple branches by linkage on a given glycan\n
  | Arguments:
  | :-
  | glycan (string): Input glycan string.\n
  | Returns:
  | :-
  | str: Modified glycan string.
  """
  if ']' not in glycan:
    return glycan
  tmp = multireplace(glycan, {'(': '*', ')': '*', '[': '(', ']': ')'})
  open_close = [(openpos, closepos) for openpos, closepos, level in matches(tmp) if level == 0 and not re.search(r"^Fuc\S{6}$", tmp[openpos:closepos])]
  for k in range(len(open_close)-1):
    if open_close[k + 1][0] - open_close[k][1] != 2:
      continue
    branch1, branch2 = glycan[open_close[k][0]:open_close[k][1]], glycan[open_close[k + 1][0]:open_close[k + 1][1]]
    tmp1, tmp2 = branch1[-2], branch2[-2]
    if tmp1 == tmp2 == '?':
      if len(branch1) > len(branch2):
        tmp1, tmp2 = [1, 2]
      else:
        tmp1, tmp2 = [2, 1]
    if tmp1 > tmp2:
      glycan = f"{glycan[:open_close[k][0]]}{branch2}][{branch1}{glycan[open_close[k + 1][1]:]}"
  return glycan


def multiple_branch_branches(glycan):
  """Reorder nested multiple branches by linkage on a given glycan\n
  | Arguments:
  | :-
  | glycan (str): Input glycan string.\n
  | Returns:
  | :-
  | str: Modified glycan string.
  """
  if ']' not in glycan:
    return glycan
  tmp = multireplace(glycan, {'(': '*', ')': '*', '[': '(', ']': ')'})
  for openpos, closepos, level in matches(tmp):
    if level == 0 and not re.search(r"^Fuc\S{6}$", tmp[openpos:closepos]):
        glycan = glycan[:openpos] + multiple_branches(glycan[openpos:closepos]) + glycan[closepos:]
  return glycan


def reorder_for_drawing(glycan, by = 'linkage'):
  """order level 0 branches by linkage, ascending\n
  | Arguments:
  | :-
  | glycan (string): glycan in IUPAC-condensed format\n
  | Returns:
  | :-
  | Returns re-ordered glycan
  """
  if ']' not in glycan:
    return glycan
  tmp = multireplace(glycan, {'(': '*', ')': '*', '[': '(', ']': ')'})
  for openpos, closepos, level in matches(tmp):
    if level == 0 and not re.search(r"^Fuc\S{6}$|^Xyl\S{6}$", tmp[openpos:closepos]):
      # Nested branches
      glycan = glycan[:openpos] + branch_order(glycan[openpos:closepos]) + glycan[closepos:]
      # Branches
      group1 = glycan[:openpos-1]
      group2 = glycan[openpos:closepos]
      branch_end = [j[-2] for j in [group1, group2]]
      if by == 'length':
        branch_len = [len(k) for k in min_process_glycans([group1, group2])]
        if branch_len[0] == branch_len[1]:
          if branch_end[0] == branch_end[1]:
            glycan = group1 + '[' + group2 + ']' + glycan[closepos+1:]
          else:
            glycan = [group1, group2][np.argmin(branch_end)] + '[' + [group1, group2][np.argmax(branch_end)] + ']' + glycan[closepos+1:]
        else:
          glycan = [group1, group2][np.argmax(branch_len)] + '[' + [group1, group2][np.argmin(branch_len)] + ']' + glycan[closepos+1:]
      elif by == 'linkage':
        if branch_end[0] == branch_end[1]:
          glycan = group1 + '[' + group2 + ']' + glycan[closepos+1:]
        else:
          glycan = [group1, group2][np.argmin(branch_end)] + '[' + [group1, group2][np.argmax(branch_end)] + ']' + glycan[closepos+1:]
  return glycan


def branch_order(glycan, by = 'linkage'):
  """order nested branches by linkage, ascending\n
  | Arguments:
  | :-
  | glycan (string): glycan in IUPAC-condensed format\n
  | Returns:
  | :-
  | Returns re-ordered glycan
  """
  tmp = multireplace(glycan, {'(': '*', ')': '*', '[': '(', ']': ')'})
  for openpos, closepos, level in matches(tmp):
    if level == 0 and not re.search(r"^Fuc\S{6}$|^Xyl\S{6}$", tmp[openpos:closepos]):
      group1 = glycan[:openpos-1]
      group2 = glycan[openpos:closepos]
      branch_end = [j[-2] for j in [group1, group2]]
      branch_end = [k.replace('z', '9') for k in branch_end]
      if by == 'length':
        branch_len = [len(k) for k in min_process_glycans([group1, group2])]
        if branch_len[0] == branch_len[1]:
          if branch_end[0] == branch_end[1]:
            glycan = group1 + '[' + group2 + ']' + glycan[closepos+1:]
          else:
            glycan = [group1, group2][np.argmin(branch_end)] + '[' + [group1, group2][np.argmax(branch_end)] + ']' + glycan[closepos+1:]
        else:
          glycan = [group1, group2][np.argmax(branch_len)] + '[' + [group1, group2][np.argmin(branch_len)] + ']' + glycan[closepos+1:]
      elif by == 'linkage':
        if branch_end[0] == branch_end[1]:
          glycan = group1 + '[' + group2 + ']' + glycan[closepos+1:]
        else:
          glycan = [group1, group2][np.argmin(branch_end)] + '[' + [group1, group2][np.argmax(branch_end)] + ']' + glycan[closepos+1:]
  return glycan


def split_node(G, node):
  """split graph at node\n
  | Arguments:
  | :-
  | G (networkx object): glycan graph
  | node (int): where to split the graph object\n
  | Returns:
  | :-
  | glycan graph (networkx object), consisting of two disjointed graphs\n
  ref: https://stackoverflow.com/questions/65853641/networkx-how-to-split-nodes-into-two-effectively-disconnecting-edges
  """
  edges = G.edges(node, data = True)
  new_edges = []
  new_nodes = []
  H = G.__class__()
  H.add_nodes_from(G.subgraph(node))
  for i, (s, t, data) in enumerate(edges):
      new_node = f"{node}_{i}"
      Ix = nx.relabel_nodes(H, {node: new_node})
      new_nodes += list(Ix.nodes(data = True))
      new_edges.append((new_node, t, data))
  G.remove_node(node)
  G.add_nodes_from(new_nodes)
  G.add_edges_from(new_edges)
  return G


def unique(sequence):
  """remove duplicates but keep order\n
  | Arguments:
  | :-
  | sequence (list): \n
  | Returns:
  | :-
  | deduplicated list, preserving original order\n
  ref: https://stackoverflow.com/questions/480214/how-do-i-remove-duplicates-from-a-list-while-preserving-order
  """
  seen = set()
  return [x for x in sequence if not (x in seen or seen.add(x))]


def get_indices(x, y):
  """for each element in list x, get index (indices if present multiple times) in list y\n
  | Arguments:
  | :-
  | x (list):
  | y (list) elements can be lists: \n
  | Returns:
  | :-
  | list of list of indices\n
  """
  return [([idx for idx, val in enumerate(x) if val == sub] if sub in x else [None]) for sub in y]


def process_bonds(linkage_list):
  """prepare linkages for printing to figure\n
  | Arguments:
  | :-
  | linkage_list (list): list (or list of lists) of linkages\n
  | Returns:
  | :-
  | processed linkage_list\n
  """
  def process_single_linkage(linkage, regex_dict):
    if '?' in linkage[0] and '?' in linkage[-1]:
        return '?'
    elif '?' in linkage[0]:
        return ' ' + linkage[-1]
    elif '?' in linkage[-1]:
        return regex_dict['alpha'].match(linkage) and '\u03B1' or regex_dict['beta'].match(linkage) and '\u03B2' or '-'
    elif regex_dict['alpha'].match(linkage):
        return '\u03B1 ' + linkage[-1]
    elif regex_dict['beta'].match(linkage):
        return '\u03B2 ' + linkage[-1]
    elif regex_dict['digit'].match(linkage):
        return linkage[0] + ' - ' + linkage[-1]
    else:
        return '-'
  regex_dict = {'alpha': re.compile(r"^a\d"), 'beta': re.compile(r"^b\d"), 'digit': re.compile(r"^\d-\d")}
  if any(isinstance(el, list) for el in linkage_list):
    return [[process_single_linkage(linkage, regex_dict) for linkage in sub_list] for sub_list in linkage_list]
  else:
    return [process_single_linkage(linkage, regex_dict) for linkage in linkage_list]


def split_monosaccharide_linkage(label_list):
  """Split the monosaccharide linkage and extract relevant information from the given label list.\n
  | Arguments:
  | :-
  | label_list (list): List of labels \n
  | Returns:
  | :-
  | Three lists - sugar, sugar_modification, and bond.
  |   - sugar (list): List of sugars
  |   - sugar_modification (list): List of sugar modifications
  |   - bond (list): List of bond information
  """
  def process_sublist(sub_list):
    sugar = sub_list[::2][::-1]
    mod = [get_modification(k) if k not in additions else '' for k in sugar]
    mod = [multireplace(k, {'O': '', '-ol': ''}) for k in mod]
    sugar = [get_core(k) if k not in additions else k for k in sugar]
    bond = sub_list[1::2][::-1]
    return sugar, mod, bond
  if any(isinstance(el, list) for el in label_list):
    return zip(*[process_sublist(sub) for sub in label_list])
  else:
    return process_sublist(label_list)


def glycan_to_skeleton(glycan_string):
  """convert glycan to skeleton of node labels according to the respective glycan graph object\n
  | Arguments:
  | :-
  | glycan_string (string): \n
  | Returns:
  | :-
  | node label skeleton of glycan (string)\n
  """
  tmp = multireplace(glycan_string, {'(': ',', ')': ',', '[': ',[,', ']': ',],'})
  elements, idx = [], 0
  for k in filter(bool, tmp.split(',')):
    if k in {'[', ']'}:
      elements.append(k)
    else:
      elements.append(str(idx))
      idx += 1
  return multireplace( '-'.join(elements), {'[-': '[', '-]': ']'})


def get_highlight_attribute(glycan_graph, motif_string, termini_list = []):
  '''
  Label nodes in parent glycan (graph) by presence in motif (string)
  | Arguments:
  | :-
  | glycan_graph (networkx object): Glycan as networkx object.
  | motif_string (string): Glycan as named motif or in IUPAC-condensed format. 
  | termini_list (list): list of monosaccharide positions (from 'terminal', 'internal', and 'flexible')\n
  | Returns:
  | :-
  | Returns networkx object of glycan, annotated with the node attribute 'highlight_labels', labeling nodes that constitute the motif ('show') or not ('hide').
  '''
  g1 = glycan_graph
  g_tmp = copy.deepcopy(g1)
  if not motif_string:
    g2 = copy.deepcopy(g1)
  else:
    try:
      g2 = glycan_to_nxGraph(motif_list.loc[motif_list.motif_name == motif_string].motif.values.tolist()[0], termini = 'provided', termini_list = termini_list) if termini_list else glycan_to_nxGraph(motif_list.loc[motif_list.motif_name == motif_string].motif.values.tolist()[0])
    except:
      g2 = glycan_to_nxGraph(motif_string, termini = 'provided', termini_list = termini_list) if termini_list else glycan_to_nxGraph(motif_string)

  g1_node_labels = nx.get_node_attributes(g1, 'string_labels')
  relevant_labels = set(list(g1_node_labels.values()) + list(nx.get_node_attributes(g2, 'string_labels').values()))
  narrow_wildcard_list = {k:[j for j in get_possible_linkages(k)] for k in relevant_labels if '?' in k}
  narrow_wildcard_list2 = {k:[j for j in get_possible_monosaccharides(k)] for k in relevant_labels if k in ['Hex', 'HexNAc', 'dHex', 'Sia', 'HexA', 'Pen', 'Monosaccharide'] or '!' in k}
  narrow_wildcard_list = {**narrow_wildcard_list, **narrow_wildcard_list2}

  if termini_list or narrow_wildcard_list:
    graph_pair = nx.algorithms.isomorphism.GraphMatcher(g_tmp, g2, node_match = categorical_node_match_wildcard('string_labels', 'unknown', narrow_wildcard_list, 'termini', 'flexible'))
  else:
    graph_pair = nx.algorithms.isomorphism.GraphMatcher(g_tmp, g2, node_match = nx.algorithms.isomorphism.categorical_node_match('string_labels', 'unknown'))
  graph_pair.subgraph_is_isomorphic()
  mapping = graph_pair.mapping
  mapping = {v: k for k, v in mapping.items()}
  mapping_show = {v: 'show' for v in list(graph_pair.mapping.keys())}

  while list(graph_pair.mapping.keys()) != []:
    g_tmp.remove_nodes_from(graph_pair.mapping.keys())
    if termini_list or narrow_wildcard_list:
      graph_pair = nx.algorithms.isomorphism.GraphMatcher(g_tmp, g2, node_match = categorical_node_match_wildcard('string_labels', 'unknown', narrow_wildcard_list, 'termini', 'flexible'))
    else:
      graph_pair = nx.algorithms.isomorphism.GraphMatcher(g_tmp, g2, node_match = nx.algorithms.isomorphism.categorical_node_match('string_labels', 'unknown'))
    graph_pair.subgraph_is_isomorphic()
    mapping = graph_pair.mapping
    mapping = {v: k for k, v in mapping.items()}
    mapping_show.update({v: 'show' for v in list(graph_pair.mapping.keys())})

  mapping_hide = {v: 'hide' for v in list(set(list(g1_node_labels.keys())) - set(list(mapping_show.keys())))}
  mapping_show.update(mapping_hide)
  
  nx.set_node_attributes(g1, dict(sorted(mapping_show.items())), 'highlight_labels')

  return g1


def process_repeat(repeat):
  """prepare input for repeat unit\n
  | Arguments:
  | :-
  | repeat (string): Repeat unit in IUPAC-condensed format, terminating either in linkage connecting units or first monosaccharide of the unit\n
  | Returns:
  | :-
  | processed repeat unit in GlycoDraw-ready IUPAC-condensed format\n
  """
  backbone = re.findall(r'.*\((?!.*\()', repeat)[0]
  repeat_connection = re.sub(r'\)(.*)', '', re.sub(r'.*\((?!.*\()', '', repeat))
  return 'blank(?1-' + repeat_connection[-1] + ')' + backbone + repeat_connection[:2] + '-?)'


def get_coordinates_and_labels(draw_this, highlight_motif, show_linkage = True, termini_list = []):
  """Extract monosaccharide labels and calculate coordinates for drawing\n
  | Arguments:
  | :-
  | draw_this (string): Glycan structure to be drawn.
  | highlight_motif (string): Glycan as named motif or in IUPAC-condensed format.
  | show_linkage (bool, optional): Flag indicating whether to show linkages. Default: True.
  | termini_list (list): list of monosaccharide positions (from 'terminal', 'internal', and 'flexible')\n
  | Returns:
  | :-
  | data_combined (list of lists)
  | contains lists with monosaccharide label, x position, y position, modification, bond, conformation
  """
  if not draw_this.startswith('['):
    draw_this = multiple_branches(multiple_branch_branches(draw_this))
    draw_this = reorder_for_drawing(draw_this)
    draw_this = multiple_branch_branches(multiple_branches(draw_this))

  graph = glycan_to_nxGraph(draw_this, termini = 'calc') if termini_list else glycan_to_nxGraph(draw_this)
  graph = get_highlight_attribute(graph, highlight_motif, termini_list = termini_list)
  node_labels = nx.get_node_attributes(graph, 'string_labels')
  highlight_labels = nx.get_node_attributes(graph, 'highlight_labels')
  edges = list(graph.edges())
  glycan = glycan_to_skeleton(draw_this)

  # Split main & branches, get labels
  levels = [2, 1, 0]
  parts = [[] for _ in levels]

  if ']' in glycan:
    glycan = glycan.replace('[', '(').replace(']', ')')
    for k, lev in enumerate(levels):
      for openpos, closepos, level in matches(glycan):
        if level == lev:
          parts[k].append(glycan[openpos:closepos])
          glycan = glycan[:openpos-1] + len(glycan[openpos-1:closepos+1])*'*' + glycan[closepos+1:]
      glycan = glycan.replace('*', '')
    parts = [[[i for i in k.split('-') if i] for k in part] for part in parts]

  main_node = glycan.split('-')
  main_node = [k for k in main_node if k != '']
  node_values = list(node_labels.values())
  highlight_values = list(highlight_labels.values())
  graph_nodes = list(graph.nodes())
  branch_branch_branch_node, branch_branch_node, branch_node = parts

  branch_label, branch_highlight = [], []
  for k in range(len(branch_node)):
    branch_label.append([node_values[j] for j in range(len(graph_nodes)) if graph_nodes[j] in map(int, branch_node[k])])
    branch_highlight.append([highlight_values[j] for j in range(len(graph_nodes)) if graph_nodes[j] in map(int, branch_node[k])])
  branch_branch_label, branch_branch_highlight = [], []
  for k in range(len(branch_branch_node)):
    branch_branch_label.append([node_values[j] for j in range(len(graph_nodes)) if graph_nodes[j] in map(int, branch_branch_node[k])])
    branch_branch_highlight.append([highlight_values[j] for j in range(len(graph_nodes)) if graph_nodes[j] in map(int, branch_branch_node[k])])
  bbb_label, bbb_highlight = [], []
  for k in range(len(branch_branch_branch_node)):
    bbb_label.append([node_values[j] for j in range(len(graph_nodes)) if graph_nodes[j] in map(int, branch_branch_branch_node[k])])
    bbb_highlight.append([highlight_values[j] for j in range(len(graph_nodes)) if graph_nodes[j] in map(int, branch_branch_branch_node[k])])
  main_label = [node_values[j] for j in range(len(graph_nodes)) if graph_nodes[j] in map(int, main_node)]
  main_highlight = [highlight_values[j] for j in range(len(graph_nodes)) if graph_nodes[j] in map(int, main_node)]

  # Split in monosaccharide & linkages
  main_sugar, main_sugar_modification, main_bond = split_monosaccharide_linkage(main_label)
  branch_sugar, branch_sugar_modification, branch_bond = split_monosaccharide_linkage(branch_label)
  branch_branch_sugar, branch_branch_sugar_modification, branch_branch_bond = split_monosaccharide_linkage(branch_branch_label)
  bbb_sugar, bbb_sugar_modification, bbb_bond = split_monosaccharide_linkage(bbb_label)

  main_sugar_label, _, main_bond_label = split_monosaccharide_linkage(main_highlight)
  branch_sugar_label, _, branch_bond_label = split_monosaccharide_linkage(branch_highlight)
  branch_branch_sugar_label, _, branch_branch_bond_label = split_monosaccharide_linkage(branch_branch_highlight)
  bbb_sugar_label, _, bbb_bond_label = split_monosaccharide_linkage(bbb_highlight)

  # Process linkages
  main_bond = process_bonds(main_bond)
  branch_bond = process_bonds(branch_bond)
  branch_branch_bond = process_bonds(branch_branch_bond)
  bbb_bond = process_bonds(bbb_bond)

  # Get connectivity
  branch_connection = []
  for x in branch_node:
    branch_connection = branch_connection + [edges[k][1] for k in range(len(edges)) if edges[k][0] == int(x[-1])]
  branch_connection = unwrap(get_indices(main_node[::2][::-1], [str(k) for k in branch_connection]))

  branch_node_old = branch_node
  if '?' not in [k[0] for k in branch_bond]:
    sorted_indices = np.argsort(branch_connection)[::-1]
    branch_sugar = [branch_sugar[k] for k in sorted_indices]
    branch_sugar_label = [branch_sugar_label[k] for k in sorted_indices]
    branch_sugar_modification = [branch_sugar_modification[k] for k in sorted_indices]
    branch_bond = [branch_bond[k] for k in sorted_indices]
    branch_bond_label = [branch_bond_label[k] for k in sorted_indices]
    branch_node = [branch_node[k] for k in sorted_indices]
    branch_connection = [branch_connection[k] for k in sorted_indices]

  branch_branch_connection = []
  for x in branch_branch_node:
    branch_branch_connection = branch_branch_connection + [edges[k][1] for k in range(len(edges)) if edges[k][0] == int(x[-1])]
  tmp = []
  for k in branch_branch_connection:
    tmp.append([(i, color.index(str(k))) for i, color in enumerate([k[::2][::-1] for k in branch_node]) if str(k) in color])
  branch_branch_connection = unwrap(tmp)

  bbb_connection = []
  for x in branch_branch_branch_node:
    bbb_connection = bbb_connection + [edges[k][1] for k in range(len(edges)) if edges[k][0] == int(x[-1])]
  tmp = []
  for k in bbb_connection:
    tmp.append([(i, color.index(str(k))) for i, color in enumerate([k[::2][::-1] for k in branch_branch_node]) if str(k) in color])
  bbb_connection = unwrap(tmp)

  # Order multiple connections on a branch level
  new_order = []
  for k in sorted(set(branch_connection), reverse = True):
    idx = unwrap(get_indices(branch_connection, [k]))
    if len(idx) == 1:
      new_order.extend(idx)
    else:
      new_order.extend([idx[i] for i in np.argsort([k[0][-1] for k in [j for j in [branch_bond[k] for k in idx]]])])
  
  branch_sugar = [branch_sugar[i] for i in new_order]
  branch_sugar_label = [branch_sugar_label[i] for i in new_order]
  branch_sugar_modification = [branch_sugar_modification[i] for i in new_order]
  branch_bond = [branch_bond[i] for i in new_order]
  branch_bond_label = [branch_bond_label[i] for i in new_order]
  branch_node = [branch_node[i] for i in new_order]
  branch_connection = [branch_connection[i] for i in new_order]

  for k in range(len(branch_branch_connection)):
    tmp = get_indices(new_order, [branch_branch_connection[k][0]])
    branch_branch_connection[k] = (tmp[0][0], branch_branch_connection[k][1])

  # Order multiple connections on a branch_branch level
  new_order = []
  for k in sorted(set(branch_branch_connection), reverse = True):
    idx = unwrap(get_indices(branch_branch_connection, [k]))
    if len(idx) == 1:
      new_order.extend(idx)
    else:
      new_order.extend([idx[i] for i in np.argsort([k[0][-1] for k in [j for j in [branch_branch_bond[k] for k in idx]]])])

  branch_branch_sugar = [branch_branch_sugar[i] for i in new_order]
  branch_branch_sugar_label = [branch_branch_sugar_label[i] for i in new_order]
  branch_branch_sugar_modification = [branch_branch_sugar_modification[i] for i in new_order]
  branch_branch_bond = [branch_branch_bond[i] for i in new_order]
  branch_branch_bond_label = [branch_branch_bond_label[i] for i in new_order]
  branch_branch_node = [branch_branch_node[i] for i in new_order]
  branch_branch_connection = [branch_branch_connection[i] for i in new_order]

  # Main chain x y
  main_length = len(main_sugar)
  main_sugar_x_pos = list(range(main_length))
  main_sugar_y_pos = [0] * main_length
  if main_sugar[-1] in {'Fuc', 'Xyl'} and branch_sugar:
    main_sugar_x_pos[-1] -= 1
    main_sugar_y_pos[-1] += -2 if main_length != 2 else 2

  # Branch x
  branch_x_pos = []
  for j, branch in enumerate(branch_sugar):
    start_x = main_sugar_x_pos[branch_connection[j]]
    tmp = [start_x + 1 + k for k in range(len(branch))]
    if branch[-1] in {'Fuc', 'Xyl'}:
        tmp[-1] -= 1
    branch_x_pos.append(tmp)

  # Branch branch x
  branch_branch_x_pos = []
  for j, branch_branch in enumerate(branch_branch_sugar):
    start_x = branch_x_pos[branch_branch_connection[j][0]][branch_branch_connection[j][1]]
    tmp = [start_x + 1 + k for k in range(len(branch_branch))]
    if branch_branch[-1] in {'Fuc', 'Xyl'}:
      tmp[-1] -= 1
    branch_branch_x_pos.append(tmp)

  # Branch branch branch x
  bbb_x_pos = []
  for j, bbb in enumerate(bbb_sugar):
    start_x = branch_branch_x_pos[bbb_connection[j][0]][bbb_connection[j][1]]
    tmp = [start_x + 1 + k for k in range(len(bbb))]
    if bbb[-1] in {'Fuc', 'Xyl'}:
      tmp[-1] -= 1
    bbb_x_pos.append(tmp)

  # Branch y
  branch_y_pos = [[0 for _ in y] for y in branch_x_pos]
  branch_branch_y_pos = [[0 for _ in y] for y in branch_branch_x_pos]
  bbb_y_pos = [[0 for _ in y] for y in bbb_x_pos]
  counter = 0
  for k, bx_pos in enumerate(branch_x_pos):
    pos_value = main_sugar_y_pos[branch_connection[k]]
    # Branch terminating in fucose
    if len(branch_sugar[k]) > 1 and branch_sugar[k][-1] in {'Fuc', 'Xyl'}:
      tmp = [pos_value+2+counter for x in bx_pos]
      tmp[-1] = tmp[-1]-2
      branch_y_pos[k] = tmp
      counter += 2
    # Remaining branches longer than one sugar
    elif branch_sugar[k] not in [['Fuc'], ['Xyl']] and len(branch_sugar[k]) > 1:
      if main_sugar[-1] not in {'Fuc', 'Xyl'}:
        branch_y_pos[k] = [pos_value+2+counter for x in bx_pos]
        counter += 2
      elif len(main_sugar) - (branch_connection[k]+1) == 1:
        branch_y_pos[k] = [pos_value+counter for x in bx_pos]
        counter += 2
      else:
        branch_y_pos[k] = [pos_value+2+counter for x in bx_pos]
        counter += 2
    # Core fucose
    elif branch_sugar[k] in [['Fuc'], ['Xyl']] and branch_connection[k] == 0:
      if main_sugar_modification[0] != '' or branch_bond[k][0][-1] == '3':
        branch_y_pos[k] = [pos_value-2 for x in bx_pos]
      else:
        branch_y_pos[k] = [pos_value+2 for x in bx_pos]
    # One monosaccharide branches
    elif len(branch_sugar[k]) == 1 and branch_sugar[k][0] not in {'Fuc', 'Xyl'}:
      if main_sugar[-1] not in {'Fuc', 'Xyl'}:
        branch_y_pos[k] = [pos_value+2+counter for x in bx_pos]
        counter += 2
      elif len(main_sugar) - (branch_connection[k]+1) == 1:
        branch_y_pos[k] = [pos_value+counter for x in bx_pos]
        counter += 2
      else:
        branch_y_pos[k] = [pos_value+2+counter for x in bx_pos]
        counter += 2
    # Fucose not on core
    else:
      branch_y_pos[k] = [pos_value-2 for x in bx_pos]

  # Branch branch y
  counter = 0
  for k, bbx_pos in enumerate(branch_branch_x_pos):
    pos_value = branch_y_pos[branch_branch_connection[k][0]][branch_branch_connection[k][1]]
    # Branch branch terminating in fucose
    if len(branch_branch_sugar[k]) > 1 and branch_branch_sugar[k][-1] in {'Fuc', 'Xyl'}:
      tmp = [pos_value+2+counter for x in bbx_pos]
      tmp[-1] = tmp[-1]-2
      branch_branch_y_pos[k] = tmp
    elif branch_node[branch_branch_connection[k][0]][::2][::-1][-2] == branch_node[branch_branch_connection[k][0]][::2][::-1][branch_branch_connection[k][1]] and branch_sugar[branch_branch_connection[k][0]][-1] == 'Fuc':
      branch_branch_y_pos[k] = [pos_value for x in bbx_pos]
    elif len(branch_branch_sugar[k]) > 1 and branch_branch_sugar[k][-1] not in {'Fuc', 'Xyl'}:
      branch_branch_y_pos[k] = [pos_value+2+counter for x in bbx_pos]
      counter += 2
    elif branch_branch_sugar[k] == ['GlcNAc']:
      branch_branch_y_pos[k] = [pos_value+2+counter for x in bbx_pos]
    elif branch_branch_sugar[k] in [['Fuc'], ['Xyl']]:
      branch_branch_y_pos[k] = [pos_value-2+counter for x in bbx_pos]
    elif min(bbx_pos) > max(branch_x_pos[branch_branch_connection[k][0]]):
      branch_branch_y_pos[k] = [pos_value+counter for x in bbx_pos]
    else:
      branch_branch_y_pos[k] = [pos_value+2+counter for x in bbx_pos]

  # Branch branch branch y
  for k, bbbx_pos in enumerate(bbb_x_pos):
    pos_value = branch_branch_y_pos[bbb_connection[k][0]][bbb_connection[k][1]]
    # Branch branch terminating in fucose
    if len(bbb_sugar[k]) > 1 and bbb_sugar[k][-1] in {'Fuc', 'Xyl'}:
      tmp = [branch_branch_y_pos[bbb_connection[k][0]][bbb_connection[k][1]]+2 for x in bbbx_pos]
      tmp[-1] = tmp[-1]-2
      bbb_y_pos[k] = tmp
    elif len(bbb_sugar[k]) > 1 and bbb_sugar[k][-1] not in {'Fuc', 'Xyl'}:
      bbb_y_pos[k] = [pos_value+2 for x in bbbx_pos]
    elif bbb_sugar[k] == ['GlcNAc']:
      bbb_y_pos[k] = [pos_value+2 for x in bbbx_pos]
    elif bbb_sugar[k] in [['Fuc'], ['Xyl']]:
      bbb_y_pos[k] = [pos_value-2 for x in bbbx_pos]
    elif min(bbbx_pos) > max(branch_branch_x_pos[bbb_connection[k][0]]):
      bbb_y_pos[k] = [pos_value for x in bbbx_pos]
    else:
      bbb_y_pos[k] = [pos_value+2 for x in bbbx_pos]

  # Adjust y spacing between branches
  splits = [main_node[::2][::-1][k] for k in [branch_connection[k] for k in unwrap(get_indices([k[::2][::-1] for k in branch_node], [k for k in [k[::2][::-1] for k in branch_node]]))]]
  tmp = unwrap(get_indices([k[::2][::-1] for k in branch_branch_node], [k for k in [k[::2][::-1] for k in branch_branch_node]]))
  for k in tmp:
    splits.append([j[::2][::-1] for j in branch_node][branch_branch_connection[k][0]][branch_branch_connection[k][1]])

  splits = unique(splits)
  filtery = []
  for k, sugar in enumerate(branch_sugar):
    if sugar in [['Fuc'], ['Xyl']]:
      filtery.append(main_node[::2][::-1][branch_connection[k]])
    if sugar[-1] == 'Fuc' and len(sugar) > 1:
      filtery.append(branch_node[k][::2][::-1][-2]) 
  for k, sugar in enumerate(branch_branch_sugar):
    if sugar in [['Fuc'], ['Xyl']]:
      filtery.append([j[::2][::-1] for j in branch_node][branch_branch_connection[k][0]][branch_branch_connection[k][1]])
  splits = [k for k in splits if k not in filtery]

  tmp_a = main_node + unwrap(branch_node_old) + unwrap(branch_branch_node) + unwrap(branch_branch_branch_node)
  tmp_b = main_label + unwrap(branch_label) + unwrap(branch_branch_label) + unwrap(bbb_label)
  for n in splits:
    graph = glycan_to_nxGraph(draw_this)
    graph2 = split_node(graph, int(n))
    edges = graph.edges()
    split_node_connections = [e[0] for e in edges if f"{n}_"in str(e[1])]
    node_crawl = [k for k in [list(nx.node_connected_component(graph2, k)) for k in split_node_connections] if int(main_node[-1]) not in k]
    new_node_crawl = [[x for x in k if '_' not in str(x)] for k in node_crawl]

    final_linkage = [tmp_b[unwrap(get_indices(tmp_a, [str(k[-1])]))[0]] for k in new_node_crawl]
    final_linkage = [k[-1] for k in final_linkage]
    new_node_crawl = [new_node_crawl[i] for i in np.argsort(final_linkage)]
    pairwise_node_crawl = list(zip(new_node_crawl, new_node_crawl[1:]))

    base_list = main_node[::2][::-1]
    branch_node_list = unwrap([k[::2][::-1] for k in branch_node_old])
    branch_branch_node_list = unwrap([k[::2][::-1] for k in branch_branch_node])
    branch_branch_branch_node_list = unwrap([k[::2][::-1] for k in branch_branch_branch_node])
    node_list = base_list + branch_node_list + branch_branch_node_list + branch_branch_branch_node_list
    y_list = main_sugar_y_pos + unwrap(branch_y_pos) + unwrap(branch_branch_y_pos) + unwrap(bbb_y_pos)
    x_list = main_sugar_x_pos + unwrap(branch_x_pos) + unwrap(branch_branch_x_pos) + unwrap(bbb_x_pos)

    for pair in pairwise_node_crawl:
      idx_A = [k for k in get_indices(node_list, [str(k) for k in pair[0]]) if k != [None]]
      idx_B = [k for k in get_indices(node_list, [str(k) for k in pair[1]]) if k != [None]]
      upper, lower = (pair[0], pair[1]) if max(y_list[k[0]] for k in idx_A) > max(y_list[k[0]] for k in idx_B) else (pair[1], pair[0])
      upper_min = min(y_list[k[0]] for k in (idx_A if upper == pair[0] else idx_B))
      lower_max = max(y_list[k[0]] for k in (idx_B if upper == pair[0] else idx_A))

      to_add = 2 - (upper_min-lower_max)
      if main_sugar[-1] not in {'Fuc', 'Xyl'} or len(main_sugar) != 2:
        for k in range(len(branch_y_pos)):
          for j in range(len(branch_y_pos[k])):
            if [k[::2][::-1] for k in branch_node][k][j] in [str(k) for k in upper] and branch_sugar[k] not in [['Fuc'], ['Xyl']]:
              branch_y_pos[k][j] = branch_y_pos[k][j] + to_add
              
      for k in range(len(branch_branch_y_pos)):
        for j in range(len(branch_branch_y_pos[k])):
          if [k[::2][::-1] for k in branch_branch_node][k][j] in [str(k) for k in upper]:
            branch_branch_y_pos[k][j] = branch_branch_y_pos[k][j] + to_add

      for k in range(len(bbb_y_pos)):
        for j in range(len(bbb_y_pos[k])):
          if [k[::2][::-1] for k in branch_branch_branch_node][k][j] in [str(k) for k in upper]:
            bbb_y_pos[k][j] = bbb_y_pos[k][j] + to_add

  # Adjust y branch_branch connections
  for j, conn in enumerate(unique(branch_branch_connection)):
    if branch_branch_sugar[j] not in [['Fuc'], ['Xyl']] and max(branch_x_pos[branch_branch_connection[j][0]]) >= branch_branch_x_pos[j][0]:
      tmp = [branch_branch_y_pos[j][0] for j in unwrap(get_indices(branch_branch_connection, [conn]))]
      y_adj = (max(tmp)-branch_y_pos[branch_branch_connection[j][0]][branch_branch_connection[j][1]+1])/2
      # For each branch
      for k in range(len(branch_x_pos)):
        # If connected
        if k == branch_branch_connection[j][0]:
          # And if smaller/equal x
          for n in range(len(branch_x_pos[k])):
            if branch_x_pos[k][n] <= branch_x_pos[branch_branch_connection[j][0]][branch_branch_connection[j][1]]:
              branch_y_pos[k][n] = branch_y_pos[k][n] + y_adj
      # For each branch branch
      for k in range(len(branch_branch_x_pos)):
        # If connected
        if branch_branch_connection[k][0] == branch_branch_connection[j][0] and branch_branch_connection[k][1] == branch_branch_connection[j][1]:
          # And if smaller/equal x
          for n in range(len(branch_branch_x_pos[k])):
            if branch_branch_x_pos[k][n] <= branch_x_pos[branch_branch_connection[j][0]][branch_branch_connection[j][1]]:
              branch_branch_y_pos[k][n] = branch_branch_y_pos[k][n] + y_adj

  # Adjust y branch connections
  # print(branch_connection)
  for k in range(len(unique(branch_connection))):
    tmp = [branch_y_pos[j][0] for j in unwrap(get_indices(branch_connection, [unique(branch_connection)[k]]))]
    if ['Fuc'] in [branch_sugar[j] for j in unwrap(get_indices(branch_connection, [unique(branch_connection)[k]]))] and branch_connection.count(unique(branch_connection)[k]) < 2 or ['Fuc'] in [branch_sugar[j] for j in unwrap(get_indices(branch_connection, [unique(branch_connection)[k]]))] and branch_connection.count(0) > 1:  # and list(set(unwrap([branch_sugar[k] for k in unwrap(get_indices(unwrap(branch_sugar), ['Fuc']))]))) == ['Fuc']:
      y_adj = 0
    elif ['Xyl'] in [branch_sugar[j] for j in unwrap(get_indices(branch_connection, [unique(branch_connection)[k]]))] and branch_connection.count(unique(branch_connection)[k]) < 2:
      y_adj = 0
    else:
      y_adj = (max(tmp)-main_sugar_y_pos[unique(branch_connection)[k]])/2
    for j in range(len(main_sugar_x_pos)):
      if main_sugar_x_pos[j] <= main_sugar_x_pos[unique(branch_connection)[k]]:
        main_sugar_y_pos[j] += y_adj
      else:
        pass
    for j in range(len(branch_x_pos)):
      if branch_connection[j] == unique(branch_connection)[k] or branch_sugar[j] in [['Fuc'], ['Xyl']]:
        for n in range(len(branch_x_pos[j])):
          if branch_x_pos[j][n] <= main_sugar_x_pos[unique(branch_connection)[k]]:
            branch_y_pos[j][n] += y_adj

  # Fix for handling 'wrong' structures with the core fucose in the main chain
  if main_sugar[-1] in {'Fuc', 'Xyl'} and len(main_sugar) == 2 and branch_sugar != []:
    to_add = branch_y_pos[0][0] - main_sugar_y_pos[0]
    main_sugar_y_pos = [k + to_add for k in main_sugar_y_pos]

  # Fix spacing
  splits = [k for k in splits if k not in filtery]
  tmp_a = main_node + unwrap(branch_node_old) + unwrap(branch_branch_node) + unwrap(branch_branch_branch_node)
  tmp_b = main_label + unwrap(branch_label) + unwrap(branch_branch_label) + unwrap(bbb_label)
  base_list = main_node[::2][::-1]
  branch_node_list = unwrap([k[::2][::-1] for k in branch_node_old])
  branch_branch_node_list = unwrap([k[::2][::-1] for k in branch_branch_node])
  branch_branch_branch_node_list = unwrap([k[::2][::-1] for k in branch_branch_branch_node])
  node_list = base_list + branch_node_list + branch_branch_node_list + branch_branch_branch_node_list
  y_list = main_sugar_y_pos+unwrap(branch_y_pos)+unwrap(branch_branch_y_pos)+unwrap(bbb_y_pos)
  x_list = main_sugar_x_pos+unwrap(branch_x_pos)+unwrap(branch_branch_x_pos)+unwrap(bbb_x_pos)
  for n in splits:
    graph = glycan_to_nxGraph(draw_this)
    graph2 = split_node(graph, int(n))
    edges = graph.edges()
    split_node_connections = [e[0] for e in edges if f"{n}_" in str(e[1])]
    node_crawl = [k for k in [list(nx.node_connected_component(graph2, k)) for k in split_node_connections] if int(main_node[-1]) not in k]
    anti_node_crawl = [k for k in [list(nx.node_connected_component(graph2, k)) for k in split_node_connections] if int(main_node[-1]) in k]
    anti_node_crawl = [re.sub(r"_\S*$", '', str(k)) for k in unwrap(anti_node_crawl)]
    new_node_crawl = [[x for x in k if '_' not in str(x)] for k in node_crawl]
    final_linkage = [tmp_b[unwrap(get_indices(tmp_a, [str(k[-1])]))[0]] for k in new_node_crawl]
    final_linkage = [k[-1] for k in final_linkage]
    new_node_crawl = [new_node_crawl[i] for i in np.argsort(final_linkage)]
    pairwise_node_crawl = list(zip(new_node_crawl, new_node_crawl[1:]))

    for pair in pairwise_node_crawl:
      idx_A = [k for k in get_indices(node_list, [str(k) for k in pair[0]]) if k != [None]]
      idx_B = [k for k in get_indices(node_list, [str(k) for k in pair[1]]) if k != [None]]
      upper, lower = (pair[0], pair[1]) if max(y_list[k[0]] for k in idx_A) > max(y_list[k[0]] for k in idx_B) else (pair[1], pair[0])
      upper_min = min(y_list[k[0]] for k in (idx_A if upper == pair[0] else idx_B))
      lower_max = max(y_list[k[0]] for k in (idx_B if upper == pair[0] else idx_A))

      if max([y_list[k[0]] for k in idx_A]) > max([y_list[k[0]] for k in idx_B]):
        upper_y, upper_x = [y_list[k[0]] for k in idx_A], [x_list[k[0]] for k in idx_A]
        lower_y, lower_x = [y_list[k[0]] for k in idx_B], [x_list[k[0]] for k in idx_B]
      else:
        upper_y, upper_x = [y_list[k[0]] for k in idx_B], [x_list[k[0]] for k in idx_B]
        lower_y, lower_x = [y_list[k[0]] for k in idx_A], [x_list[k[0]] for k in idx_A]

      diff_to_fix = []
      for x_cor in list(set(upper_x)):
        if x_cor in list(set(lower_x)):
          min_y_upper = min([upper_y[k] for k in unwrap(get_indices(upper_x, [x_cor]))])
          max_y_lower = max([lower_y[k] for k in unwrap(get_indices(lower_x, [x_cor]))])
          diff_to_fix.append(2 - (min_y_upper-max_y_lower))
      if diff_to_fix:
        to_add = max(diff_to_fix)

      str_upper = [str(k) for k in upper]
      if main_sugar[-1] != 'Fuc':
        for k in range(len(branch_y_pos)):
          for j in range(len(branch_y_pos[k])):
            if [k[::2][::-1] for k in branch_node][k][j] in str_upper:
              branch_y_pos[k][j] += to_add
            if branch_x_pos[k][j] == 0:
                branch_y_pos[k][j] += (to_add/2)
        tmp_listy = []
        for k in range(len(main_sugar)):
          if main_sugar_x_pos[k] < min([x for x in unwrap(branch_x_pos) if x > 0]):
            tmp_listy.append(main_sugar_x_pos[k])
        for k in range(len(tmp_listy)):
          main_sugar_y_pos[k] += (to_add/2)

      for list_to_update in [branch_branch_y_pos, bbb_y_pos]:
        for k in range(len(list_to_update)):
          for j in range(len(list_to_update[k])):
            if [k[::2][::-1] for k in (branch_branch_node if list_to_update is branch_branch_y_pos else branch_branch_branch_node)][k][j] in str_upper:
              list_to_update[k][j] += to_add

  main_conf = [k.group()[0] if k is not None else '' for k in [re.search('^L-|^D-', k) for k in main_sugar_modification]]
  main_sugar_modification = [re.sub('^L-|^D-', '', k) for k in main_sugar_modification]

  b_conf = [[k.group()[0] if k is not None else '' for k in j] for j in [[re.search('^L-|^D-', k) for k in j] for j in branch_sugar_modification]]
  branch_sugar_modification = [[re.sub('^L-|^D-', '', k) for k in j] for j in branch_sugar_modification]

  bb_conf = [[k.group()[0] if k is not None else '' for k in j] for j in [[re.search('^L-|^D-', k) for k in j] for j in branch_branch_sugar_modification]]
  branch_branch_sugar_modification = [[re.sub('^L-|^D-', '', k) for k in j] for j in branch_branch_sugar_modification]

  bbb_conf = [[k.group()[0] if k is not None else '' for k in j] for j in [[re.search('^L-|^D-', k) for k in j] for j in bbb_sugar_modification]]
  bbb_sugar_modification = [[re.sub('^L-|^D-', '', k) for k in j] for j in bbb_sugar_modification]

  data_combined = [
      [main_sugar, main_sugar_x_pos, main_sugar_y_pos, main_sugar_modification, main_bond, main_conf, main_sugar_label, main_bond_label],
      [branch_sugar, branch_x_pos, branch_y_pos, branch_sugar_modification, branch_bond, branch_connection, b_conf, branch_sugar_label, branch_bond_label],
      [branch_branch_sugar, branch_branch_x_pos, branch_branch_y_pos, branch_branch_sugar_modification, branch_branch_bond, branch_branch_connection, bb_conf, branch_branch_sugar_label, branch_branch_bond_label],
      [bbb_sugar, bbb_x_pos, bbb_y_pos, bbb_sugar_modification, bbb_bond, bbb_connection, bbb_conf, bbb_sugar_label, bbb_bond_label]
  ]
  return data_combined


def draw_bracket(x, y_min_max, direction = 'right', dim = 50,  highlight = 'show', deg = 0):
  """Draw a bracket shape at the specified position and dimensions.\n
  | Arguments:
  | :-
  | x (int): X coordinate of the bracket on the drawing canvas.
  | y_min_max (list): List containing the minimum and maximum Y coordinates for the bracket.
  | direction (string, optional): Direction of the bracket. Possible values are 'right' and 'left'. Default: 'right'.
  | dim (int, optional): Arbitrary dimension unit used for scaling the bracket's size. Default: 50.\n
  | Returns:
  | :-
  | None
  """
  if highlight == 'hide':
    col_dict = col_dict_transparent
  else:
    col_dict = col_dict_base
  
  stroke_opts = {'stroke_width': 0.04 * dim, 'stroke': col_dict['black']}
  x_common = 0 - (x * dim)
  y_min = 0 + (y_min_max[0] * dim) - 0.75 * dim
  y_max = 0 + (y_min_max[1] * dim) + 0.75 * dim
  # rot = 'rotate(' + str(deg) + ' ' + str(0-abs(x*2)*(dim)) + ' ' + str(0-abs(y_min)*(dim)) + ')'
  rot = 'rotate(' + str(deg) + ' ' + str(x_common) + ' ' + str(np.mean(y_min_max)) + ')'
  # Vertical

  g = draw.Group(transform = rot)
  p = draw.Path(**stroke_opts)
  p.M(x_common, y_max)
  p.L(x_common, y_min)
  # d.append(p)
  g.append(p)
  dir_mult = 1 if direction == 'right' else -1
  for y in [y_min, y_max]:
    p = draw.Path(**stroke_opts)
    p.M(x_common - (0.02 * dim * dir_mult), y)
    p.L(x_common + (0.25 * dim * dir_mult), y)
    # d.append(p)
    g.append(p)
  d.append(g)


@rescue_glycans
def GlycoDraw(draw_this, vertical = False, compact = False, show_linkage = True, dim = 50, highlight_motif = None, highlight_termini_list = [], repeat = None, repeat_range = None, filepath = None):
  """Draws a glycan structure based on the provided input.\n
  | Arguments:
  | :-
  | draw_this (string): The glycan structure or motif to be drawn.
  | vertical (bool, optional): Set to True to draw the structure vertically. Default: False.
  | compact (bool, optional): Set to True to draw the structure in a compact form. Default: False.
  | show_linkage (bool, optional): Set to False to hide the linkage information. Default: True.
  | dim (int, optional): The dimension (size) of the individual sugar units in the structure. Default: 50.
  | highlight_motif (string, optional): Glycan motif to highlight within the parent structure.
  | highlight_termini_list (list): list of monosaccharide positions (from 'terminal', 'internal', and 'flexible')
  | repeat (bool | int | str): If specified, indicate repeat unit by brackets (True: n units, int: # of units, str: range of units)
  | repeat_range (list of 2 int): List of index integers for the first and last main-chain monosaccharide in repeating unit. Monosaccharides are numbered starting from 0 (invisible placeholder = 0 in case of structure terminating in a linkage) at the reducing end. 
  | filepath (string, optional): The path to the output file to save as SVG or PDF. Default: None.\n
  """
  if any([k in draw_this for k in [';', '-D-', 'RES', '=']]):
    raise Exception
  bond_hack = False
  if 'Man(a1-?)' in draw_this and 'Man(a1-3)' not in draw_this and 'Man(a1-6)' not in draw_this:
    draw_this = 'Man(a1-6)'.join(draw_this.rsplit('Man(a1-?)', 1))
    bond_hack = True
  if repeat and not repeat_range:
    draw_this = process_repeat(draw_this)
  if draw_this[-1] == ')':
    draw_this += 'blank'
  if compact:
    show_linkage = False
  if isinstance(highlight_motif, str) and highlight_motif[0] == 'r':
    temp = get_match(highlight_motif[1:], draw_this)
    highlight_motif = temp[0] if temp else None

  # Handle floaty bits if present
  floaty_bits = []
  for openpos, closepos, level in matches(draw_this, opendelim = '{', closedelim = '}'):
      floaty_bits.append(draw_this[openpos:closepos]+'blank')
      draw_this = draw_this[:openpos-1] + len(draw_this[openpos-1:closepos+1])*'*' + draw_this[closepos+1:]
  draw_this = draw_this.replace('*', '')

  try:
    data = get_coordinates_and_labels(draw_this, show_linkage = show_linkage, highlight_motif = highlight_motif, termini_list = highlight_termini_list)
  except:
    try:
      draw_this = motif_list.loc[motif_list.motif_name == draw_this].motif.values.tolist()[0]
      data = get_coordinates_and_labels(draw_this, show_linkage = show_linkage, highlight_motif = highlight_motif, termini_list = highlight_termini_list)
    except:
      return print('Error: did you enter a real glycan or motif?')
      ys.exit(1)

  main_sugar, main_sugar_x_pos, main_sugar_y_pos, main_sugar_modification, main_bond, main_conf, main_sugar_label, main_bond_label = data[0]
  branch_sugar, branch_x_pos, branch_y_pos, branch_sugar_modification, branch_bond, branch_connection, b_conf, branch_sugar_label, branch_bond_label = data[1]
  branch_branch_sugar, branch_branch_x_pos, branch_branch_y_pos, branch_branch_sugar_modification, branch_branch_bond, branch_branch_connection, bb_conf, branch_branch_sugar_label, branch_branch_bond_label = data[2]
  bbb_sugar, bbb_x_pos, bbb_y_pos, bbb_sugar_modification, bbb_bond, bbb_connection, bbb_conf, bbb_sugar_label, bbb_bond_label = data[3]

  while bond_hack:
    for k in range(len(main_bond)):
      if main_sugar[k] + '--' + main_bond[k] == 'Man--α 6':
        main_bond[k] = 'α'
        bond_hack = False

    for branch in range(len(branch_bond)):
      for bond in range(len(branch_bond[branch])):
        if branch_sugar[branch][bond] + '--' + branch_bond[branch][bond] == 'Man--α 6':
          branch_bond[branch][bond] = 'α'
          bond_hack = False
    bond_hack = False

  if not show_linkage:
    main_bond = ['-'] * len(main_bond)
    branch_bond = [['-' for _ in y] for y in branch_bond]
    branch_branch_bond = [['-' for _ in y] for y in branch_branch_bond]
    bbb_bond = [['-' for _ in y] for y in bbb_bond]

  # Fix for drawsvg 2.0 y
  main_sugar_y_pos = [-k for k in main_sugar_y_pos]
  branch_y_pos = [[-x for x in y] for y in branch_y_pos]
  branch_branch_y_pos = [[-x for x in y] for y in branch_branch_y_pos]
  bbb_y_pos = [[-x for x in y] for y in bbb_y_pos]

  # Calculate angles for main chain Y, Z fragments
  def calculate_degree(y1, y2, x1, x2):
    slope = -1 * (y2 - y1) / ((x2 * 2) - (x1 * 2))
    return degrees(atan(slope))
  
  main_deg = [calculate_degree(main_sugar_y_pos[k], main_sugar_y_pos[k-1], main_sugar_x_pos[k], main_sugar_x_pos[k-1])
              if sugar in {'Z', 'Y'} else 0 for k, sugar in enumerate(main_sugar)]

  # Calculate angles for branch Y, Z fragments
  branch_deg = []
  for k, sugars in enumerate(branch_sugar):
    branch_deg.append([
      calculate_degree(branch_y_pos[k][j], main_sugar_y_pos[branch_connection[k]], branch_x_pos[k][j], main_sugar_x_pos[branch_connection[k]])
      if sugar in {'Z', 'Y'} and len(sugars) == 1 else
      calculate_degree(branch_y_pos[k][j], branch_y_pos[k][j-1], branch_x_pos[k][j], branch_x_pos[k][j-1])
      if sugar in {'Z', 'Y'} else 0 for j, sugar in enumerate(sugars)
      ])

  # Calculate angles for branch_branch Y, Z fragments
  branch_branch_deg = []
  for k, sugars in enumerate(branch_branch_sugar):
    branch_branch_deg.append([
      calculate_degree(branch_branch_y_pos[k][j], branch_y_pos[branch_branch_connection[k][0]][branch_branch_connection[k][1]], branch_branch_x_pos[k][j], branch_x_pos[branch_branch_connection[k][0]][branch_branch_connection[k][1]])
      if sugar in {'Z', 'Y'} and len(sugars) == 1 else
      calculate_degree(branch_branch_y_pos[k][j], branch_branch_y_pos[k][j-1], branch_branch_x_pos[k][j], branch_branch_x_pos[k][j-1])
      if sugar in {'Z', 'Y'} else 0 for j, sugar in enumerate(sugars)
      ])

  # Adjust drawing dimensions
  max_y = max(unwrap(bbb_y_pos)+unwrap(branch_branch_y_pos)+unwrap(branch_y_pos)+main_sugar_y_pos)
  min_y = min(unwrap(bbb_y_pos)+unwrap(branch_branch_y_pos)+unwrap(branch_y_pos)+main_sugar_y_pos)
  max_x = max(unwrap(bbb_x_pos)+unwrap(branch_branch_x_pos)+unwrap(branch_x_pos)+main_sugar_x_pos)

  # Canvas size
  width = ((((max_x+1)*2)-1)*dim)+dim
  if floaty_bits:
    len_one_gw = ((max([len(j) for k in min_process_glycans(floaty_bits) for j in k]) / 6) + 1) * dim
    len_multiple_gw = (max([len(k) for k in min_process_glycans(floaty_bits)], default = 0)+1) * dim
    width += max([len_one_gw, len_multiple_gw])
  if len(floaty_bits) > len(set(floaty_bits)):
     width += dim
  height = ((((max(abs(min_y), max_y)+1)*2)-1)*dim)+10+50
  x_ori = -width+(dim/2)+0.5*dim
  y_ori = (-height/2)+(((max_y-abs(min_y))/2)*dim)

  global d
  # Draw
  d2 = draw.Drawing(width, height, origin = (x_ori, y_ori))
  deg = 90 if vertical else 0
  rot = f'rotate({deg} {width / 2} {height / 2})'
  d = draw.Group(transform = rot)

  # Bond main chain
  [add_bond(main_sugar_x_pos[k+1], main_sugar_x_pos[k], main_sugar_y_pos[k+1], main_sugar_y_pos[k], main_bond[k], dim = dim, compact = compact, highlight = main_bond_label[k]) for k in range(len(main_sugar)-1)]
  # Bond branch
  [add_bond(branch_x_pos[b_idx][s_idx+1], branch_x_pos[b_idx][s_idx], branch_y_pos[b_idx][s_idx+1], branch_y_pos[b_idx][s_idx], branch_bond[b_idx][s_idx+1], dim = dim, compact = compact, highlight = branch_bond_label[b_idx][s_idx+1]) for b_idx in range(len(branch_sugar)) for s_idx in range(len(branch_sugar[b_idx])-1) if len(branch_sugar[b_idx]) > 1]
  # Bond branch to main chain
  [add_bond(branch_x_pos[k][0], main_sugar_x_pos[branch_connection[k]], branch_y_pos[k][0], main_sugar_y_pos[branch_connection[k]], branch_bond[k][0], dim = dim, compact = compact, highlight = branch_bond_label[k][0]) for k in range(len(branch_sugar))]
  # Bond branch branch
  [add_bond(branch_branch_x_pos[b_idx][s_idx+1], branch_branch_x_pos[b_idx][s_idx], branch_branch_y_pos[b_idx][s_idx+1], branch_branch_y_pos[b_idx][s_idx], branch_branch_bond[b_idx][s_idx+1], dim = dim, compact = compact, highlight = branch_branch_bond_label[b_idx][s_idx+1]) for b_idx in range(len(branch_branch_sugar)) for s_idx in range(len(branch_branch_sugar[b_idx])-1) if len(branch_branch_sugar[b_idx]) > 1]
  # Bond branch branch branch
  [add_bond(bbb_x_pos[b_idx][s_idx+1], bbb_x_pos[b_idx][s_idx], bbb_y_pos[b_idx][s_idx+1], bbb_y_pos[b_idx][s_idx], bbb_bond[b_idx][s_idx+1], dim = dim, compact = compact, highlight = bbb_bond_label[b_idx][s_idx+1]) for b_idx in range(len(bbb_sugar)) for s_idx in range(len(bbb_sugar[b_idx])-1) if len(bbb_sugar[b_idx]) > 1]
  # Bond branch_branch to branch
  [add_bond(branch_branch_x_pos[k][0], branch_x_pos[branch_branch_connection[k][0]][branch_branch_connection[k][1]], branch_branch_y_pos[k][0], branch_y_pos[branch_branch_connection[k][0]][branch_branch_connection[k][1]], branch_branch_bond[k][0], dim = dim, compact = compact, highlight = branch_branch_bond_label[k][0]) for k in range(len(branch_branch_sugar))]
  # Bond branch_branch_branch to branch_branch
  [add_bond(bbb_x_pos[k][0], branch_branch_x_pos[bbb_connection[k][0]][bbb_connection[k][1]], bbb_y_pos[k][0], branch_branch_y_pos[bbb_connection[k][0]][bbb_connection[k][1]], bbb_bond[k][0], dim = dim, compact = compact, highlight = bbb_bond_label[k][0]) for k in range(len(bbb_sugar))]
  
  # Sugar main chain
  [add_sugar(main_sugar[k], main_sugar_x_pos[k], main_sugar_y_pos[k], modification = main_sugar_modification[k], conf = main_conf[k], compact = compact, dim = dim, deg = main_deg[k], highlight = main_sugar_label[k]) for k in range(len(main_sugar))]
  # Sugar branch
  [add_sugar(branch_sugar[b_idx][s_idx], branch_x_pos[b_idx][s_idx], branch_y_pos[b_idx][s_idx], modification = branch_sugar_modification[b_idx][s_idx], conf = b_conf[b_idx][s_idx], compact = compact, dim = dim, deg = branch_deg[b_idx][s_idx], highlight = branch_sugar_label[b_idx][s_idx]) for b_idx in range(len(branch_sugar)) for s_idx in range(len(branch_sugar[b_idx]))]
  # Sugar branch_branch
  [add_sugar(branch_branch_sugar[b_idx][s_idx], branch_branch_x_pos[b_idx][s_idx], branch_branch_y_pos[b_idx][s_idx], modification = branch_branch_sugar_modification[b_idx][s_idx], conf = bb_conf[b_idx][s_idx], compact = compact, dim = dim, deg = branch_branch_deg[b_idx][s_idx], highlight = branch_branch_sugar_label[b_idx][s_idx]) for b_idx in range(len(branch_branch_sugar)) for s_idx in range(len(branch_branch_sugar[b_idx]))]
  # Sugar branch branch branch
  [add_sugar(bbb_sugar[b_idx][s_idx], bbb_x_pos[b_idx][s_idx], bbb_y_pos[b_idx][s_idx], modification = bbb_sugar_modification[b_idx][s_idx], conf = bbb_conf[b_idx][s_idx], compact = compact, dim = dim, highlight = bbb_sugar_label[b_idx][s_idx]) for b_idx in range(len(bbb_sugar)) for s_idx in range(len(bbb_sugar[b_idx]))]

  if highlight_motif == None:
    highlight = 'show'
  else:
    highlight = 'hide'
  
  if floaty_bits != []:
    fb_count = {i: floaty_bits.count(i) for i in floaty_bits}
    floaty_bits = list(set(floaty_bits))
    floaty_data = []
    for k in range(len(floaty_bits)):
      try:
        floaty_data.append(get_coordinates_and_labels(floaty_bits[k], show_linkage = show_linkage, highlight_motif = None))
      except:
        floaty_data.append(get_coordinates_and_labels('blank(-)blank', show_linkage = show_linkage, highlight_motif = None))
    y_span = max_y-min_y
    n_floats = len(floaty_bits)
    floaty_span = n_floats * 2 - 2
    y_diff = (floaty_span/2) - (y_span/2)

    for j in range(len(floaty_data)):
      floaty_sugar, floaty_sugar_x_pos, floaty_sugar_y_pos, floaty_sugar_modification, floaty_bond, floaty_conf, _, _ = floaty_data[j][0]
      floaty_sugar_label = ['show' if highlight_motif == None else 'hide' for k in floaty_sugar]
      floaty_bond_label = ['show' if highlight_motif == None else 'hide' for k in floaty_bond]
      floaty_sugar_x_pos = [floaty_sugar_x_pos[k] + max_x + 1 for k in floaty_sugar_x_pos]
      floaty_sugar_y_pos = [floaty_sugar_y_pos[k] + 2 * j - y_diff for k in floaty_sugar_y_pos]
      # Fix for drawsvg 2.0
      floaty_sugar_y_pos = [(k*-1) for k in floaty_sugar_y_pos]
      if floaty_sugar != ['blank', 'blank']:
        [add_bond(floaty_sugar_x_pos[k+1], floaty_sugar_x_pos[k], floaty_sugar_y_pos[k+1], floaty_sugar_y_pos[k], floaty_bond[k], dim = dim, compact = compact, highlight = floaty_bond_label[k]) for k in range(len(floaty_sugar)-1)]
        [add_sugar(floaty_sugar[k], floaty_sugar_x_pos[k], floaty_sugar_y_pos[k], modification = floaty_sugar_modification[k], conf = floaty_conf, compact = compact, dim = dim, highlight = floaty_sugar_label[k]) for k in range(len(floaty_sugar))]
      else:
        add_sugar('text', min(floaty_sugar_x_pos), floaty_sugar_y_pos[-1], modification = floaty_bits[j].replace('blank', ''), compact = compact, dim = dim, text_anchor = 'end', highlight = highlight)

      if fb_count[floaty_bits[j]] > 1:
        if not compact:
          add_sugar('blank', max(floaty_sugar_x_pos)+0.5, floaty_sugar_y_pos[-1]+0.75, modification = str(fb_count[floaty_bits[j]]) + 'x', compact = compact, dim = dim, highlight = highlight)
        else:
          add_sugar('blank', max(floaty_sugar_x_pos)+0.75, floaty_sugar_y_pos[-1]+1.15, modification = str(fb_count[floaty_bits[j]]) + 'x', compact = compact, dim = dim, highlight = highlight)

    bracket_x = max_x * (2 if not compact else 1.2) + 1
    bracket_y = (min_y, max_y) if not compact else ((min_y * 0.5) * 1.2, (max_y * 0.5) * 1.2)
    draw_bracket(bracket_x, bracket_y, direction = 'right', dim = dim, highlight = highlight)

  # add brackets around repeating unit
  if repeat:

    # process annotation
    repeat_annot = 'n'
    if isinstance(repeat, (str, int)):
      if repeat!= True:
        repeat_annot += ' = ' + str(repeat)
    
    # repeat range code block
    if repeat_range:
      bracket_open = (main_sugar_x_pos[repeat_range[1]]*2)+1 if not compact else (main_sugar_x_pos[repeat_range[1]]*1.2)+0.6
      bracket_close = (main_sugar_x_pos[repeat_range[0]]*2)-1 if not compact else (main_sugar_x_pos[repeat_range[0]]*1.2)-0.6
      bracket_y_open =  (main_sugar_y_pos[repeat_range[1]], main_sugar_y_pos[repeat_range[1]]) if not compact else (((np.mean(main_sugar_y_pos[repeat_range[1]]) * 0.5) * 1.2)+0.0, ((np.mean(main_sugar_y_pos[repeat_range[1]]) * 0.5) * 1.2)-0.0)
      bracket_y_close = (main_sugar_y_pos[repeat_range[0]], main_sugar_y_pos[repeat_range[0]]) if not compact else (((np.mean(main_sugar_y_pos[repeat_range[0]]) * 0.5) * 1.2)+0.0, ((np.mean(main_sugar_y_pos[repeat_range[0]]) * 0.5) * 1.2)-0.0)
      text_x = main_sugar_x_pos[repeat_range[0]]-0.5 
      text_y = main_sugar_y_pos[0]+1.05 if not compact else (main_sugar_y_pos[0]+1.03)/0.6
      draw_bracket(bracket_close, bracket_y_close, direction = 'left', dim = dim, highlight = highlight, deg = 0)
      draw_bracket(bracket_open, bracket_y_open, direction = 'right', dim = dim, highlight = highlight, deg = 0)
      add_sugar('text', text_x, text_y, modification = repeat_annot, compact = compact, dim = dim, text_anchor = 'start', highlight = highlight)

    # repeat unit code block
    else:
      open_deg = calculate_degree(main_sugar_y_pos[-1], main_sugar_y_pos[-2], main_sugar_x_pos[-1], main_sugar_x_pos[-2])
      if open_deg == 0:
        bracket_open = np.mean([k*2 for k in main_sugar_x_pos][-2:])+0.2 if not compact else np.mean([k*1.2 for k in main_sugar_x_pos][-2:])+0.15
        bracket_y_open = (np.mean(main_sugar_y_pos[-2:]), np.mean(main_sugar_y_pos[-2:])) if not compact else (((np.mean(main_sugar_y_pos[-2:]) * 0.5) * 1.2)+0.0, ((np.mean(main_sugar_y_pos[-2:]) * 0.5) * 1.2)-0.0)
        bracket_y_close = (main_sugar_y_pos[0], main_sugar_y_pos[0]) if not compact else (((np.mean(main_sugar_y_pos[0]) * 0.5) * 1.2)+0.0, ((np.mean(main_sugar_y_pos[0]) * 0.5) * 1.2)-0.0)
      else:
        bracket_open = np.mean([k*2 for k in main_sugar_x_pos][-2:])+0.0 if not compact else np.mean([k*1.2 for k in main_sugar_x_pos][-2:])+0
        bracket_y_open = (np.mean(main_sugar_y_pos[-2:]), np.mean(main_sugar_y_pos[-2:])) if not compact else (((np.mean(main_sugar_y_pos[-2:]) * 0.5) * 1.2)+0.3, ((np.mean(main_sugar_y_pos[-2:]) * 0.5) * 1.2)-0.3)
        bracket_y_close = (main_sugar_y_pos[0], main_sugar_y_pos[0]) if not compact else (((np.mean(main_sugar_y_pos[0]) * 0.5) * 1.2)+0.3, ((np.mean(main_sugar_y_pos[0]) * 0.5) * 1.2)-0.3)
      bracket_close = np.mean([k*2 for k in main_sugar_x_pos][:2])-0.2 if not compact else np.mean([k*1.2 for k in main_sugar_x_pos][:2])-0.15
      text_x = bracket_close-(0.42) if not compact else bracket_close-(0.13)
      text_y = main_sugar_y_pos[0]+1.05 if not compact else (main_sugar_y_pos[0]+1.03)/0.6
      draw_bracket(bracket_open, bracket_y_open, direction = 'right', dim = dim, highlight = highlight, deg = open_deg)
      draw_bracket(bracket_close, bracket_y_close, direction = 'left', dim = dim, highlight = highlight, deg = 0)
      add_sugar('text', text_x, text_y, modification = repeat_annot, compact = compact, dim = dim, text_anchor = 'start', highlight = highlight)

  d2.append(d)

  if filepath:
      filepath = filepath.replace('?', '_')
      data = d2.as_svg()
      data = re.sub(r'<text font-size="17.5" ', r'<text font-size="17.5" font-family="century gothic" font-weight="bold" ', data)
      data = re.sub(r'<text font-size="20.0" ', r'<text font-size="20" font-family="century gothic" ', data)
      data = re.sub(r'<text font-size="15.0" ', r'<text font-size="17.5" font-family="century gothic" font-style="italic" ', data)
      if 'svg' in filepath:
        with open(filepath, 'w') as f:
          f.write(data)
      elif 'pdf' in filepath:
        try:
          from cairosvg import svg2pdf
          svg2pdf(bytestring = data, write_to = filepath)
        except:
          raise ImportError("You're missing some draw dependencies. Either use .svg or head to https://bojarlab.github.io/glycowork/examples.html#glycodraw-code-snippets to learn more.")
  return d2


def scale_in_range(listy, a, b):
  """Normalize list of numbers in range a to b\n
  | Arguments:
  | :-
  | listy (list): list of numbers as integers/floats
  | a (integer/float): min value in normalized range
  | b (integer/float): max value in normalized range\n
  | Returns:
  | :-
  | Normalized list of numbers
  """
  min_val = min(listy)
  max_val = max(listy)
  range_val = max_val - min_val
  return [(b - a) * ((x - min_val) / range_val) + a for x in listy]


def annotate_figure(svg_input, scale_range = (25, 80), compact = False, glycan_size = 'medium', filepath = '',
                    scale_by_DE_res = None, x_thresh = 1, y_thresh = 0.05, x_metric = 'Log2FC'):
  """Modify matplotlib svg figure to replace text labels with glycan figures\n
  | Arguments:
  | :-
  | svg_input (string): absolute path including full filename for input svg figure
  | scale_range (tuple): tuple of two integers defining min/max glycan dim; default:(25,80)
  | compact (bool): if True, draw compact glycan figures; default:False
  | glycan_size (string): modify glycan size; default:'medium'; options are 'small', 'medium', 'large'
  | filepath (string): absolute path including full filename allows for saving the plot
  | scale_by_DE_res (df): result table from motif_analysis.get_differential_expression. Include to scale glycan figure size by -10logp
  | x_thresh (float): absolute x metric threshold for datapoints included for scaling, set to match get_differential_expression; default:1.0
  | y_thresh (float): corr p threshhold for datapoints included for scaling, set to match get_differential_expression; default:0.05
  | x_metric (string): x-axis metric; default:'Log2FC'; options are 'Log2FC', 'Effect size'\n
  | Returns:
  | :-
  | Modified figure svg code
  """
  glycan_size_dict = {
      'small': 'scale(0.1 0.1)  translate(0, -74)',
      'medium': 'scale(0.2 0.2)  translate(0, -55)',
      'large': 'scale(0.3 0.3)  translate(0, -49)'
      }
  glycan_scale = ''

  if scale_by_DE_res is not None:
    res_df = scale_by_DE_res.loc[(abs(scale_by_DE_res[x_metric]) > x_thresh) & (scale_by_DE_res['corr p-val'] < y_thresh)]
    y = -np.log10(res_df['corr p-val'].values.tolist())
    labels = res_df['Glycan'].values.tolist()
    glycan_scale = [y, labels]

  # Get svg code
  svg_tmp = open(svg_input, "r").read()
  # Get all text labels
  label_pattern = re.compile(r'<!--\s*(.*?)\s*-->')
  transform_pattern = re.compile(r'<g transform\s*(.*?)\s*">')
  matches = re.findall(r"<!--.*-->[\s\S]*?<\/g>", svg_tmp)
  # Prepare for appending
  svg_tmp = svg_tmp.replace('</svg>', '')
  element_id = 0
  edit_svg = False

  for match in matches:
    # Keep track of current label and position in figure
    current_label = label_pattern.findall(match)[0]
    # Check if label is glycan
    try:
      glycan_to_nxGraph(current_label)
      edit_svg = True
    except:
      pass
    try:
      glycan_to_nxGraph(motif_list.loc[motif_list.motif_name == current_label].motif.values.tolist()[0])
      edit_svg = True
    except:
      pass
    # Delete text label, append glycan figure
    if edit_svg:
      current_pos = '<g transform' + transform_pattern.findall(match)[0] + '">'
      current_pos = current_pos.replace('scale(0.1 -0.1)', glycan_size_dict[glycan_size])
      svg_tmp = svg_tmp.replace(match, '')
      if glycan_scale == '':
        d = GlycoDraw(current_label, compact = compact)
      else:
        d = GlycoDraw(current_label, compact = compact, dim = scale_in_range(glycan_scale[0], scale_range[0], scale_range[1])[glycan_scale[1].index(current_label)])
      data = d.as_svg().replace('<?xml version="1.0" encoding="UTF-8"?>\n', '')
      id_matches = re.findall(r'd\d+', data)
      # Reassign element ids to avoid duplicates
      for idx in id_matches:
        data = data.replace(idx, 'd' + str(element_id))
        element_id += 1
      svg_tmp += '\n' + current_pos + '\n' + data + '\n</g>'
      edit_svg = False
  svg_tmp += '</svg>'

  if filepath:
    try:
      from cairosvg import svg2pdf, svg2svg, svg2png
      if filepath.split('.')[-1] == 'pdf':
        svg2pdf(bytestring = svg_tmp, write_to = filepath, dpi = 300)
      elif filepath.split('.')[-1] == 'svg':
        svg2svg(bytestring = svg_tmp, write_to = filepath, dpi = 300)
      elif filepath.split('.')[-1] == 'png':
        svg2png(bytestring = svg_tmp, write_to = filepath, dpi = 300)
    except:
      raise ImportError("You're missing some draw dependencies. Either don't use filepath or head to https://bojarlab.github.io/glycowork/examples.html#glycodraw-code-snippets to learn more.")
  else:
    return svg_tmp


def plot_glycans_excel(df, folder_filepath, glycan_col_num = 0,
                       scaling_factor = 0.2):
  """plots SNFG images of glycans into new column in df and saves df as Excel file\n
  | Arguments:
  | :-
  | df (dataframe): dataframe containing glycan sequences [alternative: filepath to .csv]
  | folder_filepath (string): full filepath to the folder you want to save the output to
  | glycan_col_num (int): index of the column containing glycan sequences; default:0 (first column)
  | scaling_factor (float): how large the glycans should be; default:0.2\n
  | Returns:
  | :-
  | Saves the dataframe with glycan images as output.xlsx into folder_filepath
  """
  try:
    from cairosvg import svg2png
  except:
    raise ImportError("You're missing some draw dependencies. If you want to use this function, head to https://bojarlab.github.io/glycowork/examples.html#glycodraw-code-snippets to learn more.")
  if isinstance(df, str):
    df = pd.read_csv(df)
  if not folder_filepath.endswith('/'):
    folder_filepath += '/'
  image_column_number = df.shape[1] + 2
  df["SNFG"] = [np.nan for k in range(len(df))]
  # Convert df_out to Excel
  writer = pd.ExcelWriter(folder_filepath + "output.xlsx", engine = "openpyxl")
  df.to_excel(writer)
  # Load the workbook and get the active sheet
  workbook = writer.book
  sheet = writer.sheets["Sheet1"]
  for i, glycan_structure in enumerate(df.iloc[:, glycan_col_num]):
    if glycan_structure and glycan_structure[0]:
      if not isinstance(glycan_structure[0], str):
        glycan_structure = glycan_structure[0][0]
      # Generate glycan image using GlycoDraw
      svg_data = GlycoDraw(glycan_structure).as_svg()
      # Convert SVG data to image
      png_data = svg2png(bytestring = svg_data)
      img_stream = BytesIO(png_data)
      img = Image.open(img_stream)
      # Set the size of the image
      img_width, img_height = img.size
      img = img.resize((int(img_width * scaling_factor), int(img_height * scaling_factor)), Image.LANCZOS)  
      # Save the image to a BytesIO object
      img_stream = BytesIO()
      img.save(img_stream, format = 'PNG')
      img_stream.seek(0)
      # Create an image
      img_for_excel = OpenpyxlImage(img_stream)
      img_for_excel.width, img_for_excel.height = img.width, img.height  # Set width and height
      # Find the cell to insert the image
      cell = sheet.cell(row = i + 2, column = image_column_number)  # +2 because Excel is 1-indexed and there's a header row
      # Insert the image into the cell
      sheet.add_image(img_for_excel, cell.coordinate)
      # Resize the cell to fit the image
      column_letter = get_column_letter(image_column_number)
      sheet.column_dimensions[column_letter].width = img.width*0.75*0.25
      sheet.row_dimensions[cell.row].height = img.height*0.75
  # Save the workbook
  workbook.save(filename = folder_filepath + "output.xlsx")
